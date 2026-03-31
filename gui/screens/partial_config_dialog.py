#!/usr/bin/env python3
"""
Partial LTA Configuration Dialog
Allows users to configure partial LTA processing
"""

import tkinter as tk
from tkinter import ttk, messagebox
import os
import glob
import logging
import re
from openpyxl import load_workbook
from gui.utils.file_utils import get_lta_partial_info, save_lta_partial_config
from gui.utils.validators import validate_ds_series, normalize_ds_series, validate_location

logger = logging.getLogger(__name__)

# Same predefined locations as preparation.py
def _looks_like_weight_ds_fields_swapped(weight_str, ds_str):
    """Detect common mistake: DS série (e.g. '3129 X') in Poids and weight in DS Série."""
    w = (weight_str or "").strip()
    d = (ds_str or "").strip()
    if not w or not d:
        return False
    return bool(re.match(r"^\d+\s+[A-Za-z]$", w) and re.match(r"^\d+(\.\d+)?$", d))


PARTIAL_LOCATIONS = [
    "RYAD K.KHALED",
    "ISTAMBOUL ATATUR",
    "JEDDAH K/ABDUL A",
    "BAHREIN MOHARRAQ",
    "DOHA INT",
    "ABOU DHABI INT",
    "SHANGHAI PU DONG"
]


class PartialConfigDialog:
    """Dialog for configuring partial LTA processing"""
    
    def __init__(self, parent, lta_folder_path, folder_name):
        self.parent = parent
        self.lta_folder_path = lta_folder_path
        self.folder_name = folder_name
        self.config_saved = False
        
        # Create dialog FIRST (before any operations that might fail)
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(f"Configuration Partielle - {folder_name}")
        self.dialog.geometry("800x600")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Load existing config if available
        self.existing_config = get_lta_partial_info(lta_folder_path, folder_name)
        
        # Load LTA data from generated_excel
        self.lta_data = self._load_lta_data()
        
        if not self.lta_data:
            # Show error but keep dialog open so user can see the error
            messagebox.showerror("Erreur", "Impossible de charger les données LTA.\nVeuillez exécuter le script de préparation d'abord.")
            # Don't return - let the dialog stay open so user can see what went wrong
            # But don't call _setup_ui() if data is invalid
            return
        
        self._setup_ui()
    
    def _load_lta_data(self):
        """Load LTA data from generated_excel file"""
        try:
            lta_subfolder = os.path.join(self.lta_folder_path, self.folder_name)
            excel_files = glob.glob(os.path.join(lta_subfolder, "generated_excel*.xlsx"))
            
            if not excel_files:
                logger.error(f"No generated_excel file found in {lta_subfolder}")
                messagebox.showwarning(
                    "Fichier introuvable",
                    f"Le fichier 'generated_excel' n'a pas été trouvé dans:\n{lta_subfolder}\n\n"
                    "Veuillez exécuter la détection LTA d'abord."
                )
                return None
            
            logger.info(f"Loading LTA data from: {excel_files[0]}")
            wb = load_workbook(excel_files[0], data_only=True)
            
            # Check if Summary sheet exists
            if 'Summary' not in wb.sheetnames:
                logger.error(f"Summary sheet not found. Available sheets: {wb.sheetnames}")
                wb.close()
                messagebox.showerror(
                    "Erreur",
                    f"La feuille 'Summary' n'existe pas dans le fichier Excel.\n\n"
                    f"Feuilles disponibles: {', '.join(wb.sheetnames)}"
                )
                return None
            
            ws = wb['Summary']
            
            # Get total weight and positions from Summary sheet
            # Data is in column A (labels) and column B (values)
            total_weight = None
            total_positions = None
            
            # Search for "P,BRUT" and "P" labels in column A (rows 1-10)
            for row in range(1, 15):
                cell_a = ws[f'A{row}'].value
                if cell_a:
                    cell_a_str = str(cell_a).strip().upper()
                    if 'P,BRUT' in cell_a_str or 'P.BRUT' in cell_a_str:
                        val = ws[f'B{row}'].value
                        if val and isinstance(val, (int, float)):
                            total_weight = val
                            logger.info(f"Found total weight at B{row}: {total_weight}")
                    elif cell_a_str == 'P' and not total_positions:  # P for positions (before P,BRUT in file)
                        val = ws[f'B{row}'].value
                        if val and isinstance(val, (int, float)):
                            total_positions = val
                            logger.info(f"Found total positions at B{row}: {total_positions}")
            
            logger.info(f"Total weight: {total_weight}, Total positions: {total_positions}")
            
            # Count DUMs by checking C11, C18, C25... (DUM labels in column C)
            # DUM data structure: 
            # Row N: "DUM X" in column C
            # Row N+1: P (positions) - label in A, value in B
            # Row N+2: V (value) - label in A, value in B  
            # Row N+3: P,NET - label in A, value in B
            # Row N+4: P,BRUT (weight) - label in A, value in B
            dums = []
            # Use dynamic detection: continue until no more DUMs found (up to 50 DUMs max)
            for dum_idx in range(1, 51):  # Increased from 10 to 50 to support more DUMs
                row_num = 11 + (dum_idx - 1) * 7
                
                # Safety check: don't go beyond reasonable row limit
                if row_num > 500:
                    break
                
                cell_value = ws[f'C{row_num}'].value
                
                if cell_value and 'DUM' in str(cell_value).upper():
                    # Get DUM positions and weight from column A (labels) and B (values)
                    # P is at row_num + 1, P,BRUT is at row_num + 4
                    dum_positions_row = row_num + 1  # P is 1 row below DUM label
                    dum_weight_row = row_num + 4     # P,BRUT is 4 rows below DUM label
                    
                    dum_positions = ws[f'B{dum_positions_row}'].value or 0
                    dum_weight = ws[f'B{dum_weight_row}'].value or 0
                    
                    logger.info(f"DUM {dum_idx} (row {row_num}): weight={dum_weight}, positions={dum_positions}")
                    
                    dums.append({
                        'number': dum_idx,
                        'weight': float(dum_weight) if dum_weight else 0,
                        'positions': int(dum_positions) if dum_positions else 0
                    })
                else:
                    # No DUM found at this position, stop searching
                    break
            
            wb.close()
            
            logger.info(f"Loaded {len(dums)} DUMs")
            
            return {
                'total_weight': float(total_weight) if total_weight else 0,
                'total_positions': int(total_positions) if total_positions else 0,
                'dums': dums
            }
            
        except Exception as e:
            logger.error(f"Error loading LTA data: {e}", exc_info=True)
            messagebox.showerror(
                "Erreur",
                f"Erreur lors du chargement des données LTA:\n{str(e)}\n\n"
                f"Vérifiez le fichier Excel 'generated_excel' dans:\n{lta_subfolder}"
            )
            return None
    
    def _setup_ui(self):
        """Setup the dialog UI"""
        # Main container with scrollbar
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title = ttk.Label(
            main_frame,
            text=f"📦 Configuration LTA Partiel: {self.folder_name}",
            font=('Arial', 12, 'bold')
        )
        title.pack(pady=(0, 10))
        
        # LTA Totals
        totals_frame = ttk.LabelFrame(main_frame, text="Totaux LTA", padding="10")
        totals_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(totals_frame, text=f"Poids Total: {self.lta_data['total_weight']} kg").pack(anchor=tk.W)
        ttk.Label(totals_frame, text=f"Positions Totales: {self.lta_data['total_positions']}").pack(anchor=tk.W)
        ttk.Label(totals_frame, text=f"Nombre de DUMs: {len(self.lta_data['dums'])}").pack(anchor=tk.W)
        
        # Number of partials
        partials_frame = ttk.LabelFrame(main_frame, text="Nombre de Partiels", padding="10")
        partials_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(partials_frame, text="Nombre de vols partiels:").grid(row=0, column=0, padx=5)
        self.num_partials_var = tk.IntVar(value=2)
        num_partials_spinbox = ttk.Spinbox(
            partials_frame,
            from_=2,
            to=5,
            textvariable=self.num_partials_var,
            width=10
        )
        num_partials_spinbox.grid(row=0, column=1, padx=5)
        
        generate_btn = ttk.Button(
            partials_frame,
            text="Générer Formulaire",
            command=self._generate_partial_forms
        )
        generate_btn.grid(row=0, column=2, padx=10)
        
        # Exception case warning frame (initially hidden)
        self.exception_frame = ttk.LabelFrame(main_frame, text="⚠️ CAS D'EXCEPTION DÉTECTÉ", padding="10")
        self.exception_frame.pack(fill=tk.X, pady=5)
        self.exception_frame.pack_forget()  # Hide initially
        
        exception_info = ttk.Label(
            self.exception_frame,
            text="Le plus petit partiel est inférieur au poids du plus grand DUM (cas de complément).\n"
                 "Une seule ED sera créée. Renseignez les infos du partiel exception ci-dessous:",
            foreground="red",
            font=('Arial', 9, 'bold')
        )
        exception_info.grid(row=0, column=0, columnspan=4, sticky=tk.W, pady=(0, 10))
        
        ttk.Label(self.exception_frame, text="Référence créée à l'aéroport:", font=('Arial', 9, 'bold')).grid(
            row=1, column=0, sticky=tk.W, padx=5, pady=2
        )
        self.airport_reference_var = tk.StringVar(value="")
        airport_ref_entry = ttk.Entry(self.exception_frame, textvariable=self.airport_reference_var, width=25)
        airport_ref_entry.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(self.exception_frame, text="(ex: 157-41680645)", font=('Arial', 8, 'italic')).grid(
            row=1, column=2, sticky=tk.W, padx=5, pady=2
        )
        
        ttk.Label(self.exception_frame, text="Positions du plus petit partiel:", font=('Arial', 9, 'bold')).grid(
            row=2, column=0, sticky=tk.W, padx=5, pady=2
        )
        self.smallest_partial_positions_var = tk.StringVar(value="")
        positions_entry = ttk.Entry(self.exception_frame, textvariable=self.smallest_partial_positions_var, width=10)
        positions_entry.grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
        # Trace changes to update distribution preview when positions change
        self.smallest_partial_positions_var.trace('w', lambda *args: self._update_distribution_preview())
        ttk.Label(self.exception_frame, text="(nombre de positions)", font=('Arial', 8, 'italic')).grid(
            row=2, column=2, sticky=tk.W, padx=5, pady=2
        )
        
        # Buttons frame - Pack FIRST (before scrollable content) so it always reserves space at bottom
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(10, 5), padx=10)
        
        ttk.Button(
            buttons_frame,
            text="💾 Sauvegarder",
            command=self._save_config
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            buttons_frame,
            text="❌ Annuler",
            command=self.dialog.destroy
        ).pack(side=tk.LEFT, padx=5)
        
        # Partials container (scrollable) - Pack AFTER buttons so it fills remaining space
        self.partials_container = ttk.Frame(main_frame)
        self.partials_container.pack(fill=tk.BOTH, expand=True, pady=10)
        
        canvas = tk.Canvas(self.partials_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.partials_container, orient=tk.VERTICAL, command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor=tk.NW)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.canvas = canvas
        
        # Load existing config if available
        if self.existing_config:
            self.num_partials_var.set(len(self.existing_config['partials']))
            # Load exception case data if exists
            if self.existing_config.get('partial_type') == 'exception':
                self.airport_reference_var.set(self.existing_config.get('smallest_partial_airport_reference', ''))
                self.smallest_partial_positions_var.set(str(self.existing_config.get('smallest_partial_positions', '')))
            self._generate_partial_forms(load_existing=True)
    
    def _generate_partial_forms(self, load_existing=False):
        """Generate forms for each partial"""
        # Clear existing forms
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        self.partial_forms = []
        num_partials = self.num_partials_var.get()
        
        for i in range(num_partials):
            partial_num = i + 1
            
            # Load existing data if available
            existing_data = None
            if load_existing and self.existing_config:
                for p in self.existing_config['partials']:
                    if p['partial_number'] == partial_num:
                        existing_data = p
                        break
            
            frame = self._create_partial_form(partial_num, existing_data)
            frame.pack(fill=tk.X, pady=5, padx=10)
            
        # Bind mousewheel
        def on_mousewheel(event):
            if event.num == 5 or event.delta < 0:
                self.canvas.yview_scroll(1, "units")
            elif event.num == 4 or event.delta > 0:
                self.canvas.yview_scroll(-1, "units")
        
        self.canvas.bind("<MouseWheel>", on_mousewheel)
        self.canvas.bind("<Button-4>", on_mousewheel)
        self.canvas.bind("<Button-5>", on_mousewheel)
    
    def _create_partial_form(self, partial_num, existing_data=None):
        """Create form for a single partial"""
        frame = ttk.LabelFrame(
            self.scrollable_frame,
            text=f"Partiel {partial_num}",
            padding="10"
        )
        
        # Weight
        ttk.Label(frame, text="Poids (kg):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        weight_var = tk.StringVar(value=existing_data['weight'] if existing_data else "")
        weight_entry = ttk.Entry(frame, textvariable=weight_var, width=15)
        weight_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)
        
        # Calculated positions (read-only, will be auto-calculated)
        ttk.Label(frame, text="Positions (auto):").grid(row=0, column=2, sticky=tk.W, padx=5, pady=2)
        positions_var = tk.StringVar(value="")
        positions_label = ttk.Label(frame, textvariable=positions_var, foreground="blue")
        positions_label.grid(row=0, column=3, sticky=tk.W, padx=5, pady=2)
        
        # DS Série (one field, format "XXXX Y" like preparation - validated on save)
        _ser = (existing_data.get('ds_serie') or '').strip() if existing_data else ''
        _cle = (existing_data.get('ds_cle') or '').strip() if existing_data else ''
        _ds_initial = f"{_ser} {_cle}".strip() if (_ser or _cle) else ""
        ttk.Label(frame, text="DS Série:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        ds_serie_var = tk.StringVar(value=_ds_initial)
        ds_serie_entry = ttk.Entry(frame, textvariable=ds_serie_var, width=20, font=('Arial', 9))
        ds_serie_entry.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(frame, text="(ex: 9913 G)", font=('Arial', 8, 'italic')).grid(row=1, column=2, sticky=tk.W, padx=5, pady=2)
        
        # Loading Location: list select + "Autre" (same as preparation)
        ttk.Label(frame, text="Lieu de Chargement:", font=('Arial', 9, 'bold')).grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        loc_frame = ttk.Frame(frame)
        loc_frame.grid(row=2, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=2, padx=5)
        
        location_var = tk.StringVar(value=(existing_data.get('loading_location') or '') if existing_data else '')
        loc_mode_var = tk.StringVar(value="select")
        
        loc_combo = ttk.Combobox(
            loc_frame,
            textvariable=location_var,
            values=PARTIAL_LOCATIONS,
            width=25,
            font=('Arial', 9),
            state="readonly"
        )
        loc_combo.grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        if location_var.get() and location_var.get() in PARTIAL_LOCATIONS:
            loc_combo.set(location_var.get())
        elif location_var.get():
            loc_mode_var.set("custom")
        
        loc_custom_entry = ttk.Entry(loc_frame, textvariable=location_var, width=25, font=('Arial', 9))
        
        def _make_toggle_loc(mv, combo, custom):
            def _toggle():
                if mv.get() == "select":
                    combo.grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
                    custom.grid_remove()
                else:
                    combo.grid_remove()
                    custom.grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
            return _toggle
        
        _toggle_loc = _make_toggle_loc(loc_mode_var, loc_combo, loc_custom_entry)
        
        ttk.Radiobutton(loc_frame, text="Liste", variable=loc_mode_var, value="select", command=_toggle_loc).grid(row=0, column=1, padx=5)
        ttk.Radiobutton(loc_frame, text="Autre", variable=loc_mode_var, value="custom", command=_toggle_loc).grid(row=0, column=2, padx=5)
        _toggle_loc()
        
        # DUM Distribution Preview (read-only text widget)
        ttk.Label(frame, text="Distribution DUMs (auto):").grid(row=3, column=0, sticky=tk.NW, padx=5, pady=2)
        
        dums_text = tk.Text(frame, height=6, width=50, state='disabled', wrap=tk.WORD)
        dums_text.grid(row=3, column=1, columnspan=3, sticky=(tk.W, tk.E), padx=5, pady=2)
        
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=dums_text.yview)
        scrollbar.grid(row=3, column=4, sticky=(tk.N, tk.S))
        dums_text.configure(yscrollcommand=scrollbar.set)
        
        self.partial_forms.append({
            'partial_number': partial_num,
            'weight_var': weight_var,
            'positions_var': positions_var,
            'ds_serie_var': ds_serie_var,
            'location_var': location_var,
            'dums_text': dums_text
        })
        
        # Trace weight changes to auto-calculate and update display
        weight_var.trace('w', lambda *args: self._update_distribution_preview())
        
        return frame
    
    def _update_distribution_preview(self):
        """Update the DUM distribution preview for all partials"""
        try:
            # Check if partial_forms exists (may not be initialized yet)
            if not hasattr(self, 'partial_forms') or not self.partial_forms:
                return  # Forms not generated yet, skip update
            
            # Validate LTA data
            if not self.lta_data.get('dums') or self.lta_data.get('total_weight', 0) <= 0:
                # Show error message in preview
                for form_data in self.partial_forms:
                    form_data['positions_var'].set("0")
                    dums_text = form_data['dums_text']
                    dums_text.configure(state='normal')
                    dums_text.delete('1.0', tk.END)
                    dums_text.insert(tk.END, "⚠️ Données LTA invalides\n(Poids = 0 ou aucun DUM)")
                    dums_text.configure(state='disabled')
                return
            
            # Collect partial weights (never silently treat invalid input as 0 — that orphans DUMs)
            partial_weights = []
            parse_ok = []
            for form_data in self.partial_forms:
                raw_w = form_data['weight_var'].get().strip()
                try:
                    partial_weights.append(float(raw_w))
                    parse_ok.append(True)
                except ValueError:
                    partial_weights.append(0.0)
                    parse_ok.append(False)

            if not all(parse_ok):
                for idx, form_data in enumerate(self.partial_forms):
                    form_data['positions_var'].set("—")
                    dums_text = form_data['dums_text']
                    dums_text.configure(state='normal')
                    dums_text.delete('1.0', tk.END)
                    if not parse_ok[idx]:
                        msg = (
                            "⚠️ Poids invalide — entrez uniquement le poids en kg (nombre).\n"
                            "Ex: 304.9\n\n"
                        )
                        if _looks_like_weight_ds_fields_swapped(
                            form_data['weight_var'].get(),
                            form_data['ds_serie_var'].get(),
                        ):
                            msg += (
                                "Les champs « Poids (kg) » et « DS Série » semblent inversés.\n"
                                "Mettez le nombre en kg dans Poids et la série (ex: 3129 X) dans DS Série."
                            )
                        else:
                            msg += "Vérifiez qu’aucun texte (série DS, lettres) n’est dans le champ Poids."
                        dums_text.insert(tk.END, msg)
                    else:
                        dums_text.insert(
                            tk.END,
                            "Saisissez un poids valide pour chaque partiel pour calculer la distribution.",
                        )
                    dums_text.configure(state='disabled')
                return

            # Detect exception case: smallest partial < max DUM weight
            # (means partial is a leftover piece of a single DUM, not made of whole DUMs)
            max_dum_weight = max(dum['weight'] for dum in self.lta_data['dums'])
            num_partials = len(partial_weights)
            smallest_partial_weight = min(w for w in partial_weights if w > 0) if any(w > 0 for w in partial_weights) else 0
            # Business rule: exception case is supported for 2- or 3-partial workflows.
            is_exception_case = (
                num_partials in (2, 3)
                and smallest_partial_weight > 0
                and smallest_partial_weight < max_dum_weight
            )

            if is_exception_case:
                # Show exception frame if hidden
                if not self.exception_frame.winfo_manager():
                    self.exception_frame.pack(fill=tk.X, pady=5, before=self.partials_container)
            else:
                # Hide exception frame
                if self.exception_frame.winfo_manager():
                    self.exception_frame.pack_forget()
            
            # Get exception case positions if provided
            smallest_partial_positions = None
            smallest_partial_idx = None
            if is_exception_case:
                try:
                    smallest_partial_positions_str = self.smallest_partial_positions_var.get().strip()
                    if smallest_partial_positions_str:
                        smallest_partial_positions = int(smallest_partial_positions_str)
                        # Find which partial is the smallest
                        smallest_partial_weight = min(w for w in partial_weights if w > 0)
                        for idx, weight in enumerate(partial_weights):
                            if weight == smallest_partial_weight:
                                smallest_partial_idx = idx
                                break
                except (ValueError, AttributeError):
                    pass  # If not valid, ignore and use normal calculation
            
            # Calculate distribution
            distribution = self._calculate_dum_distribution(partial_weights, smallest_partial_idx, smallest_partial_positions)
            
            sum_partial_weights = sum(partial_weights)
            tw_lta = self.lta_data['total_weight']
            
            # Update each partial's display
            for idx, form_data in enumerate(self.partial_forms):
                if idx < len(distribution):
                    partial_dist = distribution[idx]
                    
                    # Update positions
                    form_data['positions_var'].set(str(partial_dist['positions']))
                    
                    # Update DUM list
                    dums_text = form_data['dums_text']
                    dums_text.configure(state='normal')
                    dums_text.delete('1.0', tk.END)
                    
                    if not partial_dist['dums']:
                        if partial_weights[idx] <= 0:
                            dums_text.insert(tk.END, "Aucun DUM assigné (poids partiel = 0)")
                        else:
                            dums_text.insert(tk.END, "Aucun DUM assigné")
                            if abs(sum_partial_weights - tw_lta) > max(0.5, tw_lta * 0.01):
                                dums_text.insert(
                                    tk.END,
                                    f"\n\n⚠️ Somme des partiels ({sum_partial_weights:.1f} kg) ≠ total LTA ({tw_lta:.1f} kg). "
                                    "Les poids des N partiels doivent totaliser le poids LTA pour répartir tous les DUMs.",
                                )
                            else:
                                dums_text.insert(
                                    tk.END,
                                    "\n\n⚠️ Plus de masse DUM disponible pour ce partiel — vérifiez les partiels précédents ou la cohérence des données.",
                                )
                    else:
                        for dum_info in partial_dist['dums']:
                            dum_num = dum_info['dum_number']
                            dum_weight = dum_info['weight']
                            dum_positions = dum_info['positions']
                            is_split = dum_info['is_split']
                            split_id = dum_info.get('split_id', '')
                            
                            if is_split:
                                dums_text.insert(tk.END, f"DUM {dum_num} {split_id}: {dum_weight:.1f}kg, {dum_positions}p ⚠️ PARTIEL\n")
                            else:
                                dums_text.insert(tk.END, f"DUM {dum_num}: {dum_weight:.1f}kg, {dum_positions}p\n")
                    
                    dums_text.configure(state='disabled')
        except Exception as e:
            # Silently handle preview errors to avoid disrupting user input
            logger.error(f"Error updating distribution preview: {e}", exc_info=True)
    
    def _calculate_dum_distribution(self, partial_weights, smallest_partial_idx=None, smallest_partial_positions=None):
        """
        Automatically distribute DUMs across partials based on weights.
        Sequential distribution: Fill partials in order until weight is reached.
        Last DUM may be split if needed.
        Exception case: Sequentially fills the largest partial and splits whichever
        DUM naturally falls at the weight boundary.
        
        Args:
            partial_weights: List of weights for each partial
            smallest_partial_idx: Index of smallest partial (for exception case)
            smallest_partial_positions: Manual positions for smallest partial (for exception case)
        """
        distribution = []
        
        total_lta_weight = self.lta_data['total_weight']
        total_lta_positions = self.lta_data['total_positions']
        dums = self.lta_data['dums']
        
        # Validate LTA data
        if not dums or total_lta_weight <= 0 or total_lta_positions <= 0:
            # Return empty distribution if LTA data is invalid
            for _ in partial_weights:
                distribution.append({'weight': 0, 'positions': 0, 'dums': []})
            return distribution
        
        # Check if exception case
        is_exception_case = (smallest_partial_idx is not None and smallest_partial_positions is not None)
        
        # For exception case: smallest partial takes a split from DUM 1 (airport DS MEAD),
        # then remaining DUM flow is distributed sequentially across other partials.
        if is_exception_case:
            non_smallest_indices = [i for i in range(len(partial_weights)) if i != smallest_partial_idx]
            per_partial_dums = {i: [] for i in non_smallest_indices}
            per_partial_weight = {i: 0.0 for i in non_smallest_indices}
            per_partial_positions = {i: 0 for i in non_smallest_indices}
            split_counters = {}

            def next_split_id(dum_number):
                count = split_counters.get(dum_number, 0) + 1
                split_counters[dum_number] = count
                return f"{dum_number}/{count}"

            # Smallest partial consumes DUM 1 first (exception business rule).
            rounded_smallest_weight = round(partial_weights[smallest_partial_idx], 1)
            rounded_smallest_positions = round(smallest_partial_positions)
            smallest_partial_dums = []

            current_dum_idx = 0
            remaining_dum_weight = round(dums[0]['weight'], 1) if dums else 0
            remaining_dum_positions = round(dums[0]['positions']) if dums else 0
            is_continuing_split = False
            original_dum_weight_at_start = remaining_dum_weight

            if dums and rounded_smallest_weight > 0:
                first_dum_num = dums[0]['number']
                first_dum_weight = round(dums[0]['weight'], 1)
                taken_weight = min(rounded_smallest_weight, first_dum_weight)
                smallest_partial_dums.append({
                    'dum_number': first_dum_num,
                    'weight': taken_weight,
                    'positions': rounded_smallest_positions,
                    'is_split': True,
                    'split_id': ''  # airport lot keeps base reference
                })

                remaining_after_smallest = round(first_dum_weight - taken_weight, 1)
                if remaining_after_smallest > 0:
                    current_dum_idx = 0
                    remaining_dum_weight = remaining_after_smallest
                    remaining_dum_positions = max(0, round(dums[0]['positions']) - rounded_smallest_positions)
                    is_continuing_split = True
                    original_dum_weight_at_start = first_dum_weight
                else:
                    current_dum_idx = 1
                    if current_dum_idx < len(dums):
                        remaining_dum_weight = round(dums[current_dum_idx]['weight'], 1)
                        remaining_dum_positions = round(dums[current_dum_idx]['positions'])
                        original_dum_weight_at_start = remaining_dum_weight
                    else:
                        remaining_dum_weight = 0
                        remaining_dum_positions = 0
                    is_continuing_split = False

            for order_idx, partial_idx in enumerate(non_smallest_indices):
                target_weight = round(partial_weights[partial_idx], 1)
                if target_weight <= 0:
                    continue

                weight_accumulated = 0.0
                positions_accumulated = 0

                while weight_accumulated < target_weight and current_dum_idx < len(dums):
                    weight_needed = round(target_weight - weight_accumulated, 1)
                    dum_number = dums[current_dum_idx]['number']

                    if remaining_dum_weight <= weight_needed:
                        rounded_weight = round(remaining_dum_weight, 1)
                        rounded_positions = round(remaining_dum_positions)
                        actually_split = is_continuing_split and (abs(remaining_dum_weight - original_dum_weight_at_start) > 0.1)

                        per_partial_dums[partial_idx].append({
                            'dum_number': dum_number,
                            'weight': rounded_weight,
                            'positions': rounded_positions,
                            'is_split': actually_split,
                            'split_id': next_split_id(dum_number) if actually_split else ''
                        })
                        weight_accumulated = round(weight_accumulated + rounded_weight, 1)
                        positions_accumulated += rounded_positions

                        current_dum_idx += 1
                        is_continuing_split = False
                        if current_dum_idx < len(dums):
                            remaining_dum_weight = round(dums[current_dum_idx]['weight'], 1)
                            remaining_dum_positions = round(dums[current_dum_idx]['positions'])
                            original_dum_weight_at_start = remaining_dum_weight
                    else:
                        split_weight = round(weight_needed, 1)
                        if remaining_dum_weight > 0:
                            split_positions = round((split_weight / remaining_dum_weight) * remaining_dum_positions)
                        else:
                            split_positions = 0
                        split_positions = max(0, min(split_positions, remaining_dum_positions))

                        per_partial_dums[partial_idx].append({
                            'dum_number': dum_number,
                            'weight': split_weight,
                            'positions': split_positions,
                            'is_split': True,
                            'split_id': next_split_id(dum_number)
                        })
                        weight_accumulated = round(weight_accumulated + split_weight, 1)
                        positions_accumulated += split_positions

                        remaining_dum_weight = round(remaining_dum_weight - split_weight, 1)
                        remaining_dum_positions = round(remaining_dum_positions - split_positions)
                        is_continuing_split = True
                        break

                per_partial_weight[partial_idx] = round(weight_accumulated, 1)
                per_partial_positions[partial_idx] = round(positions_accumulated)

            for partial_idx in range(len(partial_weights)):
                if partial_idx == smallest_partial_idx:
                    distribution.append({
                        'weight': rounded_smallest_weight,
                        'positions': rounded_smallest_positions,
                        'dums': smallest_partial_dums
                    })
                elif partial_idx in per_partial_dums:
                    distribution.append({
                        'weight': per_partial_weight[partial_idx],
                        'positions': per_partial_positions[partial_idx],
                        'dums': per_partial_dums[partial_idx]
                    })
                else:
                    distribution.append({'weight': 0, 'positions': 0, 'dums': []})

            return distribution
        
        # Normal case: sequential distribution (existing logic)
        # Keep split numbering per DUM (e.g. DUM 6 -> 6/1, 6/2),
        # independent of which partial index contains the split.
        split_counters = {}

        def next_split_id(dum_number):
            count = split_counters.get(dum_number, 0) + 1
            split_counters[dum_number] = count
            return f"{dum_number}/{count}"

        current_dum_idx = 0
        remaining_dum_weight = round(dums[0]['weight'], 1) if dums else 0
        remaining_dum_positions = round(dums[0]['positions']) if dums else 0
        is_continuing_split = False  # Track if we're continuing a split DUM
        original_dum_weight_at_start = remaining_dum_weight  # Track original DUM weight at start of partial
        
        for partial_idx, partial_weight in enumerate(partial_weights):
            if partial_weight <= 0:
                distribution.append({'weight': 0, 'positions': 0, 'dums': []})
                continue
            
            # Calculate positions for this partial (safe division)
            if total_lta_weight > 0:
                partial_positions = round((partial_weight * total_lta_positions) / total_lta_weight)
            else:
                partial_positions = 0
            
            partial_dums = []
            weight_accumulated = 0
            positions_accumulated = 0
            
            # Fill DUMs until we reach the target weight
            while weight_accumulated < partial_weight and current_dum_idx < len(dums):
                weight_needed = partial_weight - weight_accumulated
                
                if remaining_dum_weight <= weight_needed:
                    # Take entire remaining DUM (or remaining part of split DUM)
                    # Check if this is the ENTIRE DUM (not a split continuation)
                    # If remaining_dum_weight equals the original DUM weight at start, it's not a split
                    original_dum_weight = round(dums[current_dum_idx]['weight'], 1)
                    # Check if this is actually a split: only if we're continuing a split AND the weight is less than original
                    is_actually_split = is_continuing_split and (abs(remaining_dum_weight - original_dum_weight_at_start) > 0.1)
                    
                    # Special case: If this is smallest partial with manual positions (exception case)
                    if (smallest_partial_idx is not None and partial_idx == smallest_partial_idx and 
                        smallest_partial_positions is not None and is_continuing_split):
                        # Exception case: Use manual positions for continuing split in smallest partial
                        positions_to_use = smallest_partial_positions
                    else:
                        # Normal case: Use remaining positions
                        positions_to_use = remaining_dum_positions
                    
                    # Round weight to 1 decimal place to avoid floating point errors
                    rounded_weight = round(remaining_dum_weight, 1)
                    rounded_positions = round(positions_to_use)
                    
                    partial_dums.append({
                        'dum_number': dums[current_dum_idx]['number'],
                        'weight': rounded_weight,
                        'positions': rounded_positions,
                        'is_split': is_actually_split,  # Only true if it's actually a split continuation
                        'split_id': next_split_id(dums[current_dum_idx]['number']) if is_actually_split else ''
                    })
                    weight_accumulated += rounded_weight
                    positions_accumulated += rounded_positions
                    
                    # Move to next DUM
                    current_dum_idx += 1
                    is_continuing_split = False
                    if current_dum_idx < len(dums):
                        remaining_dum_weight = round(dums[current_dum_idx]['weight'], 1)
                        remaining_dum_positions = round(dums[current_dum_idx]['positions'])
                        original_dum_weight_at_start = remaining_dum_weight  # Reset for new DUM
                else:
                    # Split the DUM - this is the last DUM for this partial
                    # IMPORTANT: If this is the LAST partial, don't split - take entire remaining DUM
                    # This prevents the last DUM from being incorrectly marked as split
                    is_last_partial = (partial_idx == len(partial_weights) - 1)
                    
                    if is_last_partial:
                        # Last partial: take entire remaining DUM without splitting
                        # (even if weight is slightly more due to floating point)
                        rounded_weight = round(remaining_dum_weight, 1)
                        rounded_positions = round(remaining_dum_positions)
                        
                        partial_dums.append({
                            'dum_number': dums[current_dum_idx]['number'],
                            'weight': rounded_weight,
                            'positions': rounded_positions,
                            'is_split': is_continuing_split,  # Only true if it was already split from previous partial
                            'split_id': next_split_id(dums[current_dum_idx]['number']) if is_continuing_split else ''
                        })
                        weight_accumulated += rounded_weight
                        positions_accumulated += rounded_positions
                        
                        # Move to next DUM
                        current_dum_idx += 1
                        is_continuing_split = False
                        if current_dum_idx < len(dums):
                            remaining_dum_weight = round(dums[current_dum_idx]['weight'], 1)
                            remaining_dum_positions = round(dums[current_dum_idx]['positions'])
                            original_dum_weight_at_start = remaining_dum_weight
                    else:
                        # Not last partial: split the DUM as before
                        # Special case: If next partial is smallest partial (exception case), calculate differently
                        next_partial_is_smallest = (smallest_partial_idx is not None and 
                                                    partial_idx + 1 == smallest_partial_idx and 
                                                    smallest_partial_positions is not None)
                        
                        if next_partial_is_smallest:
                            # Exception case: Next partial is smallest with manual positions
                            # This partial gets the remaining positions (total - manual positions)
                            # The smallest partial will get the manual positions
                            positions_for_split = remaining_dum_positions - smallest_partial_positions
                            # Ensure non-negative
                            positions_for_split = max(0, positions_for_split)
                        else:
                            # Normal case: Calculate positions proportionally to the REMAINING DUM's weight
                            if remaining_dum_weight > 0:
                                # Calculate positions based on weight ratio of remaining DUM part
                                positions_for_split = round((weight_needed / remaining_dum_weight) * remaining_dum_positions)
                            else:
                                positions_for_split = 0
                            
                            # Ensure positions don't exceed remaining DUM positions
                            positions_for_split = min(positions_for_split, remaining_dum_positions)
                        
                        # Ensure positions are non-negative
                        positions_for_split = max(0, positions_for_split)
                        
                        # Round weight and positions to avoid floating point errors
                        rounded_weight_needed = round(weight_needed, 1)
                        rounded_positions_for_split = round(positions_for_split)
                        
                        partial_dums.append({
                            'dum_number': dums[current_dum_idx]['number'],
                            'weight': rounded_weight_needed,
                            'positions': rounded_positions_for_split,
                            'is_split': True,
                            'split_id': next_split_id(dums[current_dum_idx]['number'])
                        })
                        weight_accumulated += rounded_weight_needed
                        positions_accumulated += rounded_positions_for_split
                        
                        # Update remaining DUM (round to avoid floating point errors)
                        remaining_dum_weight = round(remaining_dum_weight - rounded_weight_needed, 1)
                        if next_partial_is_smallest:
                            # For exception case, remaining positions = manual positions
                            remaining_dum_positions = smallest_partial_positions
                        else:
                            remaining_dum_positions = round(remaining_dum_positions - rounded_positions_for_split)
                        is_continuing_split = True  # Mark that next partial continues this DUM
                        break
            
            # For exception case: use manual positions for smallest partial
            if smallest_partial_idx is not None and partial_idx == smallest_partial_idx and smallest_partial_positions is not None:
                final_positions = smallest_partial_positions
            else:
                final_positions = positions_accumulated  # Use actual accumulated positions from DUMs
            
            # Round accumulated weight to 1 decimal place
            rounded_weight_accumulated = round(weight_accumulated, 1)
            rounded_final_positions = round(final_positions)
            
            distribution.append({
                'weight': rounded_weight_accumulated,
                'positions': rounded_final_positions,
                'dums': partial_dums
            })
        
        return distribution
    
    def _save_config(self):
        """Validate and save configuration"""
        try:
            # Validate LTA data first
            if not self.lta_data.get('dums') or self.lta_data.get('total_weight', 0) <= 0:
                messagebox.showerror(
                    "Erreur",
                    "Données LTA invalides.\n\n"
                    "Le LTA doit avoir:\n"
                    "- Un poids total > 0\n"
                    "- Au moins un DUM\n\n"
                    "Vérifiez le fichier Excel du LTA."
                )
                return
            
            # Collect data from forms
            partials = []
            total_weight_check = 0
            
            # First collect partial weights to calculate distribution
            partial_weights = []
            for form_data in self.partial_forms:
                try:
                    weight = float(form_data['weight_var'].get().strip())
                    partial_weights.append(weight)
                except ValueError:
                    messagebox.showerror("Erreur", f"Poids invalide pour Partiel {form_data['partial_number']}")
                    return
            
            # Get exception case positions if provided (same logic as in _update_distribution_preview)
            smallest_partial_positions = None
            smallest_partial_idx = None
            max_dum_weight = max(dum['weight'] for dum in self.lta_data['dums'])
            num_partials = len(partial_weights)
            smallest_partial_weight = min(w for w in partial_weights if w > 0) if any(w > 0 for w in partial_weights) else 0
            is_exception_case = (
                num_partials in (2, 3)
                and smallest_partial_weight > 0
                and smallest_partial_weight < max_dum_weight
            )

            if is_exception_case:
                try:
                    smallest_partial_positions_str = self.smallest_partial_positions_var.get().strip()
                    if smallest_partial_positions_str:
                        smallest_partial_positions = int(smallest_partial_positions_str)
                        # Find which partial is the smallest
                        smallest_partial_weight = min(w for w in partial_weights if w > 0)
                        for idx, weight in enumerate(partial_weights):
                            if weight == smallest_partial_weight:
                                smallest_partial_idx = idx
                                break
                except (ValueError, AttributeError):
                    pass  # If not valid, ignore and use normal calculation
            
            # Calculate DUM distribution automatically (with exception case parameters)
            distribution = self._calculate_dum_distribution(partial_weights, smallest_partial_idx, smallest_partial_positions)
            
            # Build partials configuration using calculated distribution
            for idx, form_data in enumerate(self.partial_forms):
                partial_num = form_data['partial_number']
                
                # Validate required fields
                weight = form_data['weight_var'].get().strip()
                ds_serie_full = form_data['ds_serie_var'].get().strip()
                location = form_data['location_var'].get().strip()
                
                # Exception partial doesn't need Location (handled at airport), but DS Série IS required
                if is_exception_case and idx == smallest_partial_idx:
                    if not all([weight, ds_serie_full]):
                        messagebox.showerror(
                            "Validation",
                            f"Partiel {partial_num}: Poids et DS Série sont requis"
                        )
                        return
                else:
                    if not all([weight, ds_serie_full, location]):
                        messagebox.showerror(
                            "Validation",
                            f"Partiel {partial_num}: Tous les champs sont requis (Poids, DS Série, Lieu de Chargement)"
                        )
                        return
                
                # Validate and parse DS Série (required for all partials)
                if ds_serie_full:
                    ds_serie_normalized = normalize_ds_series(ds_serie_full)
                    is_valid, err_msg = validate_ds_series(ds_serie_normalized)
                    if not is_valid:
                        messagebox.showerror("Validation", f"Partiel {partial_num}: DS Série - {err_msg}")
                        return
                    parts = ds_serie_normalized.split()
                    ds_serie = parts[0] if parts else ""
                    ds_cle = parts[1] if len(parts) > 1 else ""
                else:
                    # Should not reach here due to required fields check above
                    ds_serie = ''
                    ds_cle = ''
                
                # Validate location (skip for exception partial)
                if not (is_exception_case and idx == smallest_partial_idx):
                    is_valid, err_msg = validate_location(location)
                    if not is_valid:
                        messagebox.showerror("Validation", f"Partiel {partial_num}: Lieu - {err_msg}")
                        return
                
                # Validate weight
                try:
                    weight_float = float(weight)
                    total_weight_check += weight_float
                except ValueError:
                    extra = ""
                    if _looks_like_weight_ds_fields_swapped(weight, ds_serie_full):
                        extra = "\n\nLes champs « Poids (kg) » et « DS Série » semblent inversés."
                    messagebox.showerror(
                        "Validation",
                        f"Partiel {partial_num}: Poids invalide (nombre en kg uniquement).{extra}"
                    )
                    return
                
                # Get DUMs from calculated distribution
                partial_dist = distribution[idx]
                selected_dums = []
                
                for dum_info in partial_dist['dums']:
                    selected_dums.append({
                        'dum_number': dum_info['dum_number'],
                        'weight': dum_info['weight'],
                        'positions': dum_info['positions'],
                        'is_split': dum_info['is_split'],
                        'split_id': dum_info.get('split_id', '')
                    })
                
                # Validate distribution has DUMs (exception partial has none - handled at airport)
                if not selected_dums and not (is_exception_case and idx == smallest_partial_idx):
                    messagebox.showerror(
                        "Validation",
                        f"Partiel {partial_num}: Aucun DUM assigné par distribution automatique"
                    )
                    return
                
                partials.append({
                    'partial_number': partial_num,
                    'weight': weight_float,
                    'positions': partial_dist['positions'],
                    'ds_serie': ds_serie,
                    'ds_cle': ds_cle,
                    'loading_location': location,
                    'dums': selected_dums
                })
            
            # Validate weight tolerance (allow 1% difference)
            weight_diff = abs(total_weight_check - self.lta_data['total_weight'])
            weight_tolerance = self.lta_data['total_weight'] * 0.01
            
            if weight_diff > weight_tolerance:
                response = messagebox.askyesno(
                    "Attention",
                    f"La somme des poids partiels ({total_weight_check} kg) ne correspond pas exactement au poids total ({self.lta_data['total_weight']} kg).\n\n"
                    f"Différence: {weight_diff:.2f} kg\n\n"
                    "Continuer quand même?"
                )
                if not response:
                    return
            
            # Detect split DUMs from distribution
            split_dums = {}
            for partial in partials:
                for dum in partial['dums']:
                    if dum['is_split']:
                        dum_num = str(dum['dum_number'])
                        if dum_num not in split_dums:
                            split_dums[dum_num] = {
                                'total_weight': 0,
                                'splits': []
                            }
                        
                        split_dums[dum_num]['total_weight'] += dum['weight']
                        split_dums[dum_num]['splits'].append({
                            'partial': partial['partial_number'],
                            'split_id': dum['split_id'],
                            'weight': dum['weight'],
                            'positions': dum['positions']
                        })
            
            # Detect exception case: smallest partial < max DUM weight
            # (means partial is a leftover complement of one DUM, not made of whole DUMs)
            max_dum_weight = max(dum['weight'] for dum in self.lta_data['dums'])
            num_partials = len(partial_weights)
            smallest_partial_weight = min(w for w in partial_weights if w > 0) if any(w > 0 for w in partial_weights) else 0
            is_exception_case = (
                num_partials in (2, 3)
                and smallest_partial_weight > 0
                and smallest_partial_weight < max_dum_weight
            )
            
            # For exception case, validate additional fields
            smallest_partial_number = None
            smallest_partial_positions = None
            airport_reference = None
            
            if is_exception_case:
                # Find which partial is the smallest
                for idx, weight in enumerate(partial_weights):
                    if weight == smallest_partial_weight:
                        smallest_partial_number = idx + 1
                        break
                
                # Validate exception case fields
                airport_reference = self.airport_reference_var.get().strip()
                smallest_partial_positions_str = self.smallest_partial_positions_var.get().strip()
                
                if not airport_reference:
                    messagebox.showerror(
                        "Validation",
                        "Cas d'exception détecté: Veuillez renseigner la référence créée à l'aéroport"
                    )
                    return
                
                if not smallest_partial_positions_str:
                    messagebox.showerror(
                        "Validation",
                        "Cas d'exception détecté: Veuillez renseigner les positions du plus petit partiel"
                    )
                    return
                
                try:
                    smallest_partial_positions = int(smallest_partial_positions_str)
                    if smallest_partial_positions <= 0:
                        raise ValueError("Positions must be positive")
                except ValueError:
                    messagebox.showerror(
                        "Validation",
                        "Positions du plus petit partiel: valeur invalide (doit être un nombre > 0)"
                    )
                    return
            
            # Build config
            config = {
                'lta_reference': self._get_lta_reference(),
                'lta_total_weight': self.lta_data['total_weight'],
                'lta_total_positions': self.lta_data['total_positions'],
                'partial_type': 'exception' if is_exception_case else 'normal',
                'partials': partials,
                'split_dums': split_dums
            }
            
            # Add exception case fields if applicable
            if is_exception_case:
                config['smallest_partial_number'] = smallest_partial_number
                config['smallest_partial_positions'] = smallest_partial_positions
                config['smallest_partial_airport_reference'] = airport_reference
            
            # Save config
            success = save_lta_partial_config(
                self.lta_folder_path,
                self.folder_name,
                config
            )
            
            if success:
                self.config_saved = True
                messagebox.showinfo("Succès", "Configuration sauvegardée!")
                self.dialog.destroy()
            else:
                messagebox.showerror("Erreur", "Impossible de sauvegarder la configuration")
                
        except Exception as e:
            logger.error(f"Error saving partial config: {e}", exc_info=True)
            messagebox.showerror("Erreur", f"Erreur lors de la sauvegarde:\n{e}")
    
    def _get_lta_reference(self):
        """Get LTA reference from LTA file and clean leading zeros"""
        try:
            lta_file_patterns = [
                f"{self.folder_name}.txt",
                f"{self.folder_name.replace(' ', '')}.txt",
                f"{self.folder_name.lower().replace(' ', '')}.txt"
            ]
            
            for pattern in lta_file_patterns:
                lta_file = os.path.join(self.lta_folder_path, pattern)
                if os.path.exists(lta_file):
                    with open(lta_file, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    if len(lines) >= 4:
                        reference = lines[3].strip()  # Line 4 (index 3)
                        # Remove /1 suffix if present
                        if reference.endswith('/1'):
                            reference = reference[:-2]
                        
                        # Clean leading zeros from the first part (before dash)
                        # Example: "065-123456" -> "65-123456"
                        if '-' in reference:
                            parts = reference.split('-', 1)
                            if len(parts) == 2:
                                # Remove leading zeros from first part
                                first_part = parts[0].lstrip('0') or '0'  # Keep at least one digit
                                reference = f"{first_part}-{parts[1]}"
                        
                        return reference
            
            return "UNKNOWN"
            
        except Exception as e:
            logger.error(f"Error getting LTA reference: {e}")
            return "UNKNOWN"
