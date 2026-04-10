#!/usr/bin/env python3
"""
Email Sender Utility
Sends generated_excel files via email when DUMs are successfully processed
"""

import smtplib
import os
import glob
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
import json

logger = logging.getLogger(__name__)

# Configuration file path
CONFIG_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config', 'email_config.json')


def load_email_config():
    """Load email configuration from config file"""
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config
        return None
    except Exception as e:
        logger.error(f"Error loading email config: {e}")
        return None


def save_email_config(config):
    """Save email configuration to config file"""
    try:
        os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logger.error(f"Error saving email config: {e}")
        return False


def send_excel_via_email(excel_file_path, lta_name, dum_number=None, recipient_email=None, lta_folder_path=None):
    """
    Send generated_excel file via email, optionally with the MAWB PDF attached.

    Args:
        excel_file_path: Path to the generated_excel*.xlsx file
        lta_name: Name of the LTA folder
        dum_number: Optional DUM number (if sending after single DUM)
        recipient_email: Optional recipient email (uses config if not provided)
        lta_folder_path: Optional path to the parent folder containing the LTA subfolder.
                         Used to locate the MAWB PDF and to read error series from Excel.

    Returns:
        bool: True if email sent successfully, False otherwise
    """
    try:
        # Load configuration
        config = load_email_config()
        if not config:
            logger.warning("Email configuration not found. Email sending disabled.")
            return False
        
        # Get recipient email(s) - support both single email and list
        if not recipient_email:
            # Try recipient_emails (list) first, then recipient_email (single)
            recipient_emails = config.get('recipient_emails')
            recipient_email = config.get('recipient_email')
            
            if recipient_emails:
                # If recipient_emails is a list, use it
                if isinstance(recipient_emails, list):
                    recipient_email = recipient_emails
                elif isinstance(recipient_emails, str):
                    # If it's a string, split by comma
                    recipient_email = [email.strip() for email in recipient_emails.split(',')]
            elif recipient_email:
                # If recipient_email is a string, convert to list for consistency
                if isinstance(recipient_email, str):
                    recipient_email = [email.strip() for email in recipient_email.split(',')]
                elif isinstance(recipient_email, list):
                    recipient_email = recipient_email
            else:
                recipient_email = None
        
        # Normalize recipient_email to list format
        if recipient_email:
            if isinstance(recipient_email, str):
                recipient_email = [email.strip() for email in recipient_email.split(',')]
            elif not isinstance(recipient_email, list):
                recipient_email = [str(recipient_email)]
        else:
            logger.error("No recipient email configured")
            return False
        
        # Email settings from config
        sender_email = config.get('sender_email')
        sender_password = config.get('sender_password')  # Gmail App Password
        smtp_server = config.get('smtp_server', 'smtp.gmail.com')
        smtp_port = config.get('smtp_port', 587)
        
        if not sender_email or not sender_password:
            logger.error("Email credentials not configured")
            return False
        
        # Check if file exists
        if not os.path.exists(excel_file_path):
            logger.error(f"Excel file not found: {excel_file_path}")
            return False

        # --- Detect ERROR series in the Excel ---
        error_series = _get_error_series_from_excel(excel_file_path)
        has_errors = len(error_series) > 0

        # --- Resolve MAWB from PDF filename for subject ---
        mawb_suffix = ""
        if lta_folder_path:
            pdf_path_for_subject = _find_mawb_pdf(lta_folder_path, lta_name)
            if pdf_path_for_subject:
                pdf_basename = os.path.basename(pdf_path_for_subject)  # e.g. "3eme LTA - 235-97495543.pdf"
                # Strip lta_name prefix and .pdf suffix to get just the MAWB
                mawb_part = pdf_basename
                if mawb_part.startswith(lta_name + " - "):
                    mawb_part = mawb_part[len(lta_name) + 3:]  # remove "<lta_name> - "
                if mawb_part.lower().endswith(".pdf"):
                    mawb_part = mawb_part[:-4]
                mawb_suffix = f" - {mawb_part}"

        # Create email message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        # Join multiple recipients with comma for email header
        msg['To'] = ', '.join(recipient_email) if isinstance(recipient_email, list) else recipient_email
        
        # Ensure no duplicate Subject header (safety check)
        if 'Subject' in msg:
            del msg['Subject']
        
        # Build subject — includes MAWB and flags errors when present
        if dum_number:
            subject = f"DUM {dum_number} Traite - {lta_name}{mawb_suffix}"
        else:
            if has_errors:
                subject = f"[ERREUR DUM] LTA Complet - {lta_name}{mawb_suffix}"
            else:
                subject = f"LTA Complet - {lta_name}{mawb_suffix}"
        msg['Subject'] = subject
        
        # Email body
        body = f"""
Bonjour,

Le fichier Excel généré pour le LTA "{lta_name}" est prêt.
"""
        if dum_number:
            body += f"\nDUM {dum_number} a été traité avec succès.\n"
        else:
            body += f"\nTous les DUMs de ce LTA ont été traités.\n"

        if has_errors:
            body += f"\n\n⚠️  ATTENTION — DUMs en erreur ({len(error_series)}):\n"
            for s in error_series:
                body += f"   • {s}\n"
            body += "\nCes DUMs nécessitent un traitement manuel.\n"
        
        body += f"""
Fichier joint: {os.path.basename(excel_file_path)}
Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Cordialement,
TAJANI EL IDRISSI Ismail
"""
        
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # Attach Excel file
        with open(excel_file_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {os.path.basename(excel_file_path)}'
        )
        msg.attach(part)

        # --- Attach MAWB PDF if available ---
        if lta_folder_path:
            # Reuse the path already resolved for the subject when possible
            pdf_path = pdf_path_for_subject if mawb_suffix else _find_mawb_pdf(lta_folder_path, lta_name)
            if pdf_path and os.path.exists(pdf_path):
                try:
                    with open(pdf_path, 'rb') as pdf_file:
                        pdf_part = MIMEBase('application', 'octet-stream')
                        pdf_part.set_payload(pdf_file.read())
                    encoders.encode_base64(pdf_part)
                    pdf_part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {os.path.basename(pdf_path)}'
                    )
                    msg.attach(pdf_part)
                    logger.info(f"Attached MAWB PDF: {pdf_path}")
                    print(f"   📎 PDF joint: {os.path.basename(pdf_path)}")
                except Exception as pdf_error:
                    logger.warning(f"Could not attach PDF: {pdf_error}")
            else:
                logger.info(f"MAWB PDF not found for {lta_name} — sending without PDF")
        
        # Send email
        recipients_str = ', '.join(recipient_email) if isinstance(recipient_email, list) else recipient_email
        logger.info(f"Sending email to {recipients_str}...")
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        text = msg.as_string()
        # sendmail accepts list of recipients
        server.sendmail(sender_email, recipient_email if isinstance(recipient_email, list) else [recipient_email], text)
        server.quit()
        
        logger.info(f"✅ Email sent successfully to {recipients_str}")
        print(f"   📧 Email envoyé avec succès à {recipients_str}")
        return True
        
    except Exception as e:
        logger.error(f"Error sending email: {e}", exc_info=True)
        print(f"   ⚠️  Erreur envoi email: {e}")
        return False


def send_excel_after_dum_success(excel_file_path, lta_name, dum_number):
    """
    Convenience function to send Excel after DUM success
    """
    return send_excel_via_email(excel_file_path, lta_name, dum_number=dum_number)


def _get_error_series_from_excel(excel_file_path):
    """
    Read the Summary sheet and return a list of DUM serie values that contain an error.
    Series are in column C at rows 12, 19, 26, 33, … (pattern: 12 + (n-1)*7).
    A value is considered an error if the string contains 'error' (case-insensitive)
    or is exactly 'error'.

    Returns:
        List[str]: e.g. ["0159942R (error)", "error"]  — empty list if none or on any failure.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file_path, data_only=True, read_only=True)
        if 'Summary' not in wb.sheetnames:
            wb.close()
            return []
        ws = wb['Summary']
        errors = []
        dum_number = 1
        consecutive_none = 0
        while dum_number <= 100:  # Safety cap — supports up to 100 DUMs per LTA
            row = 12 + (dum_number - 1) * 7
            cell = ws[f"C{row}"]
            value = cell.value
            if value is None:
                consecutive_none += 1
                if consecutive_none >= 3:
                    break  # 3 empty rows in a row → no more DUMs
                dum_number += 1
                continue
            consecutive_none = 0  # reset on non-empty cell
            value_str = str(value).strip()
            if 'error' in value_str.lower():
                errors.append(value_str)
            dum_number += 1
        wb.close()
        return errors
    except Exception as e:
        logger.warning(f"Could not read error series from Excel: {e}")
        return []


def _find_mawb_pdf(lta_folder_path, lta_name):
    """
    Locate the MAWB PDF file inside the LTA subfolder.
    Pattern: <lta_name>/<lta_name> - <MAWB>.pdf
    Falls back to any .pdf in the subfolder if naming doesn't match.

    Returns:
        str | None: absolute path to PDF, or None if not found.
    """
    try:
        lta_subfolder = os.path.join(lta_folder_path, lta_name)
        if not os.path.isdir(lta_subfolder):
            return None
        # Try the standard naming pattern first
        pdfs = glob.glob(os.path.join(lta_subfolder, f"{lta_name} - *.pdf"))
        if not pdfs:
            # Fallback: any PDF in the subfolder
            pdfs = glob.glob(os.path.join(lta_subfolder, "*.pdf"))
        return pdfs[0] if pdfs else None
    except Exception as e:
        logger.warning(f"Could not locate MAWB PDF for {lta_name}: {e}")
        return None


def send_excel_after_lta_completion(excel_file_path, lta_name, lta_folder_path=None):
    """
    Send Excel (and optionally the MAWB PDF) after LTA completion.
    Subject will flag ERROR if any DUM serie contains an error.
    """
    return send_excel_via_email(
        excel_file_path,
        lta_name,
        lta_folder_path=lta_folder_path,
    )
