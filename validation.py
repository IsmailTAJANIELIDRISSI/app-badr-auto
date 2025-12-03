#!/usr/bin/env python3
"""
Test script to verify LTA folder structure and data reading
Run this to check if your files are properly formatted before running the full automation
"""
import os
import glob
from openpyxl import load_workbook

def test_lta_folder(lta_folder_path, lta_name):
    """Test if an LTA folder has all required files and proper format"""
    print("\n" + "="*70)
    print(f"📁 TESTING: {lta_name}")
    print("="*70)
    
    errors = []
    warnings = []
    
    # Test 1: Check folder exists
    if not os.path.exists(lta_folder_path):
        errors.append(f"❌ Dossier introuvable: {lta_folder_path}")
        return errors, warnings
    
    print(f"✓ Dossier existe: {lta_folder_path}")
    
    # Test 2: Check .txt file exists (parent directory)
    parent_dir = os.path.dirname(lta_folder_path)
    txt_file_path = os.path.join(parent_dir, f"{lta_name}.txt")
    
    if not os.path.exists(txt_file_path):
        errors.append(f"❌ Fichier .txt introuvable: {txt_file_path}")
    else:
        print(f"✓ Fichier .txt existe: {txt_file_path}")
        
        # Test 2a: Read shipper name
        try:
            with open(txt_file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                shipper_name = None
                for line in lines:
                    stripped = line.strip()
                    if stripped:
                        shipper_name = stripped
                        break
                
                if shipper_name:
                    print(f"   ✓ Expéditeur: {shipper_name}")
                else:
                    warnings.append(f"⚠️  Fichier .txt vide ou sans nom d'expéditeur")
        except Exception as e:
            errors.append(f"❌ Erreur lecture .txt: {e}")
    
    # Test 3: Check summary_file exists
    summary_files = glob.glob(os.path.join(lta_folder_path, "summary_file*.xlsx"))
    
    if not summary_files:
        errors.append(f"❌ Aucun summary_file*.xlsx trouvé dans {lta_folder_path}")
    else:
        summary_file = summary_files[0]
        print(f"✓ Summary file: {os.path.basename(summary_file)}")
        
        # Test 3a: Read and validate summary_file structure
        try:
            wb = load_workbook(summary_file, data_only=True)
            
            # Find the active sheet
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
            else:
                ws = wb.active
            
            print(f"   ✓ Feuille active: {ws.title}")
            
            # Find header row
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
                cell_values = [str(cell.value).lower() if cell.value else '' for cell in row]
                if 'sheet' in ' '.join(cell_values) and 'total' in ' '.join(cell_values):
                    header_row = row_idx
                    break
            
            if not header_row:
                warnings.append(f"⚠️  En-tête du tableau non trouvé (recherche 'Sheet' + 'Total')")
            else:
                print(f"   ✓ En-tête trouvé à la ligne {header_row}")
                
                # Read headers
                headers = []
                for cell in ws[header_row]:
                    if cell.value:
                        headers.append(str(cell.value))
                
                print(f"   📊 Colonnes: {', '.join(headers)}")
                
                # Count data rows
                dum_count = 0
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if row and row[0]:  # First column not empty
                        dum_count += 1
                
                if dum_count == 0:
                    warnings.append(f"⚠️  Aucune ligne de données trouvée après l'en-tête")
                else:
                    print(f"   ✓ {dum_count} DUM(s) trouvé(s)")
                    
                    # Show sample data
                    print(f"\n   📝 Exemple de données (premier DUM):")
                    first_data_row = ws[header_row + 1]
                    for i, (header_cell, data_cell) in enumerate(zip(ws[header_row], first_data_row)):
                        if header_cell.value:
                            print(f"      {header_cell.value}: {data_cell.value}")
                            if i >= 6:  # Limit to first 7 columns
                                break
            
        except Exception as e:
            errors.append(f"❌ Erreur lecture summary_file: {e}")
    
    # Test 4: Check for generated_excel (optional, for reference)
    generated_files = glob.glob(os.path.join(lta_folder_path, "generated_excel*.xlsx"))
    if generated_files:
        print(f"✓ Generated Excel trouvé: {os.path.basename(generated_files[0])}")
    else:
        warnings.append(f"⚠️  Aucun generated_excel*.xlsx trouvé (optionnel)")
    
    # Test 5: Check for PDF files
    pdf_files = glob.glob(os.path.join(lta_folder_path, "*.pdf"))
    if pdf_files:
        print(f"✓ {len(pdf_files)} fichier(s) PDF trouvé(s)")
    else:
        warnings.append(f"⚠️  Aucun fichier PDF trouvé")
    
    # Summary
    print("\n" + "-"*70)
    if errors:
        print("❌ ERREURS BLOQUANTES:")
        for error in errors:
            print(f"   {error}")
    
    if warnings:
        print("\n⚠️  AVERTISSEMENTS:")
        for warning in warnings:
            print(f"   {warning}")
    
    if not errors and not warnings:
        print("✅ TOUT EST OK - Dossier prêt pour l'automatisation!")
    elif not errors:
        print("✅ VALIDATION RÉUSSIE - Quelques avertissements mineurs")
    else:
        print("❌ VALIDATION ÉCHOUÉE - Corriger les erreurs avant de continuer")
    
    print("="*70)
    
    return errors, warnings

def main():
    """Test all LTA folders in current directory"""
    print("="*70)
    print("  TEST DE VALIDATION DES DOSSIERS LTA")
    print("="*70)
    print("\nCe script vérifie que vos dossiers LTA sont correctement formatés")
    print("pour l'automatisation BADR.\n")
    
    # Find all LTA folders
    all_directories = [d for d in os.listdir(".") if os.path.isdir(d)]
    lta_folders = [d for d in all_directories if 'lta' in d.lower()]
    
    if not lta_folders:
        print("❌ Aucun dossier LTA trouvé dans le répertoire courant")
        print("\nAssurez-vous que:")
        print("  - Les noms de dossiers contiennent 'LTA'")
        print("  - Vous exécutez ce script depuis le bon répertoire")
        input("\nAppuyez sur Entrée pour quitter...")
        return
    
    print(f"✓ {len(lta_folders)} dossier(s) LTA trouvé(s):")
    for folder in lta_folders:
        print(f"   - {folder}")
    
    # Test each folder
    total_errors = 0
    total_warnings = 0
    results = []
    
    for folder_name in lta_folders:
        folder_path = os.path.join(".", folder_name)
        errors, warnings = test_lta_folder(folder_path, folder_name)
        total_errors += len(errors)
        total_warnings += len(warnings)
        results.append((folder_name, len(errors), len(warnings)))
    
    # Final summary
    print("\n" + "="*70)
    print("  RÉSUMÉ FINAL")
    print("="*70)
    
    for folder_name, error_count, warning_count in results:
        status = "✅" if error_count == 0 else "❌"
        print(f"{status} {folder_name}: {error_count} erreur(s), {warning_count} avertissement(s)")
    
    print("\n" + "-"*70)
    print(f"Total: {total_errors} erreur(s), {total_warnings} avertissement(s)")
    
    if total_errors == 0:
        print("\n✅ TOUS LES DOSSIERS SONT PRÊTS POUR L'AUTOMATISATION!")
        print("\nVous pouvez maintenant exécuter:")
        print("   python process_lta_folders.py")
    else:
        print("\n⚠️  Certains dossiers ont des erreurs - corrigez-les avant de continuer")
    
    print("="*70)
    
    # Don't wait for input when run from GUI
    # input("\nAppuyez sur Entrée pour quitter...")

if __name__ == "__main__":
    main()
