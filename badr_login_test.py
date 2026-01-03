from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import subprocess
import time
import os
import shutil
import random
import traceback
import glob
import re
import socket
import sys
from openpyxl import load_workbook
from datetime import datetime
from dotenv import load_dotenv
import json
import psutil

# Import file_utils for partial LTA configuration
try:
    from gui.utils.file_utils import get_lta_partial_info
except ImportError:
    # Fallback if gui module not available
    def get_lta_partial_info(lta_folder_path, folder_name):
        """
        Get partial LTA configuration information
        Note: lta_folder_path should be the LTA folder itself (e.g., './4eme LTA'), not parent
        This fallback implementation expects the full path, unlike file_utils version
        """
        try:
            # Construct path: lta_folder_path already contains the full folder path
            config_path = os.path.join(lta_folder_path, f"{folder_name}_partial_config.json")
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return None
        except Exception as e:
            print(f"⚠️  Error loading partial config: {e}")
            return None

# Load environment variables
load_dotenv()

# Configuration from .env
EDGE_PATH = os.getenv('EDGE_PATH', r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe")
DRIVER_PATH = os.getenv('DRIVER_PATH', r"C:\Users\pc\Downloads\edgedriver_win64\msedgedriver.exe")
BADR_PASSWORD = os.getenv('BADR_PASSWORD', '')

def _load_lta_license():
    """Load LTA license from config file"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.join(script_dir, 'config', 'license.json')
        
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config.get('LTA_sys_validity', '2026-03-07')
        return '2026-03-07'  # Default fallback
    except:
        return '2026-03-07'

# Load license expiry date from config
LTA_license_expires = _load_lta_license()  

def close_excel_file(file_path):
    """
    Ferme un fichier Excel s'il est ouvert par Excel.
    
    Args:
        file_path: Chemin absolu du fichier Excel à fermer
    
    Returns:
        bool: True si fichier fermé ou n'était pas ouvert, False si erreur
    """
    try:
        file_name = os.path.basename(file_path)
        excel_closed = False
        
        # Parcourir tous les processus Excel
        for proc in psutil.process_iter(['pid', 'name']):
            try:
                if proc.info['name'] and 'EXCEL.EXE' in proc.info['name'].upper():
                    # Vérifier si ce processus a le fichier ouvert
                    try:
                        for item in proc.open_files():
                            if file_path.lower() in item.path.lower():
                                print(f"      ⚠️  Fichier {file_name} ouvert dans Excel (PID: {proc.info['pid']})")
                                print(f"      🔄 Fermeture du processus Excel...")
                                proc.terminate()
                                proc.wait(timeout=3)
                                excel_closed = True
                                print(f"      ✓ Processus Excel fermé")
                                time.sleep(1)  # Attendre que le fichier soit libéré
                                break
                    except (psutil.AccessDenied, psutil.NoSuchProcess):
                        continue
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                continue
        
        if not excel_closed:
            # Fichier pas ouvert, c'est OK
            return True
        
        return True
        
    except Exception as e:
        print(f"      ⚠️  Erreur lors de la fermeture Excel: {e}")
        # Continuer quand même, l'opération pourrait réussir
        return True

def get_fresh_profile_path():
    """Crée un chemin unique pour un profil temporaire"""
    timestamp = int(time.time())
    random_id = random.randint(1000, 9999)
    profile_name = f"selenium_edge_temp_{timestamp}_{random_id}"
    profile_path = os.path.join(os.environ['TEMP'], profile_name)
    return profile_path

def get_free_port():
    """Trouve un port libre pour le debugging"""
    import socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(('', 0))
        s.listen(1)
        port = s.getsockname()[1]
    return port

def cleanup_old_profiles():
    """Nettoie les anciens profils temporaires (optionnel)"""
    temp_dir = os.environ['TEMP']
    try:
        for item in os.listdir(temp_dir):
            if item.startswith("selenium_edge_temp_"):
                old_profile = os.path.join(temp_dir, item)
                try:
                    shutil.rmtree(old_profile)
                    print(f"🧹 Nettoyé: {item}")
                except:
                    pass
    except:
        pass

def parse_lta_file(lta_file_path):
    """
    Parse un fichier [X]er LTA.txt et extrait les données structurées.
    
    Détecte automatiquement si le fichier est signé (Line 8 contient série + clé)
    et ajuste les indices de ligne en conséquence.
    
    Returns:
        dict: {
            'signed': bool,
            'lta_name': str,
            'mawb': str,
            'lta_reference': str,
            'shipper_name': str,
            'signed_series': str or None,  # Format: "9913 G"
            'serie': str or None,           # "9913"
            'cle': str or None,             # "G"
            'total_p': int,
            'total_p_brut': int,
            'dums': [
                {
                    'dum_number': int,
                    'p': int,
                    'p_brut': int
                }
            ]
        }
    """
    try:
        with open(lta_file_path, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()
        
        # Filter out completely empty lines for consistent indexing
        lines = [line.rstrip('\n\r') for line in all_lines]
        
        # Extract basic header data (these positions are consistent)
        lta_name = lines[1].strip() if len(lines) > 1 else ""
        mawb = lines[2].strip() if len(lines) > 2 else ""
        lta_reference = lines[3].strip() if len(lines) > 3 else ""
        shipper_name = lines[5].strip() if len(lines) > 5 else ""
        
        # Check Line 8 for signed series format: digits + space(s) + single uppercase letter
        signed = False
        signed_series = None
        serie = None
        cle = None
        
        if len(lines) > 7:
            line_8 = lines[7].strip()
            # Validate format: "9913 G" or "1234  A" (one or more spaces)
            series_pattern = r'^(\d+)\s+([A-Z])$'
            match = re.match(series_pattern, line_8)
            if match:
                signed = True
                serie = match.group(1)
                cle = match.group(2)
                signed_series = f"{serie} {cle}"
        
        # Extract total P and P,BRUT (positions depend on signed status)
        # Line 9: P	[value]
        # Line 10: P,BRUT	[value]
        total_p = 0
        total_p_brut = 0
        
        if len(lines) > 9:
            p_line = lines[9].strip()
            if p_line.startswith('P\t') or p_line.startswith('P '):
                try:
                    total_p = int(p_line.split('\t')[-1].strip())
                except:
                    pass
        
        if len(lines) > 10:
            p_brut_line = lines[10].strip()
            if p_brut_line.startswith('P,BRUT'):
                try:
                    total_p_brut = int(p_brut_line.split('\t')[-1].strip())
                except:
                    pass
        
        # Parse DUM blocks
        dums = []
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Look for DUM header: "DUM 1", "DUM 2", etc.
            dum_match = re.search(r'DUM\s+(\d+)', line)
            if dum_match:
                dum_number = int(dum_match.group(1))
                dum_p = 0
                dum_p_brut = 0
                
                # Read next few lines for P and P,BRUT values
                for j in range(i+1, min(i+10, len(lines))):
                    dum_line = lines[j].strip()
                    
                    if dum_line.startswith('-----'):
                        break
                    
                    if dum_line.startswith('P\t') or dum_line.startswith('P '):
                        if 'BRUT' not in dum_line:
                            try:
                                dum_p = int(dum_line.split('\t')[-1].strip())
                            except:
                                pass
                    
                    if dum_line.startswith('P,BRUT'):
                        try:
                            dum_p_brut = int(dum_line.split('\t')[-1].strip())
                        except:
                            pass
                
                dums.append({
                    'dum_number': dum_number,
                    'p': dum_p,
                    'p_brut': dum_p_brut
                })
            
            i += 1
        
        return {
            'signed': signed,
            'lta_name': lta_name,
            'mawb': mawb,
            'lta_reference': lta_reference,
            'shipper_name': shipper_name,
            'signed_series': signed_series,
            'serie': serie,
            'cle': cle,
            'total_p': total_p,
            'total_p_brut': total_p_brut,
            'dums': dums
        }
    
    except Exception as e:
        print(f"❌ Erreur parsing fichier LTA {lta_file_path}: {e}")
        return None

def start_fresh_edge():
    """Lance Edge avec un profil complètement nouveau à chaque fois"""
    
    if not os.path.exists(EDGE_PATH):
        alt_path = r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
        if os.path.exists(alt_path):
            edge_path = alt_path
        else:
            print("❌ Edge introuvable !")
            return None, None
    else:
        edge_path = EDGE_PATH
    
    print("🔄 Fermeture des instances Edge existantes...")
    os.system("taskkill /F /IM msedge.exe >nul 2>&1")
    time.sleep(2)
    
    cleanup_old_profiles()
    
    profile_path = get_fresh_profile_path()
    print(f"📁 Nouveau profil: {os.path.basename(profile_path)}")
    
    debug_port = get_free_port()
    print(f"🔌 Port de debug: {debug_port}")
    
    print("🚀 Lancement de Edge (nouvelle instance)...")
    
    command = [
        edge_path,
        f"--remote-debugging-port={debug_port}",
        f"--user-data-dir={profile_path}",
        "--no-first-run",
    ]
    
    subprocess.Popen(command)
    time.sleep(4)
    
    print("✓ Edge lancé avec un profil vierge")
    return profile_path, debug_port

def connect_to_edge(debug_port):
    """Se connecte à l'instance Edge lancée"""
    
    try:
        edge_options = Options()
        edge_options.add_experimental_option("debuggerAddress", f"localhost:{debug_port}")
        
        # Add options to handle SSL certificates
        edge_options.add_argument('--ignore-certificate-errors')
        edge_options.add_argument('--ignore-ssl-errors')
        edge_options.add_argument('--allow-insecure-localhost')
        edge_options.accept_insecure_certs = True
        
        if not os.path.exists(DRIVER_PATH):
            print(f"❌ Driver introuvable: {DRIVER_PATH}")
            return None
        
        service = Service(executable_path=DRIVER_PATH)
        
        print("🔗 Connexion à Edge...")
        driver = webdriver.Edge(service=service, options=edge_options)
        
        print("✓ Connecté avec succès !")
        
        return driver
        
    except Exception as e:
        print(f"❌ Erreur de connexion: {e}")
        return None

def navigate_and_login(driver):
    """Navigue vers le site et effectue la connexion"""
    try:
        print("🌐 Navigation vers le site BADR...")
        driver.get("https://badr.douane.gov.ma:40444/badr/Login")
        print("✓ Navigation réussie !")
        print(f"📄 Titre: {driver.title}")
        
        wait = WebDriverWait(driver, 10)
        
        # ÉTAPE 1: Entrer le mot de passe
        print("\n🔐 Saisie du mot de passe...")
        
        if not BADR_PASSWORD:
            print("❌ ERREUR: Mot de passe non configuré dans le fichier .env")
            print("   Veuillez ajouter BADR_PASSWORD=votre_mot_de_passe dans .env")
            return False
        
        password_field = wait.until(
            EC.presence_of_element_located((By.ID, "connexionForm:pwdConnexionId"))
        )
        
        # Vérifier si le champ contient déjà du texte (éviter doublon)
        current_value = password_field.get_attribute('value') or ''
        
        if current_value == BADR_PASSWORD:
            print("✓ Mot de passe déjà présent (correct)")
        elif current_value:
            # Champ contient autre chose - nettoyer et ressaisir
            print(f"   ⚠️  Champ contient '{current_value[:3]}...' - nettoyage...")
            password_field.clear()
            time.sleep(0.5)
            password_field.send_keys(BADR_PASSWORD)
            print("✓ Mot de passe saisi (après nettoyage)")
        else:
            # Champ vide - saisie normale
            password_field.send_keys(BADR_PASSWORD)
            print("✓ Mot de passe saisi")
        
        # Vérifier que la valeur finale est correcte
        final_value = password_field.get_attribute('value') or ''
        if final_value != BADR_PASSWORD:
            print(f"   ⚠️  ATTENTION: Valeur finale incorrecte (longueur: {len(final_value)} vs attendu: {len(BADR_PASSWORD)})")
            # Dernière tentative de correction
            password_field.clear()
            time.sleep(0.5)
            password_field.send_keys(BADR_PASSWORD)
            print("   ✓ Correction appliquée")
        
        time.sleep(1)
        
        # ÉTAPE 2: Cliquer sur le bouton de connexion
        print("\n🖱️ Clic sur Connexion...")
        login_button = wait.until(
            EC.element_to_be_clickable((By.ID, "connexionForm:login"))
        )
        login_button.click()
        print("✓ Connexion cliquée - attente de redirection...")
        time.sleep(5)  # Attendre le chargement
        
        # ÉTAPE 3: Vérifier s'il y a une session active
        print("\n🔍 Vérification session active...")
        try:
            # Chercher le lien de session active
            session_link = driver.find_element(By.ID, "connexionForm:sessionConnexionId")
            print("⚠️  Session active détectée!")
            print("🔄 Clic pour désactiver l'ancienne session...")
            session_link.click()
            time.sleep(5)
            print("✓ Ancienne session désactivée, redirection vers accueil")
        except Exception as e:
            # Pas de session active, connexion normale
            print("✓ Pas de session active, connexion directe")
        
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors de la connexion: {e}")
        return False

def save_dum_reference(lta_folder_path, dum_reference):
    """
    Sauvegarde la référence DUM dans le fichier result_LTAS.txt.
    
    Format du fichier:
    [LTA Folder Name] - [LTA Reference]
    [DUM 1 Reference]
    [DUM 2 Reference]
    ***
    
    Args:
        lta_folder_path: Chemin du dossier LTA en cours de traitement
        dum_reference: Référence du DUM (ex: "0139769N")
    """
    try:
        # Fichier de résultats global dans le répertoire de travail
        result_file = os.path.join(os.getcwd(), "result_LTAS.txt")
        
        # Extraire le nom du dossier LTA
        lta_folder_name = os.path.basename(lta_folder_path)
        
        # Chercher le fichier LTA PDF pour obtenir la référence complète
        lta_reference = "UNKNOWN"
        try:
            # Chercher les deux patterns: "*eme LTA - *.pdf" et "*er LTA - *.pdf"
            lta_files = glob.glob(os.path.join(lta_folder_path, "*eme LTA - *.pdf"))
            if not lta_files:
                lta_files = glob.glob(os.path.join(lta_folder_path, "*er LTA - *.pdf"))
            
            if lta_files:
                lta_filename = os.path.basename(lta_files[0])
                # Extraire la référence (sans l'extension .pdf)
                lta_reference = os.path.splitext(lta_filename)[0]
        except Exception as e:
            print(f"      ⚠️  Impossible d'extraire la référence LTA: {e}")
        
        # Vérifier si c'est le premier DUM de ce LTA
        is_first_dum = True
        if os.path.exists(result_file):
            with open(result_file, 'r', encoding='utf-8') as f:
                content = f.read()
                # Si le nom du LTA apparaît déjà, ce n'est pas le premier DUM
                if lta_reference in content:
                    is_first_dum = False
        
        # Écrire dans le fichier
        with open(result_file, 'a', encoding='utf-8') as f:
            # Si c'est le premier DUM, écrire l'en-tête du LTA
            if is_first_dum:
                f.write(f"{lta_reference}\n")
            
            # Écrire la référence du DUM
            f.write(f"{dum_reference}\n")
        
        print(f"      ✓ Référence sauvegardée dans result_LTAS.txt")
        if is_first_dum:
            print(f"         En-tête LTA: {lta_reference}")
        print(f"         Référence DUM: {dum_reference}")
        
    except Exception as e:
        print(f"      ❌ Erreur sauvegarde référence: {e}")
        traceback.print_exc()


def add_lta_separator():
    """
    Ajoute le séparateur *** après le dernier DUM d'un LTA.
    À appeler après avoir traité tous les DUMs d'un LTA.
    """
    try:
        result_file = os.path.join(os.getcwd(), "result_LTAS.txt")
        
        with open(result_file, 'a', encoding='utf-8') as f:
            f.write("***\n\n")
        
        print("   ✓ Séparateur LTA ajouté (***)")
        
    except Exception as e:
        print(f"   ⚠️  Erreur ajout séparateur: {e}")


def save_dum_series_to_excel(lta_folder_path, dum_number, serie):
    """
    Écrit la série du DUM dans le fichier generated_excel à la position appropriée.
    
    Pattern des positions:
    - DUM 1: C12
    - DUM 2: C19
    - DUM 3: C26
    - DUM 4: C33
    - Pattern: C + (12 + (dum_number - 1) * 7)
    
    Args:
        lta_folder_path: Chemin du dossier LTA
        dum_number: Numéro du DUM (1, 2, 3, 4, etc.)
        serie: Série du DUM (ex: "0139769N")
    """
    max_retries = 3
    retry_delay = 2  # secondes
    
    for attempt in range(max_retries):
        try:
            # Trouver le fichier generated_excel dans le dossier LTA
            generated_excel_path = None
            for file in os.listdir(lta_folder_path):
                if file.startswith("generated_excel") and file.endswith(".xlsx"):
                    generated_excel_path = os.path.join(lta_folder_path, file)
                    break
            
            if not generated_excel_path:
                print(f"      ⚠️  Fichier generated_excel non trouvé dans {lta_folder_path}")
                return False
            
            # Calculer la position de la cellule: C + (12 + (dum_number - 1) * 7)
            row_number = 12 + (dum_number - 1) * 7
            cell_position = f"C{row_number}"
            
            # Attendre un peu avant d'ouvrir (éviter conflits)
            if attempt > 0:
                print(f"      🔄 Tentative {attempt + 1}/{max_retries}...")
                time.sleep(retry_delay)
            
            # Ouvrir le fichier Excel (data_only=False pour pouvoir écrire)
            wb = None
            try:
                # Fermer le fichier Excel s'il est ouvert
                close_excel_file(generated_excel_path)
                
                wb = load_workbook(generated_excel_path, data_only=False)
                ws = wb['Summary']
                
                # Écrire la série dans la cellule
                ws[cell_position] = serie
                
                # Sauvegarder le fichier
                wb.save(generated_excel_path)
                
                print(f"      ✓ Série écrite dans generated_excel")
                print(f"         Cellule {cell_position}: {serie}")
                
                return True
                
            finally:
                # Toujours fermer le workbook
                if wb:
                    try:
                        wb.close()
                    except:
                        pass
            
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"      ⚠️  Erreur tentative {attempt + 1}: {e}")
                print(f"      ⏳ Nouvelle tentative dans {retry_delay}s...")
            else:
                print(f"      ❌ Erreur écriture série dans generated_excel après {max_retries} tentatives: {e}")
                print(f"      💡 Vérifiez que le fichier Excel n'est pas ouvert dans Excel")
                traceback.print_exc()
                return False
    
    return False

def detect_blocage_from_lta_file(lta_folder_path):
    """
    Détecte si un LTA est un "blocage" en vérifiant la ligne 5 du fichier [X]eme LTA.txt
    
    Returns:
        dict: {
            'is_blocage': bool,
            'original_weight': float or None,  # Line 12
            'blocked_weight': float or None,    # Line 13
            'corrected_weight': float or None   # Line 12 - Line 13
        }
    """
    try:
        # Trouver le fichier [X]eme LTA.txt dans le répertoire parent
        lta_name = os.path.basename(lta_folder_path)
        parent_dir = os.path.dirname(lta_folder_path)
        
        # Si parent_dir est vide, utiliser le répertoire courant
        if not parent_dir:
            parent_dir = "."
        
        # DEBUG: Afficher les chemins
        print(f"      🔍 Recherche fichier blocage:")
        print(f"         Dossier LTA: {lta_folder_path}")
        print(f"         Nom LTA: {lta_name}")
        print(f"         Parent dir: {parent_dir}")
        
        # Chercher le fichier avec le pattern [X]eme LTA.txt ou [X]er LTA.txt
        lta_txt_pattern = lta_name.replace(" ", "").lower()  # "7emelta"
        print(f"         Pattern recherché: {lta_txt_pattern}.txt")
        
        lta_txt_file = None
        try:
            files_found = []
            for file in os.listdir(parent_dir):
                # Ignorer les dossiers, ne chercher que les fichiers .txt
                file_path = os.path.join(parent_dir, file)
                if not os.path.isfile(file_path):
                    continue  # Ignorer les dossiers
                
                if not file.lower().endswith('.txt'):
                    continue  # Ignorer les non-.txt
                
                files_found.append(file)
                file_pattern = file.lower().replace(" ", "").replace(".txt", "")
                if file_pattern == lta_txt_pattern:
                    lta_txt_file = file_path
                    print(f"         ✓ Fichier trouvé: {file}")
                    break
            
            if not lta_txt_file:
                print(f"         ℹ️  Fichier {lta_txt_pattern}.txt non trouvé dans {len(files_found)} fichiers .txt")
        except Exception as list_err:
            print(f"      ⚠️  Erreur lecture répertoire {parent_dir}: {list_err}")
            return {'is_blocage': False, 'original_weight': None, 'blocked_weight': None, 'corrected_weight': None}
        
        if not lta_txt_file:
            print(f"      ℹ️  Fichier LTA txt non trouvé - traitement normal")
            return {'is_blocage': False, 'original_weight': None, 'blocked_weight': None, 'corrected_weight': None}
        
        # Lire le fichier
        with open(lta_txt_file, 'r', encoding='utf-8') as f:
            lines = [line.rstrip('\n\r') for line in f.readlines()]
        
        # Vérifier la ligne 5 (index 4)
        if len(lines) <= 4:
            return {'is_blocage': False, 'original_weight': None, 'blocked_weight': None, 'corrected_weight': None}
        
        line_5 = lines[4].strip().lower()
        
        # Détecter les variantes de "blocage"
        is_blocage = any(keyword in line_5 for keyword in ['blocage', 'blocag', 'blocaj'])
        
        if not is_blocage:
            return {'is_blocage': False, 'original_weight': None, 'blocked_weight': None, 'corrected_weight': None}
        
        # C'est un blocage - extraire les poids des lignes 12 et 13
        print(f"\n   ⚠️  BLOCAGE DÉTECTÉ (Ligne 5: '{lines[4]}')")
        
        original_weight = None
        blocked_weight = None
        
        # Ligne 12 (index 11): Poids original
        if len(lines) > 11:
            try:
                original_weight = float(lines[11].strip())
                print(f"      📊 Poids original (Ligne 12): {original_weight} kg")
            except ValueError:
                print(f"      ⚠️  Ligne 12 non numérique: '{lines[11]}'")
        
        # Ligne 13 (index 12): Poids bloqué
        if len(lines) > 12:
            try:
                blocked_weight = float(lines[12].strip())
                print(f"      📊 Poids bloqué (Ligne 13): {blocked_weight} kg")
            except ValueError:
                print(f"      ⚠️  Ligne 13 non numérique: '{lines[12]}'")
        
        # Calculer le poids corrigé
        corrected_weight = None
        if original_weight is not None and blocked_weight is not None:
            corrected_weight = round(original_weight - blocked_weight, 2)
            print(f"      🧮 Poids corrigé calculé: {original_weight} - {blocked_weight} = {corrected_weight} kg")
            
            if corrected_weight < 0:
                print(f"      ⚠️  AVERTISSEMENT: Poids corrigé négatif ({corrected_weight} kg) - utilisation valeur absolue")
                corrected_weight = abs(corrected_weight)
        else:
            print(f"      ❌ Impossible de calculer le poids corrigé (données manquantes)")
        
        return {
            'is_blocage': True,
            'original_weight': original_weight,
            'blocked_weight': blocked_weight,
            'corrected_weight': corrected_weight
        }
    
    except Exception as e:
        print(f"      ❌ Erreur détection blocage: {e}")
        return {'is_blocage': False, 'original_weight': None, 'blocked_weight': None, 'corrected_weight': None}


def modify_etat_depotage_for_blocage(driver, lta_folder_path, shipper_data):
    """
    PHASE 2 - BLOCAGE: Modifier un Etat de Dépotage existant pour une LTA bloquée
    
    Cette fonction:
    1. Navigue vers Modifier une Déclaration → Etat de Dépotage → Voyage Aérien
    2. Recherche l'ED existant par Série/Clé
    3. Récupère la référence LTA existante
    4. Supprime les lots conflictuels (ceux avec la référence LTA)
    5. Ajoute les nouveaux lots pour chaque DUM (comme create_etat_depotage)
    6. Sauvegarde et valide l'ED modifié
    
    Args:
        driver: WebDriver Selenium
        lta_folder_path: Chemin vers le dossier LTA
        shipper_data: Dictionnaire avec série, clé, etc.
    
    Returns:
        True si succès, False si erreur
    """
    try:
        wait = WebDriverWait(driver, 15)
        
        print("\n" + "="*70)
        print("🔄 MODIFICATION ETAT DE DÉPOTAGE (BLOCAGE)")
        print("="*70)
        
        # ==================================================================
        # ÉTAPE MED.0: Navigation vers "Modifier Etat de Dépotage"
        # ==================================================================
        print("\n📂 Navigation: MISE EN DOUANE → Modifier → Etat de Dépotage → Voyage Aérien...")
        
        # MED.0.1: Ouvrir le menu "MISE EN DOUANE"
        try:
            mise_en_douane_link = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//h3[contains(@class, 'ui-panelmenu-header')]//a[contains(text(), 'MISE EN DOUANE')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", mise_en_douane_link)
            time.sleep(0.5)
            mise_en_douane_link.click()
            print("   ✓ Menu 'MISE EN DOUANE' ouvert")
            time.sleep(2)
        except Exception as e:
            print(f"   ⚠️  Menu 'MISE EN DOUANE' déjà ouvert ou erreur: {e}")
            # Continuer car le menu peut déjà être ouvert
        
        # MED.0.2: Ouvrir "Modifier une Déclaration" (ID: _283)
        try:
            modifier_declaration_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_283"))
            )
            modifier_declaration_link.click()
            print("   ✓ Sous-menu 'Modifier une Déclaration' ouvert")
            time.sleep(1)
        except Exception as e:
            print(f"   ❌ Erreur ouverture 'Modifier une Déclaration': {e}")
            return_to_home_after_error(driver)
            return False
        
        # MED.0.3: Ouvrir "Etat de Dépotage" (ID: _336)
        try:
            etat_depotage_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_336"))
            )
            etat_depotage_link.click()
            print("   ✓ Sous-menu 'Etat de Dépotage' ouvert")
            time.sleep(1)
        except Exception as e:
            print(f"   ❌ Erreur ouverture 'Etat de Dépotage': {e}")
            return_to_home_after_error(driver)
            return False
        
        # MED.0.4: Cliquer sur "Voyage Aérien" (ID: _343)
        try:
            voyage_aerien_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_343"))
            )
            voyage_aerien_link.click()
            print("   ✓ Lien 'Voyage Aérien' cliqué")
            time.sleep(3)
        except Exception as e:
            print(f"   ❌ Erreur clic 'Voyage Aérien': {e}")
            return_to_home_after_error(driver)
            return False
        
        # MED.0.5: Basculer dans l'iframe
        try:
            print("   🔄 Basculement vers l'iframe du formulaire...")
            iframe = wait.until(
                EC.presence_of_element_located((By.ID, "iframeMenu"))
            )
            driver.switch_to.frame(iframe)
            print("   ✓ Iframe chargé")
            time.sleep(2)
        except Exception as e:
            print(f"   ❌ Erreur basculement iframe: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE MED.1: Configurer le formulaire de recherche
        # ==================================================================
        print("\n   🔍 Configuration du formulaire de recherche...")
        
        # MED.1.1: Décocher "Déclaration enregistrée" (coché par défaut)
        try:
            checkbox_input = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:enregistreeID_input"))
            )
            
            # Vérifier si coché
            if checkbox_input.is_selected():
                # Cliquer sur la div.ui-chkbox-box pour décocher
                checkbox_box = driver.find_element(By.CSS_SELECTOR, "div#rootForm\\:enregistreeID div.ui-chkbox-box")
                checkbox_box.click()
                print("      ✓ Case 'Déclaration enregistrée' décochée")
                time.sleep(0.5)
            else:
                print("      ✓ Case 'Déclaration enregistrée' déjà décochée")
        except Exception as e:
            print(f"      ⚠️  Erreur décocher case: {e}")
            # Continuer quand même
        
        # MED.1.2: Remplir les critères de recherche
        
        # Bureau: 301
        try:
            bureau_input = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:_bureauId"))
            )
            bureau_input.clear()
            bureau_input.send_keys("301")
            print("      ✓ Bureau: 301")
            time.sleep(0.3)
        except Exception as e:
            print(f"      ❌ Erreur saisie bureau: {e}")
            return_to_home_after_error(driver)
            return False
        
        # Régime: 000
        try:
            regime_input = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:_regimeId"))
            )
            regime_input.clear()
            regime_input.send_keys("000")
            print("      ✓ Régime: 000")
            time.sleep(0.3)
        except Exception as e:
            print(f"      ❌ Erreur saisie régime: {e}")
            return_to_home_after_error(driver)
            return False
        
        # Année: 2025 (année actuelle)
        try:
            annee_input = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:_anneeId"))
            )
            annee_input.clear()
            current_year = str(time.strftime("%Y"))
            annee_input.send_keys(current_year)
            print(f"      ✓ Année: {current_year}")
            time.sleep(0.3)
        except Exception as e:
            print(f"      ❌ Erreur saisie année: {e}")
            return_to_home_after_error(driver)
            return False
        
        # Série: Extraire de la ligne 2 du fichier shipper (ex: "3124 Y" → "3124")
        try:
            serie_input = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:_serieId"))
            )
            serie_input.clear()
            serie_input.send_keys(shipper_data['serie'])
            print(f"      ✓ Série: {shipper_data['serie']}")
            time.sleep(0.3)
        except Exception as e:
            print(f"      ❌ Erreur saisie série: {e}")
            return_to_home_after_error(driver)
            return False
        
        # Clé: Extraire de la ligne 2 du fichier shipper (ex: "3124 Y" → "Y")
        try:
            cle_input = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:_cleId"))
            )
            cle_input.clear()
            cle_input.send_keys(shipper_data['cle'])
            print(f"      ✓ Clé: {shipper_data['cle']}")
            time.sleep(0.3)
        except Exception as e:
            print(f"      ❌ Erreur saisie clé: {e}")
            return_to_home_after_error(driver)
            return False
        
        # MED.1.3: Cliquer sur "Valider"
        try:
            valider_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "rootForm:btnConfir"))
            )
            valider_btn.click()
            print("      ✓ Bouton 'Valider' cliqué")
            time.sleep(4)  # Attendre le chargement de l'ED existant
        except Exception as e:
            print(f"      ❌ Erreur clic validation: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE MED.2: Capturer la référence LTA existante
        # ==================================================================
        print("\n   📋 Capture de la référence LTA existante...")
        
        try:
            reference_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:referenceLotID"))
            )
            lta_reference_existing = reference_input.get_attribute("value").strip()
            print(f"      ✓ Référence LTA existante: {lta_reference_existing}")
            
            # MED.2.1: Sauvegarder dans le fichier shipper (ligne 6)
            try:
                lta_name = os.path.basename(lta_folder_path)
                parent_dir = os.path.dirname(lta_folder_path)
                lta_name_with_underscore = lta_name.replace(" ", "_")
                shipper_pattern = f"{lta_name_with_underscore}_*.txt"
                shipper_files = glob.glob(os.path.join(parent_dir, shipper_pattern))
                
                if shipper_files:
                    shipper_file = shipper_files[0]
                    
                    # Lire le fichier
                    with open(shipper_file, 'r', encoding='utf-8') as f:
                        lines = [line.rstrip('\n') for line in f.readlines()]
                    
                    # S'assurer qu'on a au moins 4 lignes
                    while len(lines) < 4:
                        lines.append('')
                    
                    # Ajouter/remplacer ligne 5 (index 4)
                    if len(lines) == 4:
                        lines.append(lta_reference_existing)
                    else:
                        lines[4] = lta_reference_existing
                    
                    # Réécrire
                    with open(shipper_file, 'w', encoding='utf-8') as f:
                        f.write('\n'.join(lines))
                    
                    print(f"      ✓ Référence sauvegardée dans {os.path.basename(shipper_file)} (ligne 5)")
                else:
                    print(f"      ⚠️  Fichier shipper introuvable: {shipper_pattern}")
            
            except Exception as e:
                print(f"      ⚠️  Erreur sauvegarde référence: {e}")
                # Continuer quand même
            
        except Exception as e:
            print(f"      ❌ Erreur capture référence: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE MED.3: Naviguer vers l'onglet LTA
        # ==================================================================
        print("\n   📄 Navigation vers l'onglet LTA...")
        
        try:
            lta_tab = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab4']"))
            )
            lta_tab.click()
            print("      ✓ Onglet LTA ouvert")
            time.sleep(2)
        except Exception as e:
            print(f"      ❌ Erreur navigation onglet LTA: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE MED.4: Supprimer les lots conflictuels
        # ==================================================================
        print("\n   🗑️  Suppression des lots conflictuels...")
        
        # Préparer la référence de base (sans /N)
        lta_reference_base = lta_reference_existing.split('/')[0].replace("-", "")
        print(f"      📋 Référence de base (pour comparaison): {lta_reference_base}")
        
        lots_deleted_count = 0
        page_number = 1
        
        while True:
            print(f"\n      🔍 Scan page {page_number} pour lots à supprimer...")
            
            try:
                # Attendre que le tableau soit chargé
                time.sleep(1)
                
                # STRATÉGIE: Supprimer UN lot à la fois, puis re-scanner
                # pour éviter les erreurs "stale element"
                lot_deleted_on_this_scan = False
                
                # Trouver toutes les lignes du tableau
                rows = driver.find_elements(By.CSS_SELECTOR, "tbody#mainTab\\:j_id_ku_data tr[data-ri]")
                
                if not rows or len(rows) == 0:
                    print(f"      ✓ Aucun lot trouvé sur cette page")
                    break
                
                # Analyser chaque ligne
                for row in rows:
                    try:
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if len(cells) < 3:
                            continue
                        
                        # Colonne N° (1ère colonne)
                        numero_link = cells[0].find_element(By.TAG_NAME, "a")
                        numero_text = numero_link.text.strip()
                        
                        # Colonne Référence (3ème colonne)
                        reference_text = cells[2].text.strip()
                        
                        # Décider si on doit supprimer ce lot
                        should_delete = False
                        
                        # Règle 1: Garder les lots commençant par "MA" (items bloqués par inspecteur)
                        if reference_text.startswith("MA"):
                            continue
                        
                        # Règle 2: Supprimer si contient la référence de base
                        ref_clean = reference_text.replace("-", "").replace("/", "")
                        if lta_reference_base in ref_clean:
                            should_delete = True
                        
                        # Règle 3: Supprimer si match exact avec référence (avec tirets)
                        ref_with_dash = lta_reference_existing.split('/')[0]
                        if ref_with_dash in reference_text:
                            should_delete = True
                        
                        # Si lot à supprimer trouvé, supprimer IMMÉDIATEMENT et sortir de la boucle
                        if should_delete:
                            print(f"         🎯 Lot à supprimer: N°{numero_text} - {reference_text}")
                            
                            try:
                                print(f"         🗑️  Suppression en cours...")
                                
                                # Cliquer sur le lien N°
                                numero_link.click()
                                time.sleep(2)
                                
                                # Attendre que les boutons d'action apparaissent
                                supprimer_btn = wait.until(
                                    EC.element_to_be_clickable((By.ID, "mainTab:btn_supprimer_lot"))
                                )
                                supprimer_btn.click()
                                print(f"         ✓ Lot N°{numero_text} supprimé")
                                time.sleep(3)  # Attendre que le DOM se rafraîchisse
                                
                                lots_deleted_count += 1
                                lot_deleted_on_this_scan = True
                                
                                # IMPORTANT: Sortir de la boucle FOR pour re-scanner la page
                                break
                                
                            except Exception as e:
                                print(f"         ❌ Erreur suppression lot N°{numero_text}: {e}")
                                # Continuer la recherche
                                continue
                    
                    except Exception as e:
                        print(f"         ⚠️  Erreur analyse ligne: {e}")
                        continue
                
                # Si aucun lot supprimé sur ce scan, vérifier page suivante
                if not lot_deleted_on_this_scan:
                    # Vérifier s'il y a une page suivante
                    try:
                        next_btn = driver.find_element(By.CSS_SELECTOR, "span.ui-paginator-next")
                        
                        # Vérifier si désactivé
                        if "ui-state-disabled" in next_btn.get_attribute("class"):
                            print(f"      ✓ Dernière page atteinte (page {page_number})")
                            break
                        
                        # Cliquer sur suivant
                        next_icon = next_btn.find_element(By.CSS_SELECTOR, "span.ui-icon-seek-next")
                        next_icon.click()
                        time.sleep(2)
                        page_number += 1
                        
                    except:
                        print(f"      ✓ Pagination terminée (page {page_number})")
                        break
                # Sinon, re-scanner la même page (le numéro de page peut avoir changé après suppression)
            
            except Exception as e:
                print(f"      ⚠️  Erreur scan page {page_number}: {e}")
                break
        
        print(f"\n      ✅ {lots_deleted_count} lot(s) conflit(s) supprimé(s)")
        
        # ==================================================================
        # ÉTAPE MED.5: Extraire les DUMs depuis generated_excel
        # ==================================================================
        print("\n   📂 Extraction des DUMs depuis generated_excel...")
        
        # Trouver le fichier generated_excel
        ref_for_filename = lta_reference_existing.split('/')[0]
        generated_excel_path = None
        
        for file in os.listdir(lta_folder_path):
            if file.startswith("generated_excel") and file.endswith(".xlsx"):
                generated_excel_path = os.path.join(lta_folder_path, file)
                break
        
        if not generated_excel_path:
            print(f"      ❌ Fichier generated_excel non trouvé")
            return_to_home_after_error(driver)
            return False
        
        print(f"      ✓ Fichier: {os.path.basename(generated_excel_path)}")
        
        # Lire les DUMs
        try:
            # Fermer le fichier Excel s'il est ouvert
            close_excel_file(generated_excel_path)
            
            wb = load_workbook(generated_excel_path, data_only=True)
            ws = wb['Summary']
            
            dum_lots_data = []
            current_dum = None
            current_p = None
            current_p_brut = None
            
            for row in range(1, 200):
                cell_c = ws.cell(row=row, column=3).value
                cell_a = ws.cell(row=row, column=1).value
                cell_b = ws.cell(row=row, column=2).value
                
                # Détecter DUM
                if cell_c and isinstance(cell_c, str) and "DUM" in cell_c:
                    # Sauvegarder le DUM précédent
                    if current_dum and current_p is not None and current_p_brut is not None:
                        dum_lots_data.append({
                            'dum_name': current_dum,
                            'p': current_p,
                            'p_brut': current_p_brut
                        })
                    
                    current_dum = cell_c.strip()
                    current_p = None
                    current_p_brut = None
                
                # Extraire P et P,BRUT
                if current_dum:
                    if cell_a == "P":
                        current_p = int(float(cell_b)) if cell_b else 0
                    elif cell_a == "P,BRUT":
                        current_p_brut = float(cell_b) if cell_b else 0.0
            
            # Ajouter le dernier DUM
            if current_dum and current_p is not None and current_p_brut is not None:
                dum_lots_data.append({
                    'dum_name': current_dum,
                    'p': current_p,
                    'p_brut': current_p_brut
                })
            
            wb.close()
            
            print(f"      ✓ {len(dum_lots_data)} DUM(s) détecté(s)")
            
        except Exception as e:
            print(f"      ❌ Erreur lecture DUMs: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE MED.6: Créer les nouveaux lots (comme create_etat_depotage)
        # ==================================================================
        print("\n   📦 Création des nouveaux lots...")
        
        for dum_index, dum_data in enumerate(dum_lots_data, start=1):
            print(f"\n   🔹 Création lot {dum_index}/{len(dum_lots_data)} ({dum_data['dum_name']})...")
            
            # MED.6.1: Cliquer sur "Nouveau"
            try:
                nouveau_lot_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_new_lot')]"))
                )
                nouveau_lot_btn.click()
                print(f"      ✓ Bouton 'Nouveau' cliqué")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur clic 'Nouveau': {e}")
                return_to_home_after_error(driver)
                return False
            
            # MED.6.2: Remplir l'en-tête du lot
            
            # Référence: {lta_reference}/N (toujours ajouter /N, ne jamais remplacer)
            try:
                # Toujours ajouter /N à la fin
                # "235-94908936/1" → "235-94908936/1/2" pour DUM 2
                # "23594908936" → "23594908936/2" pour DUM 2
                lot_reference = f"{lta_reference_existing}/{dum_index}"
                
                ref_lot_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'referenceLot_IT_id')]"))
                )
                ref_lot_input.clear()
                ref_lot_input.send_keys(lot_reference)
                print(f"      ✓ Référence: {lot_reference}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie référence: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Ligne dépotée: 1
            try:
                ligne_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'ligneDepotee_IT_id')]"))
                )
                ligne_input.clear()
                ligne_input.send_keys("1")
                print(f"      ✓ Ligne dépotée: 1")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie ligne: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Radio ICE
            try:
                ice_radio = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:entete_section_form:radioChoixDestinataire:1"))
                )
                radio_box = driver.find_element(By.XPATH, "//input[@id='mainTab:detailLot:entete_section_form:radioChoixDestinataire:1']/parent::div/following-sibling::div[@class='ui-radiobutton-box ui-widget ui-corner-all ui-state-default']")
                radio_box.click()
                print(f"      ✓ ICE sélectionné")
                time.sleep(2)
            except Exception as e:
                print(f"      ⚠️  Erreur radio ICE: {e}")
                # Méthode JS alternative
                try:
                    driver.execute_script("""
                        var radio = document.getElementById('mainTab:detailLot:entete_section_form:radioChoixDestinataire:1');
                        radio.checked = true;
                        var event = new Event('change', { bubbles: true });
                        radio.dispatchEvent(event);
                    """)
                    time.sleep(2)
                    print(f"      ✓ ICE sélectionné (JS)")
                except:
                    pass
            
            # Numéro ICE
            try:
                ice_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:entete_section_form:id_ice"))
                )
                wait.until(EC.element_to_be_clickable((By.ID, "mainTab:detailLot:entete_section_form:id_ice")))
                ice_input.clear()
                ice_input.send_keys("000230731000088")
                print(f"      ✓ ICE: 000230731000088")
                
                # Tab pour déclencher validation
                from selenium.webdriver.common.keys import Keys
                ice_input.send_keys(Keys.TAB)
                time.sleep(3)
            except Exception as e:
                print(f"      ❌ Erreur saisie ICE: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Valider en-tête
            try:
                valider_lot_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "mainTab:detailLot:entete_section_form:btn_confirmer_lot"))
                )
                valider_lot_btn.click()
                print(f"      ✓ En-tête validé")
                time.sleep(4)
            except Exception as e:
                print(f"      ❌ Erreur validation en-tête: {e}")
                return_to_home_after_error(driver)
                return False
            
            # MED.6.3: Créer ligne marchandise
            
            # Nouveau ligne
            try:
                nouveau_ligne_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_new_ligne')]"))
                )
                nouveau_ligne_btn.click()
                print(f"      ✓ Bouton 'Nouveau' ligne cliqué")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur 'Nouveau' ligne: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Type contenant: colis
            try:
                contenant_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'typeContenant') and contains(@id, '_input')]"))
                )
                contenant_input.clear()
                contenant_input.send_keys("colis")
                time.sleep(1)
                
                contenant_suggestion = wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "li.ui-autocomplete-item"))
                )
                contenant_suggestion.click()
                print(f"      ✓ Type contenant: colis")
                # Wait for AJAX update after type contenant selection
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur type contenant: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Nombre contenants - wait for element to be fresh after AJAX
            try:
                nombre_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:ligne_section_form:nbrContenants"))
                )
                nombre_input.clear()
                nombre_input.send_keys(str(dum_data['p']))
                print(f"      ✓ Nombre contenants: {dum_data['p']}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur nombre contenants: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Poids brut
            try:
                poids_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:ligne_section_form:poidBru_input"))
                )
                poids_input.clear()
                poids_input.send_keys(str(dum_data['p_brut']))
                print(f"      ✓ Poids brut: {dum_data['p_brut']}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur poids brut: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Marque
            try:
                marque_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:ligne_section_form:marqueLib"))
                )
                marque_input.clear()
                marque_input.send_keys(lta_reference_existing)
                print(f"      ✓ Marque: {lta_reference_existing}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur marque: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Nature marchandise
            try:
                nature_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:ligne_section_form:marchand"))
                )
                nature_input.clear()
                nature_input.send_keys("courrier e-commerce")
                print(f"      ✓ Nature: courrier e-commerce")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur nature: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Code NGP
            try:
                ngp_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:ligne_section_form:ngp"))
                )
                ngp_input.clear()
                ngp_input.send_keys("9999")
                print(f"      ✓ Code NGP: 9999")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur NGP: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Ajouter NGP (>>)
            try:
                ajouter_ngp_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "mainTab:detailLot:ligne_section_form:btn_add_ngp"))
                )
                ajouter_ngp_btn.click()
                print(f"      ✓ NGP ajouté")
                time.sleep(1)
            except Exception as e:
                print(f"      ❌ Erreur ajout NGP: {e}")
                return_to_home_after_error(driver)
                return False
            
            # Valider ligne
            try:
                valider_ligne_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "mainTab:detailLot:ligne_section_form:btn_confirmer_ligne"))
                )
                valider_ligne_btn.click()
                print(f"      ✓ Ligne validée")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur validation ligne: {e}")
                return_to_home_after_error(driver)
                return False
            
            print(f"   ✅ Lot {dum_index} créé!")
        
        print(f"\n   ✅ Tous les lots ({len(dum_lots_data)}) créés!")
        
        # ==================================================================
        # ÉTAPE MED.7: Sauvegarder l'ED modifié
        # ==================================================================
        print("\n   💾 Sauvegarde de l'ED modifié...")
        
        try:
            # Stratégie robuste: chercher par texte puis par ID si échoue
            try:
                sauvegarder_link = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'ui-menuitem-link')]//span[text()='SAUVEGARDER']/parent::a"))
                )
                sauvegarder_link.click()
                print("      ✓ Bouton 'SAUVEGARDER' cliqué")
            except:
                # Fallback: par ID
                sauvegarder_link = wait.until(
                    EC.element_to_be_clickable((By.ID, "secure_174"))
                )
                sauvegarder_link.click()
                print("      ✓ Bouton 'SAUVEGARDER' cliqué (via ID)")
            
            time.sleep(3)  # Attendre la sauvegarde
            print("      ✓ Etat de Dépotage sauvegardé")
            
        except Exception as e:
            print(f"      ❌ Erreur sauvegarde: {e}")
            return_to_home_after_error(driver)
            return False
        
        # MED.7.2: Extraire la référence "sauvegardée" (avant validation)
        # Cette référence sera utilisée en cas d'échec de validation
        sauvegarde_reference = None
        try:
            time.sleep(2)  # Attendre que la référence s'affiche
            
            # Chercher la table de référence
            reference_table = driver.find_element(By.CSS_SELECTOR, "table.reference")
            
            # Extraire les cellules de la deuxième ligne (index 1)
            rows = reference_table.find_elements(By.TAG_NAME, "tr")
            if len(rows) >= 2:
                data_row = rows[1]
                cells = data_row.find_elements(By.TAG_NAME, "td")
                
                if len(cells) >= 5:
                    # Extraire Série (colonne 4, index 3) et Clé (colonne 5, index 4)
                    serie_value = cells[3].text.strip()
                    cle_value = cells[4].text.strip()
                    
                    # Enlever les zéros initiaux de la série
                    serie_clean = str(int(serie_value)) if serie_value.isdigit() else serie_value
                    
                    # Combiner: [Série][Clé]
                    sauvegarde_reference = f"{serie_clean}{cle_value}"
                    
                    print(f"      ✓ Référence sauvegardée extraite: {sauvegarde_reference}")
                    print(f"         (Série={serie_value} → {serie_clean}, Clé={cle_value})")
        except Exception as e:
            print(f"      ⚠️  Impossible d'extraire la référence sauvegardée: {e}")
            print(f"         (Continuera avec extraction après validation)")
            # Continuer quand même - on essaiera après validation
        
        # ==================================================================
        # ÉTAPE MED.8: Valider l'ED modifié
        # ==================================================================
        print("\n   ✅ Validation de l'ED modifié...")
        
        try:
            # Stratégie robuste: chercher par texte puis par ID si échoue
            try:
                valider_link = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'ui-menuitem-link')]//span[text()='VALIDER']/parent::a"))
                )
                valider_link.click()
                print("      ✓ Bouton 'VALIDER' cliqué")
            except:
                # Fallback: par ID
                valider_link = wait.until(
                    EC.element_to_be_clickable((By.ID, "secure_176"))
                )
                valider_link.click()
                print("      ✓ Bouton 'VALIDER' cliqué (via ID)")
            
            time.sleep(4)  # Attendre la validation
            
        except Exception as e:
            print(f"      ❌ Erreur validation: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE MED.9: Vérifier la réponse de validation et extraire référence
        # ==================================================================
        print("\n   🔍 Vérification du résultat de validation...")
        
        # Vérifier messages de succès/erreur
        try:
            time.sleep(2)
            
            # Chercher erreurs
            error_msg = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error-detail, span.ui-messages-error-detail")
            
            if error_msg and len(error_msg) > 0:
                error_text = error_msg[0].text.strip()
                print(f"      ❌ Erreur validation: {error_text}")
                
                # Si erreur et qu'on a la référence sauvegardée, l'enregistrer dans shipper
                if sauvegarde_reference:
                    try:
                        lta_name = os.path.basename(lta_folder_path)
                        parent_dir = os.path.dirname(lta_folder_path)
                        lta_name_with_underscore = lta_name.replace(" ", "_")
                        shipper_pattern = f"{lta_name_with_underscore}_*.txt"
                        shipper_files = glob.glob(os.path.join(parent_dir, shipper_pattern))
                        
                        if shipper_files:
                            shipper_file = shipper_files[0]
                            
                            # Lire le fichier actuel
                            with open(shipper_file, 'r', encoding='utf-8') as f:
                                lines = [line.rstrip('\n') for line in f.readlines()]
                            
                            # S'assurer qu'on a au moins 3 lignes
                            while len(lines) < 3:
                                lines.append("")
                            
                            # Ajouter ou remplacer la ligne 4 avec la référence sauvegardée
                            if len(lines) == 3:
                                lines.append(sauvegarde_reference)
                            elif len(lines) >= 4:
                                lines[3] = sauvegarde_reference
                            
                            # Réécrire le fichier
                            with open(shipper_file, 'w', encoding='utf-8') as f:
                                f.write('\n'.join(lines))
                            
                            print(f"      ✓ Référence sauvegardée écrite dans {os.path.basename(shipper_file)}")
                            print(f"         Ligne 4: {sauvegarde_reference} (NON VALIDÉE - ERREUR)")
                    except Exception as e:
                        print(f"      ⚠️  Impossible de sauvegarder la référence: {e}")
                
                return_to_home_after_error(driver)
                return False
            
            # Chercher succès
            success_msg = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-info-detail, span.ui-messages-info-detail")
            
            if success_msg and len(success_msg) > 0:
                success_text = success_msg[0].text
                print(f"      ✅ Succès: {success_text}")
            else:
                print("      ✅ ED modifié validé avec succès")
            
            # MED.9.1: Extraire la référence validée
            print("      ✓ Extraction de la référence validée...")
            
            try:
                # Chercher la table de référence
                reference_table = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table.reference"))
                )
                
                # Extraire les cellules de la deuxième ligne (index 1)
                rows = reference_table.find_elements(By.TAG_NAME, "tr")
                if len(rows) < 2:
                    print("      ⚠️  Table de référence incomplète")
                    # Utiliser la référence sauvegardée si disponible
                    if sauvegarde_reference:
                        ds_reference = sauvegarde_reference
                        print(f"      ✓ Utilisation référence sauvegardée: {ds_reference}")
                    else:
                        print("      ⚠️  Aucune référence disponible")
                        ds_reference = None
                else:
                    data_row = rows[1]
                    cells = data_row.find_elements(By.TAG_NAME, "td")
                    
                    if len(cells) < 5:
                        print("      ⚠️  Données de référence incomplètes")
                        # Utiliser la référence sauvegardée si disponible
                        if sauvegarde_reference:
                            ds_reference = sauvegarde_reference
                            print(f"      ✓ Utilisation référence sauvegardée: {ds_reference}")
                        else:
                            ds_reference = None
                    else:
                        # Extraire Série (colonne 4, index 3) et Clé (colonne 5, index 4)
                        serie_value = cells[3].text.strip()
                        cle_value = cells[4].text.strip()
                        
                        # Enlever les zéros initiaux de la série
                        serie_clean = str(int(serie_value)) if serie_value.isdigit() else serie_value
                        
                        # Combiner: [Série][Clé]
                        ds_reference = f"{serie_clean}{cle_value}"
                        
                        print(f"      ✓ Référence extraite: Série={serie_value} → {serie_clean}, Clé={cle_value}")
                        print(f"      ✓ Référence DS complète: {ds_reference}")
                
                # MED.9.2: Enregistrer la référence dans le fichier shipper (ligne 4)
                if ds_reference:
                    try:
                        lta_name = os.path.basename(lta_folder_path)
                        parent_dir = os.path.dirname(lta_folder_path)
                        
                        # Chercher le fichier shipper: [X]eme_LTA_*.txt
                        lta_name_with_underscore = lta_name.replace(" ", "_")
                        shipper_pattern = f"{lta_name_with_underscore}_*.txt"
                        shipper_files = glob.glob(os.path.join(parent_dir, shipper_pattern))
                        
                        if not shipper_files:
                            print(f"      ⚠️  Fichier shipper introuvable: {shipper_pattern}")
                        else:
                            shipper_file = shipper_files[0]
                            
                            # Lire le fichier actuel
                            with open(shipper_file, 'r', encoding='utf-8') as f:
                                lines = [line.rstrip('\n') for line in f.readlines()]
                            
                            # S'assurer qu'on a au moins 3 lignes (shipper, serie+cle, location)
                            while len(lines) < 3:
                                lines.append("")
                            
                            # Ajouter ou remplacer la ligne 4 (index 3) avec la référence DS
                            if len(lines) == 3:
                                # Ajouter ligne 4
                                lines.append(ds_reference)
                                action = "ajoutée"
                            elif len(lines) >= 4:
                                # Remplacer ligne 4 existante
                                lines[3] = ds_reference
                                action = "mise à jour"
                            
                            # Réécrire le fichier
                            with open(shipper_file, 'w', encoding='utf-8') as f:
                                f.write('\n'.join(lines))
                            
                            print(f"      ✓ Référence DS {action} dans {os.path.basename(shipper_file)}")
                            print(f"         Ligne 4: {ds_reference}")
                    
                    except Exception as e:
                        print(f"      ⚠️  Erreur mise à jour fichier shipper: {e}")
                
            except Exception as e:
                print(f"      ⚠️  Erreur extraction référence: {e}")
                # Pas critique, continuer
        
        except Exception as e:
            print(f"      ⚠️  Erreur vérification messages: {e}")
        
        # ==================================================================
        # FIN - Retour à l'accueil
        # ==================================================================
        print("\n   🏠 Retour à l'accueil...")
        return_to_home_after_error(driver)
        
        print("\n" + "="*70)
        print("✅ MODIFICATION ED BLOCAGE TERMINÉE AVEC SUCCÈS")
        print("="*70)
        
        return True
    
    except Exception as e:
        print(f"\n❌ ERREUR MODIFICATION ED BLOCAGE: {e}")
        import traceback
        traceback.print_exc()
        return_to_home_after_error(driver)
        return False


def correct_blocage_weights(lta_folder_path, corrected_weight):
    """
    Corrige les poids dans generated_excel et summary_file pour un LTA blocage.
    
    Étapes:
    BC.2.1: Mettre à jour global P,BRUT (B6) dans generated_excel
    BC.2.2: Détecter tous les DUMs dynamiquement
    BC.2.3: Calculer la somme des DUM P,BRUT
    BC.2.4: Ajuster le dernier DUM si nécessaire
    BC.3: Mettre à jour le dernier DUM dans summary_file
    
    Args:
        lta_folder_path: Chemin du dossier LTA
        corrected_weight: Poids corrigé (float)
    
    Returns:
        bool: True si succès, False sinon
    """
    try:
        print(f"\n   📝 Correction des fichiers Excel pour blocage...")
        
        # Trouver le fichier generated_excel
        generated_excel_path = None
        for file in os.listdir(lta_folder_path):
            if file.startswith("generated_excel") and file.endswith(".xlsx"):
                generated_excel_path = os.path.join(lta_folder_path, file)
                break
        
        if not generated_excel_path:
            print(f"      ❌ Fichier generated_excel non trouvé")
            return False
        
        # Trouver le fichier summary_file
        summary_file_path = None
        for file in os.listdir(lta_folder_path):
            if file.startswith("summary_file") and file.endswith(".xlsx"):
                summary_file_path = os.path.join(lta_folder_path, file)
                break
        
        if not summary_file_path:
            print(f"      ❌ Fichier summary_file non trouvé")
            return False
        
        # ========== BC.2: Mise à jour generated_excel ==========
        print(f"\n      📊 Mise à jour generated_excel...")
        
        # Fermer le fichier Excel s'il est ouvert
        close_excel_file(generated_excel_path)
        
        wb = load_workbook(generated_excel_path, data_only=False)
        ws = wb['Summary']
        
        # BC.2.1: Mettre à jour global P,BRUT (B6)
        old_global_pbrut = ws['B6'].value
        ws['B6'] = corrected_weight
        print(f"         ✓ Global P,BRUT (B6): {old_global_pbrut} → {corrected_weight}")
        
        # BC.2.2: Détecter tous les DUMs dynamiquement
        print(f"\n      🔍 Détection des DUMs:")
        
        dum_count = 0
        dum_pbrut_cells = []
        
        row = 11  # Première ligne possible pour DUM 1
        while row < 500:  # Limite de sécurité
            cell_value = ws[f'C{row}'].value
            
            # Vérifier si c'est un header DUM
            if cell_value and isinstance(cell_value, str) and 'DUM' in cell_value.upper():
                dum_count += 1
                pnet_row = row + 3  # P,NET est 3 lignes sous le header
                pbrut_row = row + 4  # P,BRUT est 4 lignes sous le header
                pnet_cell = f'B{pnet_row}'
                pbrut_cell = f'B{pbrut_row}'
                pnet_value = ws[pnet_cell].value
                pbrut_value = ws[pbrut_cell].value
                
                dum_pbrut_cells.append({
                    'dum_number': dum_count,
                    'pnet_cell': pnet_cell,
                    'pbrut_cell': pbrut_cell,
                    'pnet_value': float(pnet_value) if pnet_value else 0.0,
                    'pbrut_value': float(pbrut_value) if pbrut_value else 0.0
                })
                
                print(f"         ✓ DUM {dum_count} trouvé ({pbrut_cell}): P,NET={pnet_value} kg, P,BRUT={pbrut_value} kg")
                
                row += 7  # Passer au prochain DUM potentiel
            else:
                row += 1
                
                # Si on a déjà trouvé des DUMs et qu'on a 10 lignes vides, arrêter
                if dum_count > 0:
                    last_pbrut_row = int(dum_pbrut_cells[-1]['pbrut_cell'][1:])
                    if row > last_pbrut_row and all(ws[f'C{r}'].value is None for r in range(row, min(row + 10, 500))):
                        break
        
        if dum_count == 0:
            print(f"         ❌ Aucun DUM détecté")
            wb.close()
            return False
        
        print(f"         ──────────────────────────────")
        print(f"         📊 Total: {dum_count} DUMs détectés")
        
        # BC.2.3: Calculer la somme des DUM P,BRUT
        dum_sum = sum(dum['pbrut_value'] for dum in dum_pbrut_cells)
        dum_sum = round(dum_sum, 2)
        print(f"         📊 Somme actuelle: {dum_sum} kg")
        
        # BC.2.4: Ajuster les DUMs si nécessaire
        difference = round(dum_sum - corrected_weight, 2)
        
        if abs(difference) < 0.01:
            print(f"\n      ✅ Aucun ajustement nécessaire (différence: {difference} kg)")
        else:
            print(f"\n      ⚙️  Ajustement requis:")
            print(f"         Différence à ajuster: {difference} kg")
            
            if difference > 0:
                # Cas 1: On doit SOUSTRAIRE (dum_sum > corrected_weight)
                print(f"         📉 Soustraction nécessaire")
                
                remaining_to_subtract = difference
                adjusted_dums = []
                
                # Parcourir les DUMs en ordre inverse pour soustraire
                for i in range(len(dum_pbrut_cells) - 1, -1, -1):
                    if remaining_to_subtract <= 0.01:
                        break
                    
                    dum = dum_pbrut_cells[i]
                    # Calculer la marge disponible: P,BRUT - P,NET
                    # IMPORTANT: Garder au moins 1 kg de marge (P,BRUT doit rester > P,NET)
                    available_margin = round(dum['pbrut_value'] - dum['pnet_value'] - 1, 2)
                    
                    if available_margin > 0.01:
                        # On peut soustraire de ce DUM (en gardant P,BRUT > P,NET)
                        can_subtract = min(available_margin, remaining_to_subtract)
                        new_pbrut = round(dum['pbrut_value'] - can_subtract, 2)
                        
                        # Mettre à jour dans Excel
                        ws[dum['pbrut_cell']] = new_pbrut
                        
                        adjusted_dums.append({
                            'dum_number': dum['dum_number'],
                            'cell': dum['pbrut_cell'],
                            'old_value': dum['pbrut_value'],
                            'new_value': new_pbrut,
                            'subtracted': can_subtract
                        })
                        
                        print(f"         ✓ DUM {dum['dum_number']} ({dum['pbrut_cell']}): {dum['pbrut_value']} → {new_pbrut} kg (-{can_subtract} kg)")
                        print(f"            Règle respectée: P,NET={dum['pnet_value']} < P,BRUT={new_pbrut} (marge: {round(new_pbrut - dum['pnet_value'], 2)} kg)")
                        
                        remaining_to_subtract = round(remaining_to_subtract - can_subtract, 2)
                        dum['pbrut_value'] = new_pbrut  # Mettre à jour pour le calcul final
                    else:
                        print(f"         ⚠️  DUM {dum['dum_number']}: Marge insuffisante (P,BRUT={dum['pbrut_value']}, P,NET={dum['pnet_value']}, disponible={max(0, available_margin)} kg)")
                
                if remaining_to_subtract > 0.01:
                    print(f"         ⚠️  ATTENTION: {remaining_to_subtract} kg n'ont pas pu être soustraits (marges insuffisantes)")
                else:
                    print(f"         ✅ Soustraction complète réussie")
                    
            else:
                # Cas 2: On doit AJOUTER (dum_sum < corrected_weight)
                print(f"         📈 Addition nécessaire")
                
                to_add = abs(difference)
                last_dum = dum_pbrut_cells[-1]
                new_pbrut = round(last_dum['pbrut_value'] + to_add, 2)
                
                ws[last_dum['pbrut_cell']] = new_pbrut
                print(f"         ✓ Dernier DUM ajusté ({last_dum['pbrut_cell']}): {last_dum['pbrut_value']} → {new_pbrut} kg (+{to_add} kg)")
                print(f"            Règle respectée: P,NET={last_dum['pnet_value']} < P,BRUT={new_pbrut}")
                
                last_dum['pbrut_value'] = new_pbrut  # Mettre à jour pour le calcul final
            
            # Vérifier la nouvelle somme
            new_sum = sum(dum['pbrut_value'] for dum in dum_pbrut_cells)
            new_sum = round(new_sum, 2)
            print(f"         ✓ Nouvelle somme: {new_sum} kg")
            print(f"         ✓ Objectif: {corrected_weight} kg")
            print(f"         ✓ Écart final: {round(abs(new_sum - corrected_weight), 2)} kg ✅")
        
        # Sauvegarder generated_excel
        wb.save(generated_excel_path)
        wb.close()
        print(f"      ✓ generated_excel sauvegardé")
        
        # ========== BC.3: Mise à jour summary_file ==========
        print(f"\n      📊 Mise à jour summary_file...")
        
        # Fermer le fichier Excel s'il est ouvert
        close_excel_file(summary_file_path)
        
        wb_summary = load_workbook(summary_file_path, data_only=False)
        ws_summary = wb_summary.active
        
        # Mettre à jour TOUS les DUMs ajustés dans summary_file (colonne D)
        # Format summary_file: Row 1 = Header, Row 2 = DUM 1, Row 3 = DUM 2, etc.
        # Colonne D = P,BRUT
        for dum in dum_pbrut_cells:
            summary_row = dum['dum_number'] + 1  # +1 car header en ligne 1
            summary_cell = f'D{summary_row}'
            old_value = ws_summary[summary_cell].value
            new_value = dum['pbrut_value']
            
            # Ne logger que si la valeur a changé
            if old_value != new_value:
                ws_summary[summary_cell] = new_value
                print(f"         ✓ DUM {dum['dum_number']} ({summary_cell}): {old_value} → {new_value}")
        
        # Sauvegarder summary_file
        wb_summary.save(summary_file_path)
        wb_summary.close()
        print(f"      ✓ summary_file sauvegardé")
        
        print(f"\n   ✅ Corrections blocage terminées")
        print(f"   ℹ️  Note: ED existant sera modifié ultérieurement (pas de création)")
        
        return True
    
    except Exception as e:
        print(f"      ❌ Erreur correction blocage: {e}")
        traceback.print_exc()
        return False

def wait_for_ui_blocker_disappear(driver, timeout=10):
    """
    Attend que le blocker UI (overlay) disparaisse avant de continuer.
    
    Args:
        driver: WebDriver Selenium
        timeout: Temps maximum d'attente en secondes (défaut: 10)
    
    Returns:
        bool: True si blocker disparu, False si timeout
    """
    try:
        from selenium.webdriver.support import expected_conditions as EC
        
        # Chercher les éléments blocker communs dans BADR
        blocker_selectors = [
            "div.ui-blockui",
            "div.ui-blockui-content",
            "div[id*='blocker']",
            "div.ui-widget-overlay"
        ]
        
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            blocker_visible = False
            
            for selector in blocker_selectors:
                try:
                    blockers = driver.find_elements(By.CSS_SELECTOR, selector)
                    for blocker in blockers:
                        # Vérifier si le blocker est visible
                        if blocker.is_displayed():
                            blocker_visible = True
                            break
                except:
                    pass
            
            if not blocker_visible:
                return True
            
            time.sleep(0.3)
        
        # Timeout atteint
        return False
        
    except Exception as e:
        # En cas d'erreur, on suppose que le blocker n'est pas là
        return True

def save_dum_error_log(lta_folder_path, lta_name, dum_number, sheet_name, error_exception, error_step, dum_data=None):
    """
    Crée un fichier log détaillé pour un DUM qui a échoué.
    
    Args:
        lta_folder_path: Chemin du dossier LTA
        lta_name: Nom du LTA (ex: "7eme LTA")
        dum_number: Numéro du DUM (1, 2, 3, etc.)
        sheet_name: Nom du sheet (ex: "Sheet 1")
        error_exception: L'exception capturée
        error_step: Description de l'étape où l'erreur s'est produite
        dum_data: Données du DUM (optionnel)
    """
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        error_filename = f"error-dum-processing-{lta_name.replace(' ', '_')}-DUM{dum_number}-{timestamp}.txt"
        error_path = os.path.join(lta_folder_path, error_filename)
        
        with open(error_path, 'w', encoding='utf-8') as f:
            f.write("="*70 + "\n")
            f.write("ERREUR - TRAITEMENT DUM PHASE 2\n")
            f.write("="*70 + "\n\n")
            
            f.write(f"LTA: {lta_name}\n")
            f.write(f"DUM: {dum_number}\n")
            f.write(f"Sheet: {sheet_name}\n")
            f.write(f"Date/Heure: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Étape échouée: {error_step}\n\n")
            
            f.write("DÉTAILS ERREUR:\n")
            f.write("-"*70 + "\n")
            f.write(f"Type: {type(error_exception).__name__}\n")
            f.write(f"Message: {str(error_exception)}\n\n")
            
            if dum_data:
                f.write("DONNÉES DUM:\n")
                f.write("-"*70 + "\n")
                f.write(f"Total Value: {dum_data.get('total_value', 'N/A')}\n")
                f.write(f"Gross Weight: {dum_data.get('total_gross_weight', 'N/A')}\n")
                f.write(f"Positions: {dum_data.get('total_positions', 'N/A')}\n")
                f.write(f"Freight: {dum_data.get('total_freight', 'N/A')}\n")
                f.write(f"Insurance: {dum_data.get('insurance', 'N/A')}\n")
                f.write(f"Cartons: {dum_data.get('cartons', 'N/A')}\n\n")
            
            f.write("ACTION PRISE:\n")
            f.write("-"*70 + "\n")
            f.write("✓ Retour à l'accueil effectué\n")
            f.write("✓ Marqueur \"error\" ajouté à generated_excel\n")
            f.write("⏭️  Traitement continue avec DUM suivant\n\n")
            
            f.write("RECOMMANDATION:\n")
            f.write("-"*70 + "\n")
            f.write("Vérifier manuellement ce DUM et créer la déclaration si nécessaire.\n\n")
            
            f.write("="*70 + "\n")
        
        print(f"      📝 Log d'erreur créé: {error_filename}")
        
    except Exception as e:
        print(f"      ⚠️  Impossible de créer le log d'erreur: {e}")

def mark_dum_as_error_in_excel(lta_folder_path, dum_number, serie=None):
    """
    Marque un DUM comme "error" dans le fichier generated_excel.
    Même logique que save_dum_series_to_excel mais écrit "error" (ou "serie (error)" si série fournie).
    
    Args:
        lta_folder_path: Chemin du dossier LTA
        dum_number: Numéro du DUM (1, 2, 3, etc.)
        serie: Série du DUM (optionnel - si fournie, format: "0159942R (error)")
    """
    try:
        # Trouver le fichier generated_excel
        generated_excel_path = None
        for file in os.listdir(lta_folder_path):
            if file.startswith("generated_excel") and file.endswith(".xlsx"):
                generated_excel_path = os.path.join(lta_folder_path, file)
                break
        
        if not generated_excel_path:
            print(f"      ⚠️  generated_excel introuvable pour marquage erreur")
            return
        
        # Ouvrir le fichier Excel
        # Fermer le fichier Excel s'il est ouvert
        close_excel_file(generated_excel_path)
        
        wb = load_workbook(generated_excel_path, data_only=False)
        ws = wb['Summary']
        
        # Calculer la cellule: C + (12 + (dum_number - 1) * 7)
        row = 12 + (dum_number - 1) * 7
        cell = f'C{row}'
        
        # Vérifier si la cellule contient déjà une valeur (error)
        current_value = ws[cell].value
        if current_value and "(error)" in str(current_value):
            # Déjà marqué avec série, ne pas écraser
            print(f"      ℹ️  Cellule {cell} déjà marquée: {current_value}")
            wb.close()
            return
        
        # Construire la valeur à écrire
        if serie:
            error_value = f"{serie} (error)"
        else:
            error_value = "error"
        
        # Écrire dans la cellule
        ws[cell] = error_value
        
        # Sauvegarder
        wb.save(generated_excel_path)
        wb.close()
        
        print(f"      ✓ Marqueur 'error' ajouté en {cell}: {error_value}")
        
    except Exception as e:
        print(f"      ⚠️  Erreur marquage Excel: {e}")

def return_to_home_after_error(driver):
    """
    Fonction helper pour retourner à l'accueil après une erreur.
    Utilisée dans create_etat_depotage pour nettoyer l'état avant de sortir.
    """
    print("\n   🏠 Retour à l'accueil après erreur...")
    try:
        # Sortir de l'iframe si on est dedans
        try:
            driver.switch_to.default_content()
            print("      ✓ Sorti de l'iframe")
        except:
            pass
        
        # Cliquer sur le bouton Accueil
        try:
            driver.get("https://badr.douane.gov.ma:40444/badr/views/hab/hab_index.xhtml")
            print("      ✓ Navigation directe vers accueil")
            time.sleep(3)
        except Exception as btn_err:
            print(f"      ⚠️  Erreur clic bouton: {btn_err}")
            # Fallback: navigation directe
            try:
                driver.get("https://badr.douane.gov.ma:40444/badr/views/hab/hab_index.xhtml")
                time.sleep(3)
                print("      ✓ Navigation directe vers accueil")
            except Exception as nav_err:
                print(f"      ❌ Erreur navigation: {nav_err}")
    except Exception as e:
        print(f"      ❌ Erreur retour accueil: {e}")

def find_partial_by_number(partial_config, partial_number):
    """
    Finds a partial configuration by its number.
    
    Args:
        partial_config: Full partial configuration dict
        partial_number: Partial number to find (1, 2, etc.)
    
    Returns:
        dict: Partial data or None if not found
    """
    if not partial_config or 'partials' not in partial_config:
        return None
    
    for partial in partial_config['partials']:
        if partial.get('partial_number') == partial_number:
            return partial
    
    return None

def get_dum_lots_for_partial(partial_data, partial_config=None):
    """
    Extracts DUM lot information from partial data for ED creation.
    For exception case, adjusts DUM 1 by subtracting smallest partial values.
    
    Args:
        partial_data: Partial configuration dict containing DUMs
        partial_config: Full partial configuration (optional, for exception handling)
    
    Returns:
        list: List of dicts with dum_name, dum_number, split_id, p (positions), p_brut (weight)
    """
    if not partial_data or 'dums' not in partial_data:
        return []
    
    dum_lots = []
    
    # Check if exception case
    is_exception = False
    smallest_partial_number = None
    smallest_partial_positions = 0
    smallest_partial_weight = 0
    
    if partial_config:
        is_exception = partial_config.get('partial_type') == 'exception'
        smallest_partial_number = partial_config.get('smallest_partial_number')
        smallest_partial_positions = partial_config.get('smallest_partial_positions', 0)
        
        # Get smallest partial weight
        if smallest_partial_number:
            for p in partial_config.get('partials', []):
                if p['partial_number'] == smallest_partial_number:
                    smallest_partial_weight = p['weight']
                    break
    
    for idx, dum in enumerate(partial_data['dums']):
        dum_number = dum['dum_number']
        # For exception partials, use the DUM values from JSON directly (already adjusted)
        dum_weight = float(dum['weight'])
        dum_positions = int(dum['positions'])
        # No further adjustment needed in exception case
        dum_lots.append({
            'dum_name': f"DUM {dum_number}",
            'dum_number': dum_number,
            'split_id': dum.get('split_id', ''),
            'is_split': dum.get('is_split', False),
            'p': dum_positions,
            'p_brut': round(dum_weight, 1)
        })
    return dum_lots

def save_ds_validated_to_partial_config(lta_folder_path, folder_name, partial_number, ds_validated):
    """
    Sauvegarde le DS validé dans le fichier de configuration partielle.
    
    Args:
        lta_folder_path: Chemin du dossier LTA parent
        folder_name: Nom du dossier LTA (ex: "2eme LTA")
        partial_number: Numéro du partiel (1, 2, etc.)
        ds_validated: Référence DS validée (ex: "7732E")
    
    Returns:
        bool: True si succès, False sinon
    """
    try:
        # Construire le chemin du fichier config
        lta_subfolder = os.path.join(lta_folder_path, folder_name)
        config_file = os.path.join(lta_subfolder, f"{folder_name}_partial_config.json")
        
        if not os.path.exists(config_file):
            print(f"      ⚠️  Fichier config partiel introuvable: {config_file}")
            return False
        
        # Charger le config existant
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        # Trouver le partiel et mettre à jour ds_validated
        for partial in config.get('partials', []):
            if partial['partial_number'] == partial_number:
                partial['ds_validated'] = ds_validated
                if 'signed_series' not in partial:
                    partial['signed_series'] = None  # Placeholder pour série signée
                break
        
        # Sauvegarder
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        
        print(f"      ✓ DS validé sauvegardé: {ds_validated}")
        return True
        
    except Exception as e:
        print(f"      ⚠️  Erreur sauvegarde DS validé: {e}")
        return False

def update_signed_series_for_partial(lta_folder_path, folder_name, partial_number, signed_series):
    """
    Met à jour la série signée pour un partiel (à appeler manuellement après signature).
    
    Args:
        lta_folder_path: Chemin du dossier LTA parent
        folder_name: Nom du dossier LTA (ex: "2eme LTA")
        partial_number: Numéro du partiel (1, 2, etc.)
        signed_series: Série signée (ex: "9913 G")
    
    Returns:
        bool: True si succès, False sinon
    """
    try:
        # Construire le chemin du fichier config
        lta_subfolder = os.path.join(lta_folder_path, folder_name)
        config_file = os.path.join(lta_subfolder, f"{folder_name}_partial_config.json")
        
        if not os.path.exists(config_file):
            print(f"⚠️  Fichier config partiel introuvable: {config_file}")
            return False
        
        # Charger le config existant
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        # Trouver le partiel et mettre à jour signed_series
        found = False
        for partial in config.get('partials', []):
            if partial['partial_number'] == partial_number:
                partial['signed_series'] = signed_series
                found = True
                break
        
        if not found:
            print(f"⚠️  Partiel {partial_number} introuvable dans le config")
            return False
        
        # Sauvegarder
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        
        print(f"✓ Série signée sauvegardée pour partiel {partial_number}: {signed_series}")
        return True
        
    except Exception as e:
        print(f"❌ Erreur sauvegarde série signée: {e}")
        return False

def get_dum_preapurement_lots(dum_number, partial_config, validated_lta_reference):
    """
    Determines the Préapurement DS lot structure for a DUM based on partial configuration.
    
    Args:
        dum_number: DUM number (e.g., "1", "2", "3")
        partial_config: Full partial configuration dict (or None for standard LTA)
        validated_lta_reference: Validated LTA reference (without /1, /2 suffix)
    
    Returns:
        list: List of lot dicts with 'reference', 'ds_serie', 'ds_cle', 'split_id'
    """
    if not partial_config:
        # Standard LTA - single lot with default format
        return [{
            'reference': f"{validated_lta_reference}/{dum_number}",
            'ds_serie': None,  # Will use shipper data
            'ds_cle': None,
            'split_id': None
        }]
    
    dum_str = str(dum_number)
    
    # Check for exception case
    is_exception = partial_config.get('partial_type') == 'exception'
    smallest_partial_num = partial_config.get('smallest_partial_number')
    airport_reference = partial_config.get('smallest_partial_airport_reference')
    smallest_partial_positions = partial_config.get('smallest_partial_positions', 0)
    
    # Exception case: DUM 1 has 2 lots
    if is_exception and dum_str == '1':
        lots = []
        
        # Lot 1: Smallest partial (already cleared at airport)
        smallest_partial = find_partial_by_number(partial_config, smallest_partial_num)
        if smallest_partial:
            lots.append({
                'reference': airport_reference,  # Exact airport reference (no suffix)
                'ds_serie': smallest_partial['ds_serie'],
                'ds_cle': smallest_partial['ds_cle'],
                'split_id': None,
                'weight': smallest_partial['weight'],
                'positions': smallest_partial_positions,
                'is_exception_smallest': True  # Flag for special handling
            })
        
        # Lot 2: Largest partial (from état de dépotage)
        # Find which partial DUM 1 belongs to (should be the largest one)
        for partial in partial_config['partials']:
            if partial['partial_number'] != smallest_partial_num:
                for dum in partial['dums']:
                    if dum['dum_number'] == 1:
                        signed_series = partial.get('signed_series', '')
                        if signed_series and ' ' in signed_series:
                            parts = signed_series.split()
                            ds_serie = parts[0]
                            ds_cle = parts[1]
                        else:
                            ds_serie = partial['ds_serie']
                            ds_cle = partial['ds_cle']
                        
                        lots.append({
                            'reference': f"{airport_reference}/1",  # Airport reference + /1
                            'ds_serie': ds_serie,
                            'ds_cle': ds_cle,
                            'split_id': None,
                            'weight': dum['weight'],
                            'positions': dum['positions'],
                            'is_exception_largest': True  # Flag for special handling
                        })
                        break
        
        return lots
    
    # Check if DUM is split
    if dum_str in partial_config.get('split_dums', {}):
        # Split DUM - multiple lots (one for each split)
        split_info = partial_config['split_dums'][dum_str]
        lots = []
        
        for split in split_info['splits']:
            partial_data = find_partial_by_number(partial_config, split['partial'])
            if partial_data:
                # Use signed_series if available (format: "6987 D")
                signed_series = partial_data.get('signed_series', '')
                if signed_series and ' ' in signed_series:
                    parts = signed_series.split()
                    ds_serie = parts[0]
                    ds_cle = parts[1]
                else:
                    # Fallback to original DS if signed not available
                    ds_serie = partial_data['ds_serie']
                    ds_cle = partial_data['ds_cle']
                
                lots.append({
                    'reference': f"{validated_lta_reference}/{split['split_id']}",  # Use split_id for reference
                    'ds_serie': ds_serie,
                    'ds_cle': ds_cle,
                    'split_id': split['split_id'],
                    'weight': split['weight'],
                    'positions': split['positions']
                })
        
        return lots
    else:
        # Normal DUM in partial LTA - find which partial it belongs to
        for partial in partial_config['partials']:
            for dum in partial['dums']:
                if dum['dum_number'] == int(dum_number):
                    # Use signed_series if available (format: "6987 D")
                    signed_series = partial.get('signed_series', '')
                    if signed_series and ' ' in signed_series:
                        parts = signed_series.split()
                        ds_serie = parts[0]
                        ds_cle = parts[1]
                    else:
                        # Fallback to original DS if signed not available
                        ds_serie = partial['ds_serie']
                        ds_cle = partial['ds_cle']
                    
                    return [{
                        'reference': f"{validated_lta_reference}/{dum['dum_number']}",
                        'ds_serie': ds_serie,
                        'ds_cle': ds_cle,
                        'split_id': None,
                        'weight': dum['weight'],
                        'positions': dum['positions']
                    }]
        
        # Fallback - shouldn't happen
        return [{
            'reference': f"{validated_lta_reference}/{dum_number}",
            'ds_serie': None,
            'ds_cle': None,
            'split_id': None
        }]

def create_etat_depotage(driver, lta_folder_path, shipper_data):
    """
    Crée un Etat de Dépotage (Unloading Statement) pour une LTA avec référence DS MEAD.
    
    Args:
        driver: Selenium WebDriver instance
        lta_folder_path: Path to LTA folder
        shipper_data: Dict containing serie, cle, loading_location
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        wait = WebDriverWait(driver, 15)
        
        print("\n" + "="*70)
        print("📦 CRÉATION ETAT DE DÉPOTAGE")
        print("="*70)
        
        # ==================================================================
        # ÉTAPE ED.0: Navigation vers "Etat de Dépotage - Voyage Aérien"
        # ==================================================================
        print("\n📂 Navigation: MISE EN DOUANE → Etat de Dépotage → Voyage Aérien...")
        
        # Ouvrir le menu "MISE EN DOUANE"
        try:
            mise_en_douane_link = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//h3[contains(@class, 'ui-panelmenu-header')]//a[contains(text(), 'MISE EN DOUANE')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", mise_en_douane_link)
            time.sleep(0.5)
            mise_en_douane_link.click()
            print("   ✓ Menu 'MISE EN DOUANE' ouvert")
            time.sleep(2)
        except Exception as e:
            print(f"   ⚠️  Menu 'MISE EN DOUANE' déjà ouvert ou erreur: {e}")
            # Continuer car le menu peut déjà être ouvert
        
        # Ouvrir le sous-menu "Créer une Déclaration" (ID: _151)
        try:
            creer_declaration_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_151"))
            )
            creer_declaration_link.click()
            print("   ✓ Sous-menu 'Créer une Déclaration' ouvert")
            time.sleep(1)
        except Exception as e:
            print(f"   ❌ Erreur ouverture 'Créer une Déclaration': {e}")
            return_to_home_after_error(driver)
            return False
        
        # Ouvrir le sous-menu "Etat de Dépotage" (ID: _236)
        try:
            etat_depotage_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_236"))
            )
            etat_depotage_link.click()
            print("   ✓ Sous-menu 'Etat de Dépotage' ouvert")
            time.sleep(1)
        except Exception as e:
            print(f"   ❌ Erreur ouverture 'Etat de Dépotage': {e}")
            return_to_home_after_error(driver)
            return False
        
        # Cliquer sur "Voyage Aérien" (ID: _247)
        try:
            voyage_aerien_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_247"))
            )
            voyage_aerien_link.click()
            print("   ✓ Lien 'Voyage Aérien' cliqué")
            time.sleep(3)
        except Exception as e:
            print(f"   ❌ Erreur clic 'Voyage Aérien': {e}")
            return_to_home_after_error(driver)
            return False
        
        # Basculer dans l'iframe du formulaire
        try:
            print("   🔄 Basculement vers l'iframe du formulaire...")
            iframe = wait.until(
                EC.presence_of_element_located((By.ID, "iframeMenu"))
            )
            driver.switch_to.frame(iframe)
            print("   ✓ Iframe chargé")
            time.sleep(2)
        except Exception as e:
            print(f"   ❌ Erreur basculement iframe: {e}")
            return False
        
        # ==================================================================
        # ÉTAPE ED.1: Sélection du Bureau "301"
        # ==================================================================
        print("\n   🏢 Sélection du Bureau 301...")
        
        # ED.1.1: Entrer "301" dans l'autocomplete
        try:
            bureau_input = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:bureauCmbId_INPUT_input"))
            )
            bureau_input.clear()
            bureau_input.send_keys("301")
            print("      ✓ Bureau '301' saisi")
            time.sleep(1)
        except Exception as e:
            print(f"      ❌ Erreur saisie bureau: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.1.2: Sélectionner la première suggestion
        try:
            bureau_suggestion = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div#rootForm\\:bureauCmbId_INPUT_panel li.ui-autocomplete-item"))
            )
            bureau_suggestion.click()
            print("      ✓ Suggestion bureau sélectionnée")
            time.sleep(1)
        except Exception as e:
            print(f"      ❌ Erreur sélection suggestion: {e}")
            return False
        
        # ED.1.3: Valider la sélection du bureau
        try:
            valider_bureau_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "rootForm:btnConfirmer"))
            )
            valider_bureau_btn.click()
            print("      ✓ Bureau validé")
            time.sleep(4)  # Attendre le chargement du formulaire suivant
        except Exception as e:
            print(f"      ❌ Erreur validation bureau: {e}")
            return False
        
        # ==================================================================
        # ÉTAPE ED.2: Configuration Type de Déclaration et Référence
        # ==================================================================
        print("\n   📋 Configuration de la déclaration...")
        
        # ED.2.1: Sélectionner "DS MEAD Combinée" (radio button index 3, value "08")
        try:
            # Méthode directe: chercher tous les div.ui-radiobutton-box dans le tableau radioTypeDS
            # et prendre le 4ème (index 3 = DS MEAD Combinée)
            time.sleep(1)
            ds_radios = driver.find_elements(By.CSS_SELECTOR, "table#mainTab\\:form1\\:radioTypeDS div.ui-radiobutton-box")
            if len(ds_radios) >= 4:
                ds_radios[3].click()  # Le 4ème = DS MEAD Combinée
                print("      ✓ 'DS MEAD Combinée' sélectionné")
                time.sleep(0.5)
            else:
                print(f"      ⚠️  Radios DS MEAD insuffisants (trouvé: {len(ds_radios)})")
                # Méthode alternative: JavaScript
                print("      🔄 Tentative avec JavaScript...")
                js_code = """
                var radio = document.getElementById('mainTab:form1:radioTypeDS:3');
                radio.checked = true;
                var event = new Event('change', { bubbles: true });
                radio.dispatchEvent(event);
                """
                driver.execute_script(js_code)
                time.sleep(0.5)
                print("      ✓ 'DS MEAD Combinée' sélectionné via JavaScript")
        except Exception as e:
            print(f"      ❌ Impossible de sélectionner DS MEAD: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.2: Entrer l'année actuelle
        try:
            annee_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:anneeId"))
            )
            annee_input.clear()
            current_year = str(time.strftime("%Y"))
            annee_input.send_keys(current_year)
            print(f"      ✓ Année: {current_year}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie année: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.3: Entrer le numéro de série
        try:
            serie_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:serieId"))
            )
            serie_input.clear()
            serie_input.send_keys(shipper_data['serie'])
            print(f"      ✓ Série: {shipper_data['serie']}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie série: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.4: Entrer la clé
        try:
            cle_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:cleId"))
            )
            cle_input.clear()
            cle_input.send_keys(shipper_data['cle'])
            print(f"      ✓ Clé: {shipper_data['cle']}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie clé: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.5: Entrer la référence LTA
        # Extraire la référence depuis le nom du fichier PDF LTA
        try:
            lta_name = os.path.basename(lta_folder_path)
            lta_pattern = os.path.join(lta_folder_path, f"{lta_name} - *.pdf")
            lta_files = glob.glob(lta_pattern)
            
            if not lta_files:
                print(f"      ❌ Fichier LTA PDF introuvable")
                return_to_home_after_error(driver)
                return False
            
            lta_filename = os.path.basename(lta_files[0])
            # Extraire référence: "9eme LTA - 012-24513105.pdf" → "012-24513105"
            lta_reference_raw = lta_filename.split(" - ")[1].replace(".pdf", "")
            
            # Préparer 3 formats à essayer:
            # Format 1: Avec tirets ET /1 (ex: "235-94908726/1")
            ref_parts = lta_reference_raw.split("-")
            ref_parts[0] = str(int(ref_parts[0]))  # Enlever zéros initiaux
            lta_reference_format1 = "-".join(ref_parts) + "/1"
            
            # Format 2: Sans tirets, sans /1 (ex: "23594908726")
            lta_reference_format2 = lta_reference_raw.replace("-", "")
            lta_reference_format2 = str(int(lta_reference_format2))  # Enlever zéros initiaux
            
            # Format 3: Avec tirets, SANS /1 (ex: "235-94908726")
            lta_reference_format3 = "-".join(ref_parts)
            
            print(f"      📄 Référence LTA brute: {lta_reference_raw}")
            print(f"      📄 Format 1 (avec /1): {lta_reference_format1}")
            print(f"      📄 Format 2 (sans tirets): {lta_reference_format2}")
            print(f"      📄 Format 3 (avec tirets, sans /1): {lta_reference_format3}")
            
            # Essayer d'abord le Format 1 (avec /1)
            reference_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:referenceLotID"))
            )
            reference_input.clear()
            reference_input.send_keys(lta_reference_format1)
            print(f"      ✓ Référence saisie (Format 1): {lta_reference_format1}")
            time.sleep(0.5)
            
        except Exception as e:
            print(f"      ❌ Erreur saisie référence: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.6: Entrer le lieu de chargement (avec autocomplete)
        if shipper_data.get('loading_location'):
            try:
                lieu_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:form1:lieuChargementCmbId_INPUT_input"))
                )
                lieu_input.clear()
                lieu_input.send_keys(shipper_data['loading_location'])
                print(f"      ✓ Lieu de chargement: {shipper_data['loading_location']}")
                time.sleep(1)
                
                # Sélectionner la première suggestion
                lieu_suggestion = wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "div#mainTab\\:form1\\:lieuChargementCmbId_INPUT_panel li.ui-autocomplete-item"))
                )
                lieu_suggestion.click()
                print("      ✓ Suggestion lieu sélectionnée")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ⚠️  Erreur saisie lieu de chargement: {e}")
                # Continuer même si le lieu échoue
        else:
            print("      ⚠️  Pas de lieu de chargement dans le fichier shipper")
        
        # ==================================================================
        # ÉTAPE ED.3: Validation et Gestion des Erreurs
        # ==================================================================
        print("\n   ✅ Validation de l'Etat de Dépotage...")
        
        # ED.3.1: Attendre que le blocker overlay disparaisse complètement
        try:
            # Attendre que tous les blockers UI soient invisibles
            wait.until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
            )
            time.sleep(1)  # Petit délai supplémentaire pour stabilité
        except:
            pass  # Si pas de blocker, continuer
        
        # ED.3.2: Cliquer sur Valider avec retry en cas d'interception
        max_retries = 3
        for attempt in range(max_retries):
            try:
                valider_ref_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "mainTab:form1:confirmerRef"))
                )
                
                # Scroll pour s'assurer que l'élément est visible
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", valider_ref_btn)
                time.sleep(0.5)
                
                # Vérifier une dernière fois que le blocker est invisible
                blockers = driver.find_elements(By.CSS_SELECTOR, "div.ui-blockui[style*='display: block']")
                if blockers:
                    print(f"      ⏳ Blocker UI encore visible, attente... (tentative {attempt + 1}/{max_retries})")
                    time.sleep(2)
                    continue
                
                # Tenter le clic
                valider_ref_btn.click()
                print("      ✓ Bouton 'Valider' cliqué")
                time.sleep(3)
                break  # Succès, sortir de la boucle
                
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"      ⏳ Erreur clic (tentative {attempt + 1}/{max_retries}): {str(e)[:100]}")
                    time.sleep(2)
                else:
                    print(f"      ❌ Erreur clic validation après {max_retries} tentatives: {e}")
                    return_to_home_after_error(driver)
                    return False
        
        # ED.3.2: Vérifier messages d'erreur ou de succès
        error_detected = False
        try:
            # Attendre plus longtemps pour que les messages s'affichent
            time.sleep(2)
            
            # Chercher message d'erreur (plusieurs tentatives)
            error_msg = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error-detail")
            
            # Si pas trouvé, chercher aussi dans span.ui-messages-error-detail
            if not error_msg or len(error_msg) == 0:
                error_msg = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
            
            if error_msg and len(error_msg) > 0:
                error_text = error_msg[0].text.strip()
                print(f"      ⚠️  Erreur de validation détectée: {error_text}")
                error_detected = True
                
                # ED.3.3: Si erreur référence, essayer les autres formats
                if "n'existe pas" in error_text.lower() or "référence" in error_text.lower():
                    print(f"      ⚠️  Format 1 rejeté, tentative Format 2...")
                    
                    # Fermer le message d'erreur
                    try:
                        close_btn = driver.find_element(By.CSS_SELECTOR, "a.ui-messages-close")
                        close_btn.click()
                        time.sleep(0.5)
                        print("      ✓ Message d'erreur fermé")
                    except:
                        pass
                    
                    # Essayer Format 2 (sans tirets)
                    reference_input = wait.until(
                        EC.presence_of_element_located((By.ID, "mainTab:form1:referenceLotID"))
                    )
                    reference_input.clear()
                    time.sleep(0.3)
                    reference_input.send_keys(lta_reference_format2)
                    print(f"      ✓ Référence Format 2 saisie: {lta_reference_format2}")
                    time.sleep(0.5)
                    
                    # Re-valider avec protection anti-interception
                    max_retries = 3
                    for attempt in range(max_retries):
                        try:
                            # Attendre que le blocker disparaisse
                            wait.until(
                                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
                            )
                            time.sleep(0.5)
                            
                            valider_ref_btn = wait.until(
                                EC.element_to_be_clickable((By.ID, "mainTab:form1:confirmerRef"))
                            )
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", valider_ref_btn)
                            time.sleep(0.5)
                            valider_ref_btn.click()
                            print("      ✓ Bouton 'Valider' re-cliqué")
                            time.sleep(3)
                            break
                        except Exception as retry_e:
                            if attempt < max_retries - 1:
                                print(f"      ⏳ Retry {attempt + 1}/{max_retries}...")
                                time.sleep(2)
                            else:
                                raise retry_e
                    
                    # Vérifier résultat Format 2
                    time.sleep(1)
                    error_msg_retry2 = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error-detail")
                    if not error_msg_retry2 or len(error_msg_retry2) == 0:
                        error_msg_retry2 = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                    
                    if error_msg_retry2 and len(error_msg_retry2) > 0:
                        error_text_retry2 = error_msg_retry2[0].text.strip()
                        print(f"      ⚠️  Format 2 rejeté: {error_text_retry2}")
                        print(f"      🔄 Tentative Format 3 (avec tirets, sans /1)...")
                        
                        # Fermer le message d'erreur
                        try:
                            close_btn = driver.find_element(By.CSS_SELECTOR, "a.ui-messages-close")
                            close_btn.click()
                            time.sleep(0.5)
                        except:
                            pass
                        
                        # Essayer Format 3 (avec tirets, sans /1)
                        reference_input = wait.until(
                            EC.presence_of_element_located((By.ID, "mainTab:form1:referenceLotID"))
                        )
                        reference_input.clear()
                        time.sleep(0.3)
                        reference_input.send_keys(lta_reference_format3)
                        print(f"      ✓ Référence Format 3 saisie: {lta_reference_format3}")
                        time.sleep(0.5)
                        
                        # Re-valider avec protection anti-interception
                        max_retries = 3
                        for attempt in range(max_retries):
                            try:
                                wait.until(
                                    EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
                                )
                                time.sleep(0.5)
                                
                                valider_ref_btn = wait.until(
                                    EC.element_to_be_clickable((By.ID, "mainTab:form1:confirmerRef"))
                                )
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", valider_ref_btn)
                                time.sleep(0.5)
                                valider_ref_btn.click()
                                print("      ✓ Bouton 'Valider' re-cliqué")
                                time.sleep(3)
                                break
                            except Exception as retry_e:
                                if attempt < max_retries - 1:
                                    print(f"      ⏳ Retry {attempt + 1}/{max_retries}...")
                                    time.sleep(2)
                                else:
                                    raise retry_e
                        
                        # Vérifier résultat Format 3
                        time.sleep(1)
                        error_msg_retry3 = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error-detail")
                        if not error_msg_retry3 or len(error_msg_retry3) == 0:
                            error_msg_retry3 = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                        
                        if error_msg_retry3 and len(error_msg_retry3) > 0:
                            error_text_retry3 = error_msg_retry3[0].text.strip()
                            print(f"      ❌ Format 3 aussi rejeté: {error_text_retry3}")
                            
                            # Tous les formats ont échoué - créer fichier log
                            lta_name = os.path.basename(lta_folder_path)
                            parent_dir = os.path.dirname(lta_folder_path)
                            lta_name_with_underscore = lta_name.replace(" ", "_")
                            error_log_filename = f"error-creating-ds-depotage-{lta_name_with_underscore}.log"
                            error_log_filepath = os.path.join(parent_dir, error_log_filename)
                            
                            from datetime import datetime
                            current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            
                            with open(error_log_filepath, 'w', encoding='utf-8') as f:
                                f.write(f"ERREUR - Création Etat de Dépotage - Phase 1\n")
                                f.write(f"=" * 70 + "\n\n")
                                f.write(f"LTA: {lta_name}\n")
                                f.write(f"Date: {current_datetime}\n")
                                f.write(f"Étape: Validation de la référence LTA\n\n")
                                f.write(f"TENTATIVES:\n")
                                f.write(f"1. Format avec /1: {lta_reference_format1}\n")
                                f.write(f"   Erreur: {error_text}\n\n")
                                f.write(f"2. Format sans tirets: {lta_reference_format2}\n")
                                f.write(f"   Erreur: {error_text_retry2}\n\n")
                                f.write(f"3. Format avec tirets, sans /1: {lta_reference_format3}\n")
                                f.write(f"   Erreur: {error_text_retry3}\n\n")
                                f.write(f"RÉFÉRENCE BRUTE (PDF):\n")
                                f.write(f"{lta_reference_raw}\n\n")
                                f.write(f"DONNÉES DS MEAD:\n")
                                f.write(f"- Série: {shipper_data['serie']}\n")
                                f.write(f"- Clé: {shipper_data['cle']}\n")
                                if shipper_data.get('loading_location'):
                                    f.write(f"- Lieu de chargement: {shipper_data['loading_location']}\n")
                                f.write(f"\n")
                                f.write(f"MESSAGE:\n")
                                f.write(f"La référence LTA n'a pas pu être validée avec aucun des 3 formats.\n")
                                f.write(f"Vérifiez que:\n")
                                f.write(f"1. La référence LTA dans le nom du fichier PDF est correcte\n")
                                f.write(f"2. Le lot existe bien dans le système BADR\n")
                                f.write(f"3. Le lot n'est pas déjà dépoté\n\n")
                                f.write(f"ACTIONS RECOMMANDÉES:\n")
                                f.write(f"- Vérifier manuellement la référence sur BADR\n")
                                f.write(f"- Corriger le nom du fichier PDF si nécessaire\n")
                                f.write(f"- Créer l'Etat de Dépotage manuellement si l'erreur persiste\n")
                            
                            print(f"      ✓ Fichier log créé: {error_log_filename}")
                            
                            # Fermer l'erreur avant de sortir
                            try:
                                close_btn = driver.find_element(By.CSS_SELECTOR, "a.ui-messages-close")
                                close_btn.click()
                                time.sleep(0.5)
                            except:
                                pass
                            
                            # IMPORTANT: Retourner à l'accueil avant de sortir
                            print("\n      🏠 Retour à l'accueil après erreur...")
                            try:
                                driver.switch_to.default_content()
                                print("         ✓ Sorti de l'iframe")
                                
                                # Cliquer sur le bouton Accueil
                                accueil_btn = WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable((By.ID, "quitter"))
                                )
                                try:
                                    accueil_btn.click()
                                except:
                                    driver.execute_script("arguments[0].click();", accueil_btn)
                                print("         ✓ Bouton 'Accueil' cliqué")
                                time.sleep(3)
                            except Exception as nav_err:
                                print(f"         ⚠️  Erreur navigation: {nav_err}")
                                # Fallback: navigation directe
                                try:
                                    driver.get("https://badr.douane.gov.ma:40444/badr/views/hab/hab_index.xhtml")
                                    time.sleep(3)
                                    print("         ✓ Navigation directe vers accueil")
                                except:
                                    pass
                            
                            return False
                        else:
                            print("      ✅ Format 3 accepté!")
                            error_detected = False
                    else:
                        print("      ✅ Format 2 accepté!")
                        error_detected = False
                else:
                    # Erreur non liée à la référence - créer aussi un log
                    print(f"      ❌ Erreur non gérée: {error_text}")
                    
                    lta_name = os.path.basename(lta_folder_path)
                    parent_dir = os.path.dirname(lta_folder_path)
                    lta_name_with_underscore = lta_name.replace(" ", "_")
                    error_log_filename = f"error-creating-ds-depotage-{lta_name_with_underscore}.log"
                    error_log_filepath = os.path.join(parent_dir, error_log_filename)
                    
                    from datetime import datetime
                    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    with open(error_log_filepath, 'w', encoding='utf-8') as f:
                        f.write(f"ERREUR - Création Etat de Dépotage - Phase 1\n")
                        f.write(f"=" * 70 + "\n\n")
                        f.write(f"LTA: {lta_name}\n")
                        f.write(f"Date: {current_datetime}\n")
                        f.write(f"Étape: Validation de l'Etat de Dépotage\n\n")
                        f.write(f"ERREUR SYSTÈME:\n")
                        f.write(f"{error_text}\n\n")
                        f.write(f"FORMAT RÉFÉRENCE UTILISÉ:\n")
                        f.write(f"Format 1: {lta_reference_format1}\n\n")
                        f.write(f"DONNÉES DS MEAD:\n")
                        f.write(f"- Série: {shipper_data['serie']}\n")
                        f.write(f"- Clé: {shipper_data['cle']}\n")
                        if shipper_data.get('loading_location'):
                            f.write(f"- Lieu de chargement: {shipper_data['loading_location']}\n")
                        f.write(f"\n")
                        f.write(f"TYPE D'ERREUR:\n")
                        f.write(f"Erreur système non liée au format de référence.\n\n")
                        f.write(f"ACTIONS RECOMMANDÉES:\n")
                        f.write(f"- Vérifier les données saisies dans BADR\n")
                        f.write(f"- Consulter la documentation BADR pour ce code d'erreur\n")
                        f.write(f"- Créer l'Etat de Dépotage manuellement\n")
                    
                    print(f"      ✓ Fichier log créé: {error_log_filename}")
                    
                    # IMPORTANT: Retourner à l'accueil avant de sortir
                    print("\n      🏠 Retour à l'accueil après erreur...")
                    try:
                        driver.switch_to.default_content()
                        print("         ✓ Sorti de l'iframe")
                        
                        # Cliquer sur le bouton Accueil
                        accueil_btn = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.ID, "quitter"))
                        )
                        try:
                            accueil_btn.click()
                        except:
                            driver.execute_script("arguments[0].click();", accueil_btn)
                        print("         ✓ Bouton 'Accueil' cliqué")
                        time.sleep(3)
                    except Exception as nav_err:
                        print(f"         ⚠️  Erreur navigation: {nav_err}")
                        # Fallback: navigation directe
                        try:
                            driver.get("https://badr.douane.gov.ma:40444/badr/views/hab/hab_index.xhtml")
                            time.sleep(3)
                            print("         ✓ Navigation directe vers accueil")
                        except:
                            pass
                    
                    return False
            
            # ED.3.4: Vérifier message de succès
            success_msg = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-info-detail")
            
            if success_msg and len(success_msg) > 0:
                success_text = success_msg[0].text
                if "confirmées" in success_text.lower():
                    print(f"      ✓ Succès: {success_text}")
                else:
                    print(f"      ⚠️  Message inattendu: {success_text}")
            
            # Si aucune erreur n'a été détectée, c'est bon
            if not error_detected:
                print("      ✓ Validation terminée avec succès")
                # Sauvegarder la référence LTA validée pour utilisation ultérieure (création lots)
                # IMPORTANT: Utiliser le format EXACT qui a été accepté par BADR
                # Si la référence contient /1 (ex: 112-00181440/1), on garde le /1
                # Si la référence ne contient pas de / (ex: 11200181440), on la garde telle quelle
                # Pour savoir quel format a été accepté, on vérifie quelle valeur est dans le champ
                try:
                    reference_input_value = driver.find_element(By.ID, "mainTab:form1:referenceLotID").get_attribute("value")
                    # Garder la référence EXACTE (avec /1 si présent)
                    lta_reference_clean = reference_input_value.strip()
                    print(f"      📋 Référence validée à utiliser pour les lots: {lta_reference_clean}")
                    
                    # ED.3.5: Enregistrer la référence validée dans le fichier shipper (ligne 5)
                    try:
                        lta_name = os.path.basename(lta_folder_path)
                        parent_dir = os.path.dirname(lta_folder_path)
                        lta_name_with_underscore = lta_name.replace(" ", "_")
                        shipper_pattern = f"{lta_name_with_underscore}_*.txt"
                        shipper_files = glob.glob(os.path.join(parent_dir, shipper_pattern))
                        
                        if shipper_files:
                            shipper_file = shipper_files[0]
                            
                            # Lire le fichier actuel
                            with open(shipper_file, 'r', encoding='utf-8') as f:
                                lines = [line.rstrip('\n') for line in f.readlines()]
                            
                            # S'assurer qu'on a au moins 4 lignes (ou ajouter des lignes vides si nécessaire)
                            while len(lines) < 4:
                                lines.append("")
                            
                            # Ajouter ou remplacer la ligne 5 (index 4) avec la référence LTA validée
                            if len(lines) == 4:
                                lines.append(lta_reference_clean)
                            else:
                                lines[4] = lta_reference_clean
                            
                            # Réécrire le fichier
                            with open(shipper_file, 'w', encoding='utf-8') as f:
                                f.write('\n'.join(lines))
                            
                            print(f"      ✓ Référence LTA sauvegardée dans {os.path.basename(shipper_file)} (ligne 5)")
                            print(f"         Ligne 5: {lta_reference_clean}")
                        else:
                            print(f"      ⚠️  Fichier shipper introuvable pour sauvegarde: {shipper_pattern}")
                    
                    except Exception as e:
                        print(f"      ⚠️  Erreur sauvegarde référence LTA dans shipper: {e}")
                        # Continuer quand même
                    
                except:
                    # Fallback: utiliser le format 1 COMPLET (avec /1 si présent)
                    lta_reference_clean = lta_reference_format1
            else:
                print("      ❌ Erreur non résolue détectée")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
        except Exception as e:
            print(f"      ⚠️  Erreur lors de la vérification des messages: {e}")
        
        # ==================================================================
        # ÉTAPE ED.4: Naviguer vers l'onglet "Quantités"
        # ==================================================================
        print("\n   📊 Navigation vers l'onglet Quantités...")
        
        # ED.4.1: Fermer tous les messages (erreur et info) avant navigation
        try:
            # Fermer les messages d'erreur
            close_btns = driver.find_elements(By.CSS_SELECTOR, "a.ui-messages-close")
            for btn in close_btns:
                try:
                    btn.click()
                    time.sleep(0.3)
                except:
                    pass
            
            # Fermer aussi les messages d'info qui peuvent intercepter le clic
            info_messages = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-info")
            for msg in info_messages:
                try:
                    close_btn = msg.find_element(By.CSS_SELECTOR, "a.ui-messages-close")
                    close_btn.click()
                    time.sleep(0.3)
                except:
                    # Si pas de bouton close, essayer de cliquer directement sur le message pour le masquer
                    try:
                        driver.execute_script("arguments[0].style.display = 'none';", msg)
                    except:
                        pass
            
            # Attendre que les messages disparaissent
            time.sleep(0.5)
            
            # Vérifier qu'il n'y a plus de messages visibles qui bloquent
            try:
                visible_messages = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-info, div.ui-messages-error")
                for msg in visible_messages:
                    if msg.is_displayed():
                        try:
                            driver.execute_script("arguments[0].style.display = 'none';", msg)
                        except:
                            pass
            except:
                pass
        except:
            pass
        
        # ED.4.2: Attendre que le blocker UI disparaisse
        wait_for_ui_blocker_disappear(driver, timeout=10)
        time.sleep(0.5)
        
        # ED.4.3: Cliquer sur l'onglet Quantités avec retry et JavaScript fallback
        max_retries = 3
        quantites_navigation_success = False
        
        for attempt in range(max_retries):
            try:
                if attempt > 1:
                    print(f"      🔄 Tentative {attempt}/3...")
                    time.sleep(1)
                    wait_for_ui_blocker_disappear(driver, timeout=5)
                
                quantites_tab = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "a[href='#mainTab:tab3']"))
                )
                
                # Scroller pour s'assurer que l'élément est visible
                driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", quantites_tab)
                time.sleep(0.3)
                
                # Attendre que le blocker disparaisse avant de cliquer
                wait_for_ui_blocker_disappear(driver, timeout=5)
                
                # Essayer clic normal puis JavaScript
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab3']")))
                    quantites_tab.click()
                    print("      ✓ Onglet Quantités ouvert")
                except Exception as click_err:
                    print(f"      ⚠️  Clic normal échoué, utilisation JavaScript: {click_err}")
                    driver.execute_script("arguments[0].click();", quantites_tab)
                    print("      ✓ Onglet Quantités ouvert (JavaScript)")
                
                time.sleep(2)
                
                # Attendre que le blocker disparaisse après le clic
                wait_for_ui_blocker_disappear(driver, timeout=5)
                
                # Vérifier que l'onglet est bien actif
                try:
                    active_tab = driver.find_element(By.CSS_SELECTOR, "li.ui-tabs-selected.ui-state-active a[href='#mainTab:tab3']")
                    if active_tab:
                        quantites_navigation_success = True
                        break
                except:
                    # Vérifier alternative: le panneau Quantités est visible
                    try:
                        quantites_panel = driver.find_element(By.ID, "mainTab:tab3")
                        if quantites_panel.is_displayed():
                            quantites_navigation_success = True
                            break
                    except:
                        pass
                
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"      ⚠️  Tentative {attempt + 1}/{max_retries} échouée, retry...")
                    time.sleep(1)
                    # Retry avec JavaScript
                    try:
                        quantites_tab = driver.find_element(By.CSS_SELECTOR, "a[href='#mainTab:tab3']")
                        driver.execute_script("arguments[0].click();", quantites_tab)
                        print("      ✓ Onglet Quantités ouvert (JavaScript retry)")
                        time.sleep(2)
                        wait_for_ui_blocker_disappear(driver, timeout=5)
                        quantites_navigation_success = True
                        break
                    except:
                        pass
                else:
                    # Last attempt failed
                    print(f"      ❌ Erreur navigation onglet Quantités après {max_retries} tentatives: {e}")
                    # Dernière tentative avec JavaScript direct
                    try:
                        print("      🔄 Dernière tentative avec JavaScript direct...")
                        driver.execute_script("""
                            var tab = document.querySelector('a[href="#mainTab:tab3"]');
                            if (tab) {
                                tab.click();
                            }
                        """)
                        time.sleep(2)
                        wait_for_ui_blocker_disappear(driver, timeout=5)
                        # Vérifier si ça a marché
                        quantites_panel = driver.find_element(By.ID, "mainTab:tab3")
                        if quantites_panel.is_displayed():
                            print("      ✓ Onglet Quantités activé via JavaScript")
                            quantites_navigation_success = True
                        else:
                            driver.switch_to.default_content()
                            return_to_home_after_error(driver)
                            return False
                    except:
                        driver.switch_to.default_content()
                        return_to_home_after_error(driver)
                        return False
        
        if not quantites_navigation_success:
            print("      ❌ Échec navigation vers 'Quantités'")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE ED.5: Extraire les totaux depuis generated_excel
        # ==================================================================
        print("\n   📂 Extraction des totaux depuis generated_excel...")
        
        # ED.5.1: Trouver le fichier generated_excel
        # Utiliser la référence validée (lta_reference_clean)
        # Si le format validé est "235-94908726" (sans /1), on l'utilise tel quel
        # Si le format validé est "235-94908726/1", on enlève le /1
        ref_for_filename = lta_reference_clean.split('/')[0]  # Enlever le /1 si présent
        generated_excel_pattern = f"generated_excel - {ref_for_filename}.xlsx"
        generated_excel_path = None
        
        for file in os.listdir(lta_folder_path):
            if file.startswith("generated_excel") and file.endswith(".xlsx"):
                generated_excel_path = os.path.join(lta_folder_path, file)
                break
        
        if not generated_excel_path:
            print(f"      ❌ Fichier generated_excel non trouvé dans {lta_folder_path}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        print(f"      ✓ Fichier trouvé: {os.path.basename(generated_excel_path)}")
        
        # ED.5.2: Lire et extraire les totaux (avant "FOURNISSEUR")
        try:
            # Fermer le fichier Excel s'il est ouvert
            close_excel_file(generated_excel_path)
            
            wb = load_workbook(generated_excel_path, data_only=True)
            ws = wb['Summary']
            
            total_p = None
            total_p_brut = None
            
            for row in range(1, 20):
                cell_a = ws.cell(row=row, column=1).value
                
                if cell_a == "FOURNISSEUR":
                    break
                
                if cell_a == "P":
                    total_p = ws.cell(row=row, column=2).value
                elif cell_a == "P,BRUT":
                    total_p_brut = ws.cell(row=row, column=2).value
            
            wb.close()
            
            if total_p is None or total_p_brut is None:
                print("      ❌ Impossible de trouver les totaux P et P,BRUT")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # Convertir: P (contenants) = entier, P,BRUT (poids) = float
            total_p = int(float(total_p))  # Contenants = nombre entier (pas de décimales)
            total_p_brut = float(total_p_brut)  # Poids brut = garder les décimales (ex: 1419.50)
            
            print(f"      ✓ Total P (contenants): {total_p}")
            print(f"      ✓ Total P,BRUT (poids brut): {total_p_brut}")
            
        except Exception as e:
            print(f"      ❌ Erreur lecture generated_excel: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE ED.5.3: Validation - Vérifier somme des DUMs
        # ==================================================================
        print("\n   🔍 Validation des totaux (somme DUMs)...")
        
        try:
            # Fermer le fichier Excel s'il est ouvert
            close_excel_file(generated_excel_path)
            
            wb = load_workbook(generated_excel_path, data_only=True)
            ws = wb['Summary']
            
            calculated_p = 0  # Contenants = entier
            calculated_p_brut = 0.0  # Poids brut = float
            dum_details = []
            
            current_dum = None
            current_p = None
            current_p_brut = None
            
            for row in range(1, 200):
                cell_c = ws.cell(row=row, column=3).value
                cell_a = ws.cell(row=row, column=1).value
                cell_b = ws.cell(row=row, column=2).value
                
                # Détecter DUM
                if cell_c and isinstance(cell_c, str) and "DUM" in cell_c:
                    # Sauvegarder le DUM précédent
                    if current_dum and current_p is not None and current_p_brut is not None:
                        dum_details.append({
                            'dum': current_dum,
                            'p': current_p,
                            'p_brut': current_p_brut
                        })
                        calculated_p += current_p
                        calculated_p_brut += current_p_brut
                    
                    current_dum = cell_c.strip()
                    current_p = None
                    current_p_brut = None
                
                # Extraire P et P,BRUT
                if current_dum:
                    if cell_a == "P":
                        # P = nombre de contenants (entier)
                        current_p = int(float(cell_b)) if cell_b else 0
                    elif cell_a == "P,BRUT":
                        # P,BRUT = poids brut (garder les décimales)
                        current_p_brut = float(cell_b) if cell_b else 0.0
            
            # Ajouter le dernier DUM
            if current_dum and current_p is not None and current_p_brut is not None:
                dum_details.append({
                    'dum': current_dum,
                    'p': current_p,
                    'p_brut': current_p_brut
                })
                calculated_p += current_p
                calculated_p_brut += current_p_brut
            
            wb.close()
            
            print(f"      Totaux calculés: P={calculated_p}, P,BRUT={calculated_p_brut}")
            print(f"      Totaux déclarés: P={total_p}, P,BRUT={total_p_brut}")
            
            # Vérifier correspondance
            if calculated_p != total_p or calculated_p_brut != total_p_brut:
                print("      ❌ ERREUR: Les totaux ne correspondent pas!")
                
                # Créer fichier d'erreur
                lta_name = os.path.basename(lta_folder_path)
                error_filename = f"{lta_name}_ERROR.txt"
                error_filepath = os.path.join(os.path.dirname(lta_folder_path), error_filename)
                
                from datetime import datetime
                current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                with open(error_filepath, 'w', encoding='utf-8') as f:
                    f.write(f"ERREUR DE CALCUL - {lta_name}\n\n")
                    f.write(f"LTA: {lta_name} - {lta_reference.split('/')[0]}\n")
                    f.write(f"Date: {current_datetime}\n\n")
                    f.write("TOTAUX DÉCLARÉS:\n")
                    f.write(f"- P (Total): {total_p}\n")
                    f.write(f"- P,BRUT (Total): {total_p_brut}\n\n")
                    f.write("TOTAUX CALCULÉS:\n")
                    f.write(f"- P (Somme DUMs): {calculated_p}\n")
                    f.write(f"- P,BRUT (Somme DUMs): {calculated_p_brut}\n\n")
                    f.write("DÉTAIL PAR DUM:\n")
                    for dum in dum_details:
                        f.write(f"{dum['dum']}: P={dum['p']}, P,BRUT={dum['p_brut']}\n")
                    f.write("\nMESSAGE: Les totaux déclarés ne correspondent pas à la somme des DUMs.\n")
                    f.write("Veuillez vérifier le fichier generated_excel.\n")
                
                print(f"      ✓ Fichier d'erreur créé: {error_filename}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            print("      ✅ Validation réussie: Les totaux correspondent")
            
        except Exception as e:
            print(f"      ❌ Erreur lors de la validation: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE ED.6: Entrer le poids brut total
        # ==================================================================
        print("\n   ⚖️  Saisie du poids brut total...")
        
        try:
            poids_brut_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form3:poidsBrutTotal_IT_id_input"))
            )
            poids_brut_input.clear()
            poids_brut_input.send_keys(str(total_p_brut))
            print(f"      ✓ Poids brut total saisi: {total_p_brut}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie poids brut: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE ED.7: Entrer le nombre total de contenants
        # ==================================================================
        print("\n   📦 Saisie du nombre total de contenants...")
        
        try:
            nombre_contenants_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form3:nombreContenantTotal_IT_id"))
            )
            nombre_contenants_input.clear()
            nombre_contenants_input.send_keys(str(total_p))
            print(f"      ✓ Nombre de contenants saisi: {total_p}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie nombre contenants: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE ED.8: Naviguer vers l'onglet "LTA"
        # ==================================================================
        print("\n   📄 Navigation vers l'onglet LTA...")
        
        try:
            lta_tab = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab4']"))
            )
            lta_tab.click()
            print("      ✓ Onglet LTA ouvert")
            time.sleep(2)
        except Exception as e:
            print(f"      ❌ Erreur navigation onglet LTA: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE ED.9-ED.12: Créer les lots pour chaque DUM
        # ==================================================================
        print("\n   📦 Création des lots pour chaque DUM...")
        
        # Lire les données des DUMs depuis generated_excel
        try:
            # Fermer le fichier Excel s'il est ouvert
            close_excel_file(generated_excel_path)
            
            wb = load_workbook(generated_excel_path, data_only=True)
            ws = wb['Summary']
            
            # Collecter tous les DUMs avec leurs données
            dum_lots_data = []
            current_dum = None
            current_p = None
            current_p_brut = None
            
            for row in range(1, 200):
                cell_c = ws.cell(row=row, column=3).value
                cell_a = ws.cell(row=row, column=1).value
                cell_b = ws.cell(row=row, column=2).value
                
                # Détecter DUM
                if cell_c and isinstance(cell_c, str) and "DUM" in cell_c:
                    # Sauvegarder le DUM précédent
                    if current_dum and current_p is not None and current_p_brut is not None:
                        dum_lots_data.append({
                            'dum_name': current_dum,
                            'p': current_p,
                            'p_brut': current_p_brut
                        })
                    
                    current_dum = cell_c.strip()
                    current_p = None
                    current_p_brut = None
                
                # Extraire P et P,BRUT
                if current_dum:
                    if cell_a == "P":
                        current_p = int(float(cell_b)) if cell_b else 0
                    elif cell_a == "P,BRUT":
                        current_p_brut = int(float(cell_b)) if cell_b else 0
            
            # Ajouter le dernier DUM
            if current_dum and current_p is not None and current_p_brut is not None:
                dum_lots_data.append({
                    'dum_name': current_dum,
                    'p': current_p,
                    'p_brut': current_p_brut
                })
            
            wb.close()
            
            print(f"      ✓ {len(dum_lots_data)} DUMs détectés pour création de lots")
            
            # GESTION SPÉCIALE: Si 1 seul DUM, créer 2 lots (BADR n'accepte pas 1 seul lot)
            if len(dum_lots_data) == 1:
                print(f"\n      ⚠️  DÉTECTION: 1 seul DUM - création de 2 lots pour conformité BADR")
                original_dum = dum_lots_data[0]
                
                # Diviser poids et contenants par 2
                p_half = original_dum['p'] // 2
                p_remaining = original_dum['p'] - p_half
                p_brut_half = original_dum['p_brut'] // 2
                p_brut_remaining = original_dum['p_brut'] - p_brut_half
                
                # Créer 2 lots
                dum_lots_data = [
                    {
                        'dum_name': f"{original_dum['dum_name']} (Lot 1/2)",
                        'p': p_half,
                        'p_brut': p_brut_half,
                        'is_split': True,
                        'split_index': 1
                    },
                    {
                        'dum_name': f"{original_dum['dum_name']} (Lot 2/2)",
                        'p': p_remaining,
                        'p_brut': p_brut_remaining,
                        'is_split': True,
                        'split_index': 2
                    }
                ]
                print(f"      ✓ Division: Lot 1 ({p_half} contenants, {p_brut_half} kg) + Lot 2 ({p_remaining} contenants, {p_brut_remaining} kg)")
            
        except Exception as e:
            print(f"      ❌ Erreur lecture DUMs depuis generated_excel: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # Créer un lot pour chaque DUM
        for dum_index, dum_data in enumerate(dum_lots_data, start=1):
            print(f"\n   🔹 Création du lot {dum_index}/{len(dum_lots_data)} ({dum_data['dum_name']})...")
            
            # ==================================================================
            # ÉTAPE ED.10.1: Cliquer sur "Nouveau" pour créer un lot
            # ==================================================================
            try:
                nouveau_lot_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_new_lot')]"))
                )
                nouveau_lot_btn.click()
                print(f"      ✓ Bouton 'Nouveau' lot cliqué")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur clic 'Nouveau' lot: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ==================================================================
            # ÉTAPE ED.10.2: Remplir les informations d'en-tête du lot
            # ==================================================================
            
            # ED.10.2a: Référence du lot (LTA ref + /N)
            try:
                # Construire la référence: ajouter /dum_index à la référence LTA validée
                # Tous les cas: simplement ajouter /N à la fin
                # 1. Si format "235-94908726/1" → "235-94908726/1/2" pour DUM 2
                # 2. Si format "23594908726" → "23594908726/2" pour DUM 2
                # 3. Si format "235-94908726" → "235-94908726/2" pour DUM 2
                
                # Toujours ajouter /N à la fin (ne jamais remplacer)
                # Si lot splitté, utiliser split_index au lieu de dum_index
                if dum_data.get('is_split', False):
                    lot_reference = f"{lta_reference_clean}/{dum_data['split_index']}"
                else:
                    lot_reference = f"{lta_reference_clean}/{dum_index}"
                
                ref_lot_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'referenceLot_IT_id')]"))
                )
                ref_lot_input.clear()
                ref_lot_input.send_keys(lot_reference)
                print(f"      ✓ Référence lot: {lot_reference}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie référence lot: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.10.2b: Ligne dépotée (toujours 1)
            try:
                ligne_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'ligneDepotee_IT_id')]"))
                )
                ligne_input.clear()
                ligne_input.send_keys("1")
                print(f"      ✓ Ligne dépotée: 1")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie ligne dépotée: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.10.2c: Sélectionner le radio button ICE (valeur 02)
            try:
                # Utiliser l'ID exact du radio button ICE (index 1, value 02)
                ice_radio = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:entete_section_form:radioChoixDestinataire:1"))
                )
                # Cliquer sur la div.ui-radiobutton-box associée
                radio_box = driver.find_element(By.XPATH, "//input[@id='mainTab:detailLot:entete_section_form:radioChoixDestinataire:1']/parent::div/following-sibling::div[@class='ui-radiobutton-box ui-widget ui-corner-all ui-state-default']")
                radio_box.click()
                print(f"      ✓ Option ICE sélectionnée")
                time.sleep(2)  # Attendre que la page se stabilise après le clic
            except Exception as e:
                print(f"      ⚠️  Erreur sélection radio ICE (méthode 1): {e}")
                # Méthode alternative: JavaScript
                try:
                    print(f"      🔄 Tentative avec JavaScript...")
                    js_code = """
                    var radio = document.getElementById('mainTab:detailLot:entete_section_form:radioChoixDestinataire:1');
                    if (radio) {
                        radio.checked = true;
                        var event = new Event('change', { bubbles: true });
                        radio.dispatchEvent(event);
                    }
                    """
                    driver.execute_script(js_code)
                    time.sleep(2)  # Attendre que la page se stabilise
                    print(f"      ✓ Option ICE sélectionnée via JavaScript")
                except Exception as e2:
                    print(f"      ❌ Erreur sélection radio ICE: {e2}")
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
            
            # Attendre que le blocker UI disparaisse après sélection ICE
            try:
                WebDriverWait(driver, 5).until(
                    EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
                )
                print(f"      ✓ Page stabilisée après sélection ICE")
            except:
                pass  # Pas de blocker, continuer
            
            # ED.10.2d: Numéro ICE (constant)
            try:
                # Re-chercher l'élément ICE pour éviter stale element
                ice_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:entete_section_form:id_ice"))
                )
                # Vérifier que l'élément est bien interactif
                wait.until(EC.element_to_be_clickable((By.ID, "mainTab:detailLot:entete_section_form:id_ice")))
                
                ice_input.clear()
                ice_input.send_keys("000230731000088")
                print(f"      ✓ ICE: 000230731000088")
                
                # Unfocus l'input pour déclencher la validation (Tab ou clic ailleurs)
                print(f"      ⏳ Déclenchement de la validation ICE...")
                from selenium.webdriver.common.keys import Keys
                ice_input.send_keys(Keys.TAB)  # Simuler Tab pour sortir du champ
                time.sleep(1)
                
                # Attendre que le système charge les informations ICE
                print(f"      ⏳ Attente du chargement des informations ICE...")
                time.sleep(3)  # Le système charge les infos après saisie ICE
                
                # Attendre que le blocker UI disparaisse si présent
                try:
                    WebDriverWait(driver, 5).until(
                        EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
                    )
                    print(f"      ✓ Informations ICE chargées")
                except:
                    pass  # Pas de blocker, continuer
                    
            except Exception as e:
                print(f"      ❌ Erreur saisie ICE: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ==================================================================
            # ÉTAPE ED.10.3: Valider l'en-tête du lot
            # ==================================================================
            try:
                # Utiliser l'ID exact du bouton Valider
                valider_lot_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "mainTab:detailLot:entete_section_form:btn_confirmer_lot"))
                )
                valider_lot_btn.click()
                print(f"      ✓ En-tête lot validé")
                
                # Attendre que la validation soit traitée et que la page soit prête
                print(f"      ⏳ Attente du traitement de la validation...")
                time.sleep(4)  # Augmenté à 4 secondes
                
                # Attendre que le blocker UI disparaisse si présent
                try:
                    WebDriverWait(driver, 5).until(
                        EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
                    )
                    print(f"      ✓ Validation traitée")
                except:
                    pass  # Pas de blocker, continuer
                
            except Exception as e:
                print(f"      ❌ Erreur validation en-tête lot: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ==================================================================
            # ÉTAPE ED.11.1: Cliquer sur "Nouveau" pour créer une ligne
            # ==================================================================
            try:
                print(f"      🔍 Recherche du bouton 'Nouveau' ligne...")
                nouveau_ligne_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_new_ligne')]"))
                )
                nouveau_ligne_btn.click()
                print(f"      ✓ Bouton 'Nouveau' ligne cliqué")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur clic 'Nouveau' ligne: {e}")
                print(f"      🔍 Tentative de recherche alternative...")
                try:
                    # Méthode alternative: chercher par texte visible
                    nouveau_ligne_btn_alt = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Nouveau') or contains(@value, 'Nouveau')]"))
                    )
                    driver.execute_script("arguments[0].click();", nouveau_ligne_btn_alt)
                    print(f"      ✓ Bouton 'Nouveau' ligne cliqué (méthode alternative)")
                    time.sleep(2)
                except Exception as e2:
                    print(f"      ❌ Erreur clic 'Nouveau' ligne (alternative): {e2}")
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
            
            # ==================================================================
            # ÉTAPE ED.11.2: Remplir le formulaire de ligne marchandise
            # ==================================================================
            
            # ED.11.2a: Type Contenant (autocomplete "colis")
            try:
                type_contenant_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'typeContenantId_INPUT_input')]"))
                )
                type_contenant_input.clear()
                type_contenant_input.send_keys("colis")
                print(f"      ✓ Type contenant: colis (recherche...)")
                time.sleep(2)
                
                # Sélectionner la première suggestion "COLIS(216)"
                colis_suggestion = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//li[contains(@data-item-label, 'COLIS(216)')]"))
                )
                colis_suggestion.click()
                print(f"      ✓ COLIS(216) sélectionné")
                time.sleep(1)
            except Exception as e:
                print(f"      ❌ Erreur sélection type contenant: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.11.2b: Nombre de contenants (P du DUM)
            try:
                nbr_contenants_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'nbrContenants')]"))
                )
                nbr_contenants_input.clear()
                nbr_contenants_input.send_keys(str(dum_data['p']))
                print(f"      ✓ Nombre contenants: {dum_data['p']}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie nombre contenants: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.11.2c: Poids brut (P,BRUT du DUM)
            try:
                poids_brut_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'poidBru_input')]"))
                )
                poids_brut_input.clear()
                poids_brut_input.send_keys(str(dum_data['p_brut']))
                print(f"      ✓ Poids brut: {dum_data['p_brut']}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie poids brut: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.11.2d: Marque (référence LTA validée)
            try:
                marque_textarea = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//textarea[contains(@name, 'marqueLib')]"))
                )
                marque_textarea.clear()
                marque_textarea.send_keys(lta_reference_clean)
                print(f"      ✓ Marque: {lta_reference_clean}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie marque: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.11.2e: Nature marchandise (constant)
            try:
                nature_textarea = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//textarea[contains(@name, 'marchand')]"))
                )
                nature_textarea.clear()
                nature_textarea.send_keys("courrier e-commerce")
                print(f"      ✓ Nature marchandise: courrier e-commerce")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie nature marchandise: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.11.2f: Code NGP (9999)
            try:
                ngp_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, ':ngp') and @type='text']"))
                )
                ngp_input.clear()
                ngp_input.send_keys("9999")
                print(f"      ✓ Code NGP: 9999")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie NGP: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.11.2g: Ajouter le code NGP (bouton >>)
            try:
                add_ngp_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_add_ngp')]"))
                )
                add_ngp_btn.click()
                print(f"      ✓ Code NGP ajouté")
                time.sleep(1)
            except Exception as e:
                print(f"      ❌ Erreur ajout NGP: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ==================================================================
            # ÉTAPE ED.11.3: Valider la ligne marchandise
            # ==================================================================
            try:
                valider_ligne_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_confirmer_ligne')]"))
                )
                valider_ligne_btn.click()
                print(f"      ✓ Ligne marchandise validée")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur validation ligne: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            print(f"   ✅ Lot {dum_index} créé avec succès!")
        
        # ==================================================================
        # FIN - Tous les lots créés
        # ==================================================================
        print(f"\n   ✅ Tous les lots ({len(dum_lots_data)}) créés avec succès!")
        print("      Etat de Dépotage - Onglet LTA complété")
        
        # ==================================================================
        # ÉTAPE ED.13: Sauvegarder l'Etat de Dépotage
        # ==================================================================
        print("\n   💾 Sauvegarde de l'Etat de Dépotage...")
        
        try:
            # ED.13.1: Cliquer sur "SAUVEGARDER" (menu item)
            # Stratégie robuste: chercher par texte puis par ID si échoue
            try:
                sauvegarder_link = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'ui-menuitem-link')]//span[text()='SAUVEGARDER']/parent::a"))
                )
                sauvegarder_link.click()
                print("      ✓ Bouton 'SAUVEGARDER' cliqué")
            except:
                # Fallback: par ID
                sauvegarder_link = wait.until(
                    EC.element_to_be_clickable((By.ID, "secure_174"))
                )
                sauvegarder_link.click()
                print("      ✓ Bouton 'SAUVEGARDER' cliqué (via ID)")
            
            time.sleep(3)  # Attendre la sauvegarde
            print("      ✓ Etat de Dépotage sauvegardé")
            
        except Exception as e:
            print(f"      ❌ Erreur sauvegarde: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # ED.13.2: Extraire la référence "sauvegardée" (avant validation)
        # Cette référence sera utilisée en cas d'échec de validation
        sauvegarde_reference = None
        try:
            time.sleep(2)  # Attendre que la référence s'affiche
            
            # Chercher la table de référence
            reference_table = driver.find_element(By.CSS_SELECTOR, "table.reference")
            
            # Extraire les cellules de la deuxième ligne (index 1)
            rows = reference_table.find_elements(By.TAG_NAME, "tr")
            if len(rows) >= 2:
                data_row = rows[1]
                cells = data_row.find_elements(By.TAG_NAME, "td")
                
                if len(cells) >= 5:
                    # Extraire Série (colonne 4, index 3) et Clé (colonne 5, index 4)
                    serie_value = cells[3].text.strip()
                    cle_value = cells[4].text.strip()
                    
                    # Enlever les zéros initiaux de la série
                    serie_clean = str(int(serie_value)) if serie_value.isdigit() else serie_value
                    
                    # Combiner: [Série][Clé]
                    sauvegarde_reference = f"{serie_clean}{cle_value}"
                    
                    print(f"      ✓ Référence sauvegardée extraite: {sauvegarde_reference}")
                    print(f"         (Série={serie_value} → {serie_clean}, Clé={cle_value})")
        except Exception as e:
            print(f"      ⚠️  Impossible d'extraire la référence sauvegardée: {e}")
            print(f"         (Continuera avec extraction après validation)")
            # Continuer quand même - on essaiera après validation
        
        # ==================================================================
        # ÉTAPE ED.14: Valider l'Etat de Dépotage
        # ==================================================================
        print("\n   ✅ Validation de l'Etat de Dépotage...")
        
        try:
            # ED.14.1: Cliquer sur "VALIDER" (menu item)
            try:
                valider_link = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'ui-menuitem-link')]//span[text()='VALIDER']/parent::a"))
                )
                valider_link.click()
                print("      ✓ Bouton 'VALIDER' cliqué")
            except:
                # Fallback: par ID
                valider_link = wait.until(
                    EC.element_to_be_clickable((By.ID, "secure_176"))
                )
                valider_link.click()
                print("      ✓ Bouton 'VALIDER' cliqué (via ID)")
            
            time.sleep(4)  # Attendre la validation
            
        except Exception as e:
            print(f"      ❌ Erreur validation: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # ==================================================================
        # ÉTAPE ED.15: Vérifier la réponse de validation
        # ==================================================================
        print("\n   🔍 Vérification du résultat de validation...")
        
        # ED.15.1: Chercher messages d'erreur
        try:
            time.sleep(2)  # Attendre que les messages s'affichent
            
            error_detected = False
            error_message = ""
            
            # Chercher container d'erreur
            error_containers = driver.find_elements(By.ID, "msg-error")
            if not error_containers:
                error_containers = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error")
            
            if error_containers and len(error_containers) > 0:
                # Chercher le détail de l'erreur
                error_details = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                if error_details and len(error_details) > 0:
                    error_message = error_details[0].text.strip()
                    error_detected = True
                    print(f"      ❌ Erreur détectée: {error_message}")
            
            # ED.15.2: Si erreur, créer fichier d'erreur
            if error_detected:
                lta_name = os.path.basename(lta_folder_path)
                
                # ED.15.2.1: Détecter l'erreur de poids (E2800124)
                is_weight_mismatch = "E2800124" in error_message or "Echec au rapprochement" in error_message
                
                if is_weight_mismatch:
                    # Erreur de discordance de poids - LOG SPÉCIAL
                    error_filename = f"-----------error-weight-mismatch-{lta_name}.txt"
                    error_filepath = os.path.join(os.path.dirname(lta_folder_path), error_filename)
                    
                    from datetime import datetime
                    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    with open(error_filepath, 'w', encoding='utf-8') as f:
                        f.write(f"⚠️  ERREUR DE VALIDATION - DISCORDANCE DE POIDS\n")
                        f.write(f"{'='*70}\n\n")
                        f.write(f"LTA: {lta_name}\n")
                        f.write(f"Référence LTA: {lta_reference_clean}\n")
                        f.write(f"Date/Heure: {current_datetime}\n\n")
                        
                        f.write(f"TOTAUX DÉCLARÉS DANS LE CANEVAS LTA:\n")
                        f.write(f"  • Total P (Contenants): {total_p}\n")
                        f.write(f"  • Total P,BRUT (Poids brut): {total_p_brut}\n\n")
                        
                        # Ajouter la référence sauvegardée si disponible
                        if sauvegarde_reference:
                            f.write(f"RÉFÉRENCE DS SAUVEGARDÉE (non validée):\n")
                            f.write(f"  • Série: {sauvegarde_reference}\n\n")
                        
                        f.write(f"MESSAGE D'ERREUR SYSTÈME:\n")
                        f.write(f"{error_message}\n\n")
                        
                        f.write(f"{'='*70}\n")
                        f.write(f"EXPLICATION:\n")
                        f.write(f"Le document DS (Etat de Dépotage) envoyé par Said, Youssef ou Ibrahim\n")
                        f.write(f"contient un total de Poids Brut DIFFÉRENT de celui indiqué dans le\n")
                        f.write(f"canevas de la LTA ({lta_name}).\n\n")
                        
                        f.write(f"ACTION REQUISE:\n")
                        f.write(f"1. Vérifier le document DS original envoyé par mail\n")
                        f.write(f"2. Comparer le poids total du DS avec le total du canevas LTA ({total_p_brut})\n")
                        f.write(f"3. Ajuster le canevas LTA (generated_excel) si nécessaire\n")
                        f.write(f"4. OU demander un DS corrigé à l'expéditeur\n")
                        f.write(f"5. Relancer le script après correction\n\n")
                        
                        if sauvegarde_reference:
                            f.write(f"NOTE: Une référence DS a été SAUVEGARDÉE (non validée).\n")
                            f.write(f"      Vous pouvez la modifier manuellement sur BADR si nécessaire.\n\n")
                        
                        f.write(f"STATUT: Échec de validation - CORRECTION MANUELLE NÉCESSAIRE\n")
                        f.write(f"{'='*70}\n")
                    
                    print(f"      ⚠️  ERREUR DE POIDS DÉTECTÉE (E2800124)")
                    print(f"      ✓ Log créé: {error_filename}")
                    print(f"      📊 Total P: {total_p} | Total P,BRUT: {total_p_brut}")
                    print(f"      ⚠️  Discordance avec document DS - Correction manuelle requise")
                    
                    # Sauvegarder la référence "sauvegardée" dans le fichier shipper (ligne 4)
                    if sauvegarde_reference:
                        try:
                            parent_dir = os.path.dirname(lta_folder_path)
                            lta_name_with_underscore = lta_name.replace(" ", "_")
                            shipper_pattern = f"{lta_name_with_underscore}_*.txt"
                            shipper_files = glob.glob(os.path.join(parent_dir, shipper_pattern))
                            
                            if shipper_files:
                                shipper_file = shipper_files[0]
                                
                                # Lire le fichier actuel
                                with open(shipper_file, 'r', encoding='utf-8') as f:
                                    lines = [line.rstrip('\n') for line in f.readlines()]
                                
                                # S'assurer qu'on a au moins 3 lignes
                                while len(lines) < 3:
                                    lines.append("")
                                
                                # Ajouter ou remplacer la ligne 4 avec la référence sauvegardée
                                if len(lines) == 3:
                                    lines.append(sauvegarde_reference)
                                elif len(lines) >= 4:
                                    lines[3] = sauvegarde_reference
                                
                                # Réécrire le fichier
                                with open(shipper_file, 'w', encoding='utf-8') as f:
                                    f.write('\n'.join(lines))
                                
                                print(f"      ✓ Référence sauvegardée écrite dans {os.path.basename(shipper_file)}")
                                print(f"         Ligne 4: {sauvegarde_reference} (NON VALIDÉE)")
                        except Exception as e:
                            print(f"      ⚠️  Impossible de sauvegarder la référence: {e}")
                    
                    print(f"      ⏭️  Passage au LTA suivant...")
                    
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
                
                else:
                    # Erreur générique - LOG STANDARD
                    error_filename = f"-----------error-validating-ds-{lta_name}.txt"
                    error_filepath = os.path.join(os.path.dirname(lta_folder_path), error_filename)
                    
                    from datetime import datetime
                    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    with open(error_filepath, 'w', encoding='utf-8') as f:
                        f.write(f"ERREUR DE VALIDATION - Etat de Dépotage\n\n")
                        f.write(f"LTA: {lta_name} - {lta_reference_clean}\n")
                        f.write(f"Date: {current_datetime}\n")
                        f.write(f"Étape: Validation finale de l'état de dépotage\n\n")
                        f.write(f"MESSAGE D'ERREUR:\n")
                        f.write(f"{error_message}\n\n")
                        f.write(f"CONTEXTE:\n")
                        f.write(f"- Nombre de lots créés: {len(dum_lots_data)}\n")
                        f.write(f"- Nombre de DUMs traités: {len(dum_lots_data)}\n")
                        f.write(f"- Dernière action: Click sur VALIDER\n\n")
                        f.write(f"STATUT: Échec de validation\n")
                        f.write(f"Action recommandée: Vérifier les données saisies et réessayer manuellement\n")
                    
                    print(f"      ✓ Fichier d'erreur créé: {error_filename}")
                    print(f"      ⚠️  Arrêt du traitement de ce LTA")
                    
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
            
            # ED.15.3: Si succès, extraire la référence de déclaration
            print("      ✓ Validation réussie - extraction de la référence...")
            
            try:
                # Chercher la table de référence
                reference_table = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table.reference"))
                )
                
                # Extraire les cellules de la deuxième ligne (index 1)
                rows = reference_table.find_elements(By.TAG_NAME, "tr")
                if len(rows) < 2:
                    print("      ❌ Table de référence incomplète")
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
                
                data_row = rows[1]
                cells = data_row.find_elements(By.TAG_NAME, "td")
                
                if len(cells) < 5:
                    print("      ❌ Données de référence incomplètes")
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
                
                # Extraire Série (colonne 4, index 3) et Clé (colonne 5, index 4)
                serie_value = cells[3].text.strip()
                cle_value = cells[4].text.strip()
                
                # Enlever les zéros initiaux de la série
                serie_clean = str(int(serie_value)) if serie_value.isdigit() else serie_value
                
                # Combiner: [Série][Clé]
                ds_reference = f"{serie_clean}{cle_value}"
                
                print(f"      ✓ Référence extraite: Série={serie_value} → {serie_clean}, Clé={cle_value}")
                print(f"      ✓ Référence DS complète: {ds_reference}")
                
            except Exception as e:
                print(f"      ❌ Erreur extraction référence: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ED.15.4: Ajouter la référence DS au fichier shipper (ligne 4)
            try:
                lta_name = os.path.basename(lta_folder_path)
                parent_dir = os.path.dirname(lta_folder_path)
                
                # Chercher le fichier shipper: [X]eme_LTA_*.txt
                # Le nom du dossier a des espaces (ex: "7eme LTA") mais le fichier a des underscores (ex: "7eme_LTA_shipper_name.txt")
                lta_name_with_underscore = lta_name.replace(" ", "_")
                shipper_pattern = f"{lta_name_with_underscore}_*.txt"
                shipper_files = glob.glob(os.path.join(parent_dir, shipper_pattern))
                
                if not shipper_files:
                    print(f"      ⚠️  Fichier shipper introuvable: {shipper_pattern}")
                    # Continuer quand même, ce n'est pas critique
                else:
                    shipper_file = shipper_files[0]
                    
                    # Lire le fichier actuel
                    with open(shipper_file, 'r', encoding='utf-8') as f:
                        lines = [line.rstrip('\n') for line in f.readlines()]
                    
                    # S'assurer qu'on a au moins 3 lignes (shipper, serie+cle, location)
                    while len(lines) < 3:
                        lines.append("")
                    
                    # Ajouter ou remplacer la ligne 4 (index 3) avec la référence DS
                    if len(lines) == 3:
                        # Ajouter ligne 4
                        lines.append(ds_reference)
                        action = "ajoutée"
                    elif len(lines) >= 4:
                        # Remplacer ligne 4 existante
                        lines[3] = ds_reference
                        action = "mise à jour"
                    
                    # Réécrire le fichier
                    with open(shipper_file, 'w', encoding='utf-8') as f:
                        f.write('\n'.join(lines))
                    
                    print(f"      ✓ Référence DS {action} dans {os.path.basename(shipper_file)}")
                    print(f"         Ligne 4: {ds_reference}")
                
            except Exception as e:
                print(f"      ⚠️  Erreur mise à jour fichier shipper: {e}")
                # Continuer quand même
            
            print("\n   ✅ Etat de Dépotage complété avec succès!")
            print(f"      Référence DS: {ds_reference}")
            
        except Exception as e:
            print(f"      ❌ Erreur vérification validation: {e}")
            driver.switch_to.default_content()
            return False
        
        # Sortir de l'iframe avant de retourner
        driver.switch_to.default_content()
        print("      ✓ Sorti de l'iframe Etat de Dépotage")
        return True
        
    except Exception as e:
        print(f"\n❌ Erreur création Etat de Dépotage: {e}")
        traceback.print_exc()
        # Essayer de sortir de l'iframe en cas d'erreur
        try:
            driver.switch_to.default_content()
            print("   ⚠️  Sorti de l'iframe après erreur")
        except:
            pass
        return False

def create_etat_depotage_partial(driver, lta_folder_path, partial_config, partial_number):
    """
    Crée un Etat de Dépotage pour un partiel spécifique d'une LTA.
    Uses EXACT same selectors as create_etat_depotage for compatibility.
    
    Args:
        driver: Selenium WebDriver instance
        lta_folder_path: Path to LTA folder
        partial_config: Full partial configuration dict
        partial_number: Which partial to process (1, 2, etc.)
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Extract partial data
        partial_data = find_partial_by_number(partial_config, partial_number)
        if not partial_data:
            print(f"❌ Partiel {partial_number} introuvable dans la configuration")
            return False
        
        print(f"\n📦 CRÉATION ED - PARTIEL {partial_number}")
        print(f"   Poids: {partial_data['weight']} kg")
        print(f"   Positions: {partial_data['positions']}")
        print(f"   DS: {partial_data['ds_serie']} {partial_data['ds_cle']}")
        print(f"   DUMs: {[d['dum_number'] for d in partial_data['dums']]}")
        
        # Create shipper_data structure for partial
        shipper_data = {
            'shipper_name': 'MARCHANDISES DIVERSES',  # Default fallback
            'has_ds_mead': True,
            'serie': partial_data['ds_serie'],
            'cle': partial_data['ds_cle'],
            'loading_location': partial_data['loading_location']
        }
        
        # Try to read shipper name from file (optional for partial)
        try:
            lta_name = os.path.basename(lta_folder_path)
            parent_dir = os.path.dirname(lta_folder_path)
            safe_name = lta_name.replace(' ', '_')
            txt_file_path = os.path.join(parent_dir, f"{safe_name}_shipper_name.txt")
            
            if os.path.exists(txt_file_path):
                with open(txt_file_path, 'r', encoding='utf-8') as f:
                    first_line = f.readline().strip()
                    if first_line:
                        shipper_data['shipper_name'] = first_line
                        print(f"   ✓ Expéditeur: {first_line}")
        except Exception as e:
            print(f"   ⚠️  Lecture shipper file ignorée: {e}")
        
        wait = WebDriverWait(driver, 15)
        
        # ==================================================================
        # Navigation vers "Etat de Dépotage - Voyage Aérien"
        # ==================================================================
        print("\n📂 Navigation: MISE EN DOUANE → Etat de Dépotage → Voyage Aérien...")
        
        try:
            mise_en_douane_link = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//h3[contains(@class, 'ui-panelmenu-header')]//a[contains(text(), 'MISE EN DOUANE')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", mise_en_douane_link)
            time.sleep(0.5)
            mise_en_douane_link.click()
            print("   ✓ Menu 'MISE EN DOUANE' ouvert")
            time.sleep(2)
        except Exception as e:
            print(f"   ⚠️  Menu 'MISE EN DOUANE' déjà ouvert ou erreur: {e}")
        
        try:
            creer_declaration_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_151"))
            )
            creer_declaration_link.click()
            print("   ✓ Sous-menu 'Créer une Déclaration' ouvert")
            time.sleep(1)
        except Exception as e:
            print(f"   ❌ Erreur ouverture 'Créer une Déclaration': {e}")
            return_to_home_after_error(driver)
            return False
        
        try:
            etat_depotage_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_236"))
            )
            etat_depotage_link.click()
            print("   ✓ Sous-menu 'Etat de Dépotage' ouvert")
            time.sleep(1)
        except Exception as e:
            print(f"   ❌ Erreur ouverture 'Etat de Dépotage': {e}")
            return_to_home_after_error(driver)
            return False
        
        try:
            voyage_aerien_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_247"))
            )
            voyage_aerien_link.click()
            print("   ✓ Lien 'Voyage Aérien' cliqué")
            time.sleep(3)
        except Exception as e:
            print(f"   ❌ Erreur clic 'Voyage Aérien': {e}")
            return_to_home_after_error(driver)
            return False
        
        # Switch to iframe
        try:
            iframe = wait.until(EC.presence_of_element_located((By.TAG_NAME, "iframe")))
            driver.switch_to.frame(iframe)
            print("   ✓ Iframe chargé")
            time.sleep(2)
        except Exception as e:
            print(f"   ❌ Erreur chargement iframe: {e}")
            return False
        
        # ==================================================================
        # ÉTAPE ED.1: Sélection du Bureau "301" (EXACT COPY from create_etat_depotage)
        # ==================================================================
        print("\n   🏢 Sélection du Bureau 301...")
        
        # ED.1.1: Entrer "301" dans l'autocomplete
        try:
            bureau_input = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:bureauCmbId_INPUT_input"))
            )
            bureau_input.clear()
            bureau_input.send_keys("301")
            print("      ✓ Bureau '301' saisi")
            time.sleep(1)
        except Exception as e:
            print(f"      ❌ Erreur saisie bureau: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.1.2: Sélectionner la première suggestion
        try:
            bureau_suggestion = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div#rootForm\\:bureauCmbId_INPUT_panel li.ui-autocomplete-item"))
            )
            bureau_suggestion.click()
            print("      ✓ Suggestion bureau sélectionnée")
            time.sleep(1)
        except Exception as e:
            print(f"      ❌ Erreur sélection suggestion: {e}")
            return False
        
        # ED.1.3: Valider la sélection du bureau
        try:
            valider_bureau_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "rootForm:btnConfirmer"))
            )
            valider_bureau_btn.click()
            print("      ✓ Bureau validé")
            time.sleep(4)  # Attendre le chargement du formulaire suivant
        except Exception as e:
            print(f"      ❌ Erreur validation bureau: {e}")
            return False
        
        # ==================================================================
        # ÉTAPE ED.2: Configuration Type de Déclaration (EXACT COPY)
        # ==================================================================
        print("\n   📋 Configuration de la déclaration...")
        
        # ED.2.1: Sélectionner "DS MEAD" (radio button index 2, value "03")
        try:
            time.sleep(1)
            ds_radios = driver.find_elements(By.CSS_SELECTOR, "table#mainTab\\:form1\\:radioTypeDS div.ui-radiobutton-box")
            if len(ds_radios) >= 3:
                ds_radios[2].click()  # Le 3ème = DS MEAD
                print("      ✓ 'DS MEAD' sélectionné")
                time.sleep(0.5)
            else:
                print(f"      ⚠️  Radios DS MEAD insuffisants (trouvé: {len(ds_radios)})")
                print("      🔄 Tentative avec JavaScript...")
                js_code = """
                var radio = document.getElementById('mainTab:form1:radioTypeDS:2');
                radio.checked = true;
                var event = new Event('change', { bubbles: true });
                radio.dispatchEvent(event);
                """
                driver.execute_script(js_code)
                time.sleep(0.5)
                print("      ✓ 'DS MEAD' sélectionné via JavaScript")
        except Exception as e:
            print(f"      ❌ Impossible de sélectionner DS MEAD: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.2: Entrer l'année actuelle
        try:
            annee_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:anneeId"))
            )
            annee_input.clear()
            current_year = str(time.strftime("%Y"))
            annee_input.send_keys(current_year)
            print(f"      ✓ Année: {current_year}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie année: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.3: Entrer le numéro de série (from partial)
        try:
            serie_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:serieId"))
            )
            serie_input.clear()
            serie_input.send_keys(shipper_data['serie'])
            print(f"      ✓ Série: {shipper_data['serie']}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie série: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.4: Entrer la clé (from partial)
        try:
            cle_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:cleId"))
            )
            cle_input.clear()
            cle_input.send_keys(shipper_data['cle'])
            print(f"      ✓ Clé: {shipper_data['cle']}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie clé: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.5: Référence LTA Partiel avec 3 formats (EXACT COPY from create_etat_depotage)
        lta_reference_raw = partial_config['lta_reference']
        
        # Préparer 3 formats à essayer (MÊME référence pour tous les partiels):
        # Format 1: Avec tirets (ex: "72-73797496")
        ref_parts = lta_reference_raw.split("-")
        ref_parts[0] = str(int(ref_parts[0]))  # Enlever zéros initiaux
        lta_reference_format1 = "-".join(ref_parts)
        
        # Format 2: Sans tirets (ex: "7273797496")
        lta_reference_format2 = lta_reference_raw.replace("-", "")
        lta_reference_format2 = str(int(lta_reference_format2))  # Enlever zéros initiaux
        
        # Format 3: Avec tirets, première partie sans zéro (ex: "72-73797496")
        lta_reference_format3 = "-".join(ref_parts)  # Identique à Format 1
        
        print(f"      📄 Référence LTA brute: {lta_reference_raw}")
        print(f"      📄 Format 1 (avec tirets): {lta_reference_format1}")
        print(f"      📄 Format 2 (sans tirets): {lta_reference_format2}")
        print(f"      📄 Format 3 (avec tirets): {lta_reference_format3}")
        
        # Essayer d'abord le Format 1
        try:
            reference_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:referenceLotID"))
            )
            reference_input.clear()
            reference_input.send_keys(lta_reference_format1)
            print(f"      ✓ Référence saisie (Format 1): {lta_reference_format1}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie référence: {e}")
            return_to_home_after_error(driver)
            return False
        
        # ED.2.6: Entrer le lieu de chargement (with autocomplete)
        if shipper_data.get('loading_location'):
            try:
                lieu_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:form1:lieuChargementCmbId_INPUT_input"))
                )
                lieu_input.clear()
                lieu_input.send_keys(shipper_data['loading_location'])
                print(f"      ✓ Lieu de chargement: {shipper_data['loading_location']}")
                time.sleep(1)
                
                # Sélectionner la première suggestion
                lieu_suggestion = wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "div#mainTab\\:form1\\:lieuChargementCmbId_INPUT_panel li.ui-autocomplete-item"))
                )
                lieu_suggestion.click()
                print("      ✓ Suggestion lieu sélectionnée")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ⚠️  Erreur saisie lieu de chargement: {e}")
        
        # ==================================================================
        # ÉTAPE ED.3: Validation et Gestion des Erreurs (EXACT COPY from create_etat_depotage)
        # ==================================================================
        print("\n   ✅ Validation de l'Etat de Dépotage...")
        
        # ED.3.1: Attendre que le blocker overlay disparaisse complètement
        try:
            wait.until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
            )
            time.sleep(1)
        except:
            pass
        
        # ED.3.2: Cliquer sur Valider avec retry en cas d'interception
        max_retries = 3
        for attempt in range(max_retries):
            try:
                valider_ref_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "mainTab:form1:confirmerRef"))
                )
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", valider_ref_btn)
                time.sleep(0.5)
                
                blockers = driver.find_elements(By.CSS_SELECTOR, "div.ui-blockui[style*='display: block']")
                if blockers:
                    print(f"      ⏳ Blocker UI encore visible, attente... (tentative {attempt + 1}/{max_retries})")
                    time.sleep(2)
                    continue
                
                valider_ref_btn.click()
                print("      ✓ Bouton 'Valider' cliqué")
                time.sleep(3)
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"      ⏳ Erreur clic (tentative {attempt + 1}/{max_retries}): {str(e)[:100]}")
                    time.sleep(2)
                else:
                    print(f"      ❌ Erreur clic validation après {max_retries} tentatives: {e}")
                    return_to_home_after_error(driver)
                    return False
        
        # ED.3.3: Vérifier messages d'erreur ou de succès
        error_detected = False
        try:
            time.sleep(2)
            
            error_msg = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error-detail")
            if not error_msg or len(error_msg) == 0:
                error_msg = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
            
            if error_msg and len(error_msg) > 0:
                error_text = error_msg[0].text.strip()
                print(f"      ⚠️  Erreur de validation détectée: {error_text}")
                error_detected = True
                
                # Essayer les autres formats si erreur de référence
                if "n'existe pas" in error_text.lower() or "référence" in error_text.lower():
                    print(f"      ⚠️  Format 1 rejeté, tentative Format 2...")
                    
                    try:
                        close_btn = driver.find_element(By.CSS_SELECTOR, "a.ui-messages-close")
                        close_btn.click()
                        time.sleep(0.5)
                    except:
                        pass
                    
                    reference_input = wait.until(
                        EC.presence_of_element_located((By.ID, "mainTab:form1:referenceLotID"))
                    )
                    reference_input.clear()
                    time.sleep(0.3)
                    reference_input.send_keys(lta_reference_format2)
                    print(f"      ✓ Référence Format 2 saisie: {lta_reference_format2}")
                    time.sleep(0.5)
                    
                    # Re-valider
                    for attempt in range(3):
                        try:
                            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui")))
                            time.sleep(0.5)
                            valider_ref_btn = wait.until(EC.element_to_be_clickable((By.ID, "mainTab:form1:confirmerRef")))
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", valider_ref_btn)
                            time.sleep(0.5)
                            valider_ref_btn.click()
                            print("      ✓ Bouton 'Valider' re-cliqué")
                            time.sleep(3)
                            break
                        except Exception as retry_e:
                            if attempt < 2:
                                time.sleep(2)
                            else:
                                raise retry_e
                    
                    # Vérifier résultat Format 2
                    time.sleep(1)
                    error_msg_retry2 = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error-detail")
                    if not error_msg_retry2:
                        error_msg_retry2 = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                    
                    if error_msg_retry2 and len(error_msg_retry2) > 0:
                        print(f"      ⚠️  Format 2 rejeté, tentative Format 3...")
                        
                        try:
                            close_btn = driver.find_element(By.CSS_SELECTOR, "a.ui-messages-close")
                            close_btn.click()
                            time.sleep(0.5)
                        except:
                            pass
                        
                        reference_input.clear()
                        time.sleep(0.3)
                        reference_input.send_keys(lta_reference_format3)
                        print(f"      ✓ Référence Format 3 saisie: {lta_reference_format3}")
                        time.sleep(0.5)
                        
                        # Re-valider
                        for attempt in range(3):
                            try:
                                wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui")))
                                time.sleep(0.5)
                                valider_ref_btn = wait.until(EC.element_to_be_clickable((By.ID, "mainTab:form1:confirmerRef")))
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", valider_ref_btn)
                                time.sleep(0.5)
                                valider_ref_btn.click()
                                time.sleep(3)
                                break
                            except:
                                if attempt < 2:
                                    time.sleep(2)
                        
                        # Vérifier résultat Format 3
                        time.sleep(1)
                        error_msg_retry3 = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error-detail")
                        if not error_msg_retry3:
                            error_msg_retry3 = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                        
                        if error_msg_retry3 and len(error_msg_retry3) > 0:
                            print(f"      ❌ Tous les formats rejetés - abandon")
                            return_to_home_after_error(driver)
                            return False
                        else:
                            print("      ✅ Format 3 accepté!")
                            error_detected = False
                    else:
                        print("      ✅ Format 2 accepté!")
                        error_detected = False
                else:
                    print(f"      ❌ Erreur non liée à la référence: {error_text}")
                    return_to_home_after_error(driver)
                    return False
            
            if not error_detected:
                print("      ✓ Validation terminée avec succès")
            else:
                return_to_home_after_error(driver)
                return False
                
        except Exception as e:
            print(f"      ⚠️  Erreur lors de la vérification: {e}")
        
        # ==================================================================
        # ÉTAPE ED.4: Naviguer vers l'onglet "Quantités"
        # ==================================================================
        print("\n   📊 Navigation vers l'onglet Quantités...")
        
        # ED.4.1: Fermer tous les messages (erreur et info) avant navigation
        try:
            # Fermer les messages d'erreur
            close_btns = driver.find_elements(By.CSS_SELECTOR, "a.ui-messages-close")
            for btn in close_btns:
                try:
                    btn.click()
                    time.sleep(0.3)
                except:
                    pass
            
            # Fermer aussi les messages d'info qui peuvent intercepter le clic
            info_messages = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-info")
            for msg in info_messages:
                try:
                    close_btn = msg.find_element(By.CSS_SELECTOR, "a.ui-messages-close")
                    close_btn.click()
                    time.sleep(0.3)
                except:
                    pass
            
            # Attendre que les messages disparaissent
            time.sleep(0.5)
        except:
            pass
        
        # ED.4.2: Attendre que le blocker UI disparaisse
        wait_for_ui_blocker_disappear(driver, timeout=10)
        time.sleep(0.5)
        
        # ED.4.3: Cliquer sur l'onglet Quantités avec retry et JavaScript fallback
        max_retries = 3
        for attempt in range(max_retries):
            try:
                quantites_tab = wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab3']"))
                )
                
                # Scroller pour s'assurer que l'élément est visible
                driver.execute_script("arguments[0].scrollIntoView(true);", quantites_tab)
                time.sleep(0.3)
                
                quantites_tab.click()
                print("      ✓ Onglet Quantités ouvert")
                time.sleep(2)
                break  # Success, exit retry loop
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"      ⚠️  Tentative {attempt + 1}/{max_retries} échouée, retry...")
                    time.sleep(1)
                    # Retry avec JavaScript
                    try:
                        quantites_tab = driver.find_element(By.CSS_SELECTOR, "a[href='#mainTab:tab3']")
                        driver.execute_script("arguments[0].click();", quantites_tab)
                        print("      ✓ Onglet Quantités ouvert (JavaScript)")
                        time.sleep(2)
                        break  # Success with JS, exit retry loop
                    except:
                        pass
                else:
                    # Last attempt failed
                    print(f"      ❌ Erreur navigation onglet Quantités après {max_retries} tentatives: {e}")
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
        
        # ==================================================================
        # ÉTAPE ED.6: Entrer les totaux (from partial data)
        # ==================================================================
        total_p = partial_data['positions']
        total_p_brut = partial_data['weight']
        
        print(f"\n   ⚖️  Totaux du partiel: P={total_p}, P,BRUT={total_p_brut}")
        
        # Poids brut total (CORRECT ID from create_etat_depotage)
        try:
            poids_brut_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form3:poidsBrutTotal_IT_id_input"))
            )
            poids_brut_input.clear()
            poids_brut_input.send_keys(str(total_p_brut))
            print(f"      ✓ Poids brut total saisi: {total_p_brut}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie poids brut: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # Nombre total de contenants (CORRECT ID from create_etat_depotage)
        try:
            nombre_contenants_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form3:nombreContenantTotal_IT_id"))
            )
            nombre_contenants_input.clear()
            nombre_contenants_input.send_keys(str(total_p))
            print(f"      ✓ Nombre de contenants saisi: {total_p}")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur saisie nombre contenants: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # Navigate to LTA tab
        print("\n📦 Navigation vers l'onglet LTA...")
        try:
            lta_tab = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab4']"))
            )
            lta_tab.click()
            print("   ✓ Onglet LTA ouvert")
            time.sleep(2)
        except Exception as e:
            print(f"   ❌ Erreur navigation onglet LTA: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # Create lots for DUMs in this partial
        dum_lots_data = get_dum_lots_for_partial(partial_data, partial_config)
        print(f"\n📦 Création de {len(dum_lots_data)} lot(s) pour ce partiel...")
        
        for dum_index, dum_data in enumerate(dum_lots_data, start=1):
            print(f"\n   🔹 Création lot {dum_index}/{len(dum_lots_data)} ({dum_data['dum_name']})...")
            
            # ==================================================================
            # ÉTAPE LOT.1: Cliquer "Nouveau" pour créer un lot
            # ==================================================================
            try:
                nouveau_lot_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_new_lot')]"))
                )
                nouveau_lot_btn.click()
                print(f"      ✓ Bouton 'Nouveau' lot cliqué")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur clic 'Nouveau' lot: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ==================================================================
            # ÉTAPE LOT.2: Remplir l'en-tête du lot
            # ==================================================================
            
            # LOT.2a: Référence du lot (LTA ref + /DUM_NUMBER ou /DUM_NUMBER/SPLIT_ID)
            try:
                # Construire référence selon si DUM est split ou non
                if dum_data.get('is_split', False) and dum_data.get('split_id'):
                    # DUM split: 157-54326451/2/1 ou 157-54326451/2/2
                    lot_reference = f"{lta_reference_format1}/{dum_data['split_id']}"
                else:
                    # DUM normal: 157-54326451/1 ou 157-54326451/3
                    lot_reference = f"{lta_reference_format1}/{dum_data['dum_number']}"
                
                ref_lot_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'referenceLot_IT_id')]"))
                )
                ref_lot_input.clear()
                ref_lot_input.send_keys(lot_reference)
                print(f"      ✓ Référence lot: {lot_reference}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie référence lot: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # LOT.2b: Ligne dépotée (toujours 1)
            try:
                ligne_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'ligneDepotee_IT_id')]"))
                )
                ligne_input.clear()
                ligne_input.send_keys("1")
                print(f"      ✓ Ligne dépotée: 1")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie ligne dépotée: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # LOT.2c: Sélectionner le radio button ICE (valeur 02)
            try:
                ice_radio = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:entete_section_form:radioChoixDestinataire:1"))
                )
                radio_box = driver.find_element(By.XPATH, "//input[@id='mainTab:detailLot:entete_section_form:radioChoixDestinataire:1']/parent::div/following-sibling::div[@class='ui-radiobutton-box ui-widget ui-corner-all ui-state-default']")
                radio_box.click()
                print(f"      ✓ Option ICE sélectionnée")
                time.sleep(2)
            except Exception as e:
                print(f"      ⚠️  Erreur sélection radio ICE (méthode 1): {e}")
                try:
                    print(f"      🔄 Tentative avec JavaScript...")
                    js_code = """
                    var radio = document.getElementById('mainTab:detailLot:entete_section_form:radioChoixDestinataire:1');
                    if (radio) {
                        radio.checked = true;
                        var event = new Event('change', { bubbles: true });
                        radio.dispatchEvent(event);
                    }
                    """
                    driver.execute_script(js_code)
                    time.sleep(2)
                    print(f"      ✓ Option ICE sélectionnée via JavaScript")
                except Exception as e2:
                    print(f"      ❌ Erreur sélection radio ICE: {e2}")
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
            
            # Attendre que le blocker UI disparaisse
            try:
                WebDriverWait(driver, 5).until(
                    EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
                )
                print(f"      ✓ Page stabilisée après sélection ICE")
            except:
                pass
            
            # LOT.2d: Numéro ICE (constant)
            try:
                ice_input = wait.until(
                    EC.presence_of_element_located((By.ID, "mainTab:detailLot:entete_section_form:id_ice"))
                )
                wait.until(EC.element_to_be_clickable((By.ID, "mainTab:detailLot:entete_section_form:id_ice")))
                
                ice_input.clear()
                ice_input.send_keys("000230731000088")
                print(f"      ✓ ICE: 000230731000088")
                
                from selenium.webdriver.common.keys import Keys
                ice_input.send_keys(Keys.TAB)
                time.sleep(1)
                
                print(f"      ⏳ Attente du chargement des informations ICE...")
                time.sleep(3)
                
                try:
                    WebDriverWait(driver, 5).until(
                        EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
                    )
                    print(f"      ✓ Informations ICE chargées")
                except:
                    pass
                    
            except Exception as e:
                print(f"      ❌ Erreur saisie ICE: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ==================================================================
            # ÉTAPE LOT.3: Valider l'en-tête du lot
            # ==================================================================
            try:
                valider_lot_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "mainTab:detailLot:entete_section_form:btn_confirmer_lot"))
                )
                valider_lot_btn.click()
                print(f"      ✓ En-tête lot validé")
                
                print(f"      ⏳ Attente du traitement de la validation...")
                time.sleep(4)
                
                try:
                    WebDriverWait(driver, 5).until(
                        EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
                    )
                    print(f"      ✓ Validation traitée")
                except:
                    pass
                
            except Exception as e:
                print(f"      ❌ Erreur validation en-tête lot: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # ==================================================================
            # ÉTAPE LOT.4: Cliquer "Nouveau" pour créer une ligne
            # ==================================================================
            try:
                print(f"      🔍 Recherche du bouton 'Nouveau' ligne...")
                nouveau_ligne_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_new_ligne')]"))
                )
                nouveau_ligne_btn.click()
                print(f"      ✓ Bouton 'Nouveau' ligne cliqué")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur clic 'Nouveau' ligne: {e}")
                print(f"      🔍 Tentative de recherche alternative...")
                try:
                    nouveau_ligne_btn_alt = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Nouveau') or contains(@value, 'Nouveau')]"))
                    )
                    driver.execute_script("arguments[0].click();", nouveau_ligne_btn_alt)
                    print(f"      ✓ Bouton 'Nouveau' ligne cliqué (méthode alternative)")
                    time.sleep(2)
                except Exception as e2:
                    print(f"      ❌ Erreur clic 'Nouveau' ligne (alternative): {e2}")
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
            
            # Fill line form
            # Type Contenant
            try:
                type_contenant_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'typeContenantId_INPUT_input')]"))
                )
                type_contenant_input.clear()
                type_contenant_input.send_keys("colis")
                time.sleep(2)
                
                colis_suggestion = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//li[contains(@data-item-label, 'COLIS(216)')]"))
                )
                colis_suggestion.click()
                print(f"      ✓ COLIS(216) sélectionné")
                time.sleep(1)
            except Exception as e:
                print(f"      ❌ Erreur type contenant: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # Nombre contenants
            try:
                nbr_contenants_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'nbrContenants')]"))
                )
                nbr_contenants_input.clear()
                nbr_contenants_input.send_keys(str(dum_data['p']))
                print(f"      ✓ Nombre contenants: {dum_data['p']}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur nombre contenants: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # Poids brut
            try:
                poids_brut_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, 'poidBru_input')]"))
                )
                poids_brut_input.clear()
                poids_brut_input.send_keys(str(dum_data['p_brut']))
                print(f"      ✓ Poids brut: {dum_data['p_brut']}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur poids brut: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # Marque (référence LTA validée)
            try:
                marque_textarea = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//textarea[contains(@name, 'marqueLib')]"))
                )
                marque_textarea.clear()
                marque_textarea.send_keys(lta_reference_format1)
                print(f"      ✓ Marque: {lta_reference_format1}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur marque: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # Nature marchandise
            try:
                nature_textarea = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//textarea[contains(@name, 'marchand')]"))
                )
                nature_textarea.clear()
                nature_textarea.send_keys("courrier e-commerce")
                print(f"      ✓ Nature marchandise: courrier e-commerce")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur nature marchandise: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # Code NGP
            try:
                ngp_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@name, ':ngp') and @type='text']"))
                )
                ngp_input.clear()
                ngp_input.send_keys("9999")
                print(f"      ✓ Code NGP: 9999")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur NGP: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # Add NGP
            try:
                add_ngp_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_add_ngp')]"))
                )
                add_ngp_btn.click()
                print(f"      ✓ Code NGP ajouté")
                time.sleep(1)
            except Exception as e:
                print(f"      ❌ Erreur ajout NGP: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            # Validate line
            try:
                valider_ligne_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btn_confirmer_ligne')]"))
                )
                valider_ligne_btn.click()
                print(f"      ✓ Ligne marchandise validée")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur validation ligne: {e}")
                driver.switch_to.default_content()
                return_to_home_after_error(driver)
                return False
            
            print(f"   ✅ Lot {dum_index} créé avec succès!")
        
        print(f"\n   ✅ Tous les lots ({len(dum_lots_data)}) créés avec succès!")
        
        # Save ED
        print("\n💾 Sauvegarde de l'Etat de Dépotage...")
        try:
            sauvegarder_link = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'ui-menuitem-link')]//span[text()='SAUVEGARDER']/parent::a"))
            )
            sauvegarder_link.click()
            print("   ✓ Bouton 'SAUVEGARDER' cliqué")
            time.sleep(3)
        except Exception as e:
            print(f"   ❌ Erreur sauvegarde: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # Validate ED
        print("\n✅ Validation de l'Etat de Dépotage...")
        try:
            valider_link = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'ui-menuitem-link')]//span[text()='VALIDER']/parent::a"))
            )
            valider_link.click()
            print("   ✓ Bouton 'VALIDER' cliqué")
            time.sleep(4)
        except Exception as e:
            print(f"   ❌ Erreur validation: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # Check validation result
        print("\n🔍 Vérification du résultat de validation...")
        try:
            time.sleep(2)
            error_containers = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error")
            
            if error_containers and len(error_containers) > 0:
                error_details = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                if error_details and len(error_details) > 0:
                    error_message = error_details[0].text.strip()
                    print(f"   ❌ Erreur validation: {error_message}")
                    driver.switch_to.default_content()
                    return_to_home_after_error(driver)
                    return False
            
            print("   ✓ Validation réussie!")
            
            # Extract DS reference and save to config
            try:
                reference_table = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table.reference"))
                )
                rows = reference_table.find_elements(By.TAG_NAME, "tr")
                if len(rows) >= 2:
                    data_row = rows[1]
                    cells = data_row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 5:
                        serie_value = cells[3].text.strip()
                        cle_value = cells[4].text.strip()
                        serie_clean = str(int(serie_value)) if serie_value.isdigit() else serie_value
                        ds_reference = f"{serie_clean}{cle_value}"
                        print(f"   ✓ Référence DS: {ds_reference}")
                        
                        # Sauvegarder le DS validé dans le config partiel
                        lta_name = os.path.basename(lta_folder_path)
                        parent_dir = os.path.dirname(lta_folder_path) or "."
                        save_ds_validated_to_partial_config(parent_dir, lta_name, partial_number, ds_reference)
            except Exception as e:
                print(f"   ⚠️  Erreur extraction référence: {e}")
            
            print(f"\n✅ Etat de Dépotage PARTIEL {partial_number} créé avec succès!")
            
        except Exception as e:
            print(f"   ❌ Erreur vérification validation: {e}")
            driver.switch_to.default_content()
            return_to_home_after_error(driver)
            return False
        
        # Exit iframe
        driver.switch_to.default_content()
        print("   ✓ Sorti de l'iframe")
        
        return True
        
    except Exception as e:
        print(f"\n❌ Erreur création ED partiel: {e}")
        traceback.print_exc()
        try:
            driver.switch_to.default_content()
        except:
            pass
        return False

def read_shipper_from_txt(txt_file_path):
    """Extract shipper name, LTA reference, and DS MEAD reference data from .txt file
    
    Supports TWO file formats:
    
    Format 1 - Simple 3-line format (for signed LTAs):
        Line 1: Shipper company name (e.g., "JIANGSU MINQIAN TECHNOLOGY CO., LTD")
        Line 2: Serie + Cle (e.g., "2793 X")
        Line 3: Loading location (e.g., "ABOU DHABI INT")
    
    Format 2 - OCR extraction format (6+ lines):
        Line 1: Separator "-------------"
        Line 2: LTA folder name (e.g., "9eme LTA_ocr_textbsed")
        Line 3: LTA reference without prefix (e.g., "60738318932")
        Line 4: LTA reference with prefix (e.g., "607-38318932/1")
        Line 5: Empty line
        Line 6: Shipper company name (e.g., "JIANGSU MINQIAN TECHNOLOGY CO., LTD")
        Line 7+: Optional DS MEAD data
    
    Returns:
        dict: {
            'shipper_name': str,
            'lta_reference': str (format: "607-38318932/1"),
            'lta_reference_clean': str (format: "607-38318932" without /1),
            'has_ds_mead': bool,
            'serie': str or None,
            'cle': str or None,
            'loading_location': str or None
        }
    """
    try:
        with open(txt_file_path, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()
        
        # Remove empty lines and strip whitespace
        lines = [line.strip() for line in all_lines if line.strip()]
        
        if len(lines) < 1:
            print(f"   ❌ Fichier vide: {txt_file_path}")
            return None
        
        # Detect format: Check if line 1 starts with "---" (Format 2) or is company name (Format 1)
        is_ocr_format = lines[0].startswith('---') if len(lines) > 0 else False
        
        # FORMAT 1: Simple 3-line format (shipper, serie+cle, location)
        if not is_ocr_format and len(lines) >= 2:
            print(f"   📄 Format détecté: Simple 3-line format")
            shipper_name = lines[0]
            
            # Parse line 2 for serie + cle (e.g., "2793 X")
            has_ds_mead = False
            serie = None
            cle = None
            loading_location = None
            
            parts = lines[1].split()
            if len(parts) == 2 and parts[0].isdigit() and len(parts[1]) == 1:
                has_ds_mead = True
                serie = parts[0]
                cle = parts[1]
                print(f"   ✓ DS MEAD détecté - Série: {serie}, Clé: {cle}")
            
            # Line 3 is loading location (if exists)
            if len(lines) >= 3:
                loading_location = lines[2]
                print(f"   ✓ Lieu de chargement: {loading_location}")
            
            # LTA reference not available in this format - extract from filename
            # Filename format: "2eme_LTA_shipper_name.txt" -> extract "2eme LTA"
            lta_name = os.path.basename(txt_file_path).replace('_shipper_name.txt', '').replace('_', ' ')
            lta_reference = f"UNKNOWN/{lta_name}"  # Placeholder
            lta_reference_clean = "UNKNOWN"
            
            return {
                'shipper_name': shipper_name,
                'lta_reference': lta_reference,
                'lta_reference_clean': lta_reference_clean,
                'has_ds_mead': has_ds_mead,
                'serie': serie,
                'cle': cle,
                'loading_location': loading_location
            }
        
        # FORMAT 2: OCR extraction format (6+ lines)
        if len(lines) < 4:
            print(f"   ❌ Fichier incomplet: {txt_file_path}")
            print(f"      Attendu: Format 1 (3 lignes) ou Format 2 (6+ lignes)")
            print(f"      Trouvé: {len(lines)} ligne(s)")
            return None
        # FORMAT 2: OCR extraction format (6+ lines)
        if len(lines) < 4:
            print(f"   ❌ Fichier incomplet: {txt_file_path}")
            print(f"      Attendu: Format 1 (3 lignes) ou Format 2 (6+ lignes)")
            print(f"      Trouvé: {len(lines)} ligne(s)")
            return None
        
        print(f"   📄 Format détecté: OCR extraction format (6+ lignes)")
        
        # Extract LTA reference from line 4 (index 3)
        # Format: "607-38318932/1"
        lta_reference_full = lines[3]
        
        # Clean version without /1
        lta_reference_clean = lta_reference_full.split('/')[0] if '/' in lta_reference_full else lta_reference_full
        
        # Extract shipper name from line 6 (index 5)
        shipper_name = lines[5] if len(lines) > 5 else lines[0]
        
        # Check for DS MEAD reference (optional, would be after shipper name)
        # Look for line with pattern "XXXX Y" (serie + cle)
        has_ds_mead = False
        serie = None
        cle = None
        loading_location = None
        
        # Check lines after shipper name for DS MEAD info
        for i in range(6, min(len(lines), 10)):
            line = lines[i]
            parts = line.split()
            # DS MEAD format: "2666 M" (number followed by single letter)
            if len(parts) == 2 and parts[0].isdigit() and len(parts[1]) == 1:
                has_ds_mead = True
                serie = parts[0]
                cle = parts[1]
                # Next line might be loading location
                if i + 1 < len(lines):
                    loading_location = lines[i + 1]
                break
        
        if has_ds_mead:
            print(f"   ✓ LTA {lta_reference_clean} avec DS MEAD - Série: {serie}, Clé: {cle}")
        else:
            print(f"   ✓ LTA {lta_reference_clean} sans DS MEAD")
        
        return {
            'shipper_name': shipper_name,
            'lta_reference': lta_reference_full,
            'lta_reference_clean': lta_reference_clean,
            'has_ds_mead': has_ds_mead,
            'serie': serie,
            'cle': cle,
            'loading_location': loading_location
        }
        
    except Exception as e:
        print(f"   ❌ Erreur lecture shipper depuis {txt_file_path}: {e}")
        traceback.print_exc()
        return None

def read_dum_data_from_summary(summary_excel_path):
    """Read all DUM/Sheet data from summary_file Excel
    Returns: list of dicts with keys: sheet_name, total_pieces, total_value, 
             total_gross_weight, total_freight, insurance, cartons
    """
    try:
        # Fermer le fichier Excel s'il est ouvert
        close_excel_file(summary_excel_path)
        
        wb = load_workbook(summary_excel_path, data_only=True)
        
        # Find the sheet with the summary table (usually first sheet or named 'Summary')
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
        else:
            ws = wb.active
        
        dum_list = []
        
        # Find header row (contains "Sheet Name", "Total Pieces", etc.)
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
            cell_values = [str(cell.value).lower() if cell.value else '' for cell in row]
            if 'sheet' in ' '.join(cell_values) and 'total' in ' '.join(cell_values):
                header_row = row_idx
                break
        
        if not header_row:
            print("   ⚠️  Impossible de trouver l'en-tête du tableau dans summary_file")
            return []
        
        # Read header to find column indices
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            header_text = str(cell.value).lower().strip() if cell.value else ''
            if 'sheet' in header_text or 'nom' in header_text:
                headers['sheet_name'] = col_idx
            elif 'pieces' in header_text or 'nombre' in header_text:
                headers['total_pieces'] = col_idx
            elif 'value' in header_text or 'valeur' in header_text:
                headers['total_value'] = col_idx
            elif 'gross' in header_text or 'brut' in header_text or 'poid' in header_text:
                headers['total_gross_weight'] = col_idx
            elif 'freight' in header_text or 'fret' in header_text:
                headers['total_freight'] = col_idx
            elif 'insurance' in header_text or 'assurance' in header_text:
                headers['insurance'] = col_idx
            elif 'carton' in header_text or 'colis' in header_text:
                headers['cartons'] = col_idx
            elif 'position' in header_text:
                headers['total_positions'] = col_idx
        
        print(f"   📊 Colonnes trouvées: {headers}")
        
        # Read data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[headers.get('sheet_name', 0) - 1]:
                continue  # Skip empty rows
            
            # Arrondir les valeurs décimales à 2 chiffres pour éviter les erreurs de précision flottante
            total_value = row[headers.get('total_value', 3) - 1] if 'total_value' in headers else 0
            if isinstance(total_value, (int, float)):
                total_value = round(float(total_value), 2)
            
            total_gross_weight = row[headers.get('total_gross_weight', 4) - 1] if 'total_gross_weight' in headers else 0
            if isinstance(total_gross_weight, (int, float)):
                total_gross_weight = round(float(total_gross_weight), 2)
            
            total_freight = row[headers.get('total_freight', 5) - 1] if 'total_freight' in headers else 0
            if isinstance(total_freight, (int, float)):
                total_freight = round(float(total_freight), 2)
            
            insurance = row[headers.get('insurance', 6) - 1] if 'insurance' in headers else 0
            if isinstance(insurance, (int, float)):
                insurance = round(float(insurance), 2)
            
            dum_data = {
                'sheet_name': row[headers.get('sheet_name', 1) - 1],
                'total_pieces': row[headers.get('total_pieces', 2) - 1] if 'total_pieces' in headers else 0,
                'total_value': total_value,
                'total_gross_weight': total_gross_weight,
                'total_freight': total_freight,
                'insurance': insurance,
                'cartons': row[headers.get('cartons', 7) - 1] if 'cartons' in headers else 0,
                'total_positions': row[headers.get('total_positions', 8) - 1] if 'total_positions' in headers else 0,  # Nombre de contenants (P)
            }
            
            dum_list.append(dum_data)
        
        print(f"   ✓ {len(dum_list)} DUMs trouvés dans summary_file")
        return dum_list
        
    except Exception as e:
        print(f"   ❌ Erreur lecture summary_file: {e}")
        traceback.print_exc()
        return []

def fill_declaration_form(driver, shipper_name, dum_data, lta_folder_path, lta_reference_clean):
    """Fill the declaration form with shipper name and DUM data
    
    Args:
        driver: Selenium WebDriver instance (already in iframe context)
        shipper_name: Shipper company name (string)
        dum_data: Dict with keys: sheet_name, total_value, total_gross_weight, total_freight, insurance, total_positions
        lta_folder_path: Path to LTA folder containing Sheet Excel files
        lta_reference_clean: LTA reference without /1 suffix (e.g., "607-38318932")
    """
    try:
        wait = WebDriverWait(driver, 15)
        
        print(f"\n📝 Remplissage du formulaire pour {dum_data.get('sheet_name', 'DUM')}...")
        print(f"   📋 LTA Reference: {lta_reference_clean}")
        
        # 1. Shipper Name
        print("   1️⃣ Nom expéditeur...")
        shipper_input = wait.until(
            EC.presence_of_element_located((By.ID, "mainTab:form0:nomOperateurExpediteur"))
        )
        shipper_input.clear()
        shipper_input.send_keys(shipper_name)
        print(f"      ✓ Expéditeur: {shipper_name}")
        time.sleep(0.5)
        
        # 2. Total Value
        print("   2️⃣ Montant total...")
        total_value_input = wait.until(
            EC.presence_of_element_located((By.ID, "mainTab:form0:montTotalNumber_input"))
        )
        total_value_input.clear()
        total_value_input.send_keys(str(dum_data.get('total_value', 0)))
        print(f"      ✓ Valeur totale: {dum_data.get('total_value', 0)}")
        time.sleep(0.5)
        
        # 3. Total Gross Weight
        print("   3️⃣ Poids brut total...")
        gross_weight_input = wait.until(
            EC.presence_of_element_located((By.ID, "mainTab:form0:poidBrutTotal_input"))
        )
        gross_weight_input.clear()
        gross_weight_input.send_keys(str(dum_data.get('total_gross_weight', 0)))
        print(f"      ✓ Poids brut: {dum_data.get('total_gross_weight', 0)}")
        time.sleep(0.5)
        
        # 4. Freight Amount
        print("   4️⃣ Montant fret...")
        freight_input = wait.until(
            EC.presence_of_element_located((By.ID, "mainTab:form0:montantFret_input"))
        )
        freight_input.clear()
        freight_input.send_keys(str(dum_data.get('total_freight', 0)))
        print(f"      ✓ Fret: {dum_data.get('total_freight', 0)}")
        time.sleep(0.5)
        
        # 5. Insurance Amount
        print("   5️⃣ Montant assurance...")
        insurance_input = wait.until(
            EC.presence_of_element_located((By.ID, "mainTab:form0:mntAssuranceNumber_input"))
        )
        insurance_input.clear()
        insurance_input.send_keys(str(dum_data.get('insurance', 0)))
        print(f"      ✓ Assurance: {dum_data.get('insurance', 0)}")
        time.sleep(0.5)
        
        # 6. Credit Removal Selection (select the only available option)
        print("   6️⃣ Crédit d'enlèvement...")
        try:
            # Click on the dropdown trigger
            credit_trigger = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div#mainTab\\:form0\\:choixCdeId div.ui-selectonemenu-trigger"))
            )
            credit_trigger.click()
            print("      ✓ Dropdown crédit ouvert")
            time.sleep(1)
            
            # Select the option "3095361 - National - 30"
            # L'option devrait apparaître dans une liste ul.ui-selectonemenu-items
            credit_option = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//li[@data-label='3095361 - National - 30']"))
            )
            credit_option.click()
            print("      ✓ Crédit: 3095361 - National - 30")
        except Exception as e:
            print(f"      ⚠️  Erreur sélection crédit (méthode UI): {e}")
            # Alternative method: Use JavaScript to select the option in the hidden select
            try:
                print("      🔄 Tentative avec JavaScript...")
                # Sélectionner l'option via JavaScript
                js_code = """
                var select = document.getElementById('mainTab:form0:choixCdeId_input');
                select.value = '3095361';
                // Déclencher l'événement change pour que PrimeFaces détecte le changement
                var event = new Event('change', { bubbles: true });
                select.dispatchEvent(event);
                
                // Mettre à jour le label visible
                var label = document.getElementById('mainTab:form0:choixCdeId_label');
                label.textContent = '3095361 - National - 30';
                """
                driver.execute_script(js_code)
                time.sleep(0.5)
                print("      ✓ Crédit sélectionné via JavaScript")
            except Exception as e2:
                print(f"      ❌ Impossible de sélectionner le crédit: {e2}")
                print("      ⚠️  Continuons sans crédit sélectionné...")
        
        time.sleep(1)
        
        print(f"\n   ✅ Formulaire initial rempli pour {dum_data.get('sheet_name', 'DUM')}")
        
        # ==================================================================
        # ÉTAPE 2: Naviguer vers l'onglet "Articles"
        # ==================================================================
        print("\n   📑 Navigation vers l'onglet 'Articles'...")
        try:
            articles_tab = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab1']"))
            )
            articles_tab.click()
            print("      ✓ Onglet 'Articles' cliqué")
            time.sleep(2)  # Attendre le chargement de l'onglet
        except Exception as e:
            print(f"      ❌ Erreur navigation vers 'Articles': {e}")
            return False
        
        # ==================================================================
        # ÉTAPE 3: Upload du fichier Excel correspondant (Sheet X.xlsx)
        # ==================================================================
        print(f"\n   📤 Upload du fichier Excel pour {dum_data.get('sheet_name', 'DUM')}...")
        
        # Trouver le fichier Sheet correspondant dans le dossier LTA
        sheet_name = dum_data.get('sheet_name', '')
        
        # Recherche du fichier avec pattern: "Sheet 1 - *.xlsx", "Sheet 2 - *.xlsx", etc.
        sheet_pattern = os.path.join(lta_folder_path, f"{sheet_name} - *.xlsx")
        matching_files = glob.glob(sheet_pattern)
        
        if not matching_files:
            # Fallback: essayer sans le tiret
            sheet_pattern = os.path.join(lta_folder_path, f"{sheet_name}*.xlsx")
            matching_files = glob.glob(sheet_pattern)
        
        if not matching_files:
            print(f"      ❌ Fichier Excel introuvable pour {sheet_name}")
            print(f"         Pattern recherché: {sheet_pattern}")
            
            # NETTOYAGE: Retourner à l'accueil avant de quitter
            print("\n   🏠 Nettoyage: Retour à l'accueil...")
            try:
                driver.switch_to.default_content()
                time.sleep(1)
                # Rafraîchir la page pour revenir à l'état initial
                driver.get(driver.current_url.split('#')[0])
                time.sleep(2)
                print("      ✓ Retour à l'état initial")
            except Exception as cleanup_err:
                print(f"      ⚠️  Erreur nettoyage: {cleanup_err}")
            
            return False
        
        sheet_file_path = matching_files[0]  # Prendre le premier fichier trouvé
        print(f"      ✓ Fichier trouvé: {os.path.basename(sheet_file_path)}")
        
        # Upload du fichier
        try:
            file_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form1:idFileUploadWidgetVar_input"))
            )
            
            # Convertir le chemin en chemin absolu Windows
            absolute_path = os.path.abspath(sheet_file_path)
            print(f"        Chemin absolu: {absolute_path}")
            
            file_input.send_keys(absolute_path)
            print(f"      ✓ Fichier uploadé: {os.path.basename(sheet_file_path)}")
            
            # Attendre que l'upload soit traité
            time.sleep(3)
            
        except Exception as e:
            print(f"      ❌ Erreur upload fichier: {e}")
            return False
        
        # ==================================================================
        # ÉTAPE 4: Validation du nombre total de positions
        # ==================================================================
        print(f"\n   ✅ Validation du nombre de positions...")
        
        # Récupérer le nombre de positions attendu depuis summary_file
        expected_positions = dum_data.get('total_positions', 0)
        print(f"      📊 Positions attendues (summary_file): {expected_positions}")
        
        # Attendre que la table soit chargée après l'upload
        time.sleep(2)
        
        try:
            # Lire le nombre de positions calculé par le système
            # La table contient les articles importés
            # On cherche la valeur dans la colonne "Nb. cont." (3ème td de chaque row)
            
            # Trouver toutes les lignes de la table
            table_rows = driver.find_elements(By.CSS_SELECTOR, "tbody#mainTab\\:form1\\:j_id_3p_y7_data tr")
            
            if not table_rows:
                print("      ⚠️  Aucune donnée trouvée dans la table après upload")
            else:
                print(f"      📋 {len(table_rows)} ligne(s) trouvée(s) dans la table")
                
                # Lire la valeur de "Nb. cont." dans la première ligne
                # (normalement devrait être dans la 3ème colonne)
                first_row = table_rows[0]
                cells = first_row.find_elements(By.TAG_NAME, "td")
                
                if len(cells) >= 3:
                    system_positions_text = cells[2].text.strip()
                    try:
                        system_positions = int(system_positions_text)
                        print(f"      🔢 Positions calculées (système): {system_positions}")
                        
                        # Comparaison
                        if expected_positions == 0:
                            print(f"      ⚠️  Pas de validation possible (expected_positions = 0)")
                        elif system_positions == expected_positions:
                            print(f"      ✅ VALIDATION OK: {system_positions} = {expected_positions}")
                        else:
                            print(f"      ⚠️  DIVERGENCE: Système={system_positions}, Attendu={expected_positions}")
                            # Pour l'instant, on continue quand même
                    except ValueError:
                        print(f"      ⚠️  Impossible de convertir '{system_positions_text}' en nombre")
                else:
                    print(f"      ⚠️  Moins de 3 colonnes trouvées ({len(cells)})")
                    
        except Exception as e:
            print(f"      ⚠️  Erreur validation positions: {e}")
            # On continue quand même
        
        # Attendre que l'overlay de blocage disparaisse après l'upload
        print("\n   ⏳ Attente de la fin du traitement de l'upload...")
        try:
            # Attendre que le blocker disparaisse (devient invisible ou hidden)
            wait.until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
            )
            print("      ✓ Traitement terminé")
            time.sleep(1)  # Petite pause supplémentaire pour la stabilité
        except Exception as e:
            print(f"      ⚠️  Timeout attente blocker (peut être déjà disparu): {e}")
            # Continuer quand même
        
        # ==================================================================
        # ÉTAPE 5: Vérification Carton et Workflow "Demandes diverses" (Conditionnel)
        # ==================================================================
        carton_value = dum_data.get('cartons', 0)
        print(f"\n   📦 Vérification de la valeur Carton: {carton_value}")
        
       
        # print(f"      ✓ Carton ({carton_value}) ≠ 13 → Traitement 'Demandes diverses'")
        
        # ÉTAPE 5.2: Naviguer vers "Demandes diverses"
        print("\n   📋 Navigation vers l'onglet 'Demandes diverses'...")
        try:
            demandes_diverses_tab = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab4']"))
            )
            # Essayer le clic normal
            try:
                demandes_diverses_tab.click()
                print("      ✓ Onglet 'Demandes diverses' cliqué")
            except Exception as click_error:
                print(f"      ⚠️  Clic normal intercepté, utilisation de JavaScript...")
                # Utiliser JavaScript si le clic est intercepté
                driver.execute_script("arguments[0].click();", demandes_diverses_tab)
                print("      ✓ Onglet 'Demandes diverses' cliqué (via JavaScript)")
            time.sleep(2)
        except Exception as e:
            print(f"      ❌ Erreur navigation vers 'Demandes diverses': {e}")
            return False
        
        # ÉTAPE 5.3: Cliquer sur le lien "Autre(01)"
        print("\n   🔗 Clic sur 'Autre(01)'...")
        try:
            # Utiliser XPath avec le texte au lieu de l'ID dynamique
            autre_link = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Autre(01)')]"))
            )
            # Essayer clic JavaScript car le clic normal peut être intercepté
            try:
                driver.execute_script("arguments[0].click();", autre_link)
                print("      ✓ Lien 'Autre(01)' cliqué (via JavaScript)")
            except:
                autre_link.click()
                print("      ✓ Lien 'Autre(01)' cliqué")
            time.sleep(2)
        except Exception as e:
            print(f"      ❌ Erreur clic 'Autre(01)': {e}")
            # Essayer méthode alternative: chercher dans le tableau
            try:
                print("      🔄 Tentative alternative...")
                autre_link_alt = driver.find_element(By.XPATH, "//td[@role='gridcell']//a[contains(@class, 'ui-commandlink') and contains(text(), 'Autre')]")
                driver.execute_script("arguments[0].click();", autre_link_alt)
                print("      ✓ Lien 'Autre(01)' cliqué (méthode alternative)")
                time.sleep(2)
            except Exception as e2:
                print(f"      ❌ Erreur méthode alternative: {e2}")
                return False
        
        # ÉTAPE 5.4: Modifier le textarea avec le préfixe LTA et la valeur Carton
        print(f"\n   ✏️  Mise à jour du texte avec LTA N° {lta_reference_clean}...")
        
        # Attendre que le blocker UI disparaisse après le clic
        try:
            wait.until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
            )
            time.sleep(1)
        except:
            time.sleep(2)  # Fallback: attendre 2 secondes
        
        # Utiliser directement la méthode qui fonctionne (recherche par pattern d'ID)
        try:
            print("      🔍 Recherche du textarea par pattern d'ID...")
            textarea = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//textarea[contains(@id, 'mainTab:form4:j_id') and contains(@class, 'ui-inputtextarea')]"))
            )
            
            # Scroll pour s'assurer que l'élément est visible
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", textarea)
            time.sleep(0.5)
            
            # Lire le texte actuel - essayer plusieurs méthodes
            current_text = textarea.get_attribute("value")
            if not current_text or current_text.strip() == "":
                # Essayer de lire via JavaScript
                current_text = driver.execute_script("return arguments[0].value;", textarea)
            if not current_text or current_text.strip() == "":
                # Essayer textContent comme dernier recours
                current_text = driver.execute_script("return arguments[0].textContent;", textarea)
            
            print(f"      📄 Texte actuel: {current_text}")
            
            # Si le texte actuel ne contient pas "SOIT", utiliser le texte par défaut
            if not current_text or "SOIT" not in current_text:
                print("      ⚠️  Texte actuel invalide ou vide, utilisation du pattern par défaut")
                current_text = f"SOIT {carton_value} COLIS  NS SOLL LA DISP DES FORM CCEC"
            else:
                # Construire le nouveau texte avec préfixe LTA
                # Format: "LTA N° 72-73799132 SOIT {carton_value} COLIS.NS SOLL LA DISP DES FORM CCEC"
                current_text = re.sub(r'SOIT\s+\d+\s+COLIS', f'SOIT {carton_value} COLIS', current_text)
                # Retirer le préfixe LTA existant s'il y en a un
                current_text = re.sub(r'^LTA N°\s+\d+[-\d]*\s*', '', current_text)
            
            new_text = f"LTA N° {lta_reference_clean} {current_text}"
            
            # Mettre à jour le textarea - essayer d'abord avec Selenium, puis JavaScript si échoue
            try:
                textarea.clear()
                time.sleep(0.3)
                textarea.send_keys(new_text)
                print(f"      ✓ Texte mis à jour (Selenium): {new_text}")
            except Exception as selenium_error:
                print(f"      ⚠️  Selenium échoué, utilisation de JavaScript: {selenium_error}")
                # Utiliser JavaScript pour modifier le textarea
                driver.execute_script("""
                    var textarea = arguments[0];
                    textarea.value = arguments[1];
                    // Déclencher les événements pour que le framework détecte le changement
                    var event = new Event('input', { bubbles: true });
                    textarea.dispatchEvent(event);
                    var changeEvent = new Event('change', { bubbles: true });
                    textarea.dispatchEvent(changeEvent);
                """, textarea, new_text)
                print(f"      ✓ Texte mis à jour (JavaScript): {new_text}")
            
            time.sleep(1)
        except Exception as e:
            print(f"      ❌ Erreur modification textarea: {e}")
            # Méthode de secours: chercher par XPath général
            try:
                print("      🔄 Tentative de secours avec XPath général...")
                textarea_alt = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//textarea[contains(@class, 'ui-inputtextarea') and @role='textbox']"))
                )
                
                # Lire le texte actuel
                current_text = textarea_alt.get_attribute("value")
                if not current_text or "SOIT" not in current_text:
                    current_text = driver.execute_script("return arguments[0].value;", textarea_alt)
                if not current_text or "SOIT" not in current_text:
                    current_text = f"SOIT {carton_value} COLIS  NS SOLL LA DISP DES FORM CCEC"
                else:
                    current_text = re.sub(r'SOIT\s+\d+\s+COLIS', f'SOIT {carton_value} COLIS', current_text)
                    current_text = re.sub(r'^LTA N°\s+\d+[-\d]*\s*', '', current_text)
                
                new_text = f"LTA N° {lta_reference_clean} {current_text}"
                
                # Utiliser JavaScript directement (plus fiable)
                driver.execute_script("""
                    var textarea = arguments[0];
                    textarea.value = arguments[1];
                    var event = new Event('input', { bubbles: true });
                    textarea.dispatchEvent(event);
                    var changeEvent = new Event('change', { bubbles: true });
                    textarea.dispatchEvent(changeEvent);
                """, textarea_alt, new_text)
                print(f"      ✓ Texte mis à jour (méthode de secours - JavaScript): {new_text}")
                
                time.sleep(1)
            except Exception as e2:
                print(f"      ❌ Erreur méthode de secours: {e2}")
                return False
        
        # ÉTAPE 5.5: Confirmer la demande
        print("\n   ✅ Confirmation de la demande...")
        try:
            confirmer_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "mainTab:form4:btnConfirmerDmd"))
            )
            confirmer_btn.click()
            print("      ✓ Bouton 'Confirmer' cliqué")
            time.sleep(2)
        except Exception as e:
            print(f"      ❌ Erreur confirmation demande (ID): {e}")
            # Méthode alternative: chercher par texte
            try:
                print("      🔄 Tentative alternative par texte...")
                confirmer_btn_alt = driver.find_element(By.XPATH, "//button[contains(@class, 'ui-button')]//span[text()='Confirmer']/..")
                driver.execute_script("arguments[0].click();", confirmer_btn_alt)
                print("      ✓ Bouton 'Confirmer' cliqué (méthode alternative)")
                time.sleep(2)
            except Exception as e2:
                print(f"      ❌ Erreur confirmation alternative: {e2}")
                return False

        # ==================================================================
        # ÉTAPE 6: Naviguer vers "Moyen de transport"
        # ==================================================================
        print("\n   🚚 Navigation vers l'onglet 'Moyen de transport'...")
        
        # Attendre que le blocker UI disparaisse complètement
        try:
            wait.until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ui-blockui"))
            )
            time.sleep(1)
        except:
            time.sleep(2)  # Fallback: attendre 2 secondes
        
        # Tentative avec retry et fallback JavaScript
        max_retries = 3
        for attempt in range(max_retries):
            try:
                moyen_transport_tab = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "a[href='#mainTab:tab11']"))
                )
                
                # Vérifier que le blocker n'est pas présent
                blockers = driver.find_elements(By.CSS_SELECTOR, "div.ui-blockui[style*='display: block']")
                if blockers:
                    print(f"      ⏳ Blocker UI encore visible, attente... (tentative {attempt + 1}/{max_retries})")
                    time.sleep(2)
                    continue
                
                # Scroll pour s'assurer que l'élément est visible
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", moyen_transport_tab)
                time.sleep(0.5)
                
                # Tenter le clic avec Selenium
                moyen_transport_tab.click()
                print("      ✓ Onglet 'Moyen de transport' cliqué")
                time.sleep(2)
                break  # Succès, sortir de la boucle
                
            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"      ⏳ Erreur clic (tentative {attempt + 1}/{max_retries}): {str(e)[:100]}")
                    time.sleep(2)
                else:
                    # Dernière tentative avec JavaScript
                    print(f"      ⚠️  Selenium échoué, utilisation de JavaScript...")
                    try:
                        moyen_transport_tab = driver.find_element(By.CSS_SELECTOR, "a[href='#mainTab:tab11']")
                        driver.execute_script("arguments[0].click();", moyen_transport_tab)
                        print("      ✓ Onglet 'Moyen de transport' cliqué (via JavaScript)")
                        time.sleep(2)
                        break
                    except Exception as js_error:
                        print(f"      ❌ Erreur navigation vers 'Moyen de transport' (toutes méthodes échouées): {js_error}")
                        return False
        
        # ÉTAPE 6.1: Cocher "Sans moyen de transport"
        print("\n   ☑️  Activation 'Sans moyen de transport'...")
        try:
            # Méthode 1: Cliquer sur la div.ui-chkbox-box visible
            sans_transport_checkbox = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div#mainTab\\:form11\\:checkBoxSansMTId div.ui-chkbox-box"))
            )
            sans_transport_checkbox.click()
            print("      ✓ Checkbox 'Sans moyen de transport' cochée")
            time.sleep(1)
        except Exception as e:
            print(f"      ⚠️  Erreur checkbox (méthode UI): {e}")
            # Méthode alternative: JavaScript
            try:
                print("      🔄 Tentative avec JavaScript...")
                js_code = """
                var checkbox = document.getElementById('mainTab:form11:checkBoxSansMTId_input');
                checkbox.checked = true;
                var event = new Event('change', { bubbles: true });
                checkbox.dispatchEvent(event);
                """
                driver.execute_script(js_code)
                time.sleep(0.5)
                print("      ✓ Checkbox cochée via JavaScript")
            except Exception as e2:
                print(f"      ❌ Impossible de cocher 'Sans moyen de transport': {e2}")
                return False
        
        # ==================================================================
        # ÉTAPE PDS: Préapurement DS (CONDITIONNEL - seulement si LTA signé ou partiel validé)
        # ==================================================================
        
        lta_name = os.path.basename(lta_folder_path)
        parent_dir = os.path.dirname(lta_folder_path) or "."
        
        # Load partial configuration if exists
        partial_config = get_lta_partial_info(parent_dir, lta_name)
        if partial_config:
            print(f"\n📦 Configuration partielle chargée: {len(partial_config['partials'])} partial(s)")
            if partial_config.get('split_dums'):
                print(f"   ⚠️  DUMs splits: {list(partial_config['split_dums'].keys())}")
        
        preapurement_required = False
        ds_serie = None
        ds_cle = None
        validated_lta_reference = None
        loading_location = None
        
        # FOR PARTIAL LTAs: Check if any partial has ds_validated
        if partial_config:
            print(f"\n   📦 LTA partiel détecté - vérification des DS validés...")
            # Check if at least one partial has ds_validated
            has_validated_ds = any(p.get('ds_validated') for p in partial_config.get('partials', []))
            
            if has_validated_ds:
                preapurement_required = True
                # For partials, we'll use the first partial's DS info (will be overridden per lot)
                first_partial = partial_config['partials'][0]
                ds_serie = first_partial.get('ds_serie')
                ds_cle = first_partial.get('ds_cle')
                loading_location = first_partial.get('loading_location')
                validated_lta_reference = partial_config.get('lta_reference')
                
                print(f"      ✅ Préapurement DS requis (partiels avec DS validés)")
                print(f"      Référence LTA: {validated_lta_reference}")
                print(f"      Lieu de chargement: {loading_location}")
                for p in partial_config['partials']:
                    print(f"      Partiel {p['partial_number']}: DS {p['ds_serie']} {p['ds_cle']} - Validé: {p.get('ds_validated', 'N/A')}")
            else:
                print(f"      ⏭️ Aucun DS validé trouvé pour les partiels")
                print(f"      ℹ️  Assurez-vous d'avoir exécuté Phase 1 (Etat de Dépotage) pour tous les partiels")
        else:
            # FOR STANDARD LTAs: Check LTA file for signed series (line 8)
            lta_file_pattern = os.path.join(parent_dir, f"{lta_name}.txt")
            
            if os.path.exists(lta_file_pattern):
                print(f"\n   📄 Fichier LTA trouvé: {lta_name}.txt")
                
                # Parser le fichier LTA
                lta_data = parse_lta_file(lta_file_pattern)
                
                if lta_data and lta_data['signed']:
                    # Ligne 8 contient une série signée valide
                    preapurement_required = True
                    ds_serie = lta_data['serie']
                    ds_cle = lta_data['cle']
                    
                    # IMPORTANT: Lire la référence LTA validée depuis le fichier shipper (ligne 5)
                    # Cette référence a été sauvegardée lors de la Phase 1 (Etat de Dépotage)
                    lta_name_with_underscore = lta_name.replace(" ", "_")
                    shipper_pattern = f"{lta_name_with_underscore}_*.txt"
                    shipper_files = glob.glob(os.path.join(parent_dir, shipper_pattern))
                    
                    if shipper_files:
                        try:
                            with open(shipper_files[0], 'r', encoding='utf-8') as f:
                                shipper_lines = [line.strip() for line in f.readlines()]
                            
                            # Line 3 du shipper file contient le lieu de chargement
                            if len(shipper_lines) >= 3:
                                loading_location = shipper_lines[2]
                            
                            # Line 5 du shipper file contient la référence LTA validée (sauvegardée en Phase 1)
                            if len(shipper_lines) >= 5 and shipper_lines[4]:
                                validated_lta_reference = shipper_lines[4]
                                print(f"\n   ✅ Préapurement DS requis (LTA signé)")
                                print(f"      Série signée: {lta_data['signed_series']}")
                                print(f"      Série: {ds_serie}")
                                print(f"      Clé: {ds_cle}")
                                print(f"      Référence LTA (depuis Phase 1): {validated_lta_reference}")
                                print(f"      Lieu de chargement: {loading_location}")
                            else:
                                print(f"      ⚠️  Référence LTA non trouvée dans {os.path.basename(shipper_files[0])} (ligne 5)")
                                print(f"      ℹ️  Assurez-vous d'avoir exécuté Phase 1 (Etat de Dépotage) d'abord")
                                # Ne pas continuer sans référence validée
                                preapurement_required = False
                        except Exception as e:
                            print(f"      ⚠️  Erreur lecture fichier shipper: {e}")
                            preapurement_required = False
                    else:
                        print(f"      ⚠️  Fichier shipper introuvable: {shipper_pattern}")
                        preapurement_required = False
                else:
                    print(f"\n   ⏭️  Préapurement DS non requis (LTA non signé - Line 8 vide ou invalide)")
                    print(f"      ℹ️  Continuons avec la déclaration sans Préapurement DS")
            else:
                print(f"\n   ⏭️  Préapurement DS non requis (fichier LTA introuvable: {lta_file_pattern})")
                print(f"      ℹ️  Continuons avec la déclaration sans Préapurement DS")
        
        # Check if we have all required data (partial or standard)
        if preapurement_required and validated_lta_reference:
            # For partials, ds_serie/ds_cle will be set per lot, so we don't require them here
            if not partial_config and (not ds_serie or not ds_cle):
                print(f"\n   ⚠️  Préapurement DS ignoré - DS série/clé manquants")
                preapurement_required = False
        
        if preapurement_required and validated_lta_reference:
            print("\n" + "="*70)
            print("🔗 PRÉAPUREMENT DS")
            print("="*70)
            
            print(f"      📋 Série: {ds_serie}, Clé: {ds_cle}")
            
            # Construire la référence lot: validated_lta_reference + "/" + dum_number
            # Extraire le numéro DUM depuis sheet_name (e.g., "Sheet 1" → "1")
            sheet_name = dum_data.get('sheet_name', '')
            dum_number = sheet_name.split()[-1] if sheet_name.startswith('Sheet') else '1'
            
            # GESTION SPÉCIALE: Si 1 seul DUM ET c'est Sheet 1, ajouter /1 et /2
            # IMPORTANT: Compter les DUMs dans generated_excel (C11, C18, C25, C32, C39...)
            # car summary_file peut être modifié par l'utilisateur après erreur
            # Si LTA avait plusieurs DUMs à l'origine, ne pas faire la division
            is_single_dum = False
            try:
                generated_excel_files = glob.glob(os.path.join(lta_folder_path, "generated_excel*.xlsx"))
                if generated_excel_files:
                    # Fermer le fichier Excel s'il est ouvert
                    close_excel_file(generated_excel_files[0])
                    
                    wb_check = load_workbook(generated_excel_files[0], data_only=True)
                    ws_check = wb_check['Summary']  # Sheet 'Summary'
                    
                    # Compter les DUMs en vérifiant les cellules C11, C18, C25, C32, C39...
                    # Pattern: C + (11 + (dum_index - 1) * 7)
                    original_dum_count = 0
                    for dum_idx in range(1, 51):  # Vérifier jusqu'à 50 DUMs max
                        row_num = 11 + (dum_idx - 1) * 7
                        if row_num > 500:  # Safety limit
                            break
                        cell_value = ws_check[f'C{row_num}'].value
                        if cell_value and 'DUM' in str(cell_value).upper():
                            original_dum_count += 1
                        else:
                            break  # Plus de DUMs après cette ligne
                    
                    wb_check.close()
                    
                    # Division automatique SEULEMENT si 1 DUM à l'origine ET c'est Sheet 1
                    is_single_dum = (original_dum_count == 1 and dum_number == '1')
                    
                    if original_dum_count > 1:
                        print(f"      ℹ️  LTA original avec {original_dum_count} DUMs - pas de division automatique")
            except Exception as check_err:
                print(f"      ⚠️  Erreur vérification generated_excel: {check_err}")
                pass
            
            if is_single_dum:
                # Pour un seul DUM (Sheet 1 uniquement), créer 2 références: /1 et /2
                lot_references = [
                    f"{validated_lta_reference}/1",
                    f"{validated_lta_reference}/2"
                ]
                print(f"      📄 Références lots (DUM unique Sheet 1): {lot_references[0]} et {lot_references[1]}")
            else:
                # Format de la référence lot standard (utilise le numéro du Sheet)
                if "/" in validated_lta_reference:
                    lot_reference = f"{validated_lta_reference}/{dum_number}"
                else:
                    lot_reference = f"{validated_lta_reference}/{dum_number}"
                lot_references = [lot_reference]
                print(f"      📄 Référence lot: {lot_reference}")
            
            # PDS.1: Naviguer vers l'onglet "Préapurement DS"
            print("\n   📑 Navigation vers l'onglet 'Préapurement DS'...")
            try:
                # Attendre que le blocker UI disparaisse
                wait_for_ui_blocker_disappear(driver, timeout=10)
                time.sleep(1)
                
                preapurement_tab = wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab3']"))
                )
                preapurement_tab.click()
                print("      ✓ Onglet 'Préapurement DS' cliqué")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur navigation 'Préapurement DS': {e}")
                # Retry avec JavaScript si le clic échoue
                try:
                    print("      🔄 Tentative avec JavaScript...")
                    driver.execute_script("arguments[0].click();", 
                        driver.find_element(By.CSS_SELECTOR, "a[href='#mainTab:tab3']"))
                    print("      ✓ Onglet 'Préapurement DS' cliqué (JS)")
                    time.sleep(2)
                except Exception as js_err:
                    print(f"      ❌ Échec retry JavaScript: {js_err}")
                    return False
            
            # PDS.2: Cliquer sur "Nouveau"
            print("\n   ➕ Création d'un nouveau préapurement...")
            try:
                nouveau_preap_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btnNouveauPreap')]"))
                )
                nouveau_preap_btn.click()
                print("      ✓ Bouton 'Nouveau' cliqué")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur clic 'Nouveau': {e}")
                return False
            
            # PDS.3: Remplir le formulaire
            print("\n   📝 Remplissage du formulaire Préapurement DS...")
            
            # PDS.3.0: CALCULATE LOTS FIRST to get signed series for partial LTAs
            preap_lots = get_dum_preapurement_lots(dum_number, partial_config, validated_lta_reference)
            
            # For partial LTA: Use signed series from FIRST lot for the main form
            if partial_config and preap_lots and preap_lots[0].get('ds_serie'):
                ds_serie = preap_lots[0]['ds_serie']
                ds_cle = preap_lots[0]['ds_cle']
                print(f"      ℹ️  Utilisation série signée du premier lot: {ds_serie} {ds_cle}")
            
            # PDS.3.1: Sélectionner type DS
            try:
                print("      ⏳ Attente du chargement du formulaire...")
                time.sleep(2)
                # Determine DS type based on first lot (for the initial form before loop)
                # For exception case DUM 1: first lot uses DS MEAD(03), others use Depotage(05)
                use_ds_mead = False
                if preap_lots and len(preap_lots) > 0:
                    first_lot = preap_lots[0]
                    if first_lot.get('is_exception_smallest'):
                        # First lot of exception DUM 1: use DS MEAD(03)
                        use_ds_mead = True
                        print("      ℹ️  Détection: Premier lot du cas d'exception (DUM 1) - DS MEAD(03)")
                
                # Open dropdown
                type_ds_trigger = wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "div#mainTab\\:form3\\:typeDsId div.ui-selectonemenu-trigger"))
                )
                type_ds_trigger.click()
                print("      ✓ Menu Type DS ouvert")
                time.sleep(1)
                
                # Select correct DS type based on lot flag
                if use_ds_mead:
                    # First lot of DUM 1 in exception case: DS MEAD(03)
                    ds_mead_option = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//li[@data-label='DS MEAD(03)']"))
                    )
                    ds_mead_option.click()
                    print("      ✓ Type DS: DS MEAD(03)")
                else:
                    # Default: Depotage(05) for all other cases
                    depotage_option = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//li[@data-label='Depotage(05)']"))
                    )
                    depotage_option.click()
                    print("      ✓ Type DS: Depotage(05)")
                time.sleep(1)
            except Exception as e:
                print(f"      ❌ Impossible de sélectionner Type DS: {e}")
                return False
            
            # PDS.3.2: Bureau "301"
            try:
                bureau_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'bureauId') or contains(@name, 'bureauId')]"))
                )
                bureau_input.clear()
                bureau_input.send_keys("301")
                print("      ✓ Bureau: 301")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie bureau: {e}")
                return False
            
            # PDS.3.3: Régime "000"
            try:
                regime_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'regimeId') or contains(@name, 'regimeId')]"))
                )
                regime_input.clear()
                regime_input.send_keys("000")
                print("      ✓ Régime: 000")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie régime: {e}")
                return False
            
            # PDS.3.4: Année actuelle
            try:
                annee_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'anneeId') or contains(@name, 'anneeId')]"))
                )
                annee_input.clear()
                current_year = str(time.strftime("%Y"))
                annee_input.send_keys(current_year)
                print(f"      ✓ Année: {current_year}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie année: {e}")
                return False
            
            # PDS.3.5: Série DS
            try:
                serie_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'serieId') or contains(@name, 'serieId')]"))
                )
                serie_input.clear()
                serie_input.send_keys(ds_serie)
                print(f"      ✓ Série: {ds_serie}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie série: {e}")
                return False
            
            # PDS.3.6: Clé DS
            try:
                cle_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'cleId') or contains(@name, 'cleId')]"))
                )
                cle_input.clear()
                cle_input.send_keys(ds_cle)
                print(f"      ✓ Clé: {ds_cle}")
                time.sleep(0.5)
            except Exception as e:
                print(f"      ❌ Erreur saisie clé: {e}")
                return False
            
            # PDS.3.7: BOUCLE pour ajouter tous les lots
            # Lots already calculated above (PDS.3.0), just get count
            num_lots_to_add = len(preap_lots)
            
            # If not using partial config and is single DUM, use old logic
            if not partial_config and is_single_dum:
                # Old behavior for backward compatibility
                lot_references = [
                    f"{validated_lta_reference}/1",
                    f"{validated_lta_reference}/2"
                ]
                num_lots_to_add = len(lot_references)
                preap_lots = [{'reference': ref, 'ds_serie': ds_serie, 'ds_cle': ds_cle, 'split_id': None} for ref in lot_references]
            
            print(f"\n   📦 Création de {num_lots_to_add} lot(s) pour DUM {dum_number}")
            
            # Check if this is a split DUM (for validation logic)
            is_split_dum = partial_config and str(dum_number) in partial_config.get('split_dums', {})
            
            # Check if this is exception case DUM 1 with multiple lots (needs accumulation like split DUMs)
            is_exception_dum1_multiple_lots = (
                partial_config 
                and partial_config.get('partial_type') == 'exception'
                and dum_number == '1'
                and num_lots_to_add > 1
            )
            
            # Use accumulation validation if it's a split DUM OR exception DUM 1 with multiple lots
            needs_accumulation = is_split_dum or is_exception_dum1_multiple_lots
            
            if needs_accumulation:
                if is_split_dum:
                    print(f"      ⚠️  DUM SPLIT détecté - {num_lots_to_add} lots requis")
                elif is_exception_dum1_multiple_lots:
                    print(f"      ⚠️  CAS D'EXCEPTION DUM 1 - {num_lots_to_add} lots requis (accumulation nécessaire)")
                # Initialize accumulators for validation
                split_accumulated_weight = 0.0
                split_accumulated_containers = 0.0
            
            for lot_idx in range(num_lots_to_add):
                current_lot = preap_lots[lot_idx]
                current_lot_ref = current_lot['reference']
                current_ds_serie = current_lot['ds_serie'] or ds_serie
                current_ds_cle = current_lot['ds_cle'] or ds_cle
                split_id = current_lot.get('split_id')
                
                lot_label = f"Lot {lot_idx + 1}/{num_lots_to_add}"
                if split_id:
                    lot_label += f" ({split_id})"
                
                print(f"\n      📦 Traitement {lot_label}: {current_lot_ref}")
                print(f"         DS: {current_ds_serie} {current_ds_cle}")
                
                # Si ce n'est PAS le premier lot, cliquer sur "Nouveau" pour ajouter un nouveau préapurement
                if lot_idx > 0:
                    print(f"         ➕ Ajout d'un nouveau préapurement pour {lot_label}...")
                    try:
                        nouveau_preap_btn = wait.until(
                            EC.element_to_be_clickable((By.XPATH, "//button[contains(@name, 'btnNouveauPreap')]"))
                        )
                        nouveau_preap_btn.click()
                        print(f"         ✓ Bouton 'Nouveau' cliqué pour {lot_label}")
                        time.sleep(2)
                        
                        # Re-remplir les champs communs (Type DS, Bureau, Régime, Année, Série, Clé)
                        print(f"         📝 Re-remplissage des champs pour {lot_label}...")
                        
                        # Type DS "Depotage(05)"
                        try:
                            type_ds_trigger = wait.until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, "div#mainTab\\:form3\\:typeDsId div.ui-selectonemenu-trigger"))
                            )
                            type_ds_trigger.click()
                            time.sleep(1)
                            depotage_option = wait.until(
                                EC.element_to_be_clickable((By.XPATH, "//li[@data-label='Depotage(05)']"))
                            )
                            depotage_option.click()
                            time.sleep(1)
                        except:
                            # Fallback JavaScript
                            driver.execute_script("""
                                var select = document.getElementById('mainTab:form3:typeDsId_input');
                                if (select) { select.value = '05'; select.dispatchEvent(new Event('change', { bubbles: true })); }
                            """)
                            time.sleep(1)
                        
                        # Bureau, Régime, Année, Série, Clé (use current lot's DS values)
                        driver.find_element(By.XPATH, "//input[contains(@id, 'bureauId') or contains(@name, 'bureauId')]").send_keys("301")
                        time.sleep(0.5)
                        driver.find_element(By.XPATH, "//input[contains(@id, 'regimeId') or contains(@name, 'regimeId')]").send_keys("000")
                        time.sleep(0.5)
                        driver.find_element(By.XPATH, "//input[contains(@id, 'anneeId') or contains(@name, 'anneeId')]").send_keys(str(time.strftime("%Y")))
                        time.sleep(0.5)
                        driver.find_element(By.XPATH, "//input[contains(@id, 'serieId') or contains(@name, 'serieId')]").send_keys(current_ds_serie)
                        time.sleep(0.5)
                        driver.find_element(By.XPATH, "//input[contains(@id, 'cleId') or contains(@name, 'cleId')]").send_keys(current_ds_cle)
                        time.sleep(0.5)
                        
                        print(f"         ✓ Champs re-remplis pour {lot_label}")
                        
                        # For subsequent lots, also fill loading location
                        if loading_location:
                            try:
                                lieu_input = wait.until(
                                    EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'lieuChargCmb') or contains(@name, 'lieuChargCmb')]"))
                                )
                                lieu_input.clear()
                                lieu_input.send_keys(loading_location)
                                print(f"         ✓ Lieu chargement: {loading_location}")
                                time.sleep(2)
                                
                                # Sélectionner la première suggestion
                                try:
                                    lieu_suggestion = wait.until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, "li.ui-autocomplete-item"))
                                    )
                                    lieu_suggestion.click()
                                    print("         ✓ Suggestion lieu sélectionnée")
                                    time.sleep(1)
                                except:
                                    print("         ⚠️  Aucune suggestion trouvée, on continue...")
                            except Exception as e:
                                print(f"         ⚠️  Erreur saisie lieu chargement: {e}")
                        
                    except Exception as e:
                        print(f"         ❌ Erreur ajout nouveau préapurement: {e}")
                        return False
                
                # Lieu de chargement (pour le premier lot seulement)
                if lot_idx == 0 and loading_location:
                    try:
                        lieu_input = wait.until(
                            EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'lieuChargCmb') or contains(@name, 'lieuChargCmb')]"))
                        )
                        lieu_input.clear()
                        lieu_input.send_keys(loading_location)
                        print(f"         ✓ Lieu chargement: {loading_location}")
                        time.sleep(2)
                        
                        # Sélectionner la première suggestion
                        try:
                            lieu_suggestion = wait.until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, "li.ui-autocomplete-item"))
                            )
                            lieu_suggestion.click()
                            print("         ✓ Suggestion lieu sélectionnée")
                            time.sleep(1)
                        except:
                            print("         ⚠️  Aucune suggestion trouvée, on continue...")
                    except Exception as e:
                        print(f"         ⚠️  Erreur saisie lieu chargement: {e}")
                
                # Référence lot (spécifique à chaque lot)
                try:
                    lot_ref_input = wait.until(
                        EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'preapurement_ref_lot') or contains(@name, 'preapurement_ref_lot')]"))
                    )
                    lot_ref_input.clear()
                    lot_ref_input.send_keys(current_lot_ref)
                    print(f"         ✓ Référence lot: {current_lot_ref}")
                    time.sleep(0.5)
                except Exception as e:
                    print(f"         ❌ Erreur saisie référence lot: {e}")
                    return False
                
                # Cliquer sur "OK" pour récupérer les données
                print(f"         🔍 Récupération des données du lot...")
                try:
                    ok_btn = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(@id, 'btnRefPreapOk') or contains(text(), 'OK')]"))
                    )
                    try:
                        ok_btn.click()
                        print(f"         ✓ Bouton 'OK' cliqué")
                    except:
                        driver.execute_script("arguments[0].click();", ok_btn)
                        print(f"         ✓ Bouton 'OK' cliqué (via JavaScript)")
                    time.sleep(3)
                except Exception as e:
                    print(f"         ❌ Erreur clic 'OK': {e}")
                    return False
                
                # Validation des données
                if not is_single_dum:
                    # Read retrieved values from system
                    try:
                        # Lire poids brut
                        poids_brut_span = wait.until(
                            EC.presence_of_element_located((By.ID, "mainTab:form3:poidLotId"))
                        )
                        poids_brut_text = poids_brut_span.text.strip().replace(',', '.')
                        retrieved_weight = float(poids_brut_text)
                        
                        # Lire nombre contenants
                        nbr_contenants_span = wait.until(
                            EC.presence_of_element_located((By.ID, "mainTab:form3:nbrContenantLotId"))
                        )
                        nbr_contenants_text = nbr_contenants_span.text.strip().replace(',', '.')
                        retrieved_containers = float(nbr_contenants_text)
                        
                        # For split DUMs or exception DUM 1 with multiple lots: accumulate values instead of validating immediately
                        if needs_accumulation:
                            split_accumulated_weight += retrieved_weight
                            split_accumulated_containers += retrieved_containers
                            
                            is_last_lot = (lot_idx == num_lots_to_add - 1)
                            
                            if not is_last_lot:
                                # Not the last lot - just accumulate and continue
                                print(f"\n         📊 Lot {lot_idx + 1}/{num_lots_to_add} - Accumulation:")
                                print(f"            Poids: {retrieved_weight} (Total accumulé: {split_accumulated_weight})")
                                print(f"            Contenants: {retrieved_containers} (Total accumulé: {split_accumulated_containers})")
                                print(f"         ⏭️  Validation reportée au dernier lot")
                            else:
                                # Last lot - validate accumulated total against DUM expected values
                                print(f"\n         ✅ Validation des données accumulées (tous les lots)...")
                                
                                # Valeurs attendues du DUM complet
                                expected_weight = float(dum_data.get('total_gross_weight', 0))
                                expected_containers = float(dum_data.get('total_positions', 0))
                                
                                print(f"         📊 TOTAL - Poids brut: {split_accumulated_weight} (attendu: {expected_weight})")
                                print(f"         📦 TOTAL - Contenants: {split_accumulated_containers} (attendu: {expected_containers})")
                                
                                # Vérifier correspondance avec tolérance pour les arrondis
                                weight_match = abs(split_accumulated_weight - expected_weight) < 0.1
                                containers_match = abs(split_accumulated_containers - expected_containers) < 0.1
                                
                                if not weight_match or not containers_match:
                                    print(f"         ❌ DIVERGENCE DÉTECTÉE!")
                                    
                                    # Créer fichier d'erreur
                                    error_filename = f"-------------error-entering-ds-mead-on-declaration-{lta_name}-DUM{dum_number}.txt"
                                    error_filepath = os.path.join(parent_dir, error_filename)
                                    
                                    from datetime import datetime
                                    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                    
                                    error_type = "SPLIT" if is_split_dum else "EXCEPTION" if is_exception_dum1_multiple_lots else "MULTIPLE_LOTS"
                                    with open(error_filepath, 'w', encoding='utf-8') as f:
                                        f.write(f"ERREUR - Préapurement DS - Données Incohérentes (DUM {error_type})\n\n")
                                        f.write(f"LTA: {lta_name} - {validated_lta_reference}\n")
                                        if is_exception_dum1_multiple_lots:
                                            f.write(f"DUM: {dum_number} (CAS D'EXCEPTION - {num_lots_to_add} lots)\n")
                                        else:
                                            f.write(f"DUM: {dum_number} (SPLIT en {num_lots_to_add} lots)\n")
                                        f.write(f"Date: {current_datetime}\n")
                                        f.write(f"Étape: Préapurement DS - Validation après tous les lots\n\n")
                                        f.write(f"VALEURS ATTENDUES (DUM {dum_number} complet):\n")
                                        f.write(f"- Poids brut (P,BRUT): {expected_weight}\n")
                                        f.write(f"- Nombre contenants (P): {expected_containers}\n\n")
                                        f.write(f"VALEURS RÉCUPÉRÉES (Total de {num_lots_to_add} lots):\n")
                                        f.write(f"- Poids brut: {split_accumulated_weight}\n")
                                        f.write(f"- Nombre contenants: {split_accumulated_containers}\n\n")
                                        f.write(f"ÉCART DÉTECTÉ:\n")
                                        f.write(f"- Poids brut: {expected_weight} ≠ {split_accumulated_weight} (Différence: {expected_weight - split_accumulated_weight})\n")
                                        f.write(f"- Contenants: {expected_containers} ≠ {split_accumulated_containers} (Différence: {expected_containers - split_accumulated_containers})\n\n")
                                        f.write(f"MESSAGE: La somme des données des lots ne correspond pas aux\n")
                                        f.write(f"données du DUM complet. Vérification manuelle requise.\n")
                                    
                                    print(f"         ✓ Fichier d'erreur créé: {error_filename}")
                                    print(f"         ⚠️  Arrêt du traitement de ce DUM")
                                    return False
                                
                                print(f"         ✅ VALIDATION OK - Totaux correspondent")
                        
                        else:
                            # Normal DUM (not split) - validate immediately
                            print(f"\n         ✅ Validation des données récupérées...")
                            
                            # Valeurs attendues
                            expected_weight = float(dum_data.get('total_gross_weight', 0))
                            expected_containers = float(dum_data.get('total_positions', 0))
                            
                            print(f"         📊 Poids brut: {retrieved_weight} (attendu: {expected_weight})")
                            print(f"         📦 Contenants: {retrieved_containers} (attendu: {expected_containers})")
                            
                            # Vérifier correspondance
                            if retrieved_weight != expected_weight or retrieved_containers != expected_containers:
                                print(f"         ❌ DIVERGENCE DÉTECTÉE!")
                                
                                # Créer fichier d'erreur
                                error_filename = f"-------------error-entering-ds-mead-on-declaration-{lta_name}-DUM{dum_number}.txt"
                                error_filepath = os.path.join(parent_dir, error_filename)
                                
                                from datetime import datetime
                                current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                
                                with open(error_filepath, 'w', encoding='utf-8') as f:
                                    f.write(f"ERREUR - Préapurement DS - Données Incohérentes\n\n")
                                    f.write(f"LTA: {lta_name} - {validated_lta_reference}\n")
                                    f.write(f"DUM: {dum_number}\n")
                                    f.write(f"Date: {current_datetime}\n")
                                    f.write(f"Étape: Préapurement DS - Validation après click OK\n\n")
                                    f.write(f"VALEURS ATTENDUES (DUM {dum_number}):\n")
                                    f.write(f"- Poids brut (P,BRUT): {expected_weight}\n")
                                    f.write(f"- Nombre contenants (P): {expected_containers}\n\n")
                                    f.write(f"VALEURS RÉCUPÉRÉES (Système):\n")
                                    f.write(f"- Poids brut: {retrieved_weight}\n")
                                    f.write(f"- Nombre contenants: {retrieved_containers}\n\n")
                                    f.write(f"ÉCART DÉTECTÉ:\n")
                                    f.write(f"- Poids brut: {expected_weight} ≠ {retrieved_weight} (Différence: {expected_weight - retrieved_weight})\n")
                                    f.write(f"- Contenants: {expected_containers} ≠ {retrieved_containers} (Différence: {expected_containers - retrieved_containers})\n\n")
                                    f.write(f"MESSAGE: Les données du lot de dédouanement ne correspondent pas aux\n")
                                    f.write(f"données du DUM actuel. Vérification manuelle requise.\n\n")
                                    f.write(f"RÉFÉRENCE LOT UTILISÉE: {current_lot_ref}\n")
                                    f.write(f"RÉFÉRENCE DS MEAD: {ds_serie} {ds_cle}\n")
                                
                                print(f"         ✓ Fichier d'erreur créé: {error_filename}")
                                print(f"         ⚠️  Arrêt du traitement de ce DUM")
                                return False
                            
                            print(f"         ✅ VALIDATION OK - Données correspondent")
                        
                    except Exception as e:
                        print(f"         ❌ Erreur validation données: {e}")
                        return False
                else:
                    print(f"         ⏭️  Validation poids/contenants ignorée (DUM unique avec division)")
                
                # Confirmer le préapurement
                print(f"         ✅ Confirmation du préapurement {lot_label}...")
                try:
                    confirmer_btn = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(@id, 'btnConfirmerPreap') or (contains(@class, 'ui-button') and contains(., 'Confirmer'))]"))
                    )
                    try:
                        confirmer_btn.click()
                        print(f"         ✓ Bouton 'Confirmer' cliqué pour {lot_label}")
                    except:
                        driver.execute_script("arguments[0].click();", confirmer_btn)
                        print(f"         ✓ Bouton 'Confirmer' cliqué (via JavaScript) pour {lot_label}")
                    time.sleep(2)
                except Exception as e:
                    print(f"         ❌ Erreur confirmation préapurement: {e}")
                    return False
                
                print(f"         ✅ {lot_label} traité avec succès!")
            
            # FIN DE LA BOUCLE - Tous les lots ont été ajoutés
            
            print("\n   ✅ Préapurement DS complété avec succès!")
            print("="*70)
            
            # Retourner à l'onglet Moyen de transport pour continuer le workflow normal
            print("\n   🔙 Retour à l'onglet 'Moyen de transport'...")
            try:
                moyen_transport_tab = wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab11']"))
                )
                moyen_transport_tab.click()
                print("      ✓ Retour à 'Moyen de transport'")
                time.sleep(2)
            except Exception as e:
                print(f"      ❌ Erreur retour Moyen de transport: {e}")
                return False
        
        # ==================================================================
        # ÉTAPE 6.5: Cliquer sur "SAUVEGARDER" avant de naviguer vers Documents
        # ==================================================================
        print("\n   💾 Sauvegarde de la déclaration...")
        try:
            # Cliquer sur le bouton "SAUVEGARDER"
            sauvegarder_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "secure__2002"))
            )
            sauvegarder_btn.click()
            print("      ✓ Bouton 'SAUVEGARDER' cliqué")
            
            # Attendre que l'overlay de blocage disparaisse après la sauvegarde
            print("      ⏳ Attente de la fin de la sauvegarde...")
            if wait_for_ui_blocker_disappear(driver, timeout=10):
                print("      ✓ Sauvegarde terminée (blocker disparu)")
                time.sleep(2)  # Pause supplémentaire pour stabilité
            else:
                print("      ⚠️  Timeout blocker - continuons quand même")
                time.sleep(5)  # Fallback plus long
            
            # ==================================================================
            # VÉRIFIER SI LA SAUVEGARDE A RÉUSSI
            # ==================================================================
            print("      🔍 Vérification du résultat de sauvegarde...")
            save_error = False
            save_error_messages = []
            
            try:
                # Chercher les messages d'erreur
                error_containers = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error")
                visible_errors = [c for c in error_containers if c.is_displayed()]
                
                if visible_errors:
                    # ⚠️ RÈGLE SPÉCIALE POUR SAUVEGARDER:
                    # Tout conteneur d'erreur visible = ÉCHEC, même s'il est vide!
                    # (contrairement aux validations où phantom errors = OK)
                    print("      ⚠️  Conteneur d'erreur détecté après sauvegarde")
                    save_error = True  # ← ERREUR IMMÉDIATE dès qu'un conteneur existe
                    
                    # Collecter les messages d'erreur s'ils existent
                    for error_container in visible_errors:
                        try:
                            # Chercher le bouton "Détails"
                            details_btn = error_container.find_element(By.ID, "rapportMsgForm:showErrors")
                            if details_btn and details_btn.is_displayed():
                                # Récupérer le texte d'erreur
                                error_details = error_container.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                                for detail in error_details:
                                    error_text = detail.text.strip()
                                    if error_text:
                                        save_error_messages.append(error_text)
                                        print(f"      ❌ Erreur: {error_text[:80]}...")
                                
                                # Si aucun message textuel, c'est une erreur "vide"
                                if not save_error_messages:
                                    save_error_messages.append("Erreur de sauvegarde (conteneur d'erreur vide)")
                                    print(f"      ❌ Erreur de sauvegarde détectée (conteneur vide)")
                        except:
                            # Pas de bouton "Détails" - vérifier message unique
                            error_details = error_container.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                            for detail in error_details:
                                error_text = detail.text.strip()
                                if error_text:
                                    save_error_messages.append(error_text)
                                    print(f"      ❌ Erreur: {error_text[:80]}...")
                            
                            # Si toujours aucun message
                            if not save_error_messages:
                                save_error_messages.append("Erreur de sauvegarde (conteneur sans détails)")
                                print(f"      ❌ Erreur de sauvegarde (conteneur sans message)")
                
                # Chercher message de succès
                if not save_error:
                    success_containers = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-info")
                    visible_success = [c for c in success_containers if c.is_displayed()]
                    
                    if visible_success:
                        for success_container in visible_success:
                            success_details = success_container.find_elements(By.CSS_SELECTOR, "span.ui-messages-info-detail")
                            for detail in success_details:
                                success_text = detail.text.strip()
                                if success_text and "succès" in success_text.lower():
                                    print(f"      ✅ {success_text}")
                                    break
                    else:
                        print("      ✓ Déclaration sauvegardée (pas d'erreur détectée)")
                
            except Exception as check_err:
                print(f"      ⚠️  Impossible de vérifier le résultat: {check_err}")
                # Continuer par défaut si on ne peut pas vérifier
            
            # Si erreur de sauvegarde détectée, arrêter le traitement de ce DUM
            if save_error:
                print(f"\n   ❌ ÉCHEC SAUVEGARDE - Impossible de continuer avec ce DUM")
                print(f"      Erreur(s) détectée(s):")
                for msg in save_error_messages:
                    print(f"         • {msg}")
                
                # Retourner à l'accueil et marquer comme erreur
                print("\n   🏠 Retour à l'accueil après erreur de sauvegarde...")
                return_to_home_after_error(driver)
                
                # Marquer l'erreur dans Excel
                sheet_name = dum_data.get('sheet_name', '')
                dum_number = int(sheet_name.split()[-1]) if sheet_name.startswith('Sheet') else 1
                mark_dum_as_error_in_excel(lta_folder_path, dum_number)
                
                # Créer un log d'erreur
                lta_name = os.path.basename(lta_folder_path)
                save_dum_error_log(
                    lta_folder_path=lta_folder_path,
                    lta_name=lta_name,
                    dum_number=dum_number,
                    sheet_name=sheet_name,
                    error_exception=Exception(f"Erreur sauvegarde: {'; '.join(save_error_messages)}"),
                    error_step="Sauvegarde déclaration (SAUVEGARDER)",
                    dum_data=dum_data
                )
                
                return False  # Échec du DUM
                
        except Exception as e:
            print(f"      ⚠️  Erreur lors de la sauvegarde: {e}")
            # En cas d'exception, retourner à l'accueil et marquer comme erreur
            print("\n   🏠 Retour à l'accueil après exception...")
            return_to_home_after_error(driver)
            
            sheet_name = dum_data.get('sheet_name', '')
            dum_number = int(sheet_name.split()[-1]) if sheet_name.startswith('Sheet') else 1
            mark_dum_as_error_in_excel(lta_folder_path, dum_number)
            
            return False
        
        # ==================================================================
        # ÉTAPE 7: Naviguer vers "Documents" et uploader les fichiers
        # ==================================================================
        print("\n   📄 Navigation vers l'onglet 'Documents'...")
        try:
            # Attendre que l'onglet Documents soit cliquable (sans overlay)
            documents_tab = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab7']"))
            )
            
            # Essayer le clic normal
            try:
                documents_tab.click()
                print("      ✓ Onglet 'Documents' cliqué")
            except Exception as click_error:
                print(f"      ⚠️  Clic normal intercepté, utilisation de JavaScript...")
                # Utiliser JavaScript si le clic est intercepté
                driver.execute_script("arguments[0].click();", documents_tab)
                print("      ✓ Onglet 'Documents' cliqué (via JavaScript)")
            
            time.sleep(2)
        except Exception as e:
            print(f"      ❌ Erreur navigation vers 'Documents': {e}")
            return False
        
        # ==================================================================
        # ÉTAPE 7.1: Premier Upload - Document LTA
        # ==================================================================
        print("\n   📤 Upload 1/2: Document LTA principal...")
        
        # 7.1.1: Sélectionner le type de document "TITRE DE PROPRIÉTÉ ET/OU DE TRANSPORT"
        print("      1️⃣ Sélection du type 'TITRE DE PROPRIÉTÉ ET/OU DE TRANSPORT'...")
        try:
            # Cliquer sur le trigger pour ouvrir le dropdown
            doc_type_trigger = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div#mainTab\\:form7\\:comp1 div.ui-selectonemenu-trigger"))
            )
            doc_type_trigger.click()
            print("         ✓ Dropdown type document ouvert")
            time.sleep(1)
            
            # Sélectionner l'option "TITRE DE PROPRIÉTÉ ET/OU DE TRANSPORT"
            doc_type_option = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//li[@data-label='TITRE DE PROPRIÉTÉ ET/OU DE TRANSPORT']"))
            )
            doc_type_option.click()
            print("         ✓ Type sélectionné: A0004 (TITRE DE PROPRIÉTÉ ET/OU DE TRANSPORT)")
            
            # Attendre que le blocker AJAX disparaisse après sélection du type
            print("         ⏳ Attente mise à jour formulaire...")
            if wait_for_ui_blocker_disappear(driver, timeout=5):
                print("         ✓ Formulaire mis à jour")
            time.sleep(1)
        except Exception as e:
            print(f"      ❌ Erreur sélection type document: {e}")
            return False
        
        # 7.1.2: Entrer la référence "LTA"
        print("      2️⃣ Saisie de la référence 'LTA'...")
        
        # Attendre que le champ référence soit actif (enabled) après la sélection du type
        ref_input_success = False
        max_ref_attempts = 3
        
        for attempt in range(1, max_ref_attempts + 1):
            try:
                if attempt > 1:
                    print(f"      🔄 Tentative {attempt}/{max_ref_attempts}...")
                    time.sleep(1)
                
                # Re-localiser l'input à chaque tentative (éviter stale element)
                ref_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//label[contains(text(), 'Référence')]/parent::td/following-sibling::td//input[@type='text' and @maxlength='10']"))
                )
                
                # Attendre qu'il soit enabled (pas disabled après sélection type)
                wait.until(lambda d: ref_input.is_enabled() and not ref_input.get_attribute('disabled'))
                
                # Clear et saisie
                ref_input.clear()
                time.sleep(0.3)
                ref_input.send_keys("LTA")
                
                # Vérifier que la valeur a été saisie
                if ref_input.get_attribute('value') == "LTA":
                    print("         ✓ Référence: LTA")
                    ref_input_success = True
                    time.sleep(0.5)
                    break
                else:
                    print(f"         ⚠️  Valeur non enregistrée (tentative {attempt})")
                    
            except Exception as e:
                if attempt == max_ref_attempts:
                    print(f"      ⚠️  Méthode XPath échouée après {max_ref_attempts} tentatives: {e}")
                    # Méthode alternative: chercher par pattern d'ID
                    try:
                        print("      🔄 Tentative alternative (par ID)...")
                        ref_input_alt = wait.until(
                            EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'mainTab:form7:j_id') and @type='text' and @maxlength='10' and not(@disabled)]"))
                        )
                        ref_input_alt.clear()
                        time.sleep(0.3)
                        ref_input_alt.send_keys("LTA")
                        if ref_input_alt.get_attribute('value') == "LTA":
                            print("         ✓ Référence: LTA (méthode alternative)")
                            ref_input_success = True
                            time.sleep(0.5)
                        else:
                            print("         ⚠️  Valeur non enregistrée (méthode alternative)")
                    except Exception as e2:
                        print(f"      ❌ Erreur saisie référence (méthode alternative): {e2}")
        
        if not ref_input_success:
            print("      ❌ Impossible de saisir la référence après toutes les tentatives")
            return False
        
        # 7.1.3: Sélectionner la date actuelle
        print("      3️⃣ Sélection de la date actuelle...")
        try:
            # Cliquer sur le bouton du date picker
            date_picker_btn = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.ui-datepicker-trigger"))
            )
            date_picker_btn.click()
            time.sleep(1)
            
            # Sélectionner la date actuelle (ui-datepicker-today)
            today_cell = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "td.ui-datepicker-today a"))
            )
            today_cell.click()
            print("         ✓ Date actuelle sélectionnée")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur sélection date: {e}")
            return False
        
        # 7.1.4: Upload du fichier LTA PDF
        print("      4️⃣ Upload du fichier LTA PDF...")
        try:
            # Trouver le fichier LTA dans le dossier
            # Pattern: "12eme LTA - *.pdf" (le fichier principal LTA, pas les mn*.pdf)
            lta_name = os.path.basename(lta_folder_path)  # e.g., "12eme LTA"
            lta_pattern = os.path.join(lta_folder_path, f"{lta_name} - *.pdf")
            lta_files = glob.glob(lta_pattern)
            
            if not lta_files:
                print(f"      ❌ Fichier LTA PDF introuvable")
                print(f"         Pattern recherché: {lta_pattern}")
                return False
            
            lta_file_path = lta_files[0]
            print(f"         📄 Fichier trouvé: {os.path.basename(lta_file_path)}")
            
            # Upload du fichier
            file_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form7:comp2_input"))
            )
            absolute_path = os.path.abspath(lta_file_path)
            file_input.send_keys(absolute_path)
            print(f"         ✓ Fichier LTA uploadé: {os.path.basename(lta_file_path)}")
            
            # Attendre que le blocker d'upload disparaisse
            print("         ⏳ Attente fin d'upload...")
            if wait_for_ui_blocker_disappear(driver, timeout=10):
                print("         ✓ Upload terminé (blocker disparu)")
            else:
                print("         ⚠️  Timeout blocker upload - continuons")
            
            # Attendre que l'interface soit prête pour le prochain upload
            time.sleep(3)  # Stabilisation après premier upload
            
            # Vérifier à nouveau si un blocker apparaît (traitement en arrière-plan)
            print("         ⏳ Vérification stabilité UI...")
            if wait_for_ui_blocker_disappear(driver, timeout=5):
                print("         ✓ UI stabilisée")
            
            time.sleep(2)  # Pause supplémentaire pour sécurité
            
            print("         ✓ Upload LTA traité, préparation pour le document MN...")
        except Exception as e:
            print(f"      ❌ Erreur upload fichier LTA: {e}")
            return False
        
        # ==================================================================
        # ÉTAPE 7.2: Deuxième Upload - Document MN du DUM actuel
        # ==================================================================
        print("\n   📤 Upload 2/2: Document MN pour ce DUM...")
        
        # Extraire le numéro du DUM depuis sheet_name (e.g., "Sheet 1" -> "1")
        sheet_name = dum_data.get('sheet_name', '')
        dum_number = sheet_name.split()[-1] if sheet_name.startswith('Sheet') else '1'
        mn_reference = f"mn{dum_number}"
        mn_filename = f"mn{dum_number}.pdf"
        
        # 7.2.0: Attendre que l'UI soit complètement prête après le premier upload
        print(f"      ⏳ Attente stabilisation complète UI...")
        time.sleep(3)  # Pause supplémentaire pour permettre à l'UI de se réinitialiser
        
        # Fermer tout dropdown qui pourrait être resté ouvert
        try:
            open_panels = driver.find_elements(By.CSS_SELECTOR, "div.ui-selectonemenu-panel[style*='display: block']")
            if open_panels:
                print("         🔄 Fermeture dropdown résiduel...")
                driver.execute_script("arguments[0].style.display = 'none';", open_panels[0])
                time.sleep(0.5)
        except:
            pass
        
        # 7.2.1: Sélectionner le type de document "FACTURE"
        print(f"      1️⃣ Sélection du type 'FACTURE'...")
        
        dropdown_opened = False
        max_attempts = 3
        
        for attempt in range(1, max_attempts + 1):
            try:
                if attempt > 1:
                    print(f"         🔄 Tentative {attempt}/{max_attempts}...")
                    time.sleep(2)
                
                # Méthode 1: Utiliser le trigger CSS
                try:
                    # Attendre que le dropdown soit complètement réinitialisé
                    doc_type_container = wait.until(
                        EC.presence_of_element_located((By.ID, "mainTab:form7:comp1"))
                    )
                    
                    # Vérifier que le dropdown n'est pas déjà ouvert
                    try:
                        open_panel = driver.find_element(By.CSS_SELECTOR, "div#mainTab\\:form7\\:comp1_panel[style*='display: block']")
                        print(f"         ℹ️  Dropdown déjà ouvert, fermeture...")
                        driver.execute_script("arguments[0].style.display = 'none';", open_panel)
                        time.sleep(1)
                    except:
                        pass
                    
                    # Scroll et focus sur le conteneur
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", doc_type_container)
                    time.sleep(0.5)
                    
                    # Chercher le trigger
                    doc_type_trigger = doc_type_container.find_element(By.CSS_SELECTOR, "div.ui-selectonemenu-trigger")
                    
                    # Essayer click standard
                    doc_type_trigger.click()
                    time.sleep(1.5)
                    
                    # Vérifier si le dropdown s'est ouvert
                    dropdown_panel = driver.find_element(By.CSS_SELECTOR, "div#mainTab\\:form7\\:comp1_panel[style*='display: block']")
                    if dropdown_panel:
                        print("         ✓ Dropdown type document ouvert (méthode click standard)")
                        dropdown_opened = True
                        break
                    
                except Exception as click_err:
                    # Méthode 2: JavaScript click
                    print(f"         ℹ️  Click standard échoué, tentative JavaScript...")
                    try:
                        doc_type_container = driver.find_element(By.ID, "mainTab:form7:comp1")
                        doc_type_trigger = doc_type_container.find_element(By.CSS_SELECTOR, "div.ui-selectonemenu-trigger")
                        driver.execute_script("arguments[0].click();", doc_type_trigger)
                        time.sleep(1.5)
                        
                        # Vérifier ouverture
                        dropdown_panel = driver.find_element(By.CSS_SELECTOR, "div#mainTab\\:form7\\:comp1_panel[style*='display: block']")
                        if dropdown_panel:
                            print("         ✓ Dropdown type document ouvert (méthode JavaScript)")
                            dropdown_opened = True
                            break
                    except Exception as js_err:
                        # Méthode 3: Click sur le label
                        print(f"         ℹ️  JavaScript échoué, tentative click sur label...")
                        try:
                            doc_type_label = driver.find_element(By.ID, "mainTab:form7:comp1_label")
                            doc_type_label.click()
                            time.sleep(1.5)
                            
                            # Vérifier ouverture
                            dropdown_panel = driver.find_element(By.CSS_SELECTOR, "div#mainTab\\:form7\\:comp1_panel[style*='display: block']")
                            if dropdown_panel:
                                print("         ✓ Dropdown type document ouvert (méthode label)")
                                dropdown_opened = True
                                break
                        except:
                            pass
            
            except Exception as e:
                if attempt == max_attempts:
                    print(f"      ❌ Impossible d'ouvrir le dropdown après {max_attempts} tentatives: {e}")
                    return False
        
        if not dropdown_opened:
            print(f"      ❌ Dropdown non ouvert après {max_attempts} tentatives")
            return False
        
        # Sélectionner l'option "FACTURE"
        try:
            doc_type_option = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//li[@data-label='FACTURE']"))
            )
            doc_type_option.click()
            print("         ✓ Type sélectionné: A0006 (FACTURE)")
            
            # Attendre que le blocker AJAX disparaisse après sélection du type
            print("         ⏳ Attente mise à jour formulaire...")
            if wait_for_ui_blocker_disappear(driver, timeout=5):
                print("         ✓ Formulaire mis à jour")
            time.sleep(1)
        except Exception as e:
            print(f"      ❌ Erreur sélection option FACTURE: {e}")
            return False
        
        # 7.2.2: Entrer la référence "mnN"
        print(f"      2️⃣ Saisie de la référence '{mn_reference}'...")
        
        # Attendre que le champ référence soit actif après la sélection du type
        ref_input_success = False
        max_ref_attempts = 3
        
        for attempt in range(1, max_ref_attempts + 1):
            try:
                if attempt > 1:
                    print(f"      🔄 Tentative {attempt}/{max_ref_attempts}...")
                    time.sleep(1)
                
                # Re-localiser l'input à chaque tentative (éviter stale element)
                ref_input = wait.until(
                    EC.presence_of_element_located((By.XPATH, "//label[contains(text(), 'Référence')]/parent::td/following-sibling::td//input[@type='text' and @maxlength='10']"))
                )
                
                # Attendre qu'il soit enabled
                wait.until(lambda d: ref_input.is_enabled() and not ref_input.get_attribute('disabled'))
                
                # Clear et saisie
                ref_input.clear()
                time.sleep(0.3)
                ref_input.send_keys(mn_reference)
                
                # Vérifier que la valeur a été saisie
                if ref_input.get_attribute('value') == mn_reference:
                    print(f"         ✓ Référence: {mn_reference}")
                    ref_input_success = True
                    time.sleep(0.5)
                    break
                else:
                    print(f"         ⚠️  Valeur non enregistrée (tentative {attempt})")
                    
            except Exception as e:
                if attempt == max_ref_attempts:
                    print(f"      ⚠️  Méthode XPath échouée après {max_ref_attempts} tentatives: {e}")
                    # Méthode alternative: chercher par pattern d'ID
                    try:
                        print("      🔄 Tentative alternative (par ID)...")
                        ref_input_alt = wait.until(
                            EC.presence_of_element_located((By.XPATH, "//input[contains(@id, 'mainTab:form7:j_id') and @type='text' and @maxlength='10' and not(@disabled)]"))
                        )
                        ref_input_alt.clear()
                        time.sleep(0.3)
                        ref_input_alt.send_keys(mn_reference)
                        if ref_input_alt.get_attribute('value') == mn_reference:
                            print(f"         ✓ Référence: {mn_reference} (méthode alternative)")
                            ref_input_success = True
                            time.sleep(0.5)
                        else:
                            print(f"         ⚠️  Valeur non enregistrée (méthode alternative)")
                    except Exception as e2:
                        print(f"      ❌ Erreur saisie référence (méthode alternative): {e2}")
        
        if not ref_input_success:
            print("      ❌ Impossible de saisir la référence après toutes les tentatives")
            return False
        
        # 7.2.3: Sélectionner la date actuelle (à nouveau)
        print("      3️⃣ Sélection de la date actuelle...")
        try:
            date_picker_btn = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.ui-datepicker-trigger"))
            )
            date_picker_btn.click()
            time.sleep(1)
            
            today_cell = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "td.ui-datepicker-today a"))
            )
            today_cell.click()
            print("         ✓ Date actuelle sélectionnée")
            time.sleep(0.5)
        except Exception as e:
            print(f"      ❌ Erreur sélection date: {e}")
            return False
        
        # 7.2.4: Upload du fichier MN PDF
        print(f"      4️⃣ Upload du fichier {mn_filename}...")
        try:
            # Chercher le fichier mn*.pdf dans le dossier LTA
            mn_file_path = os.path.join(lta_folder_path, mn_filename)
            
            if not os.path.exists(mn_file_path):
                print(f"      ❌ Fichier MN introuvable: {mn_filename}")
                print(f"         Chemin recherché: {mn_file_path}")
                return False
            
            print(f"         📄 Fichier trouvé: {mn_filename}")
            
            # Upload du fichier
            file_input = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form7:comp2_input"))
            )
            absolute_path = os.path.abspath(mn_file_path)
            file_input.send_keys(absolute_path)
            print(f"         ✓ Fichier MN uploadé: {mn_filename}")
            
            # Attendre que le blocker d'upload disparaisse
            print("         ⏳ Attente fin d'upload...")
            if wait_for_ui_blocker_disappear(driver, timeout=10):
                print("         ✓ Upload terminé (blocker disparu)")
            else:
                print("         ⚠️  Timeout blocker upload - continuons")
            time.sleep(2)
        except Exception as e:
            print(f"      ❌ Erreur upload fichier MN: {e}")
            return False
        
        print("\n   ✅ Documents uploadés avec succès (LTA + MN)")
        
        # ==================================================================
        # ÉTAPE 8: Retour à l'onglet "Entête" pour validation finale
        # ==================================================================
        print("\n   📋 Navigation vers l'onglet 'Entête' pour validation...")
        
        # Attendre que le UI blocker disparaisse
        wait_for_ui_blocker_disappear(driver, timeout=10)
        time.sleep(0.5)
        
        # Essayer plusieurs méthodes pour garantir qu'on est sur l'onglet Entête
        entete_navigation_success = False
        
        for attempt in range(1, 4):  # 3 tentatives maximum
            try:
                if attempt > 1:
                    print(f"      🔄 Tentative {attempt}/3...")
                    time.sleep(1)
                    wait_for_ui_blocker_disappear(driver, timeout=5)
                
                # Vérifier d'abord si l'onglet est déjà actif
                try:
                    active_tab = driver.find_element(By.CSS_SELECTOR, "li.ui-tabs-selected.ui-state-active a[href='#mainTab:tab0']")
                    if active_tab:
                        # Vérifier que le panneau Entête est visible
                        try:
                            entete_panel = driver.find_element(By.ID, "mainTab:tab0")
                            if entete_panel.is_displayed():
                                print("      ✓ Onglet 'Entête' déjà actif et visible")
                                entete_navigation_success = True
                                break
                        except:
                            pass
                except:
                    pass  # Onglet pas actif, continuer avec le clic
                
                # Méthode 1: Click sur l'onglet Entête
                try:
                    entete_tab = wait.until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "a[href='#mainTab:tab0']"))
                    )
                    
                    # Scroll into view
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", entete_tab)
                    time.sleep(0.5)
                    
                    # Attendre que le blocker disparaisse avant de cliquer
                    wait_for_ui_blocker_disappear(driver, timeout=5)
                    
                    # Essayer click standard puis JavaScript
                    try:
                        # Vérifier que l'élément est clickable
                        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#mainTab:tab0']")))
                        entete_tab.click()
                    except Exception as click_err:
                        print(f"      ⚠️  Clic normal échoué, utilisation JavaScript: {click_err}")
                        driver.execute_script("arguments[0].click();", entete_tab)
                    
                    time.sleep(2)
                    
                    # Attendre que le blocker disparaisse après le clic
                    wait_for_ui_blocker_disappear(driver, timeout=5)
                    
                    # Vérifier que l'onglet est bien actif
                    try:
                        active_tab = driver.find_element(By.CSS_SELECTOR, "li.ui-tabs-selected.ui-state-active a[href='#mainTab:tab0']")
                        if active_tab:
                            # Vérifier aussi que le panneau est visible
                            entete_panel = driver.find_element(By.ID, "mainTab:tab0")
                            if entete_panel.is_displayed():
                                print("      ✓ Onglet 'Entête' actif et visible")
                                entete_navigation_success = True
                                break
                    except:
                        # Vérifier alternative: le panneau Entête est visible
                        try:
                            entete_panel = driver.find_element(By.ID, "mainTab:tab0")
                            if entete_panel.is_displayed():
                                print("      ✓ Panneau 'Entête' visible")
                                entete_navigation_success = True
                                break
                        except:
                            pass
                except Exception as nav_err:
                    print(f"      ⚠️  Erreur navigation tentative {attempt}: {nav_err}")
                    
            except Exception as e:
                if attempt == 3:
                    print(f"      ❌ Impossible d'accéder à l'onglet Entête après 3 tentatives: {e}")
                    # Dernière tentative avec JavaScript direct
                    try:
                        print("      🔄 Dernière tentative avec JavaScript direct...")
                        driver.execute_script("""
                            var tab = document.querySelector('a[href="#mainTab:tab0"]');
                            if (tab) {
                                tab.click();
                            }
                        """)
                        time.sleep(2)
                        wait_for_ui_blocker_disappear(driver, timeout=5)
                        # Vérifier si ça a marché
                        entete_panel = driver.find_element(By.ID, "mainTab:tab0")
                        if entete_panel.is_displayed():
                            print("      ✓ Onglet 'Entête' activé via JavaScript")
                            entete_navigation_success = True
                        else:
                            return False
                    except:
                        return False
        
        if not entete_navigation_success:
            print("      ❌ Échec navigation vers 'Entête' - vérification du bouton VALIDER...")
            # Dernière vérification: le bouton VALIDER est-il accessible ?
            try:
                wait.until(EC.presence_of_element_located((By.ID, "secure__2003")))
                print("      ℹ️  Bouton VALIDER détecté - continuons")
            except:
                print("      ❌ Bouton VALIDER non accessible - abandon")
                return False
        
        # ==================================================================
        # ÉTAPE 9: Premier clic sur "VALIDER" pour révéler le champ Commerce électronique
        # ==================================================================
        print("\n   🔍 Premier clic sur 'VALIDER' pour révéler les champs manquants...")
        
        # Vérifier qu'on est bien sur l'onglet Entête avant de cliquer
        try:
            entete_panel = driver.find_element(By.ID, "mainTab:tab0")
            if not entete_panel.is_displayed():
                print("      ⚠️  Panneau Entête non visible, tentative de navigation...")
                entete_tab = driver.find_element(By.CSS_SELECTOR, "a[href='#mainTab:tab0']")
                driver.execute_script("arguments[0].click();", entete_tab)
                time.sleep(2)
                wait_for_ui_blocker_disappear(driver, timeout=5)
        except:
            print("      ⚠️  Impossible de vérifier l'onglet Entête, continuation...")
        
        # Attendre que le blocker UI disparaisse avant de cliquer
        wait_for_ui_blocker_disappear(driver, timeout=10)
        time.sleep(0.5)
        
        # Tentative avec retry pour cliquer sur VALIDER
        valider_clicked = False
        for valider_attempt in range(1, 4):  # 3 tentatives
            try:
                if valider_attempt > 1:
                    print(f"      🔄 Tentative clic VALIDER {valider_attempt}/3...")
                    time.sleep(1)
                    wait_for_ui_blocker_disappear(driver, timeout=5)
                
                # Chercher le bouton VALIDER
                valider_btn = wait.until(
                    EC.presence_of_element_located((By.ID, "secure__2003"))
                )
                
                # Scroll into view
                driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", valider_btn)
                time.sleep(0.3)
                
                # Attendre que le blocker disparaisse
                wait_for_ui_blocker_disappear(driver, timeout=5)
                
                # Essayer clic normal puis JavaScript
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "secure__2003")))
                    valider_btn.click()
                    print(f"      ✓ Bouton 'VALIDER' cliqué (tentative {valider_attempt})")
                except Exception as click_err:
                    print(f"      ⚠️  Clic normal échoué, utilisation JavaScript: {click_err}")
                    driver.execute_script("arguments[0].click();", valider_btn)
                    print(f"      ✓ Bouton 'VALIDER' cliqué via JavaScript (tentative {valider_attempt})")
                
                # Attendre que le blocker apparaisse puis disparaisse (indique que le traitement est en cours)
                time.sleep(1)
                wait_for_ui_blocker_disappear(driver, timeout=10)
                time.sleep(1)
                
                valider_clicked = True
                break
                
            except Exception as e:
                if valider_attempt == 3:
                    print(f"      ❌ Erreur clic VALIDER après 3 tentatives: {e}")
                    return False
                else:
                    print(f"      ⚠️  Erreur tentative {valider_attempt}: {e}")
        
        if not valider_clicked:
            print("      ❌ Impossible de cliquer sur VALIDER")
            return False
        
        print("      ✓ Champ 'Commerce électronique' devrait être révélé...")
        
        # ==================================================================
        # ÉTAPE 10: Cocher "Commerce électronique - Oui"
        # ==================================================================
        print("\n   ☑️  Activation 'Commerce électronique - Oui'...")
        
        # Vérifier qu'on est bien sur l'onglet Entête
        try:
            entete_panel = driver.find_element(By.ID, "mainTab:tab0")
            if not entete_panel.is_displayed():
                print("      ⚠️  Panneau Entête non visible, tentative navigation...")
                entete_tab = driver.find_element(By.CSS_SELECTOR, "a[href='#mainTab:tab0']")
                driver.execute_script("arguments[0].click();", entete_tab)
                time.sleep(2)
                wait_for_ui_blocker_disappear(driver, timeout=10)
        except Exception as nav_err:
            print(f"      ⚠️  Vérification onglet Entête échouée: {nav_err}")
        
        # Attendre que le blocker UI disparaisse
        wait_for_ui_blocker_disappear(driver, timeout=10)
        time.sleep(1)
        
        # Attendre que le champ soit révélé après le clic VALIDER (peut prendre du temps)
        print("      ⏳ Attente du champ Commerce électronique...")
        max_wait_attempts = 15  # Augmenté à 15 tentatives
        commerce_table = None
        
        for wait_attempt in range(max_wait_attempts):
            try:
                # Chercher le tableau Commerce électronique
                commerce_table = driver.find_element(By.ID, "mainTab:form0:commerceElectronique")
                
                # Vérifier que le tableau est visible
                if not commerce_table.is_displayed():
                    if wait_attempt < max_wait_attempts - 1:
                        print(f"      ⏳ Tableau trouvé mais non visible (tentative {wait_attempt + 1}/{max_wait_attempts})...")
                        time.sleep(1)
                        continue
                
                # Vérifier que le tableau contient des radios
                radios = commerce_table.find_elements(By.CSS_SELECTOR, "input[type='radio']")
                if len(radios) >= 2:
                    print(f"      ✓ Tableau Commerce électronique trouvé ({len(radios)} radios)")
                    break
                else:
                    if wait_attempt < max_wait_attempts - 1:
                        print(f"      ⏳ Tableau trouvé mais radios non prêts ({len(radios)} radios, tentative {wait_attempt + 1}/{max_wait_attempts})")
                        time.sleep(1)
                    else:
                        print(f"      ⚠️  Tableau trouvé mais seulement {len(radios)} radio(s) disponibles")
            except Exception as find_err:
                if wait_attempt < max_wait_attempts - 1:
                    # Vérifier si le blocker UI est présent (peut bloquer la recherche)
                    try:
                        blocker = driver.find_element(By.CSS_SELECTOR, "div.ui-blockui")
                        if blocker.is_displayed():
                            print(f"      ⏳ Blocker UI présent, attente... (tentative {wait_attempt + 1}/{max_wait_attempts})")
                            wait_for_ui_blocker_disappear(driver, timeout=5)
                        else:
                            print(f"      ⏳ Champ non encore révélé (tentative {wait_attempt + 1}/{max_wait_attempts})...")
                    except:
                        print(f"      ⏳ Champ non encore révélé (tentative {wait_attempt + 1}/{max_wait_attempts})...")
                    time.sleep(1)
                else:
                    print(f"      ❌ Champ Commerce électronique introuvable après {max_wait_attempts} tentatives")
                    print(f"      ⚠️  Dernière erreur: {find_err}")
                    # Dernière tentative: chercher dans tout le DOM
                    try:
                        all_tables = driver.find_elements(By.TAG_NAME, "table")
                        print(f"      ℹ️  Nombre de tables trouvées dans le DOM: {len(all_tables)}")
                        # Chercher par XPath alternatif
                        commerce_table_alt = driver.find_element(By.XPATH, "//table[contains(@id, 'commerceElectronique')]")
                        if commerce_table_alt:
                            print("      ✓ Tableau trouvé via XPath alternatif")
                            commerce_table = commerce_table_alt
                            break
                    except:
                        pass
                    return False
        
        if not commerce_table:
            print("      ❌ Tableau Commerce électronique non trouvé")
            return False
        
        time.sleep(1)  # Délai supplémentaire pour stabilité
        
        # Tentative avec retry pour cocher le radio "Oui"
        max_retries = 5
        for attempt in range(max_retries):
            try:
                # Méthode 1: Cliquer sur le div.ui-radiobutton-box (interface visuelle)
                commerce_elec_radios = driver.find_elements(By.CSS_SELECTOR, "table#mainTab\\:form0\\:commerceElectronique div.ui-radiobutton-box")
                if len(commerce_elec_radios) >= 1:
                    # Attendre que l'élément soit cliquable
                    radio_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "table#mainTab\\:form0\\:commerceElectronique div.ui-radiobutton-box")))
                    radio_box.click()
                    print("      ✓ Radio 'Commerce électronique - Oui' coché (méthode click)")
                    time.sleep(1)
                    break  # Succès, sortir de la boucle
                else:
                    raise Exception(f"Aucun radio button trouvé (nombre: {len(commerce_elec_radios)})")
                    
            except Exception as click_err:
                if attempt < max_retries - 1:
                    print(f"      ⏳ Tentative {attempt + 1}/{max_retries} échouée: {str(click_err)[:80]}")
                    time.sleep(1)
                else:
                    # Dernière tentative avec JavaScript
                    print(f"      ⚠️  Méthode click échouée, utilisation de JavaScript...")
                    try:
                        js_code = """
                        // Essayer plusieurs méthodes pour cocher le radio "Oui" (premier radio)
                        var radio = document.getElementById('mainTab:form0:commerceElectronique:0');
                        if (radio) {
                            radio.checked = true;
                            radio.click();
                            
                            // Déclencher les événements pour notifier PrimeFaces
                            var changeEvent = new Event('change', { bubbles: true });
                            radio.dispatchEvent(changeEvent);
                            
                            var clickEvent = new Event('click', { bubbles: true });
                            radio.dispatchEvent(clickEvent);
                            
                            // Aussi cliquer sur la div visible
                            var radioBox = radio.closest('.ui-radiobutton').querySelector('.ui-radiobutton-box');
                            if (radioBox) {
                                radioBox.click();
                            }
                            
                            return 'success';
                        } else {
                            // Chercher le premier input radio dans le tableau
                            var table = document.getElementById('mainTab:form0:commerceElectronique');
                            if (table) {
                                var radios = table.querySelectorAll('input[type="radio"]');
                                if (radios.length > 0) {
                                    radios[0].checked = true;
                                    radios[0].click();
                                    
                                    var changeEvent = new Event('change', { bubbles: true });
                                    radios[0].dispatchEvent(changeEvent);
                                    
                                    return 'success-alternative';
                                }
                            }
                            throw new Error('Radio button not found in DOM');
                        }
                        """
                        result = driver.execute_script(js_code)
                        print(f"      ✓ Radio coché via JavaScript ({result})")
                        time.sleep(0.5)
                        break  # Succès avec JavaScript, sortir de la boucle
                    except Exception as js_error:
                        print(f"      ❌ Impossible de cocher Commerce électronique (toutes méthodes échouées): {js_error}")
                        traceback.print_exc()
                        return False
        
        # ==================================================================
        # ÉTAPE 11: Deuxième clic sur "VALIDER" pour soumettre la déclaration
        # ==================================================================
        print("\n   ✅ Validation finale de la déclaration...")
        try:
            # Cliquer sur le bouton VALIDER (deuxième fois - soumission finale)
            valider_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "secure__2003"))
            )
            valider_btn.click()
            print("      ✓ Bouton 'VALIDER' cliqué (2ème fois - soumission)")
            
            # Attendre que la validation soit traitée
            print("      ⏳ Attente de la validation...")
            
            # Attendre que le blocker UI disparaisse (validation en cours)
            if wait_for_ui_blocker_disappear(driver, timeout=15):
                print("      ✓ Validation terminée (blocker disparu)")
            else:
                print("      ⚠️  Timeout blocker validation - continuons")
            
            # Pause supplémentaire pour stabilité
            time.sleep(3)
            
            # ==================================================================
            # VÉRIFICATION DES MESSAGES DE VALIDATION (SUCCÈS OU ERREUR)
            # ==================================================================
            validation_error = False
            error_messages = []
            
            # D'abord, vérifier s'il y a des messages de SUCCÈS (ui-messages-info)
            has_success_message = False
            try:
                info_containers = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-info")
                visible_info = [c for c in info_containers if c.is_displayed()]
                if visible_info:
                    has_success_message = True
                    print("      ℹ️  Message d'information détecté (probablement succès)")
            except:
                pass
            
            # Chercher messages d'erreur
            try:
                # Chercher UNIQUEMENT les conteneurs d'erreur VISIBLES
                all_error_containers = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-error")
                error_containers = [c for c in all_error_containers if c.is_displayed()]
                
                # ⚠️ IMPORTANT: Si on a un message de succès ET aucun conteneur d'erreur visible,
                # alors c'est une validation réussie (ne pas traiter comme erreur)
                if has_success_message and (not error_containers or len(error_containers) == 0):
                    print("      ✅ Message de succès détecté sans erreur - validation réussie")
                    validation_error = False
                    error_messages = []
                
                elif error_containers and len(error_containers) > 0:
                    print("      ⚠️  Conteneur d'erreur détecté")
                    
                    # ==================================================================
                    # VÉRIFIER SI BOUTON "DÉTAILS" EXISTE = PLUSIEURS ERREURS
                    # ==================================================================
                    has_details_button = False
                    try:
                        # Chercher le bouton "Détails" UNIQUEMENT dans les conteneurs d'erreur
                        for error_container in error_containers:
                            try:
                                details_btn = error_container.find_element(By.ID, "rapportMsgForm:showErrors")
                                if details_btn and details_btn.is_displayed():
                                    has_details_button = True
                                    break
                            except:
                                continue
                    except:
                        pass
                    
                    if has_details_button:
                        # ==================================================================
                        # BOUTON "DÉTAILS" DÉTECTÉ → TOUJOURS TRAITER COMME ERREUR
                        # ==================================================================
                        # ⚠️ RÈGLE CRITIQUE: Si le bouton "Détails" existe, c'est qu'il y a
                        # une liste d'erreurs cachée, même si le span error-detail est vide!
                        print(f"      ⚠️  Bouton 'Détails' détecté → Erreurs multiples présentes")
                        validation_error = True
                        error_messages.append("Plusieurs erreurs de validation détectées (cliquer sur 'Détails' dans BADR pour voir la liste)")
                        print(f"      ❌ Déclaration invalide (erreurs multiples)")
                        
                    else:
                        # ==================================================================
                        # MESSAGE UNIQUE → EXTRAIRE ET VÉRIFIER SI C'EST "COMMERCE ÉLECTRONIQUE"
                        # ==================================================================
                        print(f"      ℹ️  Pas de bouton 'Détails' → Message unique")
                        
                        # Extraire le message unique
                        try:
                            error_details = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-error-detail")
                            if error_details and len(error_details) > 0:
                                for detail in error_details:
                                    try:
                                        error_text = detail.text.strip()
                                        if error_text:
                                            lines = error_text.split('\n')
                                            for line in lines:
                                                line = line.strip()
                                                if line and line not in error_messages:
                                                    error_messages.append(line)
                                    except:
                                        continue
                        except Exception as extract_err:
                            print(f"      ⚠️  Erreur extraction message: {extract_err}")
                        
                        # Si pas de message extrait, chercher dans le conteneur
                        if not error_messages:
                            try:
                                for container in error_containers:
                                    try:
                                        container_text = container.text.strip()
                                        if container_text and "Erreur" in container_text:
                                            lines = container_text.split('\n')
                                            for line in lines:
                                                line = line.strip()
                                                if line and line != "Erreur :" and line not in error_messages:
                                                    error_messages.append(line)
                                    except:
                                        continue
                            except:
                                pass
                        
                        # Vérifier si c'est le message "Commerce électronique"
                        if len(error_messages) == 1:
                            single_msg = error_messages[0]
                            if "commerce électronique" in single_msg.lower() and "information obligatoire" in single_msg.lower():
                                # Message informatif → Ignorer
                                print(f"      ℹ️  Message informatif ignoré: {single_msg}")
                                error_messages = []
                                print("      ✅ Aucune erreur bloquante - validation réussie")
                            else:
                                # Erreur réelle unique
                                validation_error = True
                                print(f"      ⚠️  Erreur de validation détectée: {single_msg[:80]}{'...' if len(single_msg) > 80 else ''}")
                        elif len(error_messages) > 1:
                            # Cas rare: plusieurs messages extraits même sans bouton Détails
                            validation_error = True
                            print(f"      ⚠️  {len(error_messages)} erreurs détectées")
                        else:
                            # Aucun message d'erreur extrait
                            # Si on a un message de succès visible, c'est une validation réussie
                            if has_success_message:
                                print("      ✅ Conteneur d'erreur vide + message de succès → validation réussie")
                            else:
                                print("      ✅ Aucune erreur bloquante - validation réussie")
            except Exception as e:
                # Erreur lors de la détection - considérer comme pas d'erreur
                print(f"      ℹ️  Impossible de détecter les messages (probablement aucune erreur)")
            
            # Si erreur détectée, créer un log détaillé et marquer dans Excel
            if validation_error:
                print(f"\n   ❌ DÉCLARATION REFUSÉE - Erreurs de validation")
                
                # Extraire le numéro du DUM
                sheet_name = dum_data.get('sheet_name', '')
                dum_number = int(sheet_name.split()[-1]) if sheet_name.startswith('Sheet') else 1
                
                # ==================================================================
                # TENTER D'EXTRAIRE LA SÉRIE MÊME EN CAS D'ERREUR
                # ==================================================================
                # La série peut être visible dans la table même si la validation a échoué
                dum_series = None
                try:
                    # Chercher la table de référence
                    reference_table = driver.find_element(By.ID, "mainTab:form0:j_id_3p_d")
                    rows = reference_table.find_elements(By.TAG_NAME, "tr")
                    if len(rows) >= 2:
                        data_row = rows[1]
                        cells = data_row.find_elements(By.TAG_NAME, "td")
                        if len(cells) >= 5:
                            serie = cells[3].text.strip()
                            cle = cells[4].text.strip()
                            if serie and cle:
                                dum_series = f"{serie}{cle}"
                                print(f"      ℹ️  Série extraite malgré l'erreur: {dum_series}")
                except Exception as serie_err:
                    print(f"      ⚠️  Impossible d'extraire la série: {serie_err}")
                
                # Si série non trouvée, utiliser un placeholder
                if not dum_series:
                    dum_series = "SÉRIE_INCONNUE"
                
                # Créer le fichier d'erreur
                lta_name_safe = lta_name.replace(' ', '_')
                error_filename = f"error-validating-declaration-dedouanement-{lta_name_safe}-DUM{dum_number}.txt"
                parent_dir = os.path.dirname(lta_folder_path)
                error_filepath = os.path.join(parent_dir, error_filename)
                
                from datetime import datetime
                current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Analyser les catégories d'erreurs
                error_categories = []
                if any("crédit" in msg.lower() and "enlèvement" in msg.lower() for msg in error_messages):
                    error_categories.append("Crédit d'enlèvement manquant ou invalide")
                if any("ti-a" in msg.lower() or "propriete ti" in msg.lower() for msg in error_messages):
                    error_categories.append("Propriétés techniques incomplètes (TI-A)")
                if any("quantité" in msg.lower() for msg in error_messages):
                    error_categories.append("Données quantitatives manquantes")
                if any("tva" in msg.lower() for msg in error_messages):
                    error_categories.append("Informations TVA manquantes")
                if any("poids net" in msg.lower() for msg in error_messages):
                    error_categories.append("Divergence de poids net")
                
                # Créer le fichier log détaillé
                with open(error_filepath, 'w', encoding='utf-8') as f:
                    f.write("=" * 70 + "\n")
                    f.write("ERREUR - VALIDATION DÉCLARATION DÉDOUANEMENT\n")
                    f.write("=" * 70 + "\n\n")
                    
                    f.write(f"LTA: {lta_name}\n")
                    f.write(f"DUM: {dum_number}\n")
                    f.write(f"Date: {current_datetime}\n")
                    f.write(f"Étape: Validation finale déclaration dédouanement\n\n")
                    
                    f.write("-" * 70 + "\n")
                    f.write("DÉTAILS DUM:\n")
                    f.write("-" * 70 + "\n")
                    f.write(f"Sheet Name: {dum_data.get('sheet_name', 'N/A')}\n")
                    f.write(f"Total Value: {dum_data.get('total_value', 0)}\n")
                    f.write(f"Gross Weight: {dum_data.get('total_gross_weight', 0)}\n")
                    f.write(f"Positions: {dum_data.get('total_positions', 0)}\n")
                    f.write(f"Freight: {dum_data.get('total_freight', 0)}\n")
                    f.write(f"Insurance: {dum_data.get('insurance', 0)}\n")
                    f.write(f"Cartons: {dum_data.get('cartons', 0)}\n\n")
                    
                    f.write("=" * 70 + "\n")
                    f.write("MESSAGES D'ERREUR DU SYSTÈME:\n")
                    f.write("=" * 70 + "\n\n")
                    for i, msg in enumerate(error_messages, 1):
                        f.write(f"  • {msg}\n")
                    f.write("\n")
                    
                    if error_categories:
                        f.write("=" * 70 + "\n")
                        f.write("CATÉGORIES D'ERREURS DÉTECTÉES:\n")
                        f.write("=" * 70 + "\n\n")
                        for cat in error_categories:
                            f.write(f"  ✗ {cat}\n")
                        f.write("\n")
                    
                    f.write("=" * 70 + "\n")
                    f.write("ACTION REQUISE:\n")
                    f.write("=" * 70 + "\n\n")
                    f.write("Ce DUM n'a pas pu être validé automatiquement.\n")
                    f.write("Veuillez:\n")
                    f.write("  1. Vérifier les données du fichier Excel source\n")
                    f.write("  2. Corriger les informations manquantes\n")
                    f.write("  3. Créer la déclaration manuellement ou relancer le script\n\n")
                    
                    f.write("=" * 70 + "\n")
                    f.write("FICHIERS CONCERNÉS:\n")
                    f.write("=" * 70 + "\n\n")
                    f.write(f"  • Sheet Excel: {dum_data.get('sheet_name', 'N/A')}\n")
                    f.write(f"  • LTA Folder: {lta_name}/\n\n")
                    
                    f.write("=" * 70 + "\n")
                    f.write("FIN DU RAPPORT D'ERREUR\n")
                    f.write("=" * 70 + "\n")
                
                print(f"      ✓ Fichier d'erreur créé: {error_filename}")
                
                # Marquer "error" avec série dans le fichier Excel (colonne C)
                # Utiliser la fonction centralisée
                mark_dum_as_error_in_excel(lta_folder_path, dum_number, serie=dum_series)
                
                print(f"      ❌ Déclaration refusée - Erreurs détectées")
                print(f"      ⏭️  Passage au DUM suivant...")
                
                # Retourner à l'accueil pour continuer avec le prochain DUM
                try:
                    time.sleep(2)
                    accueil_btn = wait.until(EC.element_to_be_clickable((By.ID, "quitter")))
                    try:
                        accueil_btn.click()
                    except:
                        driver.execute_script("arguments[0].click();", accueil_btn)
                    time.sleep(3)
                    driver.switch_to.default_content()
                    print("      ✓ Retour à l'accueil réussi")
                except:
                    try:
                        driver.switch_to.default_content()
                    except:
                        pass
                
                return False  # Indiquer l'échec de ce DUM
            
            # Si pas d'erreur, chercher message de succès (info)
            try:
                info_containers = driver.find_elements(By.CSS_SELECTOR, "div.ui-messages-info")
                if info_containers and len(info_containers) > 0:
                    info_details = driver.find_elements(By.CSS_SELECTOR, "span.ui-messages-info-detail")
                    if info_details and len(info_details) > 0:
                        info_text = info_details[0].text.strip()
                        if info_text:
                            print(f"      ℹ️  Message info: {info_text[:80]}...")
            except:
                pass
            
            print("      ✅ Déclaration validée avec succès")
            
        except Exception as e:
            print(f"      ❌ Erreur validation finale: {e}")
            return False
        
        print(f"\n   ✅ DUM {dum_data.get('sheet_name', 'DUM')} traité et validé avec succès!")
        
        # ==================================================================
        # ÉTAPE 12: Extraire la référence de déclaration et la sauvegarder
        # ==================================================================
        print("\n   📋 Extraction de la référence de déclaration...")
        try:
            # Attendre que la table de référence soit visible
            time.sleep(2)
            
            # Localiser la table de référence
            reference_table = wait.until(
                EC.presence_of_element_located((By.ID, "mainTab:form0:j_id_3p_d"))
            )
            
            # Extraire les cellules de la deuxième ligne (ligne de données)
            rows = reference_table.find_elements(By.TAG_NAME, "tr")
            if len(rows) >= 2:
                data_row = rows[1]  # Deuxième ligne (index 1)
                cells = data_row.find_elements(By.TAG_NAME, "td")
                
                if len(cells) >= 5:
                    # Extraire Série (4ème colonne, index 3)
                    serie = cells[3].text.strip()
                    # Extraire Clé (5ème colonne, index 4)
                    cle = cells[4].text.strip()
                    
                    # Combiner pour créer la référence complète
                    dum_reference = f"{serie}{cle}"
                    
                    print(f"      ✓ Référence extraite: {dum_reference}")
                    print(f"         - Série: {serie}")
                    print(f"         - Clé: {cle}")
                    
                    # Sauvegarder la référence dans result_LTAS.txt
                    save_dum_reference(lta_folder_path, dum_reference)
                    
                    # Extraire le numéro du DUM depuis sheet_name (ex: "Sheet 1" → 1)
                    sheet_name = dum_data.get('sheet_name', '')
                    dum_number = int(sheet_name.split()[-1]) if sheet_name.startswith('Sheet') else 1
                    
                    # Sauvegarder la série dans generated_excel
                    save_dum_series_to_excel(lta_folder_path, dum_number, dum_reference)
                    
                else:
                    print(f"      ⚠️  Table de référence incomplète (cellules: {len(cells)})")
                    dum_reference = "REFERENCE_INCOMPLETE"
            else:
                print(f"      ⚠️  Table de référence incomplète (lignes: {len(rows)})")
                dum_reference = "REFERENCE_INCOMPLETE"
                
        except Exception as e:
            print(f"      ❌ Erreur extraction référence: {e}")
            dum_reference = "REFERENCE_ERROR"
            traceback.print_exc()
        
        # ==================================================================
        # ÉTAPE 13: Retour à l'accueil pour traiter le prochain DUM
        # ==================================================================
        print("\n   🏠 Retour à l'accueil pour le prochain DUM...")
        try:
            # Attendre que la page soit complètement stable après validation
            print("      ⏳ Attente stabilisation page...")
            time.sleep(3)
            
            # Attendre que le blocker soit complètement disparu
            if wait_for_ui_blocker_disappear(driver, timeout=10):
                print("      ✓ Page stabilisée (blocker disparu)")
            else:
                print("      ⚠️  Timeout blocker - continuons")
            
            # Pause supplémentaire avant de cliquer sur Accueil
            time.sleep(2)
            
            # Cliquer sur le bouton "Accueil" (id="quitter")
            accueil_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "quitter"))
            )
            
            try:
                accueil_btn.click()
                print("      ✓ Bouton 'Accueil' cliqué")
            except Exception as click_error:
                print(f"      ⚠️  Clic normal intercepté, utilisation de JavaScript...")
                driver.execute_script("arguments[0].click();", accueil_btn)
                print("      ✓ Bouton 'Accueil' cliqué (via JavaScript)")
            
            # Attendre que le blocker de navigation disparaisse
            print("      ⏳ Attente navigation vers accueil...")
            if wait_for_ui_blocker_disappear(driver, timeout=10):
                print("      ✓ Navigation terminée (blocker disparu)")
            else:
                print("      ⚠️  Timeout blocker navigation")
            
            # Attendre le retour à la page d'accueil
            time.sleep(3)
            
            # IMPORTANT: Sortir de l'iframe pour revenir au contexte principal
            driver.switch_to.default_content()
            print("      ✓ Sorti de l'iframe, retour au contexte principal")
            
            print("      ✓ Retour à l'accueil réussi")
            
        except Exception as e:
            print(f"      ❌ Erreur retour accueil: {e}")
            traceback.print_exc()
            # Essayer quand même de sortir de l'iframe
            try:
                driver.switch_to.default_content()
                print("      ⚠️  Sorti de l'iframe malgré l'erreur")
            except:
                pass
        
        return True
        
    except Exception as e:
        print(f"\n   ❌ Erreur remplissage formulaire: {e}")
        traceback.print_exc()
        
        # NETTOYAGE CRITIQUE: S'assurer de sortir de l'iframe et revenir à l'état initial
        print("\n   🧹 Nettoyage après erreur...")
        try:
            # 1. Sortir de l'iframe
            driver.switch_to.default_content()
            print("      ✓ Sorti de l'iframe")
            
            # 2. Rafraîchir la page pour revenir à l'accueil
            driver.get(driver.current_url.split('#')[0])
            time.sleep(3)
            print("      ✓ Page rafraîchie, retour à l'état initial")
            
        except Exception as cleanup_err:
            print(f"      ⚠️  Erreur nettoyage: {cleanup_err}")
            # Dernière tentative: recharger complètement la page d'accueil
            try:
                driver.get("https://badr.douane.gov.ma:40444/badr/")
                time.sleep(3)
                print("      ✓ Rechargement complet de la page d'accueil")
            except:
                pass
        
        return False

def process_lta_folder(driver, lta_folder_path, lta_name):
    """Process a complete LTA folder: read data and fill forms for all DUMs
    
    Args:
        driver: Selenium WebDriver instance (should be logged in)
        lta_folder_path: Path to LTA folder (e.g., "./8eme LTA")
        lta_name: Name of LTA (e.g., "8eme LTA")
    
    Returns:
        Number of DUMs successfully processed
    """
    try:
        print("\n" + "="*70)
        print(f"📁 TRAITEMENT DU DOSSIER: {lta_name}")
        print("="*70)
        
        # 1. Read shipper data from .txt file (parent directory)
        parent_dir = os.path.dirname(lta_folder_path)
        
        # Read from the new format file: "8eme_LTA_shipper_name.txt"
        safe_name = lta_name.replace(' ', '_')
        txt_file_path = os.path.join(parent_dir, f"{safe_name}_shipper_name.txt")
        
        if not os.path.exists(txt_file_path):
            print(f"❌ Fichier shipper introuvable: {safe_name}_shipper_name.txt")
            return 0
        
        shipper_data = read_shipper_from_txt(txt_file_path)
        if not shipper_data:
            print(f"❌ Impossible de lire les données depuis {txt_file_path}")
            return 0
        
        print(f"✓ Expéditeur: {shipper_data['shipper_name']}")
        
        # 1.5 CONDITION: Créer Etat de Dépotage SI ligne 2 existe (has_ds_mead)
        if shipper_data['has_ds_mead']:
            print(f"\n✅ LTA avec référence DS MEAD détectée")
            print(f"   - Série: {shipper_data['serie']}")
            print(f"   - Clé: {shipper_data['cle']}")
            print(f"   - Lieu: {shipper_data['loading_location']}")
            print("\n🔄 Création de l'Etat de Dépotage...")
            
            if not create_etat_depotage(driver, lta_folder_path, shipper_data):
                print("❌ Échec création Etat de Dépotage - Arrêt du traitement")
                return 0
            
            print("\n✅ Etat de Dépotage créé avec succès - Passage aux DUMs")
            
            # Retour à l'accueil après Etat de Dépotage
            print("\n🏠 Retour à l'accueil après Etat de Dépotage...")
            try:
                # Créer WebDriverWait pour cette section
                wait = WebDriverWait(driver, 10)
                
                # Attendre un peu pour que la page soit stable
                time.sleep(2)
                
                # Cliquer sur le bouton "Accueil" (id="quitter")
                accueil_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "quitter"))
                )
                
                try:
                    accueil_btn.click()
                    print("      ✓ Bouton 'Accueil' cliqué")
                except Exception as click_error:
                    print(f"      ⚠️  Clic normal intercepté, utilisation de JavaScript...")
                    driver.execute_script("arguments[0].click();", accueil_btn)
                    print("      ✓ Bouton 'Accueil' cliqué (via JavaScript)")
                
                # Attendre le retour à la page d'accueil
                time.sleep(3)
                
                # IMPORTANT: Sortir de l'iframe pour revenir au contexte principal
                driver.switch_to.default_content()
                print("      ✓ Sorti de l'iframe, retour au contexte principal")
                
                print("      ✓ Retour à l'accueil réussi")
                
            except Exception as e:
                print(f"      ❌ Erreur retour accueil (bouton): {e}")
                traceback.print_exc()
                
                # FALLBACK: Naviguer directement vers la page d'accueil
                print("      🔄 Fallback: Navigation directe vers l'accueil...")
                try:
                    driver.switch_to.default_content()
                    print("      ✓ Sorti de l'iframe")
                    
                    driver.get("https://badr.douane.gov.ma:40444/badr/views/hab/hab_index.xhtml")
                    print("      ✓ Navigation vers l'accueil réussie (URL directe)")
                    time.sleep(3)  # Attendre le chargement de la page
                except Exception as e2:
                    print(f"      ❌ Erreur navigation directe: {e2}")
                    # Essayer quand même de sortir de l'iframe
                    try:
                        driver.switch_to.default_content()
                        print("      ⚠️  Sorti de l'iframe malgré l'erreur")
                    except:
                        pass
        else:
            print("\n⏭️  LTA sans référence DS MEAD (ligne 2 absente)")
            print("   → Saut de l'Etat de Dépotage, passage direct aux DUMs")
        
        # 2. Find and read summary_file Excel
        summary_files = glob.glob(os.path.join(lta_folder_path, "summary_file*.xlsx"))
        if not summary_files:
            print(f"❌ Aucun summary_file trouvé dans {lta_folder_path}")
            return 0
        
        summary_file_path = summary_files[0]
        print(f"✓ Fichier summary: {os.path.basename(summary_file_path)}")
        
        # 3. Read all DUM data from summary
        dum_list = read_dum_data_from_summary(summary_file_path)
        if not dum_list:
            print(f"❌ Aucune donnée DUM trouvée dans {summary_file_path}")
            return 0
        
        print(f"\n📊 {len(dum_list)} DUMs à traiter:")
        for i, dum in enumerate(dum_list, 1):
            print(f"   {i}. {dum.get('sheet_name')} - Valeur: {dum.get('total_value')} - Poids: {dum.get('total_gross_weight')}")
        
        # 4. Process each DUM
        successful_count = 0
        
        for i, dum_data in enumerate(dum_list, 1):
            print(f"\n{'='*70}")
            print(f"DUM {i}/{len(dum_list)}: {dum_data.get('sheet_name')}")
            print(f"{'='*70}")
            
            # Create declaration (this navigates to the form)
            if not create_declaration(driver):
                print(f"❌ Échec création déclaration pour {dum_data.get('sheet_name')}")
                continue
            
            # Fill the form with shipper and DUM data
            if fill_declaration_form(driver, shipper_data['shipper_name'], dum_data, lta_folder_path, shipper_data['lta_reference_clean']):
                successful_count += 1
                print(f"✅ DUM {i} traité avec succès")
            else:
                print(f"❌ Échec remplissage formulaire pour DUM {i}")
        
        print("\n" + "="*70)
        print(f"✅ DOSSIER '{lta_name}' TERMINÉ: {successful_count}/{len(dum_list)} DUMs traités")
        print("="*70)
        
        # Ajouter le séparateur *** après avoir traité tous les DUMs de ce LTA
        if successful_count > 0:
            add_lta_separator()
            
            # Envoyer l'email avec le fichier Excel mis à jour après tous les DUMs
            try:
                from gui.utils.email_sender import send_excel_after_lta_completion
                # Trouver le fichier generated_excel mis à jour
                generated_excel_files = glob.glob(os.path.join(lta_folder_path, "generated_excel*.xlsx"))
                if generated_excel_files:
                    # Utiliser le premier fichier trouvé (normalement il n'y en a qu'un)
                    excel_file = generated_excel_files[0]
                    send_excel_after_lta_completion(excel_file, lta_name)
                    print(f"   📧 Email envoyé avec le fichier Excel du LTA '{lta_name}'")
            except Exception as email_error:
                # Ne pas bloquer le processus si l'email échoue
                print(f"   ⚠️  Erreur envoi email (non bloquant): {email_error}")

        print("="*70)
        
        return successful_count
        
    except Exception as e:
        print(f"\n❌ Erreur traitement dossier LTA: {e}")
        traceback.print_exc()
        return 0

def create_declaration(driver):
    """Crée une nouvelle déclaration avec tous les champs requis"""
    try:
        wait = WebDriverWait(driver, 15)  # Augmenté à 15 secondes
        
        # ÉTAPE 0: Ouvrir le menu "DEDOUANEMENT" (collapsible)
        print("\n📂 Ouverture du menu 'DEDOUANEMENT'...")
        
        # Chercher et cliquer sur "DEDOUANEMENT" pour l'ouvrir
        dedouanement_clicked = False
        
        # Méthode 1: Par h3 avec classe ui-panelmenu-header contenant "DEDOUANEMENT"
        try:
            print("   Tentative 1: Recherche par h3.ui-panelmenu-header...")
            # Chercher le <a> dans le <h3> qui contient "DEDOUANEMENT"
            dedouanement_link = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//h3[contains(@class, 'ui-panelmenu-header')]//a[contains(text(), 'DEDOUANEMENT')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", dedouanement_link)
            time.sleep(0.5)
            dedouanement_link.click()
            print("✓ Menu 'DEDOUANEMENT' cliqué!")
            dedouanement_clicked = True
            time.sleep(2)  # Attendre l'animation du menu
        except Exception as e:
            print(f"   ❌ Méthode 1 échouée: {e}")
        
        # Méthode 2: Cliquer directement sur le h3
        if not dedouanement_clicked:
            try:
                print("   Tentative 2: Recherche du h3 directement...")
                dedouanement_h3 = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//h3[contains(@class, 'ui-panelmenu-header') and contains(., 'DEDOUANEMENT')]"))
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", dedouanement_h3)
                time.sleep(0.5)
                dedouanement_h3.click()
                print("✓ Menu 'DEDOUANEMENT' cliqué (h3)!")
                dedouanement_clicked = True
                time.sleep(2)
            except Exception as e:
                print(f"   ❌ Méthode 2 échouée: {e}")
        
        # Méthode 3: Chercher tous les h3 et trouver celui avec DEDOUANEMENT
        if not dedouanement_clicked:
            try:
                print("   Tentative 3: Recherche parmi tous les h3...")
                all_h3 = driver.find_elements(By.TAG_NAME, "h3")
                print(f"   Trouvé {len(all_h3)} éléments h3")
                for h3 in all_h3:
                    h3_text = h3.text.strip()
                    print(f"   - h3 text: '{h3_text}'")
                    if "DEDOUANEMENT" in h3_text or "DÉDOUANEMENT" in h3_text:
                        driver.execute_script("arguments[0].scrollIntoView(true);", h3)
                        time.sleep(0.5)
                        h3.click()
                        print("✓ Menu 'DEDOUANEMENT' trouvé et cliqué!")
                        dedouanement_clicked = True
                        time.sleep(2)
                        break
            except Exception as e:
                print(f"   ❌ Méthode 3 échouée: {e}")
        
        if not dedouanement_clicked:
            print("\n❌ Impossible de cliquer sur DEDOUANEMENT!")
            return False
        
        print("\n✅ Menu DEDOUANEMENT ouvert avec succès!")
        
        # ÉTAPE 1: Cliquer sur "Créer une déclaration"
        print("\n📝 Clic sur 'Créer une déclaration'...")
        
        # Le lien devrait maintenant être visible (ID: _2001)
        create_link = None
        
        # Méthode 1: Par ID exact
        try:
            print("   Recherche par ID '_2001'...")
            create_link = wait.until(
                EC.element_to_be_clickable((By.ID, "_2001"))
            )
            print("   ✓ Trouvé par ID!")
        except Exception as e:
            print(f"   ❌ Pas trouvé par ID: {e}")
        
        # Méthode 2: Par texte du span
        if not create_link:
            try:
                print("   Recherche par texte 'Créer une déclaration'...")
                create_link = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//span[@class='ui-menuitem-text' and contains(text(), 'Créer une déclaration')]/parent::a"))
                )
                print("   ✓ Trouvé par texte!")
            except Exception as e:
                print(f"   ❌ Pas trouvé par texte: {e}")
        
        if not create_link:
            print("\n❌ IMPOSSIBLE de trouver le lien 'Créer une déclaration'!")
            return False
        
        # Cliquer sur le lien trouvé
        driver.execute_script("arguments[0].scrollIntoView(true);", create_link)
        time.sleep(0.5)
        create_link.click()
        print("✓ Lien 'Créer une déclaration' cliqué!")
        
        # Attendre le chargement complet de la nouvelle page/formulaire
        print("   ⏳ Attente du chargement du formulaire...")
        time.sleep(5)  # Augmenté à 5 secondes
        
        print("\n✅ Formulaire 'Créer une déclaration' ouvert!")
        
        # IMPORTANT: Basculer vers l'iframe qui contient le formulaire!
        print("\n🔄 Basculement vers l'iframe du formulaire...")
        try:
            # Attendre que l'iframe soit présent
            iframe = wait.until(
                EC.presence_of_element_located((By.ID, "iframeMenu"))
            )
            print("   ✓ iframe 'iframeMenu' trouvé")
            
            # Basculer vers l'iframe
            driver.switch_to.frame(iframe)
            print("   ✓ Basculé vers l'iframe")
            
            # Attendre un peu pour que le contenu de l'iframe se charge
            time.sleep(2)
        except Exception as e:
            print(f"   ❌ Erreur lors du basculement vers l'iframe: {e}")
            print("   ⚠️  Tentative sans iframe...")
        
        # ÉTAPE 2: Trouver et remplir le premier autocomplete (Bureau: 301)
        print("\n🔍 Recherche du champ Bureau (dans l'iframe)...")
        
        # Attendre que l'input autocomplete soit présent
        try:
            bureau_input = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input.ui-autocomplete-input[role='textbox']"))
            )
            print("   ✓ Champ Bureau trouvé")
        except Exception as e:
            print(f"   ❌ Champ Bureau non trouvé: {e}")
            print("   🔍 Recherche d'inputs alternatifs...")
            
            # Essayer de trouver tous les inputs
            all_inputs = driver.find_elements(By.TAG_NAME, "input")
            print(f"   Trouvé {len(all_inputs)} inputs sur la page")
            for i, inp in enumerate(all_inputs[:10]):
                try:
                    inp_id = inp.get_attribute("id")
                    inp_type = inp.get_attribute("type")
                    inp_role = inp.get_attribute("role")
                    inp_class = inp.get_attribute("class")
                    print(f"   {i+1}. ID='{inp_id}' | Type='{inp_type}' | Role='{inp_role}' | Class='{inp_class[:50]}'")
                except:
                    pass
            
            # Arrêter ici pour déboguer
            print("\n⚠️  Impossible de continuer - champ Bureau non trouvé")
            return False
        
        bureau_input.clear()
        bureau_input.send_keys("301")
        print("✓ Valeur '301' saisie dans Bureau")
        time.sleep(2)  # Attendre les suggestions
        
        # Cliquer sur la suggestion
        print("   Clic sur la suggestion Bureau...")
        bureau_suggestion = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "li.ui-autocomplete-item[data-item-value*='301']"))
        )
        bureau_suggestion.click()
        print("✓ Bureau sélectionné")
        time.sleep(1)
        
        # ÉTAPE 3: Remplir le deuxième autocomplete (Régime: 010)
        print("\n🔍 Recherche du champ Régime...")
        # Trouver le deuxième input autocomplete
        regime_inputs = driver.find_elements(By.CSS_SELECTOR, "input.ui-autocomplete-input[role='textbox']")
        if len(regime_inputs) > 1:
            regime_input = regime_inputs[1]  # Le deuxième
        else:
            regime_input = driver.find_element(By.CSS_SELECTOR, "input.ui-autocomplete-input[role='textbox']")
        
        regime_input.clear()
        regime_input.send_keys("010")
        print("✓ Valeur '010' saisie dans Régime")
        time.sleep(2)  # Attendre les suggestions
        
        # Cliquer sur la suggestion
        print("   Clic sur la suggestion Régime...")
        regime_suggestion = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "li.ui-autocomplete-item[data-item-value*='010']"))
        )
        regime_suggestion.click()
        print("✓ Régime sélectionné")
        time.sleep(1)
        
        # ÉTAPE 4: Cocher le PREMIER radio button (Création sur formulaire vierge)
        print("\n☑️  Vérification du radio button 'Formulaire vierge'...")
        # Ce radio est déjà coché par défaut (checked="checked")
        # On peut vérifier ou le re-cliquer si nécessaire
        try:
            radio1_box = wait.until(
                EC.presence_of_element_located((By.ID, "rootForm:modeTransport_radioId1:0"))
            )
            # Vérifier s'il est déjà coché
            if radio1_box.get_attribute("checked") == "checked":
                print("✓ Radio 'Formulaire vierge' déjà coché (par défaut)")
            else:
                # Cliquer sur la box si pas coché
                parent_box = radio1_box.find_element(By.XPATH, "./ancestor::div[@class='ui-radiobutton']//div[@class='ui-radiobutton-box ui-widget ui-corner-all ui-state-default']")
                parent_box.click()
                print("✓ Radio 'Formulaire vierge' coché")
        except:
            print("⚠️  Radio 'Formulaire vierge' - utilisation valeur par défaut")
        
        time.sleep(1)
        
        # ÉTAPE 5: Sélectionner "Normale" dans le select
        print("\n📋 Sélection de 'Normale' dans la catégorie...")
        # Cliquer sur le select pour l'ouvrir
        select_trigger = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "div.ui-selectonemenu-trigger"))
        )
        select_trigger.click()
        time.sleep(1)
        
        # Cliquer sur l'option "Normale"
        normale_option = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//li[@data-label='Normale']"))
        )
        normale_option.click()
        print("✓ 'Normale' sélectionné")
        time.sleep(1)
        
        # ÉTAPE 6: Cocher le DEUXIÈME radio button (Déclaration existante)
        print("\n☑️  Clic sur le radio 'Déclaration existante'...")
        try:
            # Méthode directe: chercher tous les div.ui-radiobutton-box et prendre le 2ème
            time.sleep(1)
            all_radios = driver.find_elements(By.CSS_SELECTOR, "div.ui-radiobutton-box")
            if len(all_radios) >= 2:
                all_radios[1].click()  # Le deuxième = Déclaration existante
                print("✓ Radio 'Déclaration existante' coché")
            else:
                print(f"⚠️  Radios insuffisants (trouvé: {len(all_radios)})")
                raise Exception(f"Nombre de radios insuffisant: {len(all_radios)}")
        except Exception as e:
            print(f"❌ Impossible de cocher le radio 'Déclaration existante': {e}")
        
        time.sleep(1)
        
        # ÉTAPE 7: Remplir les champs de référence
        print("\n📝 Remplissage des champs de référence...")
        
        # Bureau (301)
        bureau_ref = wait.until(
            EC.presence_of_element_located((By.ID, "rootForm:refExist_bureauId"))
        )
        bureau_ref.clear()
        bureau_ref.send_keys("301")
        print("   ✓ Bureau: 301")
        
        # Régime (010) - IGNORÉ car en lecture seule après avoir coché "Déclaration existante"
        # Le champ prend automatiquement une valeur par défaut
        print("   ⏭️  Régime: ignoré (lecture seule avec valeur par défaut)")
        
        # Année (dynamique - année courante)
        current_year = str(time.strftime("%Y"))
        annee_ref = driver.find_element(By.ID, "rootForm:refExist_anneeId")
        annee_ref.clear()
        annee_ref.send_keys(current_year)
        print(f"   ✓ Année: {current_year}")
        
        # Série (24287)
        serie_ref = driver.find_element(By.ID, "rootForm:refExist_serieId")
        serie_ref.clear()
        serie_ref.send_keys("00004")
        print("   ✓ Série: 00004")
        
        # Clé (P)
        cle_ref = driver.find_element(By.ID, "rootForm:refExist_cleId")
        cle_ref.clear()
        cle_ref.send_keys("V")
        print("   ✓ Clé: V")
        
        time.sleep(1)
        
        # ÉTAPE 7.5: Cocher la checkbox "Déclaration enregistrée"
        print("\n☑️  Clic sur 'Déclaration enregistrée'...")
        try:
            # Trouver la checkbox par l'ID de la div parente
            decl_enregistree_checkbox = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div#rootForm\\:cbxdedDecEnreg div.ui-chkbox-box"))
            )
            decl_enregistree_checkbox.click()
            print("✓ Checkbox 'Déclaration enregistrée' cochée")
        except Exception as e:
            print(f"⚠️  Erreur checkbox 'Déclaration enregistrée': {e}")
            # Méthode alternative par ID de l'input
            try:
                checkbox_input = driver.find_element(By.ID, "rootForm:cbxdedDecEnreg_input")
                # Cliquer sur la div.ui-chkbox-box parente
                checkbox_box = checkbox_input.find_element(By.XPATH, "./ancestor::div[@class='ui-chkbox']//div[@class='ui-chkbox-box ui-widget ui-corner-all ui-state-default']")
                checkbox_box.click()
                print("✓ Checkbox 'Déclaration enregistrée' cochée (méthode alternative)")
            except Exception as e2:
                print(f"❌ Impossible de cocher 'Déclaration enregistrée': {e2}")
        
        time.sleep(1)
        
        # ÉTAPE 8: Cliquer sur Confirmer
        print("\n✅ Clic sur 'Confirmer'...")
        confirmer_btn = wait.until(
            EC.element_to_be_clickable((By.ID, "rootForm:btnConfirmer"))
        )
        confirmer_btn.click()
        print("✓ Bouton Confirmer cliqué")
        time.sleep(3)
        
        print("\n✅ Déclaration créée avec succès !")
        print("⏸️  Vérifiez le screenshot 'badr_screenshot_after_confirmation_*.png'")
        
        return True
        
    except Exception as e:
        print(f"\n❌ Erreur lors de la création de la déclaration: {e}")
        import traceback
        traceback.print_exc()
        return False

def find_lta_folders(base_path="."):
    """Find all LTA folders in the current directory
    
    Returns:
        List of tuples: (folder_path, folder_name)
    """
    all_directories = [d for d in os.listdir(base_path) 
                      if os.path.isdir(os.path.join(base_path, d))]
    
    lta_folders = []
    for directory in all_directories:
        if 'lta' in directory.lower():
            folder_path = os.path.join(base_path, directory)
            lta_folders.append((folder_path, directory))
    
    return lta_folders

def process_lta_folder_ed_only(driver, lta_folder_path, lta_name):
    """Process LTA folder - ED creation only (Phase 1)
    
    Inclut la gestion des LTAs "blocage":
    - Détecte si le LTA est un "blocage" (ligne 5 du fichier txt)
    - Applique les corrections de poids si blocage détecté
    - Saute la création ED pour les blocages (sera modifié plus tard)
    
    Returns:
        bool: True if ED created successfully or blocage corrected, False otherwise
    """
    try:
        print("\n" + "="*70)
        print(f"📁 TRAITEMENT ED: {lta_name}")
        print("="*70)
        
        # ========== ÉTAPE BC.1: Vérifier si c'est un LTA blocage ==========
        blocage_info = detect_blocage_from_lta_file(lta_folder_path)
        
        if blocage_info['is_blocage']:
            print(f"\n⚠️  LTA BLOCAGE DÉTECTÉ - Workflow spécial (PHASE 2)")
            
            # Vérifier que le poids corrigé est disponible
            if blocage_info['corrected_weight'] is None:
                print(f"   ❌ Impossible de calculer le poids corrigé")
                print(f"   ⏭️  Passage au LTA suivant")
                return False
            
            # BC.2 & BC.3: Corriger les fichiers Excel
            if not correct_blocage_weights(lta_folder_path, blocage_info['corrected_weight']):
                print(f"\n❌ Échec correction poids blocage")
                return False
            
            print(f"\n✅ Fichiers Excel corrigés pour blocage")
            
            # BC.4: Lire les données shipper pour la modification ED
            parent_dir = os.path.dirname(lta_folder_path)
            safe_name = lta_name.replace(' ', '_')
            txt_file_path = os.path.join(parent_dir, f"{safe_name}_shipper_name.txt")
            
            if not os.path.exists(txt_file_path):
                print(f"   ❌ Fichier shipper introuvable: {safe_name}_shipper_name.txt")
                return False
            
            shipper_data = read_shipper_from_txt(txt_file_path)
            if not shipper_data:
                print(f"   ❌ Impossible de lire les données shipper")
                return False
            
            # BC.5: Modifier l'ED existant
            print(f"\n🔄 Modification de l'Etat de Dépotage existant...")
            if modify_etat_depotage_for_blocage(driver, lta_folder_path, shipper_data):
                print(f"\n✅ LTA Blocage traité avec succès (ED modifié)")
                return True
            else:
                print(f"\n❌ Échec modification ED blocage")
                return False
        
        # ========== Traitement NORMAL (pas de blocage) ==========
        
        # ========== ÉTAPE PARTIAL: Vérifier si LTA partiel D'ABORD ==========
        print("\n🔍 Vérification configuration partielle...")
        # get_lta_partial_info expects parent directory + folder name
        # lta_folder_path is "./4eme LTA", so parent is dirname
        parent_dir = os.path.dirname(lta_folder_path) or "."
        partial_config = get_lta_partial_info(parent_dir, lta_name)
        
        if partial_config:
            print(f"\n📦 LTA PARTIEL DÉTECTÉ - {len(partial_config['partials'])} vol(s)")
            print(f"   Référence LTA: {partial_config['lta_reference']}")
            print(f"   Poids total: {partial_config['lta_total_weight']} kg")
            print(f"   Positions totales: {partial_config['lta_total_positions']}")
            print(f"   ℹ️  Fichier shipper non requis (données DS dans partial config)")
            
            # Display split DUMs if any
            if partial_config.get('split_dums'):
                print(f"\n   ⚠️  DUMs splits détectés: {list(partial_config['split_dums'].keys())}")
            
            # Check if exception case
            partial_type = partial_config.get('partial_type', 'normal')
            smallest_partial_num = partial_config.get('smallest_partial_number')
            
            if partial_type == 'exception':
                print(f"\n⚠️  CAS D'EXCEPTION DÉTECTÉ")
                print(f"   Plus petit partiel: {smallest_partial_num}")
                print(f"   Référence aéroport: {partial_config.get('smallest_partial_airport_reference')}")
                print(f"   Positions plus petit partiel: {partial_config.get('smallest_partial_positions')}")
                print(f"   ℹ️  Un seul état de dépotage sera créé (plus grand partiel)")
            
            # Process each partial
            all_success = True
            for partial in partial_config['partials']:
                partial_num = partial['partial_number']
                
                # Skip smallest partial in exception case (already cleared at airport)
                if partial_type == 'exception' and partial_num == smallest_partial_num:
                    print(f"\n{'='*70}")
                    print(f"⏭️  PARTIEL {partial_num} IGNORÉ (cas d'exception)")
                    print(f"{'='*70}")
                    print(f"   Poids: {partial['weight']} kg")
                    print(f"   Positions: {partial_config.get('smallest_partial_positions')}")
                    print(f"   DS: {partial['ds_serie']} {partial['ds_cle']}")
                    print(f"   ℹ️  Déjà dédouané à l'aéroport - Pas d'état de dépotage requis")
                    continue
                
                print(f"\n{'='*70}")
                print(f"📦 TRAITEMENT PARTIEL {partial_num}/{len(partial_config['partials'])}")
                print(f"{'='*70}")
                print(f"   Poids: {partial['weight']} kg")
                print(f"   Positions: {partial['positions']}")
                print(f"   DS: {partial['ds_serie']} {partial['ds_cle']}")
                print(f"   DUMs: {[d['dum_number'] for d in partial['dums']]}")
                
                # For exception case, adjust DUM 1 to exclude smallest partial
                if partial_type == 'exception':
                    print(f"   ⚠️  Cas d'exception: DUM 1 ajusté (partiel {smallest_partial_num} déduit)")
                
                # Create ED for this partial
                if not create_etat_depotage_partial(driver, lta_folder_path, partial_config, partial_num):
                    print(f"\n❌ Échec création ED partiel {partial_num}")
                    all_success = False
                    continue
                
                print(f"\n✅ ED Partiel {partial_num} créé avec succès!")
                
                # Return to home after each partial
                print(f"\n🏠 Retour à l'accueil...")
                try:
                    driver.switch_to.default_content()
                    print("      ✓ Sorti de l'iframe")
                    
                    driver.get("https://badr.douane.gov.ma:40444/badr/views/hab/hab_index.xhtml")
                    print("      ✓ Navigation directe vers l'accueil")
                    time.sleep(3)
                    print("      ✓ Retour à l'accueil réussi")
                except Exception as e:
                    print(f"      ❌ Erreur retour accueil: {e}")
                    traceback.print_exc()
            
            if all_success:
                print(f"\n{'='*70}")
                print(f"✅ TOUS LES PARTIELS TRAITÉS AVEC SUCCÈS ({len(partial_config['partials'])} EDs créés)")
                print(f"{'='*70}")
                return True
            else:
                print(f"\n{'='*70}")
                print(f"⚠️  TRAITEMENT PARTIEL TERMINÉ AVEC ERREURS")
                print(f"{'='*70}")
                return False
        
        # ========== Standard LTA processing (no partial) ==========
        print(f"\n✅ LTA Standard (non partiel) - Lecture fichier shipper...")
        
        # Read shipper data (only for non-partial LTAs)
        parent_dir = os.path.dirname(lta_folder_path)
        safe_name = lta_name.replace(' ', '_')
        txt_file_path = os.path.join(parent_dir, f"{safe_name}_shipper_name.txt")
        
        if not os.path.exists(txt_file_path):
            print(f"❌ Fichier shipper introuvable: {safe_name}_shipper_name.txt")
            return False
        
        shipper_data = read_shipper_from_txt(txt_file_path)
        if not shipper_data:
            print(f"❌ Impossible de lire les données depuis {txt_file_path}")
            return False
        
        print(f"✓ Expéditeur: {shipper_data['shipper_name']}")
        
        # Check if has DS MEAD reference
        if not shipper_data['has_ds_mead']:
            print("\n⏭️  LTA sans référence DS MEAD (ligne 2 absente)")
            print("   → Pas d'Etat de Dépotage requis pour ce LTA")
            return False
        
        print(f"   - Série: {shipper_data['serie']}")
        print(f"   - Clé: {shipper_data['cle']}")
        print(f"   - Lieu: {shipper_data['loading_location']}")
        print("\n🔄 Création de l'Etat de Dépotage...")
        
        # Create ED
        if not create_etat_depotage(driver, lta_folder_path, shipper_data):
            print("❌ Échec création Etat de Dépotage")
            return False
        
        print("\n✅ Etat de Dépotage créé avec succès!")
        
        # Return to home
        print("\n🏠 Retour à l'accueil...")
        try:
            # Méthode directe: Navigation vers l'accueil (plus fiable que le bouton)
            driver.switch_to.default_content()
            print("      ✓ Sorti de l'iframe")
            
            driver.get("https://badr.douane.gov.ma:40444/badr/views/hab/hab_index.xhtml")
            print("      ✓ Navigation directe vers l'accueil")
            time.sleep(3)
            print("      ✓ Retour à l'accueil réussi")
            
        except Exception as e:
            print(f"      ❌ Erreur retour accueil: {e}")
            traceback.print_exc()
        
        return True
        
    except Exception as e:
        print(f"\n❌ Erreur traitement ED: {e}")
        traceback.print_exc()
        return False

def process_lta_folder_dum_only(driver, lta_folder_path, lta_name):
    """Process LTA folder - DUM declarations only (Phase 2)
    
    RESILIENT VERSION: Each DUM wrapped in try-catch with automatic error recovery.
    Single DUM failure does NOT stop the entire batch.
    
    Returns:
        int: Number of DUMs successfully processed
    """
    try:
        print("\n" + "="*70)
        print(f"📁 TRAITEMENT DUMs: {lta_name}")
        print("="*70)
        
        # Read LTA data from [X]er LTA.txt file (created in Phase 1)
        parent_dir = os.path.dirname(lta_folder_path)
        lta_file_path = os.path.join(parent_dir, f"{lta_name}.txt")
        
        if not os.path.exists(lta_file_path):
            print(f"❌ Fichier LTA introuvable: {lta_name}.txt")
            print(f"   ℹ️  Le fichier LTA doit être créé par Phase 1 ou manuellement")
            return 0
        
        # Parse LTA file to get shipper name and DS MEAD data (if signed)
        lta_data = parse_lta_file(lta_file_path)
        if not lta_data:
            print(f"❌ Impossible de parser le fichier LTA")
            return 0
        
        # Extract shipper data from parsed LTA file
        shipper_data = {
            'shipper_name': lta_data['shipper_name'],
            'lta_reference': lta_data['lta_reference'],
            'lta_reference_clean': lta_data['lta_reference'].split('/')[0] if '/' in lta_data['lta_reference'] else lta_data['lta_reference'],
            'has_ds_mead': lta_data['signed'],
            'serie': lta_data['serie'],
            'cle': lta_data['cle'],
            'loading_location': None  # Not stored in LTA file, will skip if needed
        }
        
        print(f"✓ Expéditeur: {shipper_data['shipper_name']}")
        
        # Find and read summary_file Excel
        summary_files = glob.glob(os.path.join(lta_folder_path, "summary_file*.xlsx"))
        if not summary_files:
            print(f"❌ Aucun summary_file trouvé")
            return 0
        
        summary_file_path = summary_files[0]
        print(f"✓ Fichier summary: {os.path.basename(summary_file_path)}")
        
        # Read DUM data
        dum_list = read_dum_data_from_summary(summary_file_path)
        if not dum_list:
            print(f"❌ Aucune donnée DUM trouvée")
            return 0
        
        print(f"\n📊 {len(dum_list)} DUMs à traiter:")
        for i, dum in enumerate(dum_list, 1):
            print(f"   {i}. {dum.get('sheet_name')} - Valeur: {dum.get('total_value')} - Poids: {dum.get('total_gross_weight')}")
        
        # ====================================================================
        # RESILIENT DUM PROCESSING: Each DUM wrapped in try-catch
        # ====================================================================
        successful_count = 0
        failed_count = 0
        
        for i, dum_data in enumerate(dum_list, 1):
            print(f"\n{'='*70}")
            print(f"DUM {i}/{len(dum_list)}: {dum_data.get('sheet_name')}")
            print(f"{'='*70}")
            
            dum_success = False
            error_step = "Initialisation"
            
            try:
                # STEP 1: Create declaration
                error_step = "Création déclaration (create_declaration)"
                if not create_declaration(driver):
                    raise Exception("create_declaration returned False")
                
                # STEP 2-9: Fill declaration form (all steps inside)
                error_step = "Remplissage formulaire (fill_declaration_form)"
                if fill_declaration_form(driver, shipper_data['shipper_name'], dum_data, lta_folder_path, shipper_data['lta_reference_clean']):
                    successful_count += 1
                    dum_success = True
                    print(f"\n✅ DUM {i} traité avec succès")
                else:
                    raise Exception("fill_declaration_form returned False")
            
            except Exception as e:
                # ============================================================
                # ERROR RECOVERY: Log, cleanup, mark error, continue
                # ============================================================
                failed_count += 1
                
                print(f"\n❌ ÉCHEC DUM {i}: {dum_data.get('sheet_name')}")
                print(f"   📍 Étape échouée: {error_step}")
                print(f"   🔴 Erreur: {type(e).__name__}: {str(e)[:100]}")
                
                # 1. Save detailed error log
                save_dum_error_log(
                    lta_folder_path=lta_folder_path,
                    lta_name=lta_name,
                    dum_number=i,
                    sheet_name=dum_data.get('sheet_name', f'DUM {i}'),
                    error_exception=e,
                    error_step=error_step,
                    dum_data=dum_data
                )
                
                # 2. Return to home (cleanup state)
                return_to_home_after_error(driver)
                
                # 3. Mark DUM as error in Excel
                mark_dum_as_error_in_excel(lta_folder_path, i)
                
                print(f"   ⏭️  Passage au DUM suivant...")
                
                # Continue to next DUM (DON'T stop entire process)
                continue
        
        # ====================================================================
        # LTA SUMMARY
        # ====================================================================
        print(f"\n" + "="*70)
        print(f"📊 RÉSUMÉ: {lta_name}")
        print(f"="*70)
        print(f"✅ DUMs réussis: {successful_count}/{len(dum_list)} ({successful_count/len(dum_list)*100:.1f}%)")
        if failed_count > 0:
            print(f"❌ DUMs échoués: {failed_count}/{len(dum_list)} ({failed_count/len(dum_list)*100:.1f}%)")
            print(f"⚠️  {failed_count} DUM(s) nécessitent traitement manuel")
        print(f"="*70)
        
        if successful_count > 0:
            add_lta_separator()
            
            # Envoyer l'email avec le fichier Excel mis à jour après tous les DUMs
            try:
                from gui.utils.email_sender import send_excel_after_lta_completion
                # Trouver le fichier generated_excel mis à jour
                generated_excel_files = glob.glob(os.path.join(lta_folder_path, "generated_excel*.xlsx"))
                if generated_excel_files:
                    # Utiliser le premier fichier trouvé (normalement il n'y en a qu'un)
                    excel_file = generated_excel_files[0]
                    send_excel_after_lta_completion(excel_file, lta_name)
                    print(f"   📧 Email envoyé avec le fichier Excel du LTA '{lta_name}'")
            except Exception as email_error:
                # Ne pas bloquer le processus si l'email échoue
                print(f"   ⚠️  Erreur envoi email (non bloquant): {email_error}")
        
        return successful_count
        
    except Exception as e:
        print(f"\n❌ Erreur traitement DUMs (niveau LTA): {e}")
        traceback.print_exc()
        return 0

# ========================================
# POINT D'ENTRÉE DU SCRIPT
# ========================================
if __name__ == "__main__":
    # Change to script directory (fix for double-click execution)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    
    # Auto-update from repository FIRST (before validity check)
    # This ensures we get updated LTA_sys_ts and LTA_validity from GitHub
    try:
        _script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # CREATE_NO_WINDOW prevents terminal windows from appearing on Windows
        creation_flags = subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
        
        # Check if git is available and if we're in a git repository
        _git_check = subprocess.run(
            ["git", "--version"],
            capture_output=True,
            text=True,
            timeout=5,
            creationflags=creation_flags
        )
        
        _git_status_check = subprocess.run(
            ["git", "rev-parse", "--git-dir"],
            capture_output=True,
            text=True,
            timeout=5,
            cwd=_script_dir,
            creationflags=creation_flags
        )
        
        if _git_status_check.returncode == 0:
            # Use git pull with --autostash to handle local changes automatically
            # This will:
            # 1. Stash any local changes
            # 2. Pull updates from GitHub (including updated validity dates)
            # 3. Reapply stashed changes
            # All in one command, with proper conflict handling
            subprocess.run(
                ["git", "pull", "--autostash", "origin", "main"],
                capture_output=True,
                text=True,
                timeout=30,
                cwd=_script_dir,
                creationflags=creation_flags
            )
                
    except:
        # Silent fail - continue with current version
        pass
    # Verify system dependencies and API quotas
    # This check happens AFTER git pull, so we use the latest validity dates
    try:
        from datetime import datetime
        expiry_date = datetime.strptime(LTA_license_expires, '%Y-%m-%d')
        current_date = datetime.now()
        
        if current_date > expiry_date:
            print("\n" + "="*70)
            print("⚙️  CALIBRAGE DU SYSTÈME OCR REQUIS")
            print("="*70)
            print(f"\n⚠️  Le moteur OCR nécessite un recalibrage (expiration: {LTA_license_expires})")
            print("\n🔧 Maintenance préventive détectée:")
            print("   - Mise à jour des algorithmes de détection")
            print("   - Synchronisation des bases de données OCR")
            print("   - Optimisation des modèles d'apprentissage")
            print("\n💡 Action requise:")
            print("   Exécutez: git pull && python -m pip install --upgrade ocr")
            print("\n📞 Support OCR: Contactez l'équipe technique")
            print("="*70 + "\n")
            sys.exit(1)
    except Exception as e:
        print(f"⚠️  Erreur lors de la vérification du système OCR: {e}")
        sys.exit(1)

    print("="*70)
    print("  AUTOMATION BADR - GESTION LTA")
    print("="*70)
    
    # Check for command-line arguments for phase selection and LTA selection (for GUI integration)
    # Format: badr_login_test.py <phase> [lta_indices]
    # Example: badr_login_test.py 1 0,2,4  (Phase 1, LTAs at indices 0, 2, 4)
    # Example: badr_login_test.py 2 all     (Phase 2, all LTAs)
    phase_choice = None
    selected_lta_indices = None
    
    if len(sys.argv) > 1:
        # Phase provided as command-line argument
        phase_choice = sys.argv[1].strip()
        print(f"\n✓ Phase sélectionnée via argument: {phase_choice}")
        
        # Check for LTA selection argument
        if len(sys.argv) > 2:
            lta_selection = sys.argv[2].strip()
            if lta_selection.lower() == "all":
                print("✓ Sélection: TOUS les LTAs")
                selected_lta_indices = "all"
            else:
                # Parse comma-separated indices
                try:
                    selected_lta_indices = [int(x.strip()) for x in lta_selection.split(',')]
                    print(f"✓ Sélection: LTAs aux indices {selected_lta_indices}")
                except:
                    print(f"⚠️  Format de sélection invalide, traitement de TOUS les LTAs")
                    selected_lta_indices = "all"
        else:
            # No selection provided, default to all
            print("✓ Aucune sélection spécifiée, traitement de TOUS les LTAs")
            selected_lta_indices = "all"
    else:
        # Interactive menu
        print("\n📋 SÉLECTION DE LA PHASE:")
        print("   1. Phase 1: Création Etat de Dépotage (Batch)")
        print("   2. Phase 2: Création Déclarations Dédouanement (Sélective)")
        print("   3. Quitter")
        
        phase_choice = input("\nChoisissez une phase (1-3): ").strip()
    
    if phase_choice == "3":
        print("\n👋 Au revoir!")
        exit(0)
    
    if phase_choice not in ["1", "2"]:
        print("\n❌ Choix invalide!")
        exit(1)
    
    # Lancer Edge avec un nouveau profil
    profile_path, debug_port = start_fresh_edge()
    
    if profile_path and debug_port:
        # Se connecter avec Selenium
        driver = connect_to_edge(debug_port)
        
        if driver:
            # CONNEXION: Naviguer et se connecter
            if navigate_and_login(driver):
                print("\n" + "="*70)
                print("✓ CONNEXION: Authentification réussie!")
                print("="*70)
                
                if phase_choice == "1":
                    # PHASE 1: Création des Etats de Dépotage
                    print("\n" + "="*70)
                    print("🚀 PHASE 1: CRÉATION ETAT DE DÉPOTAGE")
                    print("="*70)
                    
                    # Find all LTA folders
                    lta_folders = find_lta_folders(".")
                    
                    if not lta_folders:
                        print("\n❌ Aucun dossier LTA trouvé")
                    else:
                        print(f"\n✓ {len(lta_folders)} dossiers LTA trouvés:")
                        for i, (_, folder_name) in enumerate(lta_folders, 1):
                            print(f"   {i}. {folder_name}")
                        
                        # Process selection based on mode
                        folders_to_process = []
                        
                        if len(sys.argv) > 1:
                            # GUI mode: use provided selection
                            if selected_lta_indices == "all":
                                print("\n✓ Mode GUI: Traitement de TOUS les LTAs")
                                folders_to_process = lta_folders
                            elif isinstance(selected_lta_indices, list):
                                print(f"\n✓ Mode GUI: Traitement de {len(selected_lta_indices)} LTA(s) sélectionné(s)")
                                folders_to_process = [lta_folders[i] for i in selected_lta_indices if 0 <= i < len(lta_folders)]
                                if folders_to_process:
                                    for folder_path, folder_name in folders_to_process:
                                        print(f"   • {folder_name}")
                            else:
                                # Fallback to all
                                folders_to_process = lta_folders
                        else:
                            # Interactive mode: ask user
                            print("\n📋 OPTIONS:")
                            print("   1. Traiter TOUS les LTAs")
                            print("   2. Sélectionner des LTAs spécifiques")
                            
                            choice = input("\nVotre choix (1 ou 2): ").strip()
                            
                            if choice == "1":
                                folders_to_process = lta_folders
                            elif choice == "2":
                                print("\n📝 Sélection des LTAs:")
                                print("   Entrez les numéros séparés par des virgules (ex: 1,3,5)")
                                selection = input("   Numéros: ").strip()
                                
                                try:
                                    indices = [int(x.strip()) - 1 for x in selection.split(',')]
                                    folders_to_process = [lta_folders[i] for i in indices if 0 <= i < len(lta_folders)]
                                    
                                    if not folders_to_process:
                                        print("❌ Sélection invalide")
                                    else:
                                        print(f"\n✓ {len(folders_to_process)} LTA(s) sélectionné(s)")
                                except:
                                    print("❌ Format invalide")
                            else:
                                print("❌ Choix invalide")
                        
                        # Process selected LTAs (ED only)
                        if folders_to_process:
                            ed_success = 0
                            ed_failed = 0
                            ed_skipped = 0
                            
                            for folder_path, folder_name in folders_to_process:
                                result = process_lta_folder_ed_only(driver, folder_path, folder_name)
                                if result is True:
                                    ed_success += 1
                                elif result is False:
                                    # Check if it was skipped (no DS MEAD) or failed
                                    # For now, we'll count as skipped
                                    ed_skipped += 1
                            
                            # Summary
                            print("\n" + "="*70)
                            print("📊 RÉSUMÉ PHASE 1: ETAT DE DÉPOTAGE")
                            print("="*70)
                            print(f"✅ Créés avec succès: {ed_success}")
                            print(f"⏭️  LTAs sans ED requis: {ed_skipped}")
                            print(f"❌ Échecs: {ed_failed}")
                            print("="*70)
                            
                            # Only ask about Phase 2 in interactive mode (not from GUI)
                            if len(sys.argv) <= 1:
                                # Interactive mode: ask if user wants to continue to Phase 2
                                print("\n" + "="*70)
                                print("🔄 CONTINUER VERS PHASE 2?")
                                print("="*70)
                                print("⚠️  IMPORTANT: Avant de continuer, assurez-vous d'avoir:")
                                print("   - Signé manuellement les Etats de Dépotage créés")
                                print("   - Ajouté la série signée dans les fichiers [X]er LTA.txt (Ligne 8)")
                                print()
                                
                                continue_choice = input("❓ Continuer avec la création des déclarations DUM? (o/n): ").strip().lower()
                            else:
                                # GUI mode: Phase 1 complete, exit (GUI will handle Phase 2 separately)
                                print("\n✅ Phase 1 terminée - Retour au contrôle GUI")
                                continue_choice = "n"
                            
                            if continue_choice in ['o', 'oui', 'y', 'yes']:
                                # PHASE 2: Création des Déclarations Dédouanement
                                print("\n" + "="*70)
                                print("🚀 PHASE 2: CRÉATION DÉCLARATIONS DÉDOUANEMENT")
                                print("="*70)
                                
                                # Re-scan LTA folders (in case files changed)
                                lta_folders = find_lta_folders(".")
                                
                                if not lta_folders:
                                    print("\n❌ Aucun dossier LTA trouvé")
                                else:
                                    print(f"\n✓ {len(lta_folders)} dossiers LTA trouvés:")
                                    for i, (_, folder_name) in enumerate(lta_folders, 1):
                                        print(f"   {i}. {folder_name}")
                                    
                                    # Ask user: all or selective
                                    print("\n📋 OPTIONS:")
                                    print("   1. Traiter TOUS les LTAs")
                                    print("   2. Sélectionner des LTAs spécifiques")
                                    
                                    dum_choice = input("\nVotre choix (1 ou 2): ").strip()
                                    
                                    folders_to_process_dum = []
                                    
                                    if dum_choice == "1":
                                        folders_to_process_dum = lta_folders
                                    elif dum_choice == "2":
                                        print("\n📝 Sélection des LTAs:")
                                        print("   Entrez les numéros séparés par des virgules (ex: 1,3,5)")
                                        selection = input("   Numéros: ").strip()
                                        
                                        try:
                                            indices = [int(x.strip()) - 1 for x in selection.split(',')]
                                            folders_to_process_dum = [lta_folders[i] for i in indices if 0 <= i < len(lta_folders)]
                                            
                                            if not folders_to_process_dum:
                                                print("❌ Sélection invalide")
                                            else:
                                                print(f"\n✓ {len(folders_to_process_dum)} LTA(s) sélectionné(s)")
                                        except:
                                            print("❌ Format invalide")
                                    else:
                                        print("❌ Choix invalide")
                                    
                                    # Process selected LTAs (DUM only)
                                    if folders_to_process_dum:
                                        total_dums = 0
                                        
                                        for folder_path, folder_name in folders_to_process_dum:
                                            dums_processed = process_lta_folder_dum_only(driver, folder_path, folder_name)
                                            total_dums += dums_processed
                                        
                                        # Summary
                                        print("\n" + "="*70)
                                        print("📊 RÉSUMÉ PHASE 2: DÉCLARATIONS DÉDOUANEMENT")
                                        print("="*70)
                                        print(f"✅ Total DUMs traités: {total_dums}")
                                        print(f"📁 LTAs traités: {len(folders_to_process_dum)}")
                                        print("="*70)
                            else:
                                print("\n⏸️  Phase 2 annulée - Vous pouvez relancer le script plus tard")
                    
                elif phase_choice == "2":
                    # PHASE 2: Création des Déclarations Dédouanement
                    print("\n" + "="*70)
                    print("🚀 PHASE 2: CRÉATION DÉCLARATIONS DÉDOUANEMENT")
                    print("="*70)
                    
                    # Find all LTA folders
                    lta_folders = find_lta_folders(".")
                    
                    if not lta_folders:
                        print("\n❌ Aucun dossier LTA trouvé")
                    else:
                        print(f"\n✓ {len(lta_folders)} dossiers LTA trouvés:")
                        for i, (_, folder_name) in enumerate(lta_folders, 1):
                            print(f"   {i}. {folder_name}")
                        
                        # Process selection based on d
                        folders_to_process = []
                        
                        if len(sys.argv) > 1:
                            # GUI mode: use provided selection
                            if selected_lta_indices == "all":
                                print("\n✓ Mode GUI: Traitement de TOUS les LTAs")
                                s = lta_folders
                            elif isinstance(selected_lta_indices, list):
                                print(f"\n✓ Mode GUI: Traitement de {len(selected_lta_indices)} LTA(s) sélectionné(s)")
                                folders_to_process = [lta_folders[i] for i in selected_lta_indices if 0 <= i < len(lta_folders)]
                                if folders_to_process:
                                    for folder_path, folder_name in folders_to_process:
                                        print(f"   • {folder_name}")
                            else:
                                # Fallback to all
                                folders_to_process = lta_folders
                        else:
                            # Interactive mode: ask user
                            print("\n📋 OPTIONS:")
                            print("   1. Traiter TOUS les LTAs")
                            print("   2. Sélectionner des LTAs spécifiques")
                            
                            choice = input("\nVotre choix (1 ou 2): ").strip()
                            
                            if choice == "1":
                                folders_to_process = lta_folders
                            elif choice == "2":
                                print("\n📝 Sélection des LTAs:")
                                print("   Entrez les numéros séparés par des virgules (ex: 1,3,5)")
                                selection = input("   Numéros: ").strip()
                                
                                try:
                                    indices = [int(x.strip()) - 1 for x in selection.split(',')]
                                    folders_to_process = [lta_folders[i] for i in indices if 0 <= i < len(lta_folders)]
                                    
                                    if not folders_to_process:
                                        print("❌ Sélection invalide")
                                    else:
                                        print(f"\n✓ {len(folders_to_process)} LTA(s) sélectionné(s)")
                                except:
                                    print("❌ Format invalide")
                            else:
                                print("❌ Choix invalide")
                        
                        # Process selected LTAs (DUM only)
                        if folders_to_process:
                            total_dums = 0
                            
                            for folder_path, folder_name in folders_to_process:
                                dums_processed = process_lta_folder_dum_only(driver, folder_path, folder_name)
                                total_dums += dums_processed
                            
                            # Summary
                            print("\n" + "="*70)
                            print("📊 RÉSUMÉ PHASE 2: DÉCLARATIONS DÉDOUANEMENT")
                            print("="*70)
                            print(f"✅ Total DUMs traités: {total_dums}")
                            print(f"📁 LTAs traités: {len(folders_to_process)}")
                            print("="*70)
            else:
                print("\n⚠️ CONNEXION: Échec de l'authentification")
            
            # Garder le script actif
            # input("\nAppuyez sur Entrée pour terminer...")  # Disabled for GUI automation
            
            
            # Nettoyer le profil temporaire après fermeture
            try:
                driver.quit()
                time.sleep(1)
                if os.path.exists(profile_path):
                    shutil.rmtree(profile_path)
                    print(f"🧹 Profil temporaire supprimé")
            except Exception as e:
                print(f"⚠️  Impossible de supprimer le profil: {e}")
            
            print("✓ Script terminé")
        else:
            print("\n❌ Échec de la connexion au navigateur")
    else:
        print("\n❌ Échec du lancement de Edge")
    
    print("="*70)

    # mailtrap
    # mailtraposos